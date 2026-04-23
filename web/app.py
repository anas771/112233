from __future__ import annotations

import csv
import io
import json
import os
import sqlite3
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path
from statistics import mean
from typing import Any

from flask import Flask, Response, flash, redirect, render_template, request, url_for
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    from fpdf import FPDF
    HAS_FPDF = True
except Exception:
    HAS_FPDF = False

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    HAS_ARABIC_TOOLS = True
except Exception:
    HAS_ARABIC_TOOLS = False

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    HAS_MATPLOTLIB = True
except Exception:
    HAS_MATPLOTLIB = False


BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))
ASSETS_DIR = BASE_DIR / "assets"
REPORT_FONT_PATH = ASSETS_DIR / "Amiri-Regular.ttf"

from core.database import (
    DB_PATH,
    ensure_schema as shared_ensure_schema,
    get_conn,
    get_setting,
    set_setting,
)
from core.health_engine import analyze_daily_health
from core.health_repository import load_daily_input, load_history_before, save_daily_analysis
from core.import_engine import (
    apply_mapping_edits,
    build_staging,
    commit_run,
    detect_profile,
    parse_files,
)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("POULTRY_WEB_SECRET", "poultry-web-local-secret")


def _legacy_web_ensure_schema() -> None:
    schema = """
    CREATE TABLE IF NOT EXISTS system_settings (
        key TEXT PRIMARY KEY,
        value TEXT
    );

    CREATE TABLE IF NOT EXISTS warehouses (
        id    INTEGER PRIMARY KEY AUTOINCREMENT,
        name  TEXT NOT NULL UNIQUE,
        notes TEXT DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS batches (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        warehouse_id   INTEGER NOT NULL REFERENCES warehouses(id),
        batch_num      TEXT    DEFAULT '',
        date_in        TEXT    NOT NULL,
        date_out       TEXT    NOT NULL,
        days           INTEGER DEFAULT 0,
        chicks         INTEGER NOT NULL,
        chick_price    REAL    DEFAULT 0,
        chick_val      REAL    DEFAULT 0,
        feed_qty       REAL    DEFAULT 0,
        feed_val       REAL    DEFAULT 0,
        feed_trans     REAL    DEFAULT 0,
        sawdust_qty    REAL    DEFAULT 0,
        sawdust_val    REAL    DEFAULT 0,
        water_val      REAL    DEFAULT 0,
        gas_qty        REAL    DEFAULT 0,
        gas_val        REAL    DEFAULT 0,
        drugs_val      REAL    DEFAULT 0,
        wh_expenses    REAL    DEFAULT 0,
        house_exp      REAL    DEFAULT 0,
        breeders_pay   REAL    DEFAULT 0,
        qat_pay        REAL    DEFAULT 0,
        rent_val       REAL    DEFAULT 0,
        light_val      REAL    DEFAULT 0,
        sup_wh_pay     REAL    DEFAULT 0,
        sup_co_pay     REAL    DEFAULT 0,
        sup_sale_pay   REAL    DEFAULT 0,
        admin_val      REAL    DEFAULT 0,
        vaccine_pay    REAL    DEFAULT 0,
        delivery_val   REAL    DEFAULT 0,
        mixing_val     REAL    DEFAULT 0,
        wash_val       REAL    DEFAULT 0,
        other_costs    REAL    DEFAULT 0,
        total_cost     REAL    DEFAULT 0,
        cust_qty       INTEGER DEFAULT 0,
        cust_val       REAL    DEFAULT 0,
        mkt_qty        INTEGER DEFAULT 0,
        mkt_val        REAL    DEFAULT 0,
        offal_val      REAL    DEFAULT 0,
        feed_sale      REAL    DEFAULT 0,
        feed_trans_r   REAL    DEFAULT 0,
        drug_return    REAL    DEFAULT 0,
        gas_return     REAL    DEFAULT 0,
        total_rev      REAL    DEFAULT 0,
        total_sold     INTEGER DEFAULT 0,
        total_dead     INTEGER DEFAULT 0,
        mort_rate      REAL    DEFAULT 0,
        avg_weight     REAL    DEFAULT 0,
        fcr            REAL    DEFAULT 0,
        avg_price      REAL    DEFAULT 0,
        net_result     REAL    DEFAULT 0,
        share_pct      REAL    DEFAULT 65,
        share_val      REAL    DEFAULT 0,
        notes          TEXT    DEFAULT '',
        created_at     TEXT,
        consumed_birds INTEGER DEFAULT 0,
        partner_name   TEXT    DEFAULT '',
        feed_sale_qty  REAL    DEFAULT 0,
        feed_trans_r_qty REAL  DEFAULT 0,
        feed_rem_qty   REAL    DEFAULT 0,
        feed_rem_val   REAL    DEFAULT 0,
        fiscal_year    INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS daily_records (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id    INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
        rec_date    TEXT    NOT NULL,
        day_num     INTEGER DEFAULT 0,
        dead_count  INTEGER DEFAULT 0,
        feed_kg     REAL    DEFAULT 0,
        water_ltr   REAL    DEFAULT 0,
        notes       TEXT    DEFAULT '',
        UNIQUE(batch_id, rec_date)
    );

    CREATE TABLE IF NOT EXISTS farm_sales (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id  INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
        sale_date TEXT    DEFAULT '',
        sale_type TEXT    DEFAULT 'آجل',
        customer  TEXT,
        qty       INTEGER DEFAULT 0,
        price     REAL    DEFAULT 0,
        total_val REAL    DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS market_sales (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id  INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
        sale_date TEXT    DEFAULT '',
        office    TEXT,
        qty_sent  INTEGER DEFAULT 0,
        deaths    INTEGER DEFAULT 0,
        qty_sold  INTEGER DEFAULT 0,
        net_val   REAL    DEFAULT 0,
        inv_num   TEXT
    );

    CREATE TABLE IF NOT EXISTS cost_types (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        code       TEXT    NOT NULL UNIQUE,
        name_ar    TEXT    NOT NULL,
        category   TEXT    DEFAULT 'أخرى',
        has_qty    INTEGER DEFAULT 0,
        unit       TEXT,
        sort_order INTEGER DEFAULT 99,
        is_active  INTEGER DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS batch_costs (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id     INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
        cost_type_id INTEGER NOT NULL REFERENCES cost_types(id),
        qty          REAL    DEFAULT 0,
        amount       REAL    DEFAULT 0,
        notes        TEXT    DEFAULT '',
        UNIQUE(batch_id, cost_type_id)
    );

    CREATE TABLE IF NOT EXISTS revenue_types (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        code       TEXT    NOT NULL UNIQUE,
        name_ar    TEXT    NOT NULL,
        category   TEXT    DEFAULT 'مبيعات',
        has_qty    INTEGER DEFAULT 0,
        unit       TEXT,
        sort_order INTEGER DEFAULT 99,
        is_active  INTEGER DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS batch_revenues (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id        INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
        revenue_type_id INTEGER NOT NULL REFERENCES revenue_types(id),
        qty             REAL    DEFAULT 0,
        amount          REAL    DEFAULT 0,
        notes           TEXT    DEFAULT '',
        UNIQUE(batch_id, revenue_type_id)
    );
    """

    with get_conn() as conn:
        conn.executescript(schema)
        conn.executemany(
            """
            INSERT OR IGNORE INTO cost_types(code, name_ar, category, has_qty, unit, sort_order)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                ("chick_val", "الكتاكيت", "مواد", 1, "حبة", 1),
                ("feed_val", "العلف", "مواد", 1, "طن", 2),
                ("drugs_val", "علاجات وأدوية", "صحة", 0, None, 3),
                ("gas_val", "الغاز", "مرافق", 1, "اسطوانة", 4),
                ("other_costs", "مصاريف أخرى", "أخرى", 0, None, 5),
            ],
        )
        conn.executemany(
            """
            INSERT OR IGNORE INTO revenue_types(code, name_ar, category, has_qty, unit, sort_order)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                ("offal_val", "مبيعات ذبيل", "مبيعات", 0, None, 1),
                ("feed_sale", "مبيعات علف", "مبيعات", 1, "كيس", 2),
                ("drug_return", "مرتجع علاجات", "مرتجعات", 0, None, 3),
            ],
        )
        conn.commit()


def to_float(value: str | None, default: float = 0.0) -> float:
    if value is None:
        return default
    value = value.strip()
    if not value:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def to_int(value: str | None, default: int = 0) -> int:
    if value is None:
        return default
    value = value.strip()
    if not value:
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _collect_treatment_inputs() -> list[dict[str, str]]:
    catalog_ids = request.form.getlist("treatment_catalog_id[]")
    custom_names = request.form.getlist("treatment_name[]")
    custom_active = request.form.getlist("treatment_active[]")
    custom_classes = request.form.getlist("treatment_class[]")
    dose_values = request.form.getlist("treatment_dose[]")
    notes_values = request.form.getlist("treatment_notes[]")

    payload: list[dict[str, str]] = []
    max_rows = max(
        len(catalog_ids),
        len(custom_names),
        len(custom_active),
        len(custom_classes),
        len(dose_values),
        len(notes_values),
    )
    for idx in range(max_rows):
        catalog_id = catalog_ids[idx] if idx < len(catalog_ids) else ""
        cid = to_int(catalog_id, 0)
        name = custom_names[idx].strip() if idx < len(custom_names) else ""
        active = custom_active[idx].strip() if idx < len(custom_active) else ""
        cls = custom_classes[idx].strip().lower() if idx < len(custom_classes) else ""
        if cid <= 0 and not name and not active:
            continue
        payload.append(
            {
                "catalog_id": str(cid),
                "product_name": name,
                "active_ingredient": active,
                "treatment_class": cls,
                "dose_text": dose_values[idx].strip() if idx < len(dose_values) else "",
                "notes": notes_values[idx].strip() if idx < len(notes_values) else "",
            }
        )
    return payload


def _infer_treatment_class(product_name: str, active_ingredient: str) -> str:
    value = f"{product_name} {active_ingredient}".lower()
    if not value.strip():
        return "supportive"
    if any(k in value for k in ("amprolium", "toltrazuril", "diclazuril", "sulfaquinoxaline", "sulfachloropyrazine", "coccid")):
        return "anticoccidial"
    if any(k in value for k in ("tylosin", "doxy", "oxy", "florfenicol", "enrofloxacin", "amoxicillin", "colistin", "linco", "trimethoprim", "sulfadiazine")):
        return "antibiotic"
    if any(k in value for k in ("bromhexine", "expectorant", "menthol", "mucolytic", "resp")):
        return "respiratory_support"
    if any(k in value for k in ("probiotic", "bacillus", "yeast")):
        return "probiotic"
    if any(k in value for k in ("vitamin", "electrolyte", "ad3e", "ascorbic", "vit c", "multivit")):
        return "vitamin_electrolyte"
    if any(k in value for k in ("immune", "glucan", "immun")):
        return "immune_support"
    if any(k in value for k in ("liver", "silymarin", "hepa")):
        return "liver_support"
    return "supportive"


def _save_daily_treatments(
    conn: sqlite3.Connection,
    *,
    batch_id: int,
    rec_date: str,
    record_id: int,
    treatment_inputs: list[dict[str, str]],
) -> None:
    conn.execute("DELETE FROM daily_treatments WHERE batch_id=? AND rec_date=?", (batch_id, rec_date))
    if not treatment_inputs:
        return

    for item in treatment_inputs:
        product_name = (item.get("product_name") or "").strip()
        active_ingredient = (item.get("active_ingredient") or "").strip().lower()
        treatment_class = (item.get("treatment_class") or "").strip().lower()
        catalog_id = to_int(item.get("catalog_id"), 0)

        catalog_row = conn.execute(
            """
            SELECT id, product_name, active_ingredient, treatment_class
            FROM treatment_catalog
            WHERE id=? AND is_active=1
            """,
            (catalog_id,),
        ).fetchone()
        saved_catalog_id: int | None = None
        if catalog_row:
            saved_catalog_id = int(catalog_row["id"])
            if not product_name:
                product_name = (catalog_row["product_name"] or "").strip()
            if not active_ingredient:
                active_ingredient = (catalog_row["active_ingredient"] or "").strip().lower()
            if not treatment_class:
                treatment_class = (catalog_row["treatment_class"] or "").strip().lower()

        if not product_name:
            product_name = active_ingredient or "custom_treatment"
        if not treatment_class:
            treatment_class = _infer_treatment_class(product_name, active_ingredient)

        conn.execute(
            """
            INSERT INTO daily_treatments(
                batch_id, record_id, rec_date, catalog_id,
                product_name, active_ingredient, treatment_class, dose_text, notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                record_id,
                rec_date,
                saved_catalog_id,
                product_name,
                active_ingredient,
                treatment_class,
                item.get("dose_text", ""),
                item.get("notes", ""),
            ),
        )


def _run_and_store_daily_analysis(conn: sqlite3.Connection, *, batch_id: int, rec_date: str) -> None:
    daily_input = load_daily_input(conn, batch_id, rec_date)
    if not daily_input:
        return
    history = load_history_before(conn, batch_id, rec_date, days=3)
    result = analyze_daily_health(daily_input, history)
    save_daily_analysis(conn, batch_id, rec_date, result)


def recalc_batch(batch_id: int) -> None:
    with get_conn() as conn:
        b = conn.execute("SELECT * FROM batches WHERE id=?", (batch_id,)).fetchone()
        if not b:
            return

        fixed_cost = sum(
            float(b[k] or 0)
            for k in [
                "chick_val",
                "feed_val",
                "feed_trans",
                "sawdust_val",
                "water_val",
                "gas_val",
                "drugs_val",
                "wh_expenses",
                "house_exp",
                "breeders_pay",
                "qat_pay",
                "rent_val",
                "light_val",
                "sup_wh_pay",
                "sup_co_pay",
                "sup_sale_pay",
                "admin_val",
                "vaccine_pay",
                "delivery_val",
                "mixing_val",
                "wash_val",
                "other_costs",
            ]
        )
        dyn_cost = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM batch_costs WHERE batch_id=?", (batch_id,)
        ).fetchone()[0]
        total_cost = max(fixed_cost, float(dyn_cost or 0))

        farm = conn.execute(
            "SELECT COALESCE(SUM(total_val),0), COALESCE(SUM(qty),0) FROM farm_sales WHERE batch_id=?",
            (batch_id,),
        ).fetchone()
        market = conn.execute(
            "SELECT COALESCE(SUM(net_val),0), COALESCE(SUM(qty_sold),0) FROM market_sales WHERE batch_id=?",
            (batch_id,),
        ).fetchone()
        extra_rev = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM batch_revenues WHERE batch_id=?", (batch_id,)
        ).fetchone()[0]
        total_rev = float(farm[0] or 0) + float(market[0] or 0) + float(extra_rev or 0)
        total_sold = int((farm[1] or 0) + (market[1] or 0))
        total_dead = int(
            conn.execute(
                "SELECT COALESCE(SUM(dead_count),0) FROM daily_records WHERE batch_id=?",
                (batch_id,),
            ).fetchone()[0]
            or 0
        )
        chicks = int(b["chicks"] or 0)
        mort_rate = (total_dead / chicks * 100) if chicks > 0 else 0
        net_result = total_rev - total_cost

        conn.execute(
            """
            UPDATE batches
            SET total_cost=?, total_rev=?, total_sold=?, total_dead=?, mort_rate=?, net_result=?
            WHERE id=?
            """,
            (total_cost, total_rev, total_sold, total_dead, mort_rate, net_result, batch_id),
        )
        conn.commit()


def get_filters() -> tuple[list[sqlite3.Row], list[int]]:
    with get_conn() as conn:
        warehouses = conn.execute("SELECT id, name FROM warehouses ORDER BY name").fetchall()
        years = [
            int(row["fy"])
            for row in conn.execute(
                """
                SELECT DISTINCT COALESCE(fiscal_year, CAST(substr(date_in, 1, 4) AS INTEGER)) AS fy
                FROM batches
                WHERE COALESCE(fiscal_year, CAST(substr(date_in, 1, 4) AS INTEGER)) > 0
                ORDER BY fy DESC
                """
            ).fetchall()
        ]
    return warehouses, years


def make_where_clause(warehouse_id: int | None, fiscal_year: int | None) -> tuple[str, list[int]]:
    where: list[str] = []
    params: list[int] = []
    if warehouse_id:
        where.append("b.warehouse_id=?")
        params.append(warehouse_id)
    if fiscal_year:
        where.append("COALESCE(b.fiscal_year, CAST(substr(b.date_in, 1, 4) AS INTEGER))=?")
        params.append(fiscal_year)
    return (" WHERE " + " AND ".join(where)) if where else "", params


def create_database_backup() -> Path:
    backups_dir = BASE_DIR / "backups"
    backups_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = backups_dir / f"poultry_web_backup_{timestamp}.db"
    with sqlite3.connect(DB_PATH) as src_conn, sqlite3.connect(destination) as dst_conn:
        src_conn.backup(dst_conn)
    return destination


def warehouse_summary_rows(where_clause: str, params: list[int]) -> list[dict[str, object]]:
    query = f"""
        SELECT
            w.name AS warehouse_name,
            COUNT(b.id) AS batches_count,
            COALESCE(SUM(b.chicks), 0) AS chicks_total,
            COALESCE(SUM(b.total_cost), 0) AS cost_total,
            COALESCE(SUM(b.total_rev), 0) AS revenue_total,
            COALESCE(SUM(b.net_result), 0) AS net_total,
            COALESCE(AVG(b.mort_rate), 0) AS avg_mortality
        FROM warehouses w
        LEFT JOIN batches b ON b.warehouse_id = w.id
        {where_clause}
        GROUP BY w.id, w.name
        ORDER BY w.name
    """
    with get_conn() as conn:
        rows = conn.execute(query, params).fetchall()
    return [dict(row) for row in rows]


def year_summary_rows() -> list[dict[str, object]]:
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                COALESCE(fiscal_year, CAST(substr(date_in, 1, 4) AS INTEGER)) AS fiscal_year,
                COUNT(*) AS batches_count,
                COALESCE(SUM(total_cost), 0) AS cost_total,
                COALESCE(SUM(total_rev), 0) AS revenue_total,
                COALESCE(SUM(net_result), 0) AS net_total
            FROM batches
            GROUP BY COALESCE(fiscal_year, CAST(substr(date_in, 1, 4) AS INTEGER))
            HAVING fiscal_year > 0
            ORDER BY fiscal_year DESC
            """
        ).fetchall()
    return [dict(row) for row in rows]


def make_csv_response(filename: str, headers: list[str], rows: list[list[object]]) -> Response:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    csv_text = "\ufeff" + buffer.getvalue()
    return Response(
        csv_text,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def make_excel_response(filename: str, sheet_name: str, headers: list[str], rows: list[list[object]]) -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (sheet_name or "Report")[:31]
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row))

    sheet.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header or ""))
        for row_idx in range(2, len(rows) + 2):
            value = sheet.cell(row=row_idx, column=col_idx).value
            max_length = max(max_length, len(str(value or "")))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 45)

    output = io.BytesIO()
    workbook.save(output)
    payload = output.getvalue()
    return Response(
        payload,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def make_pdf_response(filename: str, payload: bytes) -> Response:
    return Response(
        payload,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def ar_text(value: object) -> str:
    text = str(value or "")
    if HAS_ARABIC_TOOLS:
        try:
            return get_display(arabic_reshaper.reshape(text))
        except Exception:
            return text
    return text


def reports_payload(warehouse_id: int | None, fiscal_year: int | None) -> dict[str, object]:
    where_clause, params = make_where_clause(warehouse_id, fiscal_year)
    warehouses, years = get_filters()
    warehouse_rows = warehouse_summary_rows(where_clause, params)

    with get_conn() as conn:
        stats = conn.execute(
            f"""
            SELECT
                COUNT(*) AS batches_count,
                COALESCE(SUM(b.chicks), 0) AS chicks_total,
                COALESCE(SUM(b.total_dead), 0) AS total_dead,
                COALESCE(SUM(b.total_sold), 0) AS total_sold,
                COALESCE(SUM(b.total_cost), 0) AS cost_total,
                COALESCE(SUM(b.total_rev), 0) AS revenue_total,
                COALESCE(SUM(b.net_result), 0) AS net_total,
                COALESCE(AVG(b.fcr), 0) AS avg_fcr,
                COALESCE(AVG(b.mort_rate), 0) AS avg_mortality
            FROM batches b
            {where_clause}
            """,
            params,
        ).fetchone()

        annual_rows = [
            dict(row)
            for row in conn.execute(
                f"""
                SELECT
                    COALESCE(b.fiscal_year, CAST(substr(b.date_in, 1, 4) AS INTEGER)) AS fiscal_year,
                    COUNT(*) AS batches_count,
                    COALESCE(SUM(b.total_cost), 0) AS cost_total,
                    COALESCE(SUM(b.total_rev), 0) AS revenue_total,
                    COALESCE(SUM(b.net_result), 0) AS net_total
                FROM batches b
                {where_clause}
                GROUP BY COALESCE(b.fiscal_year, CAST(substr(b.date_in, 1, 4) AS INTEGER))
                HAVING fiscal_year > 0
                ORDER BY fiscal_year DESC
                """,
                params,
            ).fetchall()
        ]

        batches_rows = [
            dict(row)
            for row in conn.execute(
                f"""
                SELECT b.*, w.name AS warehouse_name
                FROM batches b
                JOIN warehouses w ON w.id = b.warehouse_id
                {where_clause}
                ORDER BY b.date_in DESC, b.id DESC
                LIMIT 25
                """,
                params,
            ).fetchall()
        ]

        sales_by_type_rows = [
            dict(row)
            for row in conn.execute(
                f"""
                SELECT
                    COALESCE(fs.sale_type, 'Unknown') AS sale_type,
                    COALESCE(SUM(fs.qty), 0) AS qty_total,
                    COALESCE(SUM(fs.total_val), 0) AS amount_total
                FROM farm_sales fs
                JOIN batches b ON b.id = fs.batch_id
                {where_clause}
                GROUP BY COALESCE(fs.sale_type, 'Unknown')
                ORDER BY amount_total DESC, qty_total DESC
                """,
                params,
            ).fetchall()
        ]

        market_summary_row = dict(
            conn.execute(
                f"""
                SELECT
                    COALESCE(SUM(ms.qty_sent), 0) AS qty_sent_total,
                    COALESCE(SUM(ms.deaths), 0) AS deaths_total,
                    COALESCE(SUM(ms.qty_sold), 0) AS qty_sold_total,
                    COALESCE(SUM(ms.net_val), 0) AS net_val_total
                FROM market_sales ms
                JOIN batches b ON b.id = ms.batch_id
                {where_clause}
                """,
                params,
            ).fetchone()
        )

        cost_types_rows = [
            dict(row)
            for row in conn.execute(
                f"""
                SELECT
                    ct.name_ar AS type_name,
                    ct.category AS category,
                    COALESCE(SUM(bc.qty), 0) AS qty_total,
                    COALESCE(SUM(bc.amount), 0) AS amount_total
                FROM batch_costs bc
                JOIN cost_types ct ON ct.id = bc.cost_type_id
                JOIN batches b ON b.id = bc.batch_id
                {where_clause}
                GROUP BY ct.id, ct.name_ar, ct.category
                HAVING COALESCE(SUM(bc.qty), 0) > 0 OR COALESCE(SUM(bc.amount), 0) > 0
                ORDER BY amount_total DESC, qty_total DESC
                """,
                params,
            ).fetchall()
        ]

        revenue_types_rows = [
            dict(row)
            for row in conn.execute(
                f"""
                SELECT
                    rt.name_ar AS type_name,
                    rt.category AS category,
                    COALESCE(SUM(br.qty), 0) AS qty_total,
                    COALESCE(SUM(br.amount), 0) AS amount_total
                FROM batch_revenues br
                JOIN revenue_types rt ON rt.id = br.revenue_type_id
                JOIN batches b ON b.id = br.batch_id
                {where_clause}
                GROUP BY rt.id, rt.name_ar, rt.category
                HAVING COALESCE(SUM(br.qty), 0) > 0 OR COALESCE(SUM(br.amount), 0) > 0
                ORDER BY amount_total DESC, qty_total DESC
                """,
                params,
            ).fetchall()
        ]

        daily_summary_row = dict(
            conn.execute(
                f"""
                SELECT
                    COUNT(dr.id) AS entries_count,
                    COALESCE(SUM(dr.dead_count), 0) AS dead_total,
                    COALESCE(SUM(dr.feed_kg), 0) AS feed_total,
                    COALESCE(SUM(dr.water_ltr), 0) AS water_total,
                    COALESCE(AVG(dr.dead_count), 0) AS avg_daily_deaths
                FROM daily_records dr
                JOIN batches b ON b.id = dr.batch_id
                {where_clause}
                """,
                params,
            ).fetchone()
        )

        best_batch_row = conn.execute(
            f"""
            SELECT b.id, b.batch_num, w.name AS warehouse_name, COALESCE(b.net_result, 0) AS net_result
            FROM batches b
            JOIN warehouses w ON w.id = b.warehouse_id
            {where_clause}
            ORDER BY net_result DESC, b.id DESC
            LIMIT 1
            """,
            params,
        ).fetchone()

        worst_batch_row = conn.execute(
            f"""
            SELECT b.id, b.batch_num, w.name AS warehouse_name, COALESCE(b.net_result, 0) AS net_result
            FROM batches b
            JOIN warehouses w ON w.id = b.warehouse_id
            {where_clause}
            ORDER BY net_result ASC, b.id DESC
            LIMIT 1
            """,
            params,
        ).fetchone()

    stats_data = dict(stats) if stats else {}
    cost_total = float(stats_data.get('cost_total') or 0)
    revenue_total = float(stats_data.get('revenue_total') or 0)
    net_total = float(stats_data.get('net_total') or 0)
    chicks_total = float(stats_data.get('chicks_total') or 0)
    dead_total = float(stats_data.get('total_dead') or 0)
    sold_total = float(stats_data.get('total_sold') or 0)

    report_summary = {
        'warehouses_count': len(warehouse_rows),
        'annual_years_count': len(annual_rows),
        'shown_batches_count': int(stats_data.get('batches_count') or 0),
        'cost_total': cost_total,
        'revenue_total': revenue_total,
        'net_total': net_total,
        'chicks_total': chicks_total,
        'total_dead': dead_total,
        'total_sold': sold_total,
    }

    fcr_values = [float(row.get('fcr') or 0) for row in batches_rows if float(row.get('fcr') or 0) > 0]
    mort_values = [float(row.get('mort_rate') or 0) for row in batches_rows]
    margin_pct = (net_total / revenue_total * 100) if revenue_total > 0 else 0.0
    death_rate_pct = (dead_total / chicks_total * 100) if chicks_total > 0 else 0.0
    sold_rate_pct = (sold_total / chicks_total * 100) if chicks_total > 0 else 0.0

    top_warehouse = max(warehouse_rows, key=lambda x: float(x.get('net_total') or 0), default=None)
    best_batch = dict(best_batch_row) if best_batch_row else None
    worst_batch = dict(worst_batch_row) if worst_batch_row else None

    top_cost_type = cost_types_rows[0] if cost_types_rows else None
    top_revenue_type = revenue_types_rows[0] if revenue_types_rows else None
    farm_sales_total = sum(float(row.get('amount_total') or 0) for row in sales_by_type_rows)
    extra_revenues_total = sum(float(row.get('amount_total') or 0) for row in revenue_types_rows)
    market_sales_total = float(market_summary_row.get('net_val_total') or 0)
    daily_entries = int(daily_summary_row.get('entries_count') or 0)

    dashboard_insights = {
        'margin_pct': margin_pct,
        'avg_fcr': mean(fcr_values) if fcr_values else 0,
        'avg_mortality': mean(mort_values) if mort_values else 0,
        'death_rate_pct': death_rate_pct,
        'sold_rate_pct': sold_rate_pct,
        'top_warehouse_name': top_warehouse['warehouse_name'] if top_warehouse else '?',
        'top_warehouse_net': float(top_warehouse['net_total']) if top_warehouse else 0,
        'best_batch_label': (best_batch.get('batch_num') or best_batch.get('id')) if best_batch else '?',
        'best_batch_net': float(best_batch['net_result']) if best_batch else 0,
        'worst_batch_label': (worst_batch.get('batch_num') or worst_batch.get('id')) if worst_batch else '?',
        'worst_batch_net': float(worst_batch['net_result']) if worst_batch else 0,
        'farm_sales_total': farm_sales_total,
        'market_sales_total': market_sales_total,
        'extra_revenues_total': extra_revenues_total,
        'top_cost_type_name': top_cost_type['type_name'] if top_cost_type else '?',
        'top_cost_type_amount': float(top_cost_type['amount_total']) if top_cost_type else 0,
        'top_revenue_type_name': top_revenue_type['type_name'] if top_revenue_type else '?',
        'top_revenue_type_amount': float(top_revenue_type['amount_total']) if top_revenue_type else 0,
        'daily_entries_count': daily_entries,
        'daily_dead_total': float(daily_summary_row.get('dead_total') or 0),
        'daily_feed_total': float(daily_summary_row.get('feed_total') or 0),
        'daily_water_total': float(daily_summary_row.get('water_total') or 0),
        'avg_daily_deaths': float(daily_summary_row.get('avg_daily_deaths') or 0),
    }

    return {
        'where_clause': where_clause,
        'params': params,
        'warehouses': warehouses,
        'years': years,
        'warehouse_rows': warehouse_rows,
        'annual_rows': annual_rows,
        'batches_rows': batches_rows,
        'sales_by_type_rows': sales_by_type_rows,
        'market_summary_row': market_summary_row,
        'cost_types_rows': cost_types_rows,
        'revenue_types_rows': revenue_types_rows,
        'daily_summary_row': daily_summary_row,
        'report_summary': report_summary,
        'dashboard_insights': dashboard_insights,
    }


@app.template_filter("num")
def num_filter(value: float | int | None, digits: int = 0) -> str:
    try:
        if digits == 0:
            return f"{int(float(value or 0)):,}"
        return f"{float(value or 0):,.{digits}f}"
    except Exception:
        return "0"


def _safe_int_setting(value: str, default: int, low: int, high: int) -> int:
    try:
        parsed = int(float((value or "").strip()))
    except Exception:
        return default
    return max(low, min(high, parsed))


def _safe_float_setting(value: str, default: float, low: float, high: float) -> float:
    try:
        parsed = float((value or "").strip())
    except Exception:
        return default
    return max(low, min(high, parsed))


def _sanitize_ui_setting(key: str, value: str, default: str) -> str:
    raw = (value or "").strip()
    if key == "ui_font_scale":
        return str(_safe_int_setting(raw, _safe_int_setting(default, 110, 90, 150), 90, 150))
    if key == "ui_line_height":
        return f"{_safe_float_setting(raw, _safe_float_setting(default, 1.65, 1.3, 2.0), 1.3, 2.0):.2f}"
    if key == "ui_density":
        clean = raw.lower()
        return clean if clean in {"compact", "balanced", "comfortable"} else default
    if key == "ui_separator_strength":
        clean = raw.lower()
        return clean if clean in {"light", "normal", "strong"} else default
    if key == "ui_font_family":
        clean = raw.lower()
        return clean if clean in {"system", "amiri"} else default
    return raw or default


def _resolve_ui_settings() -> dict[str, str]:
    font_scale = _sanitize_ui_setting("ui_font_scale", get_setting("ui_font_scale", "110"), "110")
    line_height = _sanitize_ui_setting("ui_line_height", get_setting("ui_line_height", "1.65"), "1.65")
    density = _sanitize_ui_setting("ui_density", get_setting("ui_density", "balanced"), "balanced")
    separator = _sanitize_ui_setting("ui_separator_strength", get_setting("ui_separator_strength", "strong"), "strong")
    font_family = _sanitize_ui_setting("ui_font_family", get_setting("ui_font_family", "system"), "system")

    scale_value = _safe_int_setting(font_scale, 110, 90, 150)
    line_value = _safe_float_setting(line_height, 1.65, 1.3, 2.0)
    font_px = round(16.0 * (scale_value / 100.0), 2)

    return {
        "font_scale": str(scale_value),
        "line_height": f"{line_value:.2f}",
        "density": density,
        "separator": separator,
        "font_family": font_family,
        "body_classes": f"density-{density} sep-{separator} font-{font_family}",
        "inline_style": f"--ui-font-size:{font_px}px;--ui-line-height:{line_value:.2f};",
    }


@app.context_processor
def inject_ui_settings() -> dict[str, dict[str, str]]:
    return {"ui": _resolve_ui_settings()}


@app.route("/")
def dashboard():
    warehouse_id = to_int(request.args.get("warehouse_id"), 0) or None
    fiscal_year = to_int(request.args.get("fiscal_year"), 0) or None
    where_clause, params = make_where_clause(warehouse_id, fiscal_year)
    warehouses, years = get_filters()
    with get_conn() as conn:
        stats = conn.execute(
            f"""
            SELECT
              (SELECT COUNT(*) FROM warehouses) AS warehouses_count,
              COUNT(*) AS batches_count,
              COALESCE(SUM(b.total_cost),0) AS total_cost,
              COALESCE(SUM(b.total_rev),0) AS total_rev,
              COALESCE(SUM(b.net_result),0) AS total_net,
              COALESCE(SUM(b.chicks),0) AS chicks_total,
              COALESCE(AVG(b.mort_rate),0) AS avg_mortality
            FROM batches b
            {where_clause}
            """
            ,
            params,
        ).fetchone()
        latest_batches = conn.execute(
            f"""
            SELECT b.*, w.name AS warehouse_name
            FROM batches b
            JOIN warehouses w ON w.id=b.warehouse_id
            {where_clause}
            ORDER BY b.date_in DESC, b.id DESC
            LIMIT 12
            """,
            params,
        ).fetchall()
    return render_template(
        "dashboard.html",
        stats=stats,
        latest_batches=latest_batches,
        filters={"warehouse_id": warehouse_id or "", "fiscal_year": fiscal_year or ""},
        warehouses=warehouses,
        years=years,
    )


@app.route("/warehouses", methods=["GET", "POST"])
def warehouses():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        notes = request.form.get("notes", "").strip()
        if not name:
            flash("اسم العنبر مطلوب.", "error")
            return redirect(url_for("warehouses"))
        with get_conn() as conn:
            conn.execute("INSERT INTO warehouses(name, notes) VALUES (?, ?)", (name, notes))
            conn.commit()
        flash("تم إضافة العنبر بنجاح.", "success")
        return redirect(url_for("warehouses"))

    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT w.*, COUNT(b.id) AS batches_count
            FROM warehouses w
            LEFT JOIN batches b ON b.warehouse_id = w.id
            GROUP BY w.id
            ORDER BY w.id DESC
            """
        ).fetchall()
    return render_template("warehouses.html", warehouses=rows)


@app.get("/warehouses/export.csv")
def export_warehouses_csv():
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT w.id, w.name, w.notes, COUNT(b.id) AS batches_count
            FROM warehouses w
            LEFT JOIN batches b ON b.warehouse_id = w.id
            GROUP BY w.id
            ORDER BY w.id DESC
            """
        ).fetchall()
    payload = [[r["id"], r["name"], r["batches_count"], r["notes"] or ""] for r in rows]
    return make_csv_response(
        "warehouses.csv",
        ["المعرف", "اسم العنبر", "عدد الدفعات", "ملاحظات"],
        payload,
    )


@app.get("/warehouses/export.xlsx")
def export_warehouses_excel():
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT w.id, w.name, w.notes, COUNT(b.id) AS batches_count
            FROM warehouses w
            LEFT JOIN batches b ON b.warehouse_id = w.id
            GROUP BY w.id
            ORDER BY w.id DESC
            """
        ).fetchall()
    payload = [[r["id"], r["name"], r["batches_count"], r["notes"] or ""] for r in rows]
    return make_excel_response(
        "warehouses.xlsx",
        "العنابر",
        ["المعرف", "اسم العنبر", "عدد الدفعات", "ملاحظات"],
        payload,
    )


@app.post("/warehouses/<int:warehouse_id>/delete")
def delete_warehouse(warehouse_id: int):
    with get_conn() as conn:
        linked = conn.execute("SELECT COUNT(*) FROM batches WHERE warehouse_id=?", (warehouse_id,)).fetchone()[0]
        if linked:
            flash("لا يمكن حذف العنبر لأنه مرتبط بدفعات.", "error")
            return redirect(url_for("warehouses"))
        conn.execute("DELETE FROM warehouses WHERE id=?", (warehouse_id,))
        conn.commit()
    flash("تم حذف العنبر.", "success")
    return redirect(url_for("warehouses"))


@app.route("/batches")
def batches():
    warehouse_id = to_int(request.args.get("warehouse_id"), 0) or None
    fiscal_year = to_int(request.args.get("fiscal_year"), 0) or None
    where_clause, params = make_where_clause(warehouse_id, fiscal_year)
    warehouses, years = get_filters()
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT b.*, w.name AS warehouse_name
            FROM batches b
            JOIN warehouses w ON w.id=b.warehouse_id
            {where_clause}
            ORDER BY b.date_in DESC, b.id DESC
            """,
            params,
        ).fetchall()
    return render_template(
        "batches.html",
        batches=rows,
        warehouses=warehouses,
        years=years,
        filters={"warehouse_id": warehouse_id or "", "fiscal_year": fiscal_year or ""},
    )


@app.route("/batches/new", methods=["GET", "POST"])
def new_batch():
    with get_conn() as conn:
        warehouse_rows = conn.execute("SELECT * FROM warehouses ORDER BY name").fetchall()

    if request.method == "POST":
        warehouse_id = to_int(request.form.get("warehouse_id"))
        batch_num = request.form.get("batch_num", "").strip()
        date_in = request.form.get("date_in", "").strip()
        date_out = request.form.get("date_out", "").strip()
        chicks = to_int(request.form.get("chicks"), 0)
        notes = request.form.get("notes", "").strip()
        days = to_int(request.form.get("days"), 0)
        chick_price = to_float(request.form.get("chick_price"), 0.0)

        if not warehouse_id or not date_in or not date_out or chicks <= 0:
            flash("أكمل الحقول الأساسية: العنبر + تاريخ الدخول + تاريخ الخروج + عدد الكتاكيت.", "error")
            return render_template(
                "batch_form.html",
                warehouses=warehouse_rows,
                batch=None,
                form=request.form,
                action=url_for("new_batch"),
                title="إضافة دفعة",
            )

        chick_val = chick_price * chicks
        with get_conn() as conn:
            batch_id = conn.execute(
                """
                INSERT INTO batches(
                    warehouse_id,batch_num,date_in,date_out,days,chicks,chick_price,chick_val,notes,created_at,fiscal_year
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    warehouse_id,
                    batch_num,
                    date_in,
                    date_out,
                    days,
                    chicks,
                    chick_price,
                    chick_val,
                    notes,
                    date.today().isoformat(),
                    int(date_in[:4]) if len(date_in) >= 4 else 0,
                ),
            ).lastrowid
            conn.commit()
        recalc_batch(batch_id)
        flash("تم إنشاء الدفعة.", "success")
        return redirect(url_for("batch_detail", batch_id=batch_id))

    return render_template(
        "batch_form.html",
        warehouses=warehouse_rows,
        batch=None,
        form={},
        action=url_for("new_batch"),
        title="إضافة دفعة",
    )


@app.route("/batches/<int:batch_id>/edit", methods=["GET", "POST"])
def edit_batch(batch_id: int):
    with get_conn() as conn:
        batch = conn.execute("SELECT * FROM batches WHERE id=?", (batch_id,)).fetchone()
        warehouses_rows = conn.execute("SELECT * FROM warehouses ORDER BY name").fetchall()
    if not batch:
        flash("الدفعة غير موجودة.", "error")
        return redirect(url_for("batches"))

    if request.method == "POST":
        with get_conn() as conn:
            conn.execute(
                """
                UPDATE batches
                SET warehouse_id=?, batch_num=?, date_in=?, date_out=?, days=?, chicks=?, chick_price=?, notes=?
                WHERE id=?
                """,
                (
                    to_int(request.form.get("warehouse_id")),
                    request.form.get("batch_num", "").strip(),
                    request.form.get("date_in", "").strip(),
                    request.form.get("date_out", "").strip(),
                    to_int(request.form.get("days")),
                    to_int(request.form.get("chicks")),
                    to_float(request.form.get("chick_price")),
                    request.form.get("notes", "").strip(),
                    batch_id,
                ),
            )
            conn.commit()
        recalc_batch(batch_id)
        flash("تم تحديث الدفعة.", "success")
        return redirect(url_for("batch_detail", batch_id=batch_id))

    return render_template(
        "batch_form.html",
        warehouses=warehouses_rows,
        batch=batch,
        form=dict(batch),
        action=url_for("edit_batch", batch_id=batch_id),
        title=f"تعديل الدفعة #{batch_id}",
    )


@app.post("/batches/<int:batch_id>/delete")
def delete_batch(batch_id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM batches WHERE id=?", (batch_id,))
        conn.commit()
    flash("تم حذف الدفعة.", "success")
    return redirect(url_for("batches"))


@app.route("/batches/<int:batch_id>")
def batch_detail(batch_id: int):
    with get_conn() as conn:
        batch = conn.execute(
            """
            SELECT b.*, w.name AS warehouse_name
            FROM batches b JOIN warehouses w ON w.id=b.warehouse_id
            WHERE b.id=?
            """,
            (batch_id,),
        ).fetchone()
        if not batch:
            flash("الدفعة غير موجودة.", "error")
            return redirect(url_for("batches"))
        daily_count = conn.execute("SELECT COUNT(*) FROM daily_records WHERE batch_id=?", (batch_id,)).fetchone()[0]
        farm_count = conn.execute("SELECT COUNT(*) FROM farm_sales WHERE batch_id=?", (batch_id,)).fetchone()[0]
        market_count = conn.execute("SELECT COUNT(*) FROM market_sales WHERE batch_id=?", (batch_id,)).fetchone()[0]
        records = conn.execute(
            """
            SELECT rec_date, day_num, dead_count, feed_kg, water_ltr
            FROM daily_records
            WHERE batch_id=?
            ORDER BY rec_date DESC
            LIMIT 5
            """,
            (batch_id,),
        ).fetchall()
    return render_template(
        "batch_detail.html",
        batch=batch,
        daily_count=daily_count,
        farm_count=farm_count,
        market_count=market_count,
        recent_daily=records,
    )


@app.route("/batches/<int:batch_id>/daily", methods=["GET", "POST"])
def daily_records(batch_id: int):
    with get_conn() as conn:
        batch = conn.execute("SELECT id, batch_num FROM batches WHERE id=?", (batch_id,)).fetchone()
    if not batch:
        flash("الدفعة غير موجودة.", "error")
        return redirect(url_for("batches"))

    if request.method == "POST":
        rec_date = request.form.get("rec_date", "").strip()
        treatment_inputs = _collect_treatment_inputs()
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO daily_records(
                    batch_id, rec_date, day_num, dead_count, culls_count, feed_kg, water_ltr,
                    temp_min_c, temp_max_c, humidity_min_pct, humidity_max_pct, clinical_signs_text, notes
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(batch_id, rec_date)
                DO UPDATE SET
                    day_num=excluded.day_num,
                    dead_count=excluded.dead_count,
                    culls_count=excluded.culls_count,
                    feed_kg=excluded.feed_kg,
                    water_ltr=excluded.water_ltr,
                    temp_min_c=excluded.temp_min_c,
                    temp_max_c=excluded.temp_max_c,
                    humidity_min_pct=excluded.humidity_min_pct,
                    humidity_max_pct=excluded.humidity_max_pct,
                    clinical_signs_text=excluded.clinical_signs_text,
                    notes=excluded.notes
                """,
                (
                    batch_id,
                    rec_date,
                    to_int(request.form.get("day_num")),
                    to_int(request.form.get("dead_count")),
                    to_int(request.form.get("culls_count")),
                    to_float(request.form.get("feed_kg")),
                    to_float(request.form.get("water_ltr")),
                    to_float(request.form.get("temp_min_c")),
                    to_float(request.form.get("temp_max_c")),
                    to_float(request.form.get("humidity_min_pct")),
                    to_float(request.form.get("humidity_max_pct")),
                    request.form.get("clinical_signs_text", "").strip(),
                    request.form.get("notes", "").strip(),
                ),
            )
            record_row = conn.execute(
                "SELECT id FROM daily_records WHERE batch_id=? AND rec_date=?",
                (batch_id, rec_date),
            ).fetchone()
            if record_row:
                _save_daily_treatments(
                    conn,
                    batch_id=batch_id,
                    rec_date=rec_date,
                    record_id=int(record_row["id"]),
                    treatment_inputs=treatment_inputs,
                )
                _run_and_store_daily_analysis(conn, batch_id=batch_id, rec_date=rec_date)
            conn.commit()
        recalc_batch(batch_id)
        flash("تم حفظ السجل اليومي.", "success")
        return redirect(url_for("daily_records", batch_id=batch_id))

    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM daily_records WHERE batch_id=? ORDER BY rec_date DESC", (batch_id,)).fetchall()
        catalog = conn.execute(
            """
            SELECT id, product_name, active_ingredient, treatment_class
            FROM treatment_catalog
            WHERE is_active=1
            ORDER BY sort_order, product_name
            """
        ).fetchall()
        treatment_rows = conn.execute(
            """
            SELECT rec_date, product_name, treatment_class, dose_text
            FROM daily_treatments
            WHERE batch_id=?
            ORDER BY rec_date DESC, id ASC
            """,
            (batch_id,),
        ).fetchall()
    treatments_map: dict[str, list[dict[str, str]]] = {}
    for row in treatment_rows:
        rec_date = row["rec_date"]
        if rec_date not in treatments_map:
            treatments_map[rec_date] = []
        treatments_map[rec_date].append(
            {
                "product_name": row["product_name"] or "",
                "treatment_class": row["treatment_class"] or "",
                "dose_text": row["dose_text"] or "",
            }
        )
    return render_template(
        "daily_records.html",
        batch=batch,
        rows=rows,
        treatment_catalog=catalog,
        treatments_map=treatments_map,
    )


@app.get("/batches/<int:batch_id>/daily/export.csv")
def export_daily_records_csv(batch_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT rec_date, day_num, dead_count, culls_count, feed_kg, water_ltr,
                   temp_min_c, temp_max_c, humidity_min_pct, humidity_max_pct,
                   clinical_signs_text, analysis_status, risk_score, notes
            FROM daily_records
            WHERE batch_id=?
            ORDER BY rec_date
            """,
            (batch_id,),
        ).fetchall()
    payload = [
        [
            r["rec_date"],
            r["day_num"],
            r["dead_count"],
            r["culls_count"],
            r["feed_kg"],
            r["water_ltr"],
            r["temp_min_c"],
            r["temp_max_c"],
            r["humidity_min_pct"],
            r["humidity_max_pct"],
            r["clinical_signs_text"] or "",
            r["analysis_status"] or "",
            r["risk_score"],
            r["notes"] or "",
        ]
        for r in rows
    ]
    return make_csv_response(
        f"batch_{batch_id}_daily_records.csv",
        [
            "التاريخ",
            "رقم اليوم",
            "النفوق",
            "الاستبعاد",
            "العلف (كجم)",
            "الماء (لتر)",
            "أدنى حرارة",
            "أعلى حرارة",
            "أدنى رطوبة",
            "أعلى رطوبة",
            "العلامات الإكلينيكية",
            "حالة التحليل",
            "درجة المخاطر",
            "ملاحظات",
        ],
        payload,
    )


@app.get("/batches/<int:batch_id>/daily/export.xlsx")
def export_daily_records_excel(batch_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT rec_date, day_num, dead_count, culls_count, feed_kg, water_ltr,
                   temp_min_c, temp_max_c, humidity_min_pct, humidity_max_pct,
                   clinical_signs_text, analysis_status, risk_score, notes
            FROM daily_records
            WHERE batch_id=?
            ORDER BY rec_date
            """,
            (batch_id,),
        ).fetchall()
    payload = [
        [
            r["rec_date"],
            r["day_num"],
            r["dead_count"],
            r["culls_count"],
            r["feed_kg"],
            r["water_ltr"],
            r["temp_min_c"],
            r["temp_max_c"],
            r["humidity_min_pct"],
            r["humidity_max_pct"],
            r["clinical_signs_text"] or "",
            r["analysis_status"] or "",
            r["risk_score"],
            r["notes"] or "",
        ]
        for r in rows
    ]
    return make_excel_response(
        f"batch_{batch_id}_daily_records.xlsx",
        "السجل_اليومي",
        [
            "التاريخ",
            "رقم اليوم",
            "النفوق",
            "الاستبعاد",
            "العلف (كجم)",
            "الماء (لتر)",
            "أدنى حرارة",
            "أعلى حرارة",
            "أدنى رطوبة",
            "أعلى رطوبة",
            "العلامات الإكلينيكية",
            "حالة التحليل",
            "درجة المخاطر",
            "ملاحظات",
        ],
        payload,
    )


@app.post("/daily/<int:record_id>/delete")
def delete_daily_record(record_id: int):
    with get_conn() as conn:
        row = conn.execute("SELECT batch_id, rec_date FROM daily_records WHERE id=?", (record_id,)).fetchone()
        if not row:
            flash("السجل غير موجود.", "error")
            return redirect(url_for("batches"))
        batch_id = row["batch_id"]
        rec_date = row["rec_date"]
        conn.execute("DELETE FROM daily_treatments WHERE batch_id=? AND rec_date=?", (batch_id, rec_date))
        conn.execute("DELETE FROM daily_records WHERE id=?", (record_id,))
        conn.commit()
    recalc_batch(batch_id)
    flash("تم حذف السجل اليومي.", "success")
    return redirect(url_for("daily_records", batch_id=batch_id))


@app.route("/batches/<int:batch_id>/farm-sales", methods=["GET", "POST"])
def farm_sales(batch_id: int):
    if request.method == "POST":
        qty = to_int(request.form.get("qty"))
        price = to_float(request.form.get("price"))
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO farm_sales(batch_id, sale_date, sale_type, customer, qty, price, total_val)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    request.form.get("sale_date", "").strip(),
                    request.form.get("sale_type", "").strip() or "آجل",
                    request.form.get("customer", "").strip(),
                    qty,
                    price,
                    qty * price,
                ),
            )
            conn.commit()
        recalc_batch(batch_id)
        flash("تم حفظ مبيعة العنبر.", "success")
        return redirect(url_for("farm_sales", batch_id=batch_id))

    with get_conn() as conn:
        batch = conn.execute("SELECT id, batch_num FROM batches WHERE id=?", (batch_id,)).fetchone()
        rows = conn.execute("SELECT * FROM farm_sales WHERE batch_id=? ORDER BY id DESC", (batch_id,)).fetchall()
    return render_template("farm_sales.html", batch=batch, rows=rows)


@app.post("/farm-sales/<int:sale_id>/delete")
def delete_farm_sale(sale_id: int):
    with get_conn() as conn:
        row = conn.execute("SELECT batch_id FROM farm_sales WHERE id=?", (sale_id,)).fetchone()
        if not row:
            flash("المبيعة غير موجودة.", "error")
            return redirect(url_for("batches"))
        batch_id = row["batch_id"]
        conn.execute("DELETE FROM farm_sales WHERE id=?", (sale_id,))
        conn.commit()
    recalc_batch(batch_id)
    flash("تم حذف مبيعة العنبر.", "success")
    return redirect(url_for("farm_sales", batch_id=batch_id))


@app.route("/batches/<int:batch_id>/market-sales", methods=["GET", "POST"])
def market_sales(batch_id: int):
    if request.method == "POST":
        qty_sent = to_int(request.form.get("qty_sent"))
        deaths = to_int(request.form.get("deaths"))
        qty_sold = to_int(request.form.get("qty_sold"), max(0, qty_sent - deaths))
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO market_sales(batch_id, sale_date, office, qty_sent, deaths, qty_sold, net_val, inv_num)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    request.form.get("sale_date", "").strip(),
                    request.form.get("office", "").strip(),
                    qty_sent,
                    deaths,
                    qty_sold,
                    to_float(request.form.get("net_val")),
                    request.form.get("inv_num", "").strip(),
                ),
            )
            conn.commit()
        recalc_batch(batch_id)
        flash("تم حفظ مبيعة السوق.", "success")
        return redirect(url_for("market_sales", batch_id=batch_id))

    with get_conn() as conn:
        batch = conn.execute("SELECT id, batch_num FROM batches WHERE id=?", (batch_id,)).fetchone()
        rows = conn.execute("SELECT * FROM market_sales WHERE batch_id=? ORDER BY id DESC", (batch_id,)).fetchall()
    return render_template("market_sales.html", batch=batch, rows=rows)


@app.get("/batches/<int:batch_id>/market-sales/export.csv")
def export_market_sales_csv(batch_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT sale_date, office, qty_sent, deaths, qty_sold, net_val, inv_num
            FROM market_sales
            WHERE batch_id=?
            ORDER BY id DESC
            """,
            (batch_id,),
        ).fetchall()
    payload = [
        [r["sale_date"], r["office"] or "", r["qty_sent"], r["deaths"], r["qty_sold"], r["net_val"], r["inv_num"] or ""]
        for r in rows
    ]
    return make_csv_response(
        f"batch_{batch_id}_market_sales.csv",
        ["تاريخ البيع", "المكتب", "الكمية المرسلة", "وفيات السوق", "المباع", "صافي الفاتورة", "رقم الفاتورة"],
        payload,
    )


@app.get("/batches/<int:batch_id>/market-sales/export.xlsx")
def export_market_sales_excel(batch_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT sale_date, office, qty_sent, deaths, qty_sold, net_val, inv_num
            FROM market_sales
            WHERE batch_id=?
            ORDER BY id DESC
            """,
            (batch_id,),
        ).fetchall()
    payload = [
        [r["sale_date"], r["office"] or "", r["qty_sent"], r["deaths"], r["qty_sold"], r["net_val"], r["inv_num"] or ""]
        for r in rows
    ]
    return make_excel_response(
        f"batch_{batch_id}_market_sales.xlsx",
        "مبيعات_السوق",
        ["تاريخ البيع", "المكتب", "الكمية المرسلة", "وفيات السوق", "المباع", "صافي الفاتورة", "رقم الفاتورة"],
        payload,
    )


@app.post("/market-sales/<int:sale_id>/delete")
def delete_market_sale(sale_id: int):
    with get_conn() as conn:
        row = conn.execute("SELECT batch_id FROM market_sales WHERE id=?", (sale_id,)).fetchone()
        if not row:
            flash("المبيعة غير موجودة.", "error")
            return redirect(url_for("batches"))
        batch_id = row["batch_id"]
        conn.execute("DELETE FROM market_sales WHERE id=?", (sale_id,))
        conn.commit()
    recalc_batch(batch_id)
    flash("تم حذف مبيعة السوق.", "success")
    return redirect(url_for("market_sales", batch_id=batch_id))


@app.route("/batches/<int:batch_id>/costs", methods=["GET", "POST"])
def batch_costs(batch_id: int):
    if request.method == "POST":
        with get_conn() as conn:
            types = conn.execute("SELECT * FROM cost_types WHERE is_active=1 ORDER BY sort_order").fetchall()
            for t in types:
                qty = to_float(request.form.get(f"qty_{t['id']}"), 0.0)
                amount = to_float(request.form.get(f"amount_{t['id']}"), 0.0)
                if qty == 0 and amount == 0:
                    conn.execute("DELETE FROM batch_costs WHERE batch_id=? AND cost_type_id=?", (batch_id, t["id"]))
                else:
                    conn.execute(
                        """
                        INSERT INTO batch_costs(batch_id,cost_type_id,qty,amount)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(batch_id,cost_type_id)
                        DO UPDATE SET qty=excluded.qty, amount=excluded.amount
                        """,
                        (batch_id, t["id"], qty, amount),
                    )
            conn.commit()
        recalc_batch(batch_id)
        flash("تم تحديث التكاليف.", "success")
        return redirect(url_for("batch_costs", batch_id=batch_id))

    with get_conn() as conn:
        batch = conn.execute("SELECT id, batch_num FROM batches WHERE id=?", (batch_id,)).fetchone()
        types = conn.execute("SELECT * FROM cost_types WHERE is_active=1 ORDER BY sort_order").fetchall()
        values = {
            row["cost_type_id"]: row
            for row in conn.execute("SELECT * FROM batch_costs WHERE batch_id=?", (batch_id,)).fetchall()
        }
    return render_template("batch_costs.html", batch=batch, types=types, values=values)


@app.get("/batches/<int:batch_id>/costs/export.csv")
def export_batch_costs_csv(batch_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT ct.code, ct.name_ar, ct.category, COALESCE(bc.qty, 0) AS qty, COALESCE(bc.amount, 0) AS amount
            FROM cost_types ct
            LEFT JOIN batch_costs bc ON bc.cost_type_id = ct.id AND bc.batch_id = ?
            WHERE ct.is_active = 1
            ORDER BY ct.sort_order
            """,
            (batch_id,),
        ).fetchall()
    payload = [[r["code"], r["name_ar"], r["category"], r["qty"], r["amount"]] for r in rows]
    return make_csv_response(
        f"batch_{batch_id}_costs.csv",
        ["الكود", "اسم البند", "الفئة", "الكمية", "القيمة"],
        payload,
    )


@app.get("/batches/<int:batch_id>/costs/export.xlsx")
def export_batch_costs_excel(batch_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT ct.code, ct.name_ar, ct.category, COALESCE(bc.qty, 0) AS qty, COALESCE(bc.amount, 0) AS amount
            FROM cost_types ct
            LEFT JOIN batch_costs bc ON bc.cost_type_id = ct.id AND bc.batch_id = ?
            WHERE ct.is_active = 1
            ORDER BY ct.sort_order
            """,
            (batch_id,),
        ).fetchall()
    payload = [[r["code"], r["name_ar"], r["category"], r["qty"], r["amount"]] for r in rows]
    return make_excel_response(
        f"batch_{batch_id}_costs.xlsx",
        "التكاليف",
        ["الكود", "اسم البند", "الفئة", "الكمية", "القيمة"],
        payload,
    )


@app.route("/batches/<int:batch_id>/revenues", methods=["GET", "POST"])
def batch_revenues(batch_id: int):
    if request.method == "POST":
        with get_conn() as conn:
            types = conn.execute("SELECT * FROM revenue_types WHERE is_active=1 ORDER BY sort_order").fetchall()
            for t in types:
                qty = to_float(request.form.get(f"qty_{t['id']}"), 0.0)
                amount = to_float(request.form.get(f"amount_{t['id']}"), 0.0)
                if qty == 0 and amount == 0:
                    conn.execute("DELETE FROM batch_revenues WHERE batch_id=? AND revenue_type_id=?", (batch_id, t["id"]))
                else:
                    conn.execute(
                        """
                        INSERT INTO batch_revenues(batch_id,revenue_type_id,qty,amount)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(batch_id,revenue_type_id)
                        DO UPDATE SET qty=excluded.qty, amount=excluded.amount
                        """,
                        (batch_id, t["id"], qty, amount),
                    )
            conn.commit()
        recalc_batch(batch_id)
        flash("تم تحديث الإيرادات.", "success")
        return redirect(url_for("batch_revenues", batch_id=batch_id))

    with get_conn() as conn:
        batch = conn.execute("SELECT id, batch_num FROM batches WHERE id=?", (batch_id,)).fetchone()
        types = conn.execute("SELECT * FROM revenue_types WHERE is_active=1 ORDER BY sort_order").fetchall()
        values = {
            row["revenue_type_id"]: row
            for row in conn.execute("SELECT * FROM batch_revenues WHERE batch_id=?", (batch_id,)).fetchall()
        }
    return render_template("batch_revenues.html", batch=batch, types=types, values=values)


@app.get("/batches/<int:batch_id>/revenues/export.csv")
def export_batch_revenues_csv(batch_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT rt.code, rt.name_ar, rt.category, COALESCE(br.qty, 0) AS qty, COALESCE(br.amount, 0) AS amount
            FROM revenue_types rt
            LEFT JOIN batch_revenues br ON br.revenue_type_id = rt.id AND br.batch_id = ?
            WHERE rt.is_active = 1
            ORDER BY rt.sort_order
            """,
            (batch_id,),
        ).fetchall()
    payload = [[r["code"], r["name_ar"], r["category"], r["qty"], r["amount"]] for r in rows]
    return make_csv_response(
        f"batch_{batch_id}_revenues.csv",
        ["الكود", "اسم البند", "الفئة", "الكمية", "القيمة"],
        payload,
    )


@app.get("/batches/<int:batch_id>/revenues/export.xlsx")
def export_batch_revenues_excel(batch_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT rt.code, rt.name_ar, rt.category, COALESCE(br.qty, 0) AS qty, COALESCE(br.amount, 0) AS amount
            FROM revenue_types rt
            LEFT JOIN batch_revenues br ON br.revenue_type_id = rt.id AND br.batch_id = ?
            WHERE rt.is_active = 1
            ORDER BY rt.sort_order
            """,
            (batch_id,),
        ).fetchall()
    payload = [[r["code"], r["name_ar"], r["category"], r["qty"], r["amount"]] for r in rows]
    return make_excel_response(
        f"batch_{batch_id}_revenues.xlsx",
        "الإيرادات",
        ["الكود", "اسم البند", "الفئة", "الكمية", "القيمة"],
        payload,
    )


@app.route("/reports")
def reports():
    warehouse_id = to_int(request.args.get("warehouse_id"), 0) or None
    fiscal_year = to_int(request.args.get("fiscal_year"), 0) or None
    data = reports_payload(warehouse_id, fiscal_year)

    return render_template(
        "reports.html",
        warehouse_rows=data["warehouse_rows"],
        annual_rows=data["annual_rows"],
        batches_rows=data["batches_rows"],
        sales_by_type_rows=data["sales_by_type_rows"],
        market_summary_row=data["market_summary_row"],
        cost_types_rows=data["cost_types_rows"],
        revenue_types_rows=data["revenue_types_rows"],
        daily_summary_row=data["daily_summary_row"],
        warehouses=data["warehouses"],
        years=data["years"],
        filters={"warehouse_id": warehouse_id or "", "fiscal_year": fiscal_year or ""},
        report_summary=data["report_summary"],
        dashboard_insights=data["dashboard_insights"],
    )


def build_reports_export_payload(
    report_type: str,
    data: dict[str, object],
    where_clause: str,
    params: list[int],
) -> tuple[str, str, list[str], list[list[object]]]:
    if report_type == "warehouses":
        rows = data["warehouse_rows"]
        export_rows = [
            [
                row["warehouse_name"],
                row["batches_count"],
                row["chicks_total"],
                row["cost_total"],
                row["revenue_total"],
                row["net_total"],
                row["avg_mortality"],
            ]
            for row in rows
        ]
        return (
            "warehouses_summary",
            "ملخص_العنابر",
            ["العنبر", "عدد الدفعات", "إجمالي الكتاكيت", "إجمالي التكاليف", "إجمالي الإيرادات", "الصافي", "متوسط النفوق"],
            export_rows,
        )

    if report_type == "annual":
        rows = data["annual_rows"]
        export_rows = [
            [
                row["fiscal_year"],
                row["batches_count"],
                row["cost_total"],
                row["revenue_total"],
                row["net_total"],
            ]
            for row in rows
        ]
        return (
            "annual_summary",
            "ملخص_سنوي",
            ["السنة المالية", "عدد الدفعات", "إجمالي التكاليف", "إجمالي الإيرادات", "الصافي"],
            export_rows,
        )

    if report_type == "batches":
        with get_conn() as conn:
            rows = conn.execute(
                f"""
                SELECT b.id, b.batch_num, w.name AS warehouse_name, b.date_in, b.date_out,
                       b.chicks, b.total_dead, b.mort_rate, b.fcr, b.total_cost, b.total_rev, b.net_result
                FROM batches b
                JOIN warehouses w ON w.id=b.warehouse_id
                {where_clause}
                ORDER BY b.date_in DESC, b.id DESC
                """,
                params,
            ).fetchall()
        export_rows = [
            [
                row["id"],
                row["batch_num"],
                row["warehouse_name"],
                row["date_in"],
                row["date_out"],
                row["chicks"],
                row["total_dead"],
                row["mort_rate"],
                row["fcr"],
                row["total_cost"],
                row["total_rev"],
                row["net_result"],
            ]
            for row in rows
        ]
        return (
            "batches_report",
            "تقرير_الدفعات",
            [
                "المعرف",
                "رقم الدفعة",
                "العنبر",
                "تاريخ الدخول",
                "تاريخ الخروج",
                "الكتاكيت",
                "إجمالي النفوق",
                "نسبة النفوق",
                "معامل التحويل",
                "إجمالي التكلفة",
                "إجمالي الإيراد",
                "الصافي",
            ],
            export_rows,
        )

    if report_type == "sales_types":
        rows = data["sales_by_type_rows"]
        export_rows = [[row["sale_type"], row["qty_total"], row["amount_total"]] for row in rows]
        return ("sales_by_type", "أنواع_المبيعات", ["نوع البيع", "إجمالي الكمية", "إجمالي القيمة"], export_rows)

    if report_type == "cost_types":
        rows = data["cost_types_rows"]
        export_rows = [[row["type_name"], row["category"], row["qty_total"], row["amount_total"]] for row in rows]
        return ("cost_types_summary", "أنواع_التكاليف", ["اسم البند", "الفئة", "إجمالي الكمية", "إجمالي القيمة"], export_rows)

    if report_type == "revenue_types":
        rows = data["revenue_types_rows"]
        export_rows = [[row["type_name"], row["category"], row["qty_total"], row["amount_total"]] for row in rows]
        return ("revenue_types_summary", "أنواع_الإيرادات", ["اسم البند", "الفئة", "إجمالي الكمية", "إجمالي القيمة"], export_rows)

    if report_type == "daily_summary":
        row = data["daily_summary_row"]
        export_rows = [[row["entries_count"], row["dead_total"], row["feed_total"], row["water_total"], row["avg_daily_deaths"]]]
        return (
            "daily_summary",
            "ملخص_يومي",
            ["عدد السجلات", "إجمالي النفوق", "إجمالي العلف", "إجمالي الماء", "متوسط النفوق اليومي"],
            export_rows,
        )

    raise ValueError("unsupported")


@app.get("/reports/export/<string:report_type>.csv")
def export_reports_csv(report_type: str):
    warehouse_id = to_int(request.args.get("warehouse_id"), 0) or None
    fiscal_year = to_int(request.args.get("fiscal_year"), 0) or None
    data = reports_payload(warehouse_id, fiscal_year)
    where_clause = data["where_clause"]
    params = data["params"]

    try:
        filename_base, _, headers, export_rows = build_reports_export_payload(report_type, data, where_clause, params)
    except ValueError:
        flash("Unsupported report type.", "error")
        return redirect(url_for("reports"))
    return make_csv_response(f"{filename_base}.csv", headers, export_rows)


@app.get("/reports/export/<string:report_type>.xlsx")
def export_reports_excel(report_type: str):
    warehouse_id = to_int(request.args.get("warehouse_id"), 0) or None
    fiscal_year = to_int(request.args.get("fiscal_year"), 0) or None
    data = reports_payload(warehouse_id, fiscal_year)
    where_clause = data["where_clause"]
    params = data["params"]

    try:
        filename_base, sheet_name, headers, export_rows = build_reports_export_payload(report_type, data, where_clause, params)
    except ValueError:
        flash("Unsupported report type.", "error")
        return redirect(url_for("reports"))
    return make_excel_response(f"{filename_base}.xlsx", sheet_name, headers, export_rows)


def build_report_chart_images(data: dict[str, object]) -> list[str]:
    if not HAS_MATPLOTLIB:
        return []

    chart_paths: list[str] = []
    try:
        insights = data.get("dashboard_insights", {})
        warehouse_rows = list(data.get("warehouse_rows", []))
        cost_types_rows = list(data.get("cost_types_rows", []))

        warehouse_labels = [str(r.get("warehouse_name") or "Unknown") for r in warehouse_rows[:8]]
        warehouse_costs = [float(r.get("cost_total") or 0) for r in warehouse_rows[:8]]
        warehouse_revenues = [float(r.get("revenue_total") or 0) for r in warehouse_rows[:8]]

        if warehouse_labels:
            fig, ax = plt.subplots(figsize=(8.5, 3.2))
            x = list(range(len(warehouse_labels)))
            ax.bar([i - 0.2 for i in x], warehouse_costs, width=0.4, label="Costs")
            ax.bar([i + 0.2 for i in x], warehouse_revenues, width=0.4, label="Revenues")
            ax.set_title("Warehouse Financial Comparison")
            ax.set_xticks(x)
            ax.set_xticklabels(warehouse_labels, rotation=20, ha="right")
            ax.legend()
            ax.grid(axis="y", alpha=0.2)
            fig.tight_layout()
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            fig.savefig(tmp.name, dpi=150)
            plt.close(fig)
            chart_paths.append(tmp.name)

        rev_vals = [
            float(insights.get("farm_sales_total") or 0),
            float(insights.get("market_sales_total") or 0),
            float(insights.get("extra_revenues_total") or 0),
        ]
        if sum(rev_vals) > 0:
            fig, ax = plt.subplots(figsize=(8.5, 3.2))
            ax.pie(rev_vals, labels=["Farm", "Market", "Extra"], autopct="%1.1f%%", startangle=90)
            ax.set_title("Revenue Split")
            fig.tight_layout()
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            fig.savefig(tmp.name, dpi=150)
            plt.close(fig)
            chart_paths.append(tmp.name)

        top_cost = cost_types_rows[:8]
        if top_cost:
            labels = [str(r.get("type_name") or "N/A") for r in top_cost]
            vals = [float(r.get("amount_total") or 0) for r in top_cost]
            fig, ax = plt.subplots(figsize=(8.5, 3.2))
            ax.plot(labels, vals, marker="o")
            ax.fill_between(range(len(vals)), vals, alpha=0.15)
            ax.set_title("Top Cost Types")
            ax.tick_params(axis="x", rotation=25)
            ax.grid(axis="y", alpha=0.2)
            fig.tight_layout()
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            fig.savefig(tmp.name, dpi=150)
            plt.close(fig)
            chart_paths.append(tmp.name)
    except Exception:
        return chart_paths

    return chart_paths


@app.get("/reports/export/report.pdf")
def export_reports_pdf():
    if not HAS_FPDF:
        flash("PDF library is not available in this environment.", "error")
        return redirect(url_for("reports"))

    warehouse_id = to_int(request.args.get("warehouse_id"), 0) or None
    fiscal_year = to_int(request.args.get("fiscal_year"), 0) or None
    include_dashboard = request.args.get("include_dashboard", "1") == "1"
    data = reports_payload(warehouse_id, fiscal_year)

    pdf = FPDF()
    pdf.add_page()
    has_arabic_font = REPORT_FONT_PATH.exists()
    if has_arabic_font:
        pdf.add_font("Arabic", "", str(REPORT_FONT_PATH), uni=True)
        pdf.set_font("Arabic", "", 13)
    else:
        pdf.set_font("Arial", "", 12)

    company_name = get_setting("company_name", "شركة الدواجن")
    pdf.cell(0, 8, ar_text(f"{company_name} - التقرير التنفيذي"), new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Arabic" if has_arabic_font else "Arial", "", 10)
    pdf.cell(
        0,
        7,
        ar_text(
            f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')} | فلتر العنبر: {warehouse_id or 'الكل'} | السنة المالية: {fiscal_year or 'الكل'}"
        ),
        new_x="LMARGIN",
        new_y="NEXT",
        align="C",
    )
    pdf.ln(3)

    summary = data["report_summary"]
    chart_paths: list[str] = []
    if include_dashboard:
        pdf.set_font("Arabic" if has_arabic_font else "Arial", "", 11)
        pdf.cell(0, 7, ar_text("مؤشرات لوحة التحكم"), new_x="LMARGIN", new_y="NEXT", align="R")
        pdf.set_font("Arabic" if has_arabic_font else "Arial", "", 10)
        insights = data["dashboard_insights"]
        kpis = [
            f"إجمالي التكاليف: {summary['cost_total']:,.2f}",
            f"إجمالي الإيرادات: {summary['revenue_total']:,.2f}",
            f"صافي النتيجة: {summary['net_total']:,.2f}",
            f"هامش الربح: {insights['margin_pct']:.2f}%",
            f"متوسط FCR: {insights['avg_fcr']:.3f}",
            f"متوسط النفوق: {insights['avg_mortality']:.2f}%",
            f"مبيعات العنبر: {insights['farm_sales_total']:,.2f}",
            f"مبيعات السوق: {insights['market_sales_total']:,.2f}",
            f"إيرادات إضافية: {insights['extra_revenues_total']:,.2f}",
        ]
        for line in kpis:
            pdf.cell(0, 6, ar_text(line), new_x="LMARGIN", new_y="NEXT", align="R")

        chart_paths = build_report_chart_images(data)
        for chart_path in chart_paths:
            if pdf.get_y() > 220:
                pdf.add_page()
            pdf.ln(2)
            pdf.image(chart_path, x=12, y=pdf.get_y(), w=185)
            pdf.ln(58)

    pdf.ln(3)
    pdf.set_font("Arabic" if has_arabic_font else "Arial", "", 11)
    pdf.cell(0, 7, ar_text("ملخص العنابر"), new_x="LMARGIN", new_y="NEXT", align="R")
    pdf.set_font("Arabic" if has_arabic_font else "Arial", "", 9)
    headers = ["العنبر", "الدفعات", "التكاليف", "الإيرادات", "الصافي"]
    widths = [54, 20, 35, 35, 35]
    for h, w in zip(headers, widths):
        pdf.cell(w, 7, ar_text(h), border=1, align="C")
    pdf.ln()
    for row in data["warehouse_rows"]:
        vals = [
            row["warehouse_name"],
            f"{int(row['batches_count'] or 0)}",
            f"{float(row['cost_total'] or 0):,.0f}",
            f"{float(row['revenue_total'] or 0):,.0f}",
            f"{float(row['net_total'] or 0):,.0f}",
        ]
        for v, w in zip(vals, widths):
            pdf.cell(w, 6, ar_text(v), border=1, align="C")
        pdf.ln()

    pdf.ln(3)
    pdf.set_font("Arabic" if has_arabic_font else "Arial", "", 11)
    pdf.cell(0, 7, ar_text("أفضل 25 دفعة"), new_x="LMARGIN", new_y="NEXT", align="R")
    pdf.set_font("Arabic" if has_arabic_font else "Arial", "", 9)
    headers = ["الدفعة", "العنبر", "الدخول", "الخروج", "FCR", "الصافي"]
    widths = [20, 45, 30, 30, 20, 34]
    for h, w in zip(headers, widths):
        pdf.cell(w, 7, ar_text(h), border=1, align="C")
    pdf.ln()
    for row in data["batches_rows"]:
        vals = [
            row.get("batch_num") or row.get("id"),
            row.get("warehouse_name"),
            row.get("date_in"),
            row.get("date_out"),
            f"{float(row.get('fcr') or 0):.3f}",
            f"{float(row.get('net_result') or 0):,.0f}",
        ]
        for v, w in zip(vals, widths):
            pdf.cell(w, 6, ar_text(v), border=1, align="C")
        pdf.ln()

    out = pdf.output(dest="S")
    pdf_bytes = out.encode("latin-1") if isinstance(out, str) else bytes(out)
    suffix = "with_dashboard" if include_dashboard else "without_dashboard"
    filename = f"reports_{suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    for chart_path in chart_paths:
        try:
            os.unlink(chart_path)
        except Exception:
            pass

    return make_pdf_response(filename, pdf_bytes)


def _collect_import_input_files() -> list[str]:
    files: list[str] = []
    upload_items = request.files.getlist("files")
    if upload_items:
        tmp_dir = Path(tempfile.gettempdir()) / "poultry_import_uploads"
        tmp_dir.mkdir(parents=True, exist_ok=True)
        for idx, item in enumerate(upload_items):
            if not item or not getattr(item, "filename", ""):
                continue
            suffix = Path(str(item.filename)).suffix.lower()
            if suffix not in {".xlsm", ".xlsx", ".xls"}:
                continue
            saved_path = tmp_dir / f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{idx}{suffix}"
            item.save(saved_path)
            files.append(str(saved_path))

    folder_path = (request.form.get("folder_path") or "").strip()
    if folder_path:
        folder = Path(folder_path)
        if folder.exists() and folder.is_dir():
            for p in sorted(folder.glob("*.xlsm")):
                if not p.name.startswith("~$"):
                    files.append(str(p))
            for p in sorted(folder.glob("*.xlsx")):
                if not p.name.startswith("~$"):
                    files.append(str(p))
    return list(dict.fromkeys(files))


@app.get("/imports")
def imports_home():
    with get_conn() as conn:
        profiles = conn.execute(
            """
            SELECT id, name, source_key, is_default
            FROM import_profiles
            WHERE is_active=1
            ORDER BY is_default DESC, id ASC
            """
        ).fetchall()
        batches = conn.execute(
            """
            SELECT b.id, b.batch_num, w.name AS warehouse_name, b.date_in
            FROM batches b
            JOIN warehouses w ON w.id=b.warehouse_id
            ORDER BY b.date_in DESC, b.id DESC
            LIMIT 300
            """
        ).fetchall()
    return render_template("imports.html", profiles=profiles, batches=batches)


@app.post("/imports/upload")
def imports_upload():
    files = _collect_import_input_files()
    if not files:
        flash("يرجى اختيار ملف Excel أو إدخال مسار مجلد.", "error")
        return redirect(url_for("imports_home"))

    profile_id = to_int(request.form.get("profile_id"), 0) or None
    batch_mode = (request.form.get("batch_mode") or "create").strip().lower()
    if batch_mode not in {"create", "update"}:
        batch_mode = "create"
    target_batch_id = to_int(request.form.get("target_batch_id"), 0) or None

    selected_source_key = None
    if profile_id:
        with get_conn() as conn:
            profile_row = conn.execute(
                "SELECT source_key FROM import_profiles WHERE id=? AND is_active=1",
                (profile_id,),
            ).fetchone()
        if profile_row and profile_row["source_key"]:
            selected_source_key = str(profile_row["source_key"])
        else:
            profile_id = None

    profile_match = detect_profile(files, source_key=selected_source_key)
    if not profile_match.get("matched"):
        source_key = str(profile_match.get("source_key") or selected_source_key or "poultry_v4")
        if source_key != "poultry_v4":
            bad_files = [f for f in profile_match.get("files", []) if not f.get("ok")]
            if bad_files:
                details = ", ".join(
                    f"{item.get('file_name')}: {item.get('error') or 'invalid workbook'}" for item in bad_files
                )
                flash(f"تعذر قراءة بعض الملفات في وضع Generic Excel: {details}", "error")
            else:
                flash("تعذر التحقق من ملفات Excel في الوضع العام.", "error")
            return redirect(url_for("imports_home"))
        bad_files = [f for f in profile_match.get("files", []) if not f.get("ok")]
        if bad_files:
            details = ", ".join(
                f"{item.get('file_name')}: {', '.join(item.get('missing_sheets') or [])}" for item in bad_files
            )
            flash(f"الملف غير مطابق لقالب poultry_v4. التفاصيل: {details}", "error")
        else:
            flash("تعذر التحقق من القالب. تأكد من ملفات الاستيراد.", "error")
        return redirect(url_for("imports_home"))

    resolved_profile_id = profile_id or profile_match.get("profile_id")
    try:
        payload = parse_files(files, resolved_profile_id)
        run_id = build_staging(
            payload,
            resolved_profile_id,
            source_ui="web",
            created_by="web",
        )
    except Exception as exc:
        flash(f"فشل التحليل: {exc}", "error")
        return redirect(url_for("imports_home"))
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE import_runs
            SET batch_mode=?, target_batch_id=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (batch_mode, target_batch_id, run_id),
        )
        if target_batch_id:
            conn.execute(
                "UPDATE import_run_files SET target_batch_id=? WHERE run_id=?",
                (target_batch_id, run_id),
            )
        conn.commit()
    if payload.get("errors"):
        flash("تم تجاوز بعض الملفات بسبب أخطاء في التحليل. راجع شاشة النتيجة.", "warning")
    return redirect(url_for("imports_run_review", run_id=run_id))


@app.get("/imports/run/<int:run_id>")
def imports_run_review(run_id: int):
    with get_conn() as conn:
        run = conn.execute(
            """
            SELECT r.*, p.name AS profile_name
            FROM import_runs r
            LEFT JOIN import_profiles p ON p.id=r.profile_id
            WHERE r.id=?
            """,
            (run_id,),
        ).fetchone()
        if not run:
            flash("عملية الاستيراد غير موجودة.", "error")
            return redirect(url_for("imports_home"))

        run_files = conn.execute(
            "SELECT * FROM import_run_files WHERE run_id=? ORDER BY id",
            (run_id,),
        ).fetchall()
        raw_lines = conn.execute(
            """
            SELECT l.*, rf.file_name
            FROM import_run_lines l
            JOIN import_run_files rf ON rf.id=l.run_file_id
            WHERE l.run_id=? AND l.line_kind='candidate' AND l.mapping_status<>'ignored'
            ORDER BY rf.file_name, l.source_sheet, l.source_row, l.id
            """,
            (run_id,),
        ).fetchall()
        cost_types = conn.execute(
            "SELECT code, name_ar, category, unit, has_qty FROM cost_types WHERE is_active=1 ORDER BY sort_order, id"
        ).fetchall()
        revenue_types = conn.execute(
            "SELECT code, name_ar, category, unit, has_qty FROM revenue_types WHERE is_active=1 ORDER BY sort_order, id"
        ).fetchall()
        unresolved_count = conn.execute(
            """
            SELECT COUNT(*) AS c
            FROM import_run_lines
            WHERE run_id=? AND line_kind='candidate' AND mapping_status='unmapped'
            """,
            (run_id,),
        ).fetchone()

    lines: list[dict[str, Any]] = []
    for row in raw_lines:
        row_dict = dict(row)
        payload_data: dict[str, Any] = {}
        try:
            payload_data = json.loads(str(row_dict.get("payload_json") or "{}"))
        except Exception:
            payload_data = {}
        row_dict["payload"] = payload_data
        row_dict["source_label_raw"] = payload_data.get("source_label_raw") or row_dict.get("source_label") or ""
        lines.append(row_dict)

    return render_template(
        "import_review.html",
        run=run,
        run_files=run_files,
        lines=lines,
        cost_types=cost_types,
        revenue_types=revenue_types,
        unresolved_count=int(unresolved_count["c"] or 0),
    )


@app.post("/imports/run/<int:run_id>/mapping")
def imports_run_mapping(run_id: int):
    line_ids = [to_int(x, 0) for x in request.form.getlist("line_id")]
    edits: list[dict[str, Any]] = []
    for line_id in line_ids:
        if line_id <= 0:
            continue
        kind = (request.form.get(f"kind_{line_id}") or "").strip()
        existing_code = (request.form.get(f"existing_code_{line_id}") or "").strip()
        new_name = (request.form.get(f"new_name_{line_id}") or "").strip()
        category = (request.form.get(f"category_{line_id}") or "").strip()
        unit = (request.form.get(f"unit_{line_id}") or "").strip()
        has_qty = 1 if request.form.get(f"has_qty_{line_id}") else 0

        if kind == "ignore" or not kind:
            edits.append({"line_id": line_id, "action": "ignore", "target_kind": "ignore"})
        elif existing_code:
            edits.append(
                {
                    "line_id": line_id,
                    "action": "existing",
                    "target_kind": kind,
                    "target_code": existing_code,
                    "category": category,
                    "unit": unit,
                    "has_qty": has_qty,
                }
            )
        elif new_name:
            edits.append(
                {
                    "line_id": line_id,
                    "action": "new",
                    "target_kind": kind,
                    "target_name": new_name,
                    "category": category,
                    "unit": unit,
                    "has_qty": has_qty,
                }
            )
        else:
            edits.append({"line_id": line_id, "action": "ignore", "target_kind": "ignore"})

    try:
        apply_mapping_edits(run_id, edits)
    except Exception as exc:
        flash(f"فشل حفظ التصنيفات: {exc}", "error")
        return redirect(url_for("imports_run_review", run_id=run_id))
    flash("تم حفظ تصنيفات الربط.", "success")
    return redirect(url_for("imports_run_review", run_id=run_id))


@app.post("/imports/run/<int:run_id>/commit")
def imports_run_commit(run_id: int):
    with get_conn() as conn:
        run = conn.execute("SELECT batch_mode, target_batch_id FROM import_runs WHERE id=?", (run_id,)).fetchone()
    if not run:
        flash("عملية الاستيراد غير موجودة.", "error")
        return redirect(url_for("imports_home"))

    batch_mode = (request.form.get("batch_mode") or run["batch_mode"] or "create").strip().lower()
    if batch_mode not in {"create", "update"}:
        batch_mode = "create"
    merge_mode = (request.form.get("merge_mode") or "replace").strip().lower()
    if merge_mode not in {"replace", "merge"}:
        merge_mode = "replace"
    target_batch_id = to_int(request.form.get("target_batch_id"), 0) or (int(run["target_batch_id"]) if run["target_batch_id"] else None)

    try:
        report = commit_run(
            run_id=run_id,
            batch_mode=batch_mode,
            merge_mode=merge_mode,
            target_batch_id=target_batch_id,
        )
    except Exception as exc:
        flash(f"فشل التنفيذ: {exc}", "error")
        return redirect(url_for("imports_run_review", run_id=run_id))
    set_setting(f"import_report_{run_id}", json.dumps(report, ensure_ascii=False))
    if report.get("status") in {"failed", "partial_failed"}:
        flash("تم التنفيذ مع وجود ملفات فشلت. راجع التقرير.", "warning")
    else:
        flash("تم تنفيذ الاستيراد بنجاح.", "success")
    return redirect(url_for("imports_run_result", run_id=run_id))


@app.get("/imports/run/<int:run_id>/result")
def imports_run_result(run_id: int):
    with get_conn() as conn:
        run = conn.execute(
            """
            SELECT r.*, p.name AS profile_name
            FROM import_runs r
            LEFT JOIN import_profiles p ON p.id=r.profile_id
            WHERE r.id=?
            """,
            (run_id,),
        ).fetchone()
        if not run:
            flash("عملية الاستيراد غير موجودة.", "error")
            return redirect(url_for("imports_home"))
        run_files = conn.execute(
            "SELECT * FROM import_run_files WHERE run_id=? ORDER BY id",
            (run_id,),
        ).fetchall()
    report = {}
    try:
        report = json.loads(get_setting(f"import_report_{run_id}", "{}"))
    except Exception:
        report = {}
    return render_template("import_result.html", run=run, run_files=run_files, report=report)


@app.route("/settings", methods=["GET", "POST"])
def settings():
    keys = [
        ("company_name", "اسم الشركة", "شركة الدواجن"),
        ("farm_address", "العنوان", ""),
        ("contact_number", "رقم التواصل", ""),
        ("currency", "العملة", "ريال"),
        ("manager_name", "المدير العام", ""),
        ("finance_name", "المدير المالي", ""),
        ("auditor_name", "المراجع", ""),
        ("tg_token", "Telegram Bot Token", ""),
        ("tg_chat_id", "Telegram Chat ID", ""),
        ("tg_auto", "التفعيل التلقائي", "0"),
        ("ui_font_scale", "حجم الخط", "110"),
        ("ui_line_height", "تباعد الأسطر", "1.65"),
        ("ui_density", "كثافة العرض", "balanced"),
        ("ui_separator_strength", "قوة الفواصل", "strong"),
        ("ui_font_family", "نوع الخط", "system"),
    ]

    if request.method == "POST":
        for key, _, default in keys:
            value = request.form.get(key, default).strip()
            if key.startswith("ui_"):
                value = _sanitize_ui_setting(key, value, default)
            set_setting(key, value)
        flash("تم حفظ الإعدادات.", "success")
        return redirect(url_for("settings"))

    values = {key: get_setting(key, default) for key, _, default in keys}
    return render_template("settings.html", values=values)


@app.post("/settings/backup")
def create_backup():
    backup_path = create_database_backup()
    flash(f"تم إنشاء نسخة احتياطية: {backup_path}", "success")
    return redirect(url_for("settings"))


def create_app() -> Flask:
    shared_ensure_schema()
    return app


if __name__ == "__main__":
    shared_ensure_schema()
    debug_mode = os.environ.get("POULTRY_WEB_DEBUG", "0") == "1"
    app.run(host="127.0.0.1", port=5000, debug=debug_mode)
