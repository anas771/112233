import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook

import core.database as core_db
import core.import_engine as import_engine


class ImportEngineTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = Path(tempfile.gettempdir()) / "poultry_import_tests"
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.db_path = self.temp_dir / f"import_engine_{uuid4().hex}.db"
        self.excel_path = self.temp_dir / f"sample_{uuid4().hex}.xlsx"
        self.original_engine_db_path = import_engine.DB_PATH

        core_db.ensure_schema(self.db_path)
        import_engine.DB_PATH = self.db_path
        self._build_sample_workbook(self.excel_path)

    def tearDown(self) -> None:
        import_engine.DB_PATH = self.original_engine_db_path
        for suffix in ("", "-wal", "-shm"):
            p = Path(f"{self.db_path}{suffix}")
            if p.exists():
                try:
                    p.unlink()
                except PermissionError:
                    pass
        if self.excel_path.exists():
            try:
                self.excel_path.unlink()
            except PermissionError:
                pass

    def _build_sample_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "اجمالي التكاليف"
        ws_feed = wb.create_sheet("تكاليف العلف")
        ws_sales = wb.create_sheet("بيان المبيعات")
        ws_daily = wb.create_sheet("ورقة1")

        ws_summary["A3"] = "العلف - طن"
        ws_summary["B3"] = 2
        ws_summary["C3"] = 5000
        ws_summary["I3"] = "علف مباع"
        ws_summary["J3"] = 800

        ws_feed["A8"] = 120
        ws_feed["B8"] = 2
        ws_feed["C8"] = "علف متبقي في العنبر"

        ws_sales["A2"] = "اسم العميل"
        ws_sales["B2"] = "العدد"
        ws_sales["C2"] = "السعر"
        ws_sales["D2"] = "اجمالي-اجل"
        ws_sales["H2"] = "اسم المكتب-السائق"
        ws_sales["I2"] = "الكمية"
        ws_sales["J2"] = "الوفيات"
        ws_sales["K2"] = "المباع"
        ws_sales["L2"] = "صافي الفاتورة"
        ws_sales["N2"] = "التاريخ"
        ws_sales["A3"] = "عميل اختبار"
        ws_sales["B3"] = 100
        ws_sales["C3"] = 2000
        ws_sales["D3"] = 200000
        ws_sales["H3"] = "مكتب اختبار"
        ws_sales["I3"] = 80
        ws_sales["J3"] = 1
        ws_sales["K3"] = 79
        ws_sales["L3"] = 170000
        ws_sales["N3"] = datetime(2026, 1, 5)

        ws_daily["A1"] = "عدد الدجاج"
        ws_daily["B1"] = 1000
        ws_daily["I1"] = "اسم المزرعه :"
        ws_daily["K1"] = "عنبر اختبار"
        ws_daily["I2"] = "رقم الدفعة :"
        ws_daily["K2"] = "B-100"
        ws_daily["A4"] = "التاريخ"
        ws_daily["B4"] = "العمر (يوم)"
        ws_daily["C4"] = "العدد"
        ws_daily["D4"] = "الوفيات"
        ws_daily["F4"] = "العلف المستهلك /كيس"
        ws_daily["A5"] = datetime(2026, 1, 1)
        ws_daily["B5"] = 1
        ws_daily["C5"] = 1000
        ws_daily["D5"] = 5
        ws_daily["F5"] = 2
        ws_daily["A6"] = datetime(2026, 1, 2)
        ws_daily["B6"] = 2
        ws_daily["C6"] = 995
        ws_daily["D6"] = 4
        ws_daily["F6"] = 2.5

        wb.save(path)

    def test_detect_and_parse(self) -> None:
        profile_check = import_engine.detect_profile([str(self.excel_path)])
        self.assertTrue(profile_check["matched"])
        payload = import_engine.parse_files([str(self.excel_path)], profile_check["profile_id"])
        self.assertEqual(len(payload["files"]), 1)
        one_file = payload["files"][0]
        self.assertGreater(len(one_file["daily_records"]), 0)
        self.assertGreater(len(one_file["farm_sales"]), 0)
        self.assertGreater(len(one_file["candidates"]), 0)

    def test_build_staging_and_commit_create(self) -> None:
        profile_check = import_engine.detect_profile([str(self.excel_path)])
        payload = import_engine.parse_files([str(self.excel_path)], profile_check["profile_id"])
        run_id = import_engine.build_staging(payload, profile_check["profile_id"], source_ui="test", created_by="test")
        report = import_engine.commit_run(run_id, batch_mode="create", merge_mode="replace", target_batch_id=None)

        self.assertEqual(report["status"], "committed")
        self.assertEqual(report["committed_files"], 1)

        with core_db.get_conn(self.db_path) as conn:
            batch_count = conn.execute("SELECT COUNT(*) AS c FROM batches").fetchone()["c"]
            daily_count = conn.execute("SELECT COUNT(*) AS c FROM daily_records").fetchone()["c"]
            farm_count = conn.execute("SELECT COUNT(*) AS c FROM farm_sales").fetchone()["c"]
            cost_rows = conn.execute("SELECT COUNT(*) AS c FROM batch_costs").fetchone()["c"]
            rev_rows = conn.execute("SELECT COUNT(*) AS c FROM batch_revenues").fetchone()["c"]

        self.assertEqual(batch_count, 1)
        self.assertGreater(daily_count, 0)
        self.assertGreater(farm_count, 0)
        self.assertGreater(cost_rows, 0)
        self.assertGreater(rev_rows, 0)


if __name__ == "__main__":
    unittest.main()
