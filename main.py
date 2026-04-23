"""
منظومة إدارة عنابر الدجاج اللاحم — النسخة المطورة 4.5
Poultry Farm Management System — Enhanced v4.5
SQLite + Tkinter/ttkbootstrap + Matplotlib — يعمل على Windows بدون إنترنت
التحديث: إعدادات شاملة (عملة، عنوان، هواتف) وحذف عنابر مباشر ديناميكي
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
import shutil
import textwrap
import hashlib
import subprocess
import urllib.parse
import urllib.request
from datetime import datetime, date, timedelta
from pathlib import Path
from desktop_import_wizard import open_import_wizard
from core.database import ensure_schema as shared_ensure_schema

# ── دعم المظهر العصري ──────────────────────────────────────────
try:
    import ttkbootstrap as ttkb
    from ttkbootstrap.constants import *
    WindowBase = ttkb.Window
    ToplevelBase = ttkb.Toplevel
    HAS_TTKB = True
except ImportError:
    WindowBase = tk.Tk
    ToplevelBase = tk.Toplevel
    HAS_TTKB = False

# ── مكونات واجهة ذكية تدعم التنسيق الحديث والوضع الليلي ──────────
class UIFrame(ttkb.Frame if HAS_TTKB else tk.Frame):
    def __init__(self, master=None, **kwargs):
        if HAS_TTKB:
            for k in ['bg', 'fg', 'pady', 'padx', 'relief', 'bd', 'highlightbackground', 'highlightthickness']: kwargs.pop(k, None)
        super().__init__(master, **kwargs)

class UILabel(ttkb.Label if HAS_TTKB else tk.Label):
    def __init__(self, master=None, **kwargs):
        if HAS_TTKB:
            for k in ['bg', 'fg', 'activebackground', 'activeforeground', 'relief', 'bd', 'padx', 'pady']: kwargs.pop(k, None)
        super().__init__(master, **kwargs)

class UIButton(ttkb.Button if HAS_TTKB else tk.Button):
    def __init__(self, master=None, **kwargs):
        if HAS_TTKB:
            for k in ['bg', 'fg', 'activebackground', 'activeforeground', 'relief', 'bd', 'font', 'padx', 'pady']: kwargs.pop(k, None)
            txt = kwargs.get('text', '')
            if 'حذف' in txt or '🗑' in txt: kwargs.setdefault('bootstyle', 'danger')
            elif 'إلغاء' in txt: kwargs.setdefault('bootstyle', 'secondary')
            elif 'حفظ' in txt or 'إضافة' in txt or '➕' in txt: kwargs.setdefault('bootstyle', 'success')
            elif 'تعديل' in txt or '✏️' in txt: kwargs.setdefault('bootstyle', 'warning')
            elif 'PDF' in txt: kwargs.setdefault('bootstyle', 'info')
            else: kwargs.setdefault('bootstyle', 'primary')
        super().__init__(master, **kwargs)

class UIEntry(ttkb.Entry if HAS_TTKB else tk.Entry):
    def __init__(self, master=None, **kwargs):
        if HAS_TTKB:
            for k in ['bg', 'fg', 'insertbackground', 'relief', 'bd', 'highlightthickness', 'highlightbackground', 'padx', 'pady']: kwargs.pop(k, None)
        super().__init__(master, **kwargs)

class UILabelFrame(ttkb.Labelframe if HAS_TTKB else tk.LabelFrame):
    def __init__(self, master=None, **kwargs):
        if HAS_TTKB:
            for k in ['bg', 'fg', 'pady', 'padx', 'relief', 'bd', 'font', 'labelanchor']: kwargs.pop(k, None)
            kwargs.setdefault('bootstyle', 'primary')
        super().__init__(master, **kwargs)

# ── دعم الرسوم البيانية ومكتبات التصدير ─────────────────────────────────────────
try:
    import matplotlib
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    matplotlib.rc('font', family='Arial')
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    HAS_ARABIC = True
except ImportError:
    HAS_ARABIC = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── المسار الرئيسي للتطبيق ──────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, "poultry_data.db")
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
AMIRI_FONT_PATH = os.path.join(ASSETS_DIR, "Amiri-Regular.ttf")
LOGO_PATH = os.path.join(ASSETS_DIR, "logo.png")

# ── ألوان التطبيق ────────────────────────────────────────────────
CLR = {
    "bg":       "#f0f2f5",
    "header":   "#1F4E79",
    "nav":      "#2E75B6",
    "white":    "#ffffff",
    "profit":   "#27680a",
    "loss":     "#c00000",
    "warn":     "#bf9000",
    "profit_bg":"#e2efda",
    "loss_bg":  "#fce4d6",
    "warn_bg":  "#fff2cc",
    "info_bg":  "#dce6f1",
    "border":   "#c8d4e0",
    "text":     "#222222",
    "text2":    "#555555",
    "accent":   "#0070c0",
    "daily_bg": "#f0f7ff",
}

# ── معايير Ross 308 العالمية (يومياً 1-65) ───────────────────────
# علف (غرام/طائر)
ROSS_DAILY_F = {d: f for d, f in zip(range(1, 66), [16.0, 19.3, 22.7, 26.0, 29.3, 32.7, 36.0, 41.1, 46.3, 51.4, 56.6, 61.7, 66.9, 72.0, 78.1, 84.3, 90.4, 96.6, 102.7, 108.9, 115.0, 121.1, 127.3, 133.4, 139.6, 145.7, 151.9, 158.0, 163.3, 168.6, 173.9, 179.1, 184.4, 189.7, 195.0, 199.1, 203.3, 207.4, 211.6, 215.7, 219.9, 224.0, 226.6, 229.1, 231.7, 234.3, 236.9, 239.4, 242.0, 243.1, 244.3, 245.4, 246.6, 247.7, 248.9, 250.0, 250.6, 251.3, 251.9, 252.6, 253.2, 253.9, 254.5, 256.3, 258.0])}
# نافق (%)
ROSS_DAILY_M = {d: m for d, m in zip(range(1, 66), [0.15, 0.14, 0.13, 0.12, 0.12, 0.11, 0.1, 0.1, 0.09, 0.09, 0.09, 0.09, 0.08, 0.08, 0.08, 0.07, 0.07, 0.07, 0.07, 0.06, 0.06, 0.06, 0.06, 0.06, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05])}

FT_TITLE  = ("Arial", 14, "bold")
FT_HEADER = ("Arial", 12, "bold")
FT_BODY   = ("Arial", 11)
FT_SMALL  = ("Arial", 10)
FT_TINY   = ("Arial", 9)
 
def center_window(win):
    """توسيط النافذة في منتصف الشاشة"""
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    x = (win.winfo_screenwidth() // 2) - (width // 2)
    y = (win.winfo_screenheight() // 2) - (height // 2)
    win.geometry(f'{width}x{height}+{x}+{y}')

# ════════════════════════════════════════════════════════════════
# مدير قاعدة البيانات
# ════════════════════════════════════════════════════════════════
class DBManager:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        self._init_db()

    def get_conn(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def execute_script(self, script):
        with self.get_conn() as conn:
            conn.executescript(script)

    def fetch_all(self, query, params=()):
        with self.get_conn() as conn:
            rows = conn.execute(query, params).fetchall()
            return [dict(row) for row in rows]

    def fetch_one(self, query, params=()):
        with self.get_conn() as conn:
            row = conn.execute(query, params).fetchone()
            return dict(row) if row else None

    def execute(self, query, params=()):
        with self.get_conn() as conn:
            cursor = conn.execute(query, params)
            conn.commit()
            return cursor.lastrowid

    def get_setting(self, key, default=""):
        row = self.fetch_one("SELECT value FROM system_settings WHERE key=?", (key,))
        return row["value"] if row else default

    def set_setting(self, key, value):
        self.execute(""" INSERT INTO system_settings (key, value) 
            VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value
        """, (key, value))

    def _init_db(self):
        self.execute_script(""" PRAGMA journal_mode=WAL;

        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );

        CREATE TABLE IF NOT EXISTS warehouses (
            id    INTEGER PRIMARY KEY AUTOINCREMENT,
            name  TEXT NOT NULL UNIQUE,
            notes TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS batches (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            warehouse_id   INTEGER NOT NULL REFERENCES warehouses(id),
            batch_num      TEXT    DEFAULT '',
            date_in        TEXT    NOT NULL,
            date_out       TEXT    NOT NULL,
            days           INTEGER DEFAULT 0,
            chicks         INTEGER NOT NULL,
            chick_price    REAL    DEFAULT 0,
            chick_val      REAL    DEFAULT 0,
            feed_qty       REAL    DEFAULT 0,
            feed_val       REAL    DEFAULT 0,
            feed_trans     REAL    DEFAULT 0,
            sawdust_qty    REAL    DEFAULT 0,
            sawdust_val    REAL    DEFAULT 0,
            water_val      REAL    DEFAULT 0,
            gas_qty        REAL    DEFAULT 0,
            gas_val        REAL    DEFAULT 0,
            drugs_val      REAL    DEFAULT 0,
            wh_expenses    REAL    DEFAULT 0,
            house_exp      REAL    DEFAULT 0,
            breeders_pay   REAL    DEFAULT 0,
            qat_pay        REAL    DEFAULT 0,
            rent_val       REAL    DEFAULT 0,
            light_val      REAL    DEFAULT 0,
            sup_wh_pay     REAL    DEFAULT 0,
            sup_co_pay     REAL    DEFAULT 0,
            sup_sale_pay   REAL    DEFAULT 0,
            admin_val      REAL    DEFAULT 0,
            vaccine_pay    REAL    DEFAULT 0,
            delivery_val   REAL    DEFAULT 0,
            mixing_val     REAL    DEFAULT 0,
            wash_val       REAL    DEFAULT 0,
            other_costs    REAL    DEFAULT 0,
            total_cost     REAL    DEFAULT 0,
            cust_qty       INTEGER DEFAULT 0,
            cust_val       REAL    DEFAULT 0,
            mkt_qty        INTEGER DEFAULT 0,
            mkt_val        REAL    DEFAULT 0,
            offal_val      REAL    DEFAULT 0,
            feed_sale      REAL    DEFAULT 0,
            feed_trans_r   REAL    DEFAULT 0,
            drug_return    REAL    DEFAULT 0,
            gas_return     REAL    DEFAULT 0,
            total_rev      REAL    DEFAULT 0,
            total_sold     INTEGER DEFAULT 0,
            total_dead     INTEGER DEFAULT 0,
            mort_rate      REAL    DEFAULT 0,
            avg_weight     REAL    DEFAULT 0,
            fcr            REAL    DEFAULT 0,
            avg_price      REAL    DEFAULT 0,
            net_result     REAL    DEFAULT 0,
            share_pct      REAL    DEFAULT 65,
            share_val      REAL    DEFAULT 0,
            notes          TEXT    DEFAULT '',
            created_at     TEXT,
            consumed_birds INTEGER DEFAULT 0,
            partner_name   TEXT    DEFAULT '',
            feed_sale_qty  REAL    DEFAULT 0,
            feed_trans_r_qty REAL  DEFAULT 0,
            feed_rem_qty   REAL    DEFAULT 0,
            feed_rem_val   REAL    DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS daily_records (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id    INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
            rec_date    TEXT    NOT NULL,
            day_num     INTEGER DEFAULT 0,
            dead_count  INTEGER DEFAULT 0,
            feed_kg     REAL    DEFAULT 0,
            water_ltr   REAL    DEFAULT 0,
            notes       TEXT    DEFAULT '',
            UNIQUE(batch_id, rec_date)
        );

        CREATE TABLE IF NOT EXISTS farm_sales (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id  INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
            sale_date TEXT    DEFAULT '',
            sale_type TEXT    DEFAULT 'آجل',
            customer  TEXT,
            qty       INTEGER DEFAULT 0,
            price     REAL    DEFAULT 0,
            total_val REAL    DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS market_sales (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id  INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
            sale_date TEXT    DEFAULT '',
            office    TEXT,
            qty_sent  INTEGER DEFAULT 0,
            deaths    INTEGER DEFAULT 0,
            qty_sold  INTEGER DEFAULT 0,
            net_val   REAL    DEFAULT 0,
            inv_num   TEXT
        );

        CREATE TABLE IF NOT EXISTS batch_standards (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id    INTEGER NOT NULL UNIQUE REFERENCES batches(id) ON DELETE CASCADE,
            target_fcr  REAL DEFAULT 1.8,
            mort_w1 REAL DEFAULT 0.15, mort_w2 REAL DEFAULT 0.08, mort_w3 REAL DEFAULT 0.06,
            mort_w4 REAL DEFAULT 0.05, mort_w5 REAL DEFAULT 0.05, mort_w6 REAL DEFAULT 0.05,
            mort_w7 REAL DEFAULT 0.05, mort_w8 REAL DEFAULT 0.05,
            feed_w1 REAL DEFAULT 20,  feed_w2 REAL DEFAULT 45,  feed_w3 REAL DEFAULT 80,
            feed_w4 REAL DEFAULT 115, feed_w5 REAL DEFAULT 145, feed_w6 REAL DEFAULT 165,
            feed_w7 REAL DEFAULT 175, feed_w8 REAL DEFAULT 180
        );

        CREATE TABLE IF NOT EXISTS batch_daily_standards (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id    INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
            day_num     INTEGER NOT NULL,
            mort_std    REAL DEFAULT 0.05,
            feed_std    REAL DEFAULT 100,
            UNIQUE(batch_id, day_num)
        );

        CREATE TABLE IF NOT EXISTS cost_types (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            code       TEXT    NOT NULL UNIQUE,
            name_ar    TEXT    NOT NULL,
            category   TEXT    DEFAULT 'أخرى',
            has_qty    INTEGER DEFAULT 0,
            unit       TEXT,
            sort_order INTEGER DEFAULT 99,
            is_active  INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS batch_costs (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id     INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
            cost_type_id INTEGER NOT NULL REFERENCES cost_types(id),
            qty          REAL    DEFAULT 0,
            amount       REAL    DEFAULT 0,
            notes        TEXT    DEFAULT '',
            UNIQUE(batch_id, cost_type_id)
        );

        CREATE TABLE IF NOT EXISTS revenue_types (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            code       TEXT    NOT NULL UNIQUE,
            name_ar    TEXT    NOT NULL,
            category   TEXT    DEFAULT 'مبيعات',
            has_qty    INTEGER DEFAULT 0,
            unit       TEXT,
            sort_order INTEGER DEFAULT 99,
            is_active  INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS batch_revenues (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id        INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
            revenue_type_id INTEGER NOT NULL REFERENCES revenue_types(id),
            qty             REAL    DEFAULT 0,
            amount          REAL    DEFAULT 0,
            notes           TEXT    DEFAULT '',
            UNIQUE(batch_id, revenue_type_id)
        );
        """)

        for col_name, col_type in [
            ("fcr",              "REAL DEFAULT 0"),
            ("avg_weight",       "REAL DEFAULT 0"),
            ("batch_num",        "TEXT DEFAULT ''"),
            ("consumed_birds",   "INTEGER DEFAULT 0"),
            ("partner_name",     "TEXT DEFAULT ''"),
            ("feed_sale_qty",    "REAL DEFAULT 0"),
            ("feed_trans_r_qty", "REAL DEFAULT 0"),
            ("feed_rem_qty",     "REAL DEFAULT 0"),
            ("feed_rem_val",     "REAL DEFAULT 0"),
        ]:
            try: self.execute(f"ALTER TABLE batches ADD COLUMN {col_name} {col_type}")
            except: pass

        try: self.execute("ALTER TABLE farm_sales ADD COLUMN sale_type TEXT DEFAULT 'آجل'")
        except: pass
        try: self.execute("ALTER TABLE farm_sales ADD COLUMN sale_date TEXT DEFAULT ''")
        except: pass
        try: self.execute("ALTER TABLE market_sales ADD COLUMN sale_date TEXT DEFAULT ''")
        except: pass
        try: self.execute("ALTER TABLE batches ADD COLUMN fiscal_year INTEGER DEFAULT 0")
        except: pass

        self._seed_cost_types()
        self._seed_revenue_types()
        self._create_views()

    def _seed_cost_types(self):
        defaults = [
            ("chick_val",    "الكتاكيت",        "مواد",   1, "حبة",      1),
            ("feed_val",     "العلف",            "مواد",   1, "طن",       2),
            ("feed_trans",   "أجور نقل علف",     "نقل",    0, None,       3),
            ("sawdust_val",  "النشارة",          "مواد",   1, "م³",       4),
            ("water_val",    "الماء",            "مرافق",  0, None,       5),
            ("gas_val",      "الغاز",            "مرافق",  1, "أسطوانة",  6),
            ("drugs_val",    "علاجات وأدوية",    "صحة",    0, None,       7),
            ("wh_expenses",  "مصاريف عنبر",      "تشغيل",  0, None,       8),
            ("house_exp",    "مصاريف بيت",       "تشغيل",  0, None,       9),
            ("breeders_pay", "أجور مربيين",      "رواتب",  0, None,      10),
            ("qat_pay",      "قات مربيين",       "رواتب",  0, None,      11),
            ("rent_val",     "إيجار عنبر",       "عقارات", 0, None,      12),
            ("light_val",    "إضاءة وكهرباء",    "مرافق",  0, None,      13),
            ("sup_wh_pay",   "مشرف عنبر",        "إشراف",  0, None,      14),
            ("sup_co_pay",   "مشرف شركة",        "إشراف",  0, None,      15),
            ("sup_sale_pay", "مشرف بيع",         "إشراف",  0, None,      16),
            ("admin_val",    "إدارة وحسابات",    "إدارة",  0, None,      17),
            ("vaccine_pay",  "أجور لقاحات",      "صحة",    0, None,      18),
            ("delivery_val", "توصيل خدمات",      "أخرى",   0, None,      19),
            ("mixing_val",   "حمالة وخلط",       "أخرى",   0, None,      20),
            ("wash_val",     "تغسيل عنبر",       "أخرى",   0, None,      21),
            ("other_costs",  "مصاريف أخرى",      "أخرى",   0, None,      22),
        ]
        with self.get_conn() as conn:
            conn.executemany(""" INSERT OR IGNORE INTO cost_types
                    (code, name_ar, category, has_qty, unit, sort_order)
                VALUES (?, ?, ?, ?, ?, ?)
            """, defaults)
            conn.commit()

    def _seed_revenue_types(self):
        defaults = [
            ("offal_val",    "مبيعات ذبيل",     "مبيعات",  0, None,  1),
            ("feed_sale",    "مبيعات علف",       "مبيعات",  1, "كيس",  2),
            ("feed_trans_r", "علف منقول لعنابر", "تحويل",   1, "كيس",  3),
            ("feed_rem_val", "علف متبقي",        "مخزون",   1, "كيس",  4),
            ("drug_return",  "مرتجع علاجات",     "مرتجعات", 0, None,  5),
            ("gas_return",   "نقل غاز/نشارة",    "مرتجعات", 0, None,  6),
        ]
        with self.get_conn() as conn:
            conn.executemany(""" INSERT OR IGNORE INTO revenue_types
                    (code, name_ar, category, has_qty, unit, sort_order)
                VALUES (?, ?, ?, ?, ?, ?)
            """, defaults)
            conn.commit()

    def _create_views(self):
        self.execute_script(""" DROP VIEW IF EXISTS v_batches;
        CREATE VIEW v_batches AS
            SELECT b.*, w.name AS warehouse_name,
                   COALESCE(b.fiscal_year, CAST(strftime('%Y', b.date_in) AS INTEGER)) AS fy
            FROM   batches b
            JOIN   warehouses w ON b.warehouse_id = w.id;

        DROP VIEW IF EXISTS v_batch_costs_summary;
        CREATE VIEW v_batch_costs_summary AS
            SELECT bc.batch_id, ct.code, ct.name_ar, ct.category,
                   ct.has_qty, ct.unit, bc.qty, bc.amount, ct.sort_order
            FROM   batch_costs bc
            JOIN   cost_types ct ON bc.cost_type_id = ct.id;

        DROP VIEW IF EXISTS v_batch_revenues_summary;
        CREATE VIEW v_batch_revenues_summary AS
            SELECT br.batch_id, rt.code, rt.name_ar, rt.category,
                   rt.has_qty, rt.unit, br.qty, br.amount, rt.sort_order
            FROM   batch_revenues br
            JOIN   revenue_types rt ON br.revenue_type_id = rt.id;
        """)

    def get_cost_types(self, active_only=True):
        q = "SELECT * FROM cost_types"
        if active_only: q += " WHERE is_active=1"
        return self.fetch_all(q + " ORDER BY sort_order")

    def get_revenue_types(self, active_only=True):
        q = "SELECT * FROM revenue_types"
        if active_only: q += " WHERE is_active=1"
        return self.fetch_all(q + " ORDER BY sort_order")

    def get_batch_costs(self, batch_id):
        rows = self.fetch_all(""" SELECT ct.code, ct.name_ar, ct.category, ct.has_qty, ct.unit,
                   COALESCE(bc.qty, 0) AS qty,
                   COALESCE(bc.amount, 0) AS amount
            FROM   cost_types ct
            LEFT JOIN batch_costs bc
                   ON bc.cost_type_id = ct.id AND bc.batch_id = ?
            WHERE  ct.is_active = 1
            ORDER BY ct.sort_order
        """, (batch_id,))
        return {r["code"]: dict(r) for r in rows}

    def get_batch_revenues(self, batch_id):
        rows = self.fetch_all(""" SELECT rt.code, rt.name_ar, rt.category, rt.has_qty, rt.unit,
                   COALESCE(br.qty, 0) AS qty,
                   COALESCE(br.amount, 0) AS amount
            FROM   revenue_types rt
            LEFT JOIN batch_revenues br
                   ON br.revenue_type_id = rt.id AND br.batch_id = ?
            WHERE  rt.is_active = 1
            ORDER BY rt.sort_order
        """, (batch_id,))
        return {r["code"]: dict(r) for r in rows}

    def save_batch_costs(self, batch_id, costs_dict):
        with self.get_conn() as conn:
            for code, data in costs_dict.items():
                amount = float(data.get("amount") or 0)
                qty    = float(data.get("qty") or 0)
                ct = conn.execute("SELECT id FROM cost_types WHERE code=?", (code,)).fetchone()
                if not ct: continue
                if amount == 0 and qty == 0:
                    conn.execute("DELETE FROM batch_costs WHERE batch_id=? AND cost_type_id=?", (batch_id, ct["id"]))
                else:
                    conn.execute(""" INSERT INTO batch_costs (batch_id, cost_type_id, qty, amount)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(batch_id, cost_type_id)
                        DO UPDATE SET qty=excluded.qty, amount=excluded.amount
                    """, (batch_id, ct["id"], qty, amount))
            conn.commit()

    def save_batch_revenues(self, batch_id, revenues_dict):
        with self.get_conn() as conn:
            for code, data in revenues_dict.items():
                amount = float(data.get("amount") or 0)
                qty    = float(data.get("qty") or 0)
                rt = conn.execute("SELECT id FROM revenue_types WHERE code=?", (code,)).fetchone()
                if not rt: continue
                if amount == 0 and qty == 0:
                    conn.execute("DELETE FROM batch_revenues WHERE batch_id=? AND revenue_type_id=?", (batch_id, rt["id"]))
                else:
                    conn.execute(""" INSERT INTO batch_revenues (batch_id, revenue_type_id, qty, amount)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(batch_id, revenue_type_id)
                        DO UPDATE SET qty=excluded.qty, amount=excluded.amount
                    """, (batch_id, rt["id"], qty, amount))
            conn.commit()

    def add_cost_type(self, code, name_ar, category="أخرى", has_qty=False, unit=None):
        max_sort = self.fetch_one("SELECT COALESCE(MAX(sort_order),0)+1 AS s FROM cost_types")["s"]
        return self.execute(""" INSERT OR IGNORE INTO cost_types
                (code, name_ar, category, has_qty, unit, sort_order)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (code, name_ar, category, int(has_qty), unit, max_sort))

    def add_revenue_type(self, code, name_ar, category="مبيعات", has_qty=False, unit=None):
        max_sort = self.fetch_one("SELECT COALESCE(MAX(sort_order),0)+1 AS s FROM revenue_types")["s"]
        return self.execute(""" INSERT OR IGNORE INTO revenue_types
                (code, name_ar, category, has_qty, unit, sort_order)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (code, name_ar, category, int(has_qty), unit, max_sort))

db = DBManager(DB_PATH)
try:
    shared_ensure_schema(Path(DB_PATH))
except Exception:
    pass

# ════════════════════════════════════════════════════════════════
# نظام الحماية والتفعيل (Licensing System)
# ════════════════════════════════════════════════════════════════
SECRET_SALT = "POULTRY_PROTECTION_2026_V1"

class LicenseManager:
    @staticmethod
    def get_hwid():
        """استخراج معرف فريد للجهاز بناءً على المعالج والقرص الصلب"""
        try:
            # محاولة جلب الرقم التسلسلي للمعالج
            cpu = subprocess.check_output("wmic cpu get processorid", shell=True).decode().split('\n')[1].strip()
            # محاولة جلب الرقم التسلسلي للوحة الأم أو القرص
            board = subprocess.check_output("wmic baseboard get serialnumber", shell=True).decode().split('\n')[1].strip()
            if not board: # fallback to disk
                 board = subprocess.check_output("wmic diskdrive get serialnumber", shell=True).decode().split('\n')[1].strip()
            
            raw_id = f"{cpu}-{board}"
            return hashlib.sha256(raw_id.encode()).hexdigest()[:16].upper()
        except:
            # في حال الفشل نستخدم اسم المستخدم واسم الجهاز كبديل ضعيف
            alt_id = f"{os.getlogin()}-{os.getenv('COMPUTERNAME')}"
            return hashlib.sha256(alt_id.encode()).hexdigest()[:16].upper()

    @staticmethod
    def generate_key(hwid):
        """توليد كود التفعيل بناءً على هوية الجهاز والمفتاح السري"""
        return hashlib.sha256((hwid + SECRET_SALT).encode()).hexdigest()[:12].upper()

    @staticmethod
    def check_license():
        """التحقق من حالة الترخيص (فترة تجريبية أو تفعيل)"""
        # 1. التفعيل الكامل
        stored_key = db.get_setting("activation_key", "")
        hwid = LicenseManager.get_hwid()
        expected_key = LicenseManager.generate_key(hwid)
        
        if stored_key == expected_key:
            # تحقق من تاريخ الانتهاء (سنة واحدة)
            activation_date_str = db.get_setting("activation_date", "")
            if activation_date_str:
                act_date = datetime.strptime(activation_date_str, "%Y-%m-%d").date()
                if date.today() <= act_date + timedelta(days=365):
                    return "ACTIVE", None
                else:
                    return "EXPIRED", hwid
            return "ACTIVE", None # إذا لم يوجد تاريخ، نعتبره مفعلاً للأمان (أو يمكن تعيينه الآن)

        # 2. فترة التجربة (أسبوع واحد)
        install_date_str = db.get_setting("install_date", "")
        if not install_date_str:
            install_date_str = date.today().isoformat()
            db.set_setting("install_date", install_date_str)
        
        install_date = datetime.strptime(install_date_str, "%Y-%m-%d").date()
        days_passed = (date.today() - install_date).days
        
        if days_passed <= 7:
            return "TRIAL", 7 - days_passed
        
        return "REQUIRED", hwid

class ActivationWindow(ToplevelBase):
    def __init__(self, master, hwid, status="REQUIRED"):
        super().__init__(master)
        self.title("تنشيط النسخة - نظام إدارة الدواجن")
        self.geometry("500x400")
        center_window(self)
        self.hwid = hwid
        self.status = status
        self.success = False
        self.grab_set()
        self._build()

    def _build(self):
        # Header
        hdr = UIFrame(self, bg=CLR["header"], pady=15)
        hdr.pack(fill="x")
        title = "طلب تفعيل البرنامج" if self.status != "EXPIRED" else "انتهت صلاحية النسخة"
        UILabel(hdr, text=title, font=FT_TITLE, bg=CLR["header"], fg="white").pack()

        main = UIFrame(self, padding=20)
        main.pack(fill="both", expand=True)

        if self.status == "TRIAL":
            UILabel(main, text="انتهت الفترة التجريبية (7 أيام).", font=FT_BODY, fg=CLR["loss"]).pack(pady=5)
        elif self.status == "EXPIRED":
            UILabel(main, text="انتهى ترخيص السنة الواحدة. يرجى التجديد.", font=FT_BODY, fg=CLR["loss"]).pack(pady=5)
        
        UILabel(main, text="يرجى إرسال كود الجهاز للمالك للحصول على التفعيل:", font=FT_SMALL).pack(pady=(10,0))
        
        # HWID Display
        hwid_frame = UIFrame(main, relief="solid", borderwidth=1, padding=10)
        hwid_frame.pack(fill="x", pady=10)
        e_hwid = UIEntry(hwid_frame, font=("Courier", 12, "bold"), justify="center")
        e_hwid.insert(0, self.hwid)
        e_hwid.configure(state="readonly")
        e_hwid.pack(fill="x")

        # Key Entry
        UILabel(main, text="أدخل كود التفعيل هنا:", font=FT_HEADER).pack(pady=(15,0))
        self.v_key = tk.StringVar()
        self.e_key = UIEntry(main, textvariable=self.v_key, font=("Arial", 14), justify="center", width=20)
        self.e_key.pack(pady=10)
        self.e_key.focus_set()

        # Buttons
        btn_frm = UIFrame(main)
        btn_frm.pack(fill="x", pady=20)
        
        UIButton(btn_frm, text="✅ تفعيل الآن", command=self._activate, bootstyle="success", width=15).pack(side="right", padx=5)
        UIButton(btn_frm, text="❌ إغلاق", command=self.destroy, bootstyle="secondary", width=10).pack(side="left", padx=5)

    def _activate(self):
        key = self.v_key.get().strip().upper()
        expected = LicenseManager.generate_key(self.hwid)
        
        if key == expected:
            db.set_setting("activation_key", key)
            db.set_setting("activation_date", date.today().isoformat())
            messagebox.showinfo("نجاح", "تم تفعيل البرنامج بنجاح لمدة سنة كاملة!\nيرجى إعادة تشغيل البرنامج.", parent=self)
            self.success = True
            self.destroy()
        else:
            messagebox.showerror("خطأ", "كود التفعيل غير صحيح. يرجى التأكد من الكود والمحاولة مرة أخرى.", parent=self)

def check_and_run():
    # التحقق من الترخيص قبل فتح النافذة الرئيسية
    status, val = LicenseManager.check_license()
    
    if status == "ACTIVE":
        # البرنامج مفعّل، نفتح مباشرة
        app = MainWindow()
        app.mainloop()
    elif status == "TRIAL":
        # لا يزال في فترة التجربة
        res = messagebox.askyesno("فترة تجريبية", f"أنت تعمل حالياً في الفترة التجريبية. متبقي لك {val} أيام.\nهل تريد متابعة العمل بالنسخة التجريبية؟", icon='info')
        if res:
            app = MainWindow()
            app.mainloop()
        else:
            # إذا اختار لا، نفتح له نافذة التفعيل ليرى كود الجهاز
            root = WindowBase()
            root.withdraw()
            win = ActivationWindow(root, LicenseManager.get_hwid(), status="TRIAL_REMAINING")
            root.wait_window(win)
            if win.success:
                app = MainWindow()
                app.mainloop()
            else:
                root.destroy()
    else:
        # مطلوب تفعيل (انتهت التجربة أو انتهت السنة)
        root = WindowBase()
        root.withdraw()
        win = ActivationWindow(root, LicenseManager.get_hwid(), status=status)
        root.wait_window(win)
        if win.success:
            # بدلاً من فتح MainWindow مباشرة، الأفضل طلب إعادة التشغيل للتأكد من الحالة
            # ولكن لتجربة مستخدم أفضل سنفتحها
            app = MainWindow()
            app.mainloop()
        else:
            root.destroy()

def fmt_num(n, dec=0):
    try:
        n = float(n) if n else 0
        return f"{int(n):,}" if dec == 0 else f"{n:,.{dec}f}"
    except: return "—"

def tafqeet(n):
    """تحويل رقم صحيح إلى كتابة بالحروف العربية"""
    n = int(abs(n))
    if n == 0: return "صفر"
    ones = ["","واحد","اثنان","ثلاثة","أربعة","خمسة","ستة","سبعة","ثمانية","تسعة",
            "عشرة","أحد عشر","اثنا عشر","ثلاثة عشر","أربعة عشر","خمسة عشر",
            "ستة عشر","سبعة عشر","ثمانية عشر","تسعة عشر"]
    tens  = ["","عشرة","عشرون","ثلاثون","أربعون","خمسون","ستون","سبعون","ثمانون","تسعون"]
    hunds = ["","مائة","مئتان","ثلاثمائة","أربعمائة","خمسمائة","ستمائة","سبعمائة","ثمانمائة","تسعمائة"]
    def _say(num):
        if num == 0: return ""
        if num < 20: return ones[num]
        if num < 100:
            t, o = divmod(num, 10)
            return tens[t] if o == 0 else ones[o] + " و" + tens[t]
        h, r = divmod(num, 100)
        return hunds[h] if r == 0 else hunds[h] + " و" + _say(r)
    parts = []
    billions, n  = divmod(n, 1_000_000_000)
    millions,  n = divmod(n, 1_000_000)
    thousands, n = divmod(n, 1_000)
    if billions:  parts.append(_say(billions)  + " مليار")
    if millions:  parts.append(_say(millions)  + " مليون")
    if thousands: parts.append(_say(thousands) + " ألف")
    if n:         parts.append(_say(n))
    return " و".join(parts)


def lbl_entry(parent, text, row, col, width=16, readonly=False, colspan=1):
    UILabel(parent, text=text, font=FT_SMALL, bg=CLR["bg"], fg=CLR["text2"], anchor="e").grid(row=row, column=col, sticky="e", padx=(6,2), pady=8)
    v = tk.StringVar()
    state = "readonly" if readonly else "normal"
    bg = "#e9ecef" if readonly else CLR["white"]
    e = UIEntry(parent, textvariable=v, width=width, font=FT_BODY, state=state, bg=bg, relief="solid", highlightthickness=1, highlightbackground=CLR["border"])
    e.grid(row=row, column=col+1, sticky="ew", padx=(2,12), pady=8, columnspan=colspan)
    e.configure(justify="right")
    return v

def prepare_text(text):
    if not text: return ""
    if HAS_ARABIC: return get_display(arabic_reshaper.reshape(str(text)))
    return str(text)

def num_to_words_ar(n):
    """تحويل الأرقام إلى كلمات باللغة العربية (تفقيط مبسط)"""
    if n == 0: return "صفر"
    
    units = ["", "واحد", "اثنان", "ثلاثة", "أربعة", "خمسة", "ستة", "سبعة", "ثمانية", "تسعة", "عشرة",
             "أحد عشر", "اثنا عشر", "ثلاثة عشر", "أربعة عشر", "خمسة عشر", "ستة عشر", "سبعة عشر", "ثمانية عشر", "تسعة عشر"]
    tens = ["", "", "عشرون", "ثلاثون", "أربعون", "خمسون", "ستون", "سبعون", "ثمانون", "تسعون"]
    hundreds = ["", "مائة", "مائتان", "ثلاثمائة", "أربعمائة", "خمسمائة", "ستمائة", "سبعمائة", "ثمانمائة", "تسعمائة"]
    
    def _convert(n):
        if n < 20: return units[n]
        if n < 100:
            u = n % 10
            return (units[u] + " و" if u else "") + tens[n // 10]
        if n < 1000:
            remainder = n % 100
            return hundreds[n // 100] + (" و" if remainder else "") + _convert(remainder)
        if n < 2000:
            remainder = n % 1000
            return "ألف" + (" و" if remainder else "") + _convert(remainder)
        if n < 3000:
            remainder = n % 1000
            return "ألفان" + (" و" if remainder else "") + _convert(remainder)
        if n < 11000:
            remainder = n % 1000
            return _convert(n // 1000) + " آلاف" + (" و" if remainder else "") + _convert(remainder)
        if n < 1000000:
            remainder = n % 1000
            return _convert(n // 1000) + " ألف" + (" و" if remainder else "") + _convert(remainder)
        if n < 2000000:
            remainder = n % 1000000
            return "مليون" + (" و" if remainder else "") + _convert(remainder)
        if n < 3000000:
            remainder = n % 1000000
            return "مليونان" + (" و" if remainder else "") + _convert(remainder)
        if n < 11000000:
            remainder = n % 1000000
            return _convert(n // 1000000) + " ملايين" + (" و" if remainder else "") + _convert(remainder)
        remainder = n % 1000000
        return _convert(n // 1000000) + " مليون" + (" و" if remainder else "") + _convert(remainder)

    try:
        res = _convert(int(abs(n)))
        # تحسينات بسيطة للنحو العربي
        res = res.replace("اثنان ألف", "ألفان")
        res = res.replace("اثنان مليون", "مليونان")
        return res.strip()
    except: return str(n)

def make_backup():
    if not os.path.exists(DB_PATH): return None
    bk_dir = os.path.join(BASE_DIR, "backups")
    os.makedirs(bk_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(bk_dir, f"poultry_data_{ts}.db")
    with sqlite3.connect(DB_PATH) as src_conn, sqlite3.connect(dest) as dst_conn:
        src_conn.backup(dst_conn)
    files = sorted([f for f in os.listdir(bk_dir) if f.startswith("poultry_data_") and f.endswith(".db")])
    while len(files) > 10: os.remove(os.path.join(bk_dir, files.pop(0)))
    return dest

def restore_db(backup_path):
    """استرجاع قاعدة البيانات من ملف خارجي"""
    if not os.path.exists(backup_path): return False, "الملف غير موجود"
    try:
        # التأكد من أنه ملف sqlite صالح (اختياري ولكن مفيد)
        with sqlite3.connect(backup_path) as conn:
            conn.execute("SELECT name FROM sqlite_master LIMIT 1")
        
        with sqlite3.connect(backup_path) as src_conn, sqlite3.connect(DB_PATH) as dst_conn:
            src_conn.backup(dst_conn)
        return True, "تم الاسترجاع بنجاح"
    except Exception as e:
        return False, str(e)

def calc_active_birds_for_batch(batch_id, chicks, up_to_date):
    """حساب عدد التدفق الحالي للطيور الحية في تاريخ معين"""
    dead = db.fetch_one("SELECT COALESCE(SUM(dead_count),0) AS d FROM daily_records WHERE batch_id=? AND rec_date<=?", (batch_id, up_to_date))["d"]
    farm_sold = db.fetch_one("SELECT COALESCE(SUM(qty),0) AS q FROM farm_sales WHERE batch_id=? AND sale_date!='' AND sale_date<=?", (batch_id, up_to_date))["q"]
    mkt_sold = db.fetch_one("SELECT COALESCE(SUM(qty_sold),0) AS q FROM market_sales WHERE batch_id=? AND sale_date!='' AND sale_date<=?", (batch_id, up_to_date))["q"]
    consumed = db.fetch_one("SELECT consumed_birds FROM batches WHERE id=?", (batch_id,))["consumed_birds"] or 0
    return max(0, chicks - dead - farm_sold - mkt_sold - consumed)

def send_telegram(token, chat_id, message):
    """إرسال رسالة تليجرام باستخدام urllib القياسية"""
    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        data = urllib.parse.urlencode({"chat_id": chat_id, "text": message, "parse_mode": "HTML"}).encode("utf-8")
        req = urllib.request.Request(url, data=data)
        with urllib.request.urlopen(req, timeout=10) as response:
            return True, response.read().decode()
    except Exception as e:
        return False, str(e)

def check_and_send_smart_alert(batch_id, rec_date, day_num, dead, feed, notes):
    """تحليل البيانات والمقارنة مع معايير Ross 308 وإرسال تنبيه إذا لزم الأمر"""
    token = db.get_setting("tel_token")
    chat_id = db.get_setting("tel_chat_id")
    if not token or not chat_id: return False

    b = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (batch_id,))
    if not b: return False
    
    # جلب المعايير من الجدول اليومي الجديد أولاً
    daily_std = db.fetch_one("SELECT mort_std, feed_std FROM batch_daily_standards WHERE batch_id=? AND day_num=?", (batch_id, day_num))
    if daily_std:
        m_val = daily_std["mort_std"]
        f_val = daily_std["feed_std"]
    else:
        # إذا لم يوجد تخصيص لليوم، نستخدم القيمة القياسية لـ Ross 308
        m_val = ROSS_DAILY_M.get(day_num, 0.05)
        f_val = ROSS_DAILY_F.get(day_num, 100)

    active = calc_active_birds_for_batch(batch_id, b["chicks"], rec_date)
    exp_dead = active * m_val / 100
    exp_feed = active * f_val / 1000
    
    alerts = []
    if dead > exp_dead * 1.5 and dead > 0:
        alerts.append(f"⚰ <b>انحراف النافق:</b> {dead} (الطبيعي {exp_dead:.1f})")
    if feed > 0 and feed < exp_feed * 0.85:
        alerts.append(f"🌡 <b>انخفاض العلف:</b> {feed:.1f} كجم (المستهدف {exp_feed:.1f})")
    elif feed > exp_feed * 1.2:
        alerts.append(f"🌾 <b>زيادة/هدر علف:</b> {feed:.1f} كجم (المستهدف {exp_feed:.1f})")

    if alerts:
        msg = [f"🚨 <b>تنبيه فني ذكي: {b['warehouse_name']}</b>",
               f"📅 التاريخ: {rec_date} (اليوم {day_num})",
               f"📋 الحالة: " + " | ".join(alerts),
               f"🐥 الكتل النشطة: {active:,} طائر"]
        if notes: msg.append(f"📝 ملاحظة: {notes}")
        msg.append(f"\n<i>تم التحليل بناءً على معايير Ross 308.</i>")
        return send_telegram(token, chat_id, "\n".join(msg))
    return False

def build_daily_telegram_report():
    """تجهيز محتوى التقرير اليومي لجميع العنابر النشطة"""
    report_date = date.today().isoformat()
    yesterday   = (date.today() - timedelta(days=1)).isoformat()
    
    active_batches = db.fetch_all(
        "SELECT * FROM v_batches WHERE date_in<=? AND date_out>=? ORDER BY warehouse_name",
        (report_date, report_date))
    
    if not active_batches:
        return f"📊 <b>تقرير الدواجن اليومي</b>\nتاريخ: {report_date}\n\nلا توجد عنابر نشطة حالياً."

    lines = [f"📊 <b>تقرير الدواجن اليومي</b>\nتاريخ: {report_date}\n" + "─" * 20]
    
    for b in active_batches:
        bid = b["id"]
        # يوم التربية
        try:
            d_in = datetime.strptime(b["date_in"], "%Y-%m-%d")
            day_num = (datetime.now() - d_in).days + 1
        except: day_num = "?"
        
        active = calc_active_birds_for_batch(bid, b["chicks"] or 0, report_date)
        
        # إحصائيات أمس (أو آخر سجل)
        rec = db.fetch_one(
            "SELECT dead_count, feed_kg FROM daily_records WHERE batch_id=? AND rec_date=?",
            (bid, yesterday))
        
        dead_y = rec["dead_count"] if rec else 0
        feed_y = rec["feed_kg"]    if rec else 0
        
        line = (f"🏠 <b>{b['warehouse_name']}</b> (دفعة {b['batch_num'] or bid})\n"
                f"📅 اليوم: {day_num} | 🐥 الكتل: {active:,}\n"
                f"⚰️ نافق أمس: {dead_y} | 🌾 علف أمس: {feed_y} كجم\n")
        lines.append(line)
        
    lines.append("<i>تم الإرسال تلقائياً من المنظومة المطورة.</i>")
    return "\n".join(lines)

class DailyRecordsWindow(ToplevelBase):
    def __init__(self, master, batch_id, batch_info):
        super().__init__(master)
        self.batch_id   = batch_id
        self.batch_info = batch_info
        b_num = batch_info.get("batch_num") or batch_id
        self.title(f"السجلات اليومية — {batch_info.get('warehouse_name','')} — دفعة {b_num}")
        self.geometry("1300x680")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.grab_set()
        self._load_standards()
        self._build()
        self._load()

    def _load_standards(self):
        # تحميل Target FCR من الجدول القديم للتوافق
        std_old = db.fetch_one("SELECT target_fcr FROM batch_standards WHERE batch_id=?", (self.batch_id,))
        self.target_fcr = std_old["target_fcr"] if std_old else 1.8
        
        # تحميل المعايير اليومية
        rows = db.fetch_all("SELECT day_num, mort_std, feed_std FROM batch_daily_standards WHERE batch_id=? ORDER BY day_num", (self.batch_id,))
        if rows:
            self.mort_std = {r["day_num"]: r["mort_std"] for r in rows}
            self.feed_std = {r["day_num"]: r["feed_std"] for r in rows}
        else:
            # إذا لم توجد، نستخدم الافتراضيات العالمية ونقوم بحفظها لتسهيل التعديل
            self.mort_std = dict(ROSS_DAILY_M)
            self.feed_std = dict(ROSS_DAILY_F)
            with db.get_conn() as conn:
                for d in range(1, 66):
                    conn.execute("INSERT OR IGNORE INTO batch_daily_standards (batch_id, day_num, mort_std, feed_std) VALUES (?, ?, ?, ?)",
                                 (self.batch_id, d, self.mort_std.get(d, 0.05), self.feed_std.get(d, 100)))
                conn.commit()

    def _get_week(self, day_num):
        return min(max(1, (int(day_num or 1) - 1) // 7 + 1), 8)

    def _calc_active_birds(self, up_to_date):
        try:
            chicks = self.batch_info.get("chicks", 0) or 0
            dead = db.fetch_one("SELECT COALESCE(SUM(dead_count),0) AS d FROM daily_records WHERE batch_id=? AND rec_date<=?", (self.batch_id, up_to_date))["d"]
            
            # محاولة جلب المبيعات مع التعامل مع احتمال غياب عمود sale_date في بعض النسخ القديمة
            try:
                farm_sold = db.fetch_one("SELECT COALESCE(SUM(qty),0) AS q FROM farm_sales WHERE batch_id=? AND sale_date!='' AND sale_date<=?", (self.batch_id, up_to_date))["q"]
                mkt_sold = db.fetch_one("SELECT COALESCE(SUM(qty_sold),0) AS q FROM market_sales WHERE batch_id=? AND sale_date!='' AND sale_date<=?", (self.batch_id, up_to_date))["q"]
            except:
                # إذا فشل الاستعلام بسبب نقص الأعمدة، نعتبر المبيعات صفراً لهذا الحساب التقني
                farm_sold = 0
                mkt_sold = 0
                
            return max(0, chicks - dead - farm_sold - mkt_sold)
        except:
            return self.batch_info.get("chicks", 0) or 0

    def _build(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=8)
        hdr.pack(fill="x")
        b_num = self.batch_info.get("batch_num") or self.batch_id
        UILabel(hdr, text=f"📅 السجلات اليومية الفنية — الدفعة رقم {b_num}", font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)
        UIButton(hdr, text="⚙️ معايير FCR والنافق", font=FT_SMALL, bg="#2c6fad", fg="white", relief="flat", cursor="hand2", command=self._open_standards).pack(side="left", padx=10)

        inp = UILabelFrame(self, text="إضافة / تعديل سجل يومي", font=FT_HEADER, bg=CLR["daily_bg"], fg=CLR["accent"], padx=10, pady=6)
        inp.pack(fill="x", padx=10, pady=6)

        UILabel(inp, text="التاريخ:", font=FT_SMALL, bg=CLR["daily_bg"]).grid(row=0,column=0, sticky="e", padx=4)
        self.v_date = tk.StringVar(value=date.today().isoformat())
        UIEntry(inp, textvariable=self.v_date, width=13, font=FT_BODY, relief="solid").grid(row=0,column=1, padx=4)

        UILabel(inp, text="اليوم رقم:", font=FT_SMALL, bg=CLR["daily_bg"]).grid(row=0,column=2, sticky="e", padx=4)
        self.v_daynum = tk.StringVar()
        UIEntry(inp, textvariable=self.v_daynum, width=5, font=FT_BODY, relief="solid").grid(row=0,column=3, padx=4)

        UILabel(inp, text="النافق الواقعي:", font=FT_SMALL, bg=CLR["daily_bg"]).grid(row=0,column=4, sticky="e", padx=4)
        self.v_dead = tk.StringVar(value="0")
        UIEntry(inp, textvariable=self.v_dead, width=7, font=FT_BODY, relief="solid").grid(row=0,column=5, padx=4)

        UILabel(inp, text="العلف الواقعي (كجم):", font=FT_SMALL, bg=CLR["daily_bg"]).grid(row=0,column=6, sticky="e", padx=4)
        self.v_feed = tk.StringVar(value="0")
        UIEntry(inp, textvariable=self.v_feed, width=9, font=FT_BODY, relief="solid").grid(row=0,column=7, padx=4)

        UILabel(inp, text="ملاحظة:", font=FT_SMALL, bg=CLR["daily_bg"]).grid(row=1,column=0, sticky="e", padx=4, pady=4)
        self.v_notes = tk.StringVar()
        UIEntry(inp, textvariable=self.v_notes, width=45, font=FT_BODY, relief="solid").grid(row=1,column=1, columnspan=6, padx=4, sticky="ew")

        btn_frm = UIFrame(inp, bg=CLR["daily_bg"])
        btn_frm.grid(row=1, column=7, padx=4)
        UIButton(btn_frm, text="💾 حفظ", font=FT_BODY, bg=CLR["nav"], fg="white", cursor="hand2", relief="flat", padx=8, command=self._save_record).pack(side="right", padx=2)
        UIButton(btn_frm, text="🗑 حذف", font=FT_BODY, bg=CLR["loss_bg"], fg=CLR["loss"], cursor="hand2", relief="flat", padx=8, command=self._del_record).pack(side="right", padx=2)

        cols = ("التاريخ","اليوم","الأسبوع","الكتل النشطة","نافق واقعي","نافق طبيعي","فرق النافق","علف واقعي كجم","علف مطلوب كجم","فرق العلف","تراكم النافق","تراكم العلف","ملاحظة")
        frm = UIFrame(self, bg=CLR["bg"])
        frm.pack(fill="both", expand=True, padx=10, pady=4)
        self.tree = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")

        widths = [95,55,65,95,90,90,85,100,100,85,100,100,160]
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c, anchor="center")
            self.tree.column(c, width=w, anchor="center")

        self.tree.tag_configure("over_dead",  background="#fce4d6")
        self.tree.tag_configure("over_feed",  background="#fff2cc")
        self.tree.tag_configure("both_over",  background="#f4cccc")
        self.tree.tag_configure("normal",     background="#f0f9ea")

        sb = ttk.Scrollbar(frm, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="left", fill="y")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        sumfrm = UIFrame(self, bg=CLR["info_bg"], pady=5, padx=10)
        sumfrm.pack(fill="x", padx=10, pady=3)
        self.lbl_summary = UILabel(sumfrm, text="", font=FT_SMALL, bg=CLR["info_bg"], fg=CLR["accent"])
        self.lbl_summary.pack(side="right")

        bot = UIFrame(self, bg=CLR["bg"], pady=4)
        bot.pack(fill="x")
        UIButton(bot, text="📥 تصدير Excel فني", font=FT_BODY, bg=CLR["profit_bg"], fg=CLR["profit"], cursor="hand2", relief="flat", padx=10, pady=4, command=self._export_excel).pack(side="right", padx=10)
        UIButton(bot, text="📄 تصدير PDF فني", font=FT_BODY, bg=CLR["info_bg"], fg=CLR["accent"], cursor="hand2", relief="flat", padx=10, pady=4, command=self._export_pdf).pack(side="right", padx=5)

    def _load(self, focus_date=None):
        rows = db.fetch_all("SELECT rec_date, day_num, dead_count, feed_kg, notes FROM daily_records WHERE batch_id=? ORDER BY rec_date", (self.batch_id,))
        self.tree.delete(*self.tree.get_children())

        chicks = self.batch_info.get("chicks", 0) or 0
        cum_dead = 0; cum_feed = 0.0
        total_exp_dead = 0; total_exp_feed = 0.0

        for r in rows:
            day_num = r["day_num"] or 1
            week    = self._get_week(day_num)
            active  = self._calc_active_birds(r["rec_date"])

            exp_dead = round(active * self.mort_std.get(day_num, 0.05) / 100, 1)
            exp_feed = round(active * self.feed_std.get(day_num, 100) / 1000, 1)

            actual_dead = r["dead_count"]
            actual_feed = r["feed_kg"]

            diff_dead = actual_dead - exp_dead
            diff_feed = round(actual_feed - exp_feed, 1)

            cum_dead += actual_dead
            cum_feed += actual_feed
            total_exp_dead += exp_dead
            total_exp_feed += exp_feed

            dead_over = diff_dead > 0
            feed_over = diff_feed > 0
            if dead_over and feed_over: tag = "both_over"
            elif dead_over:             tag = "over_dead"
            elif feed_over:             tag = "over_feed"
            else:                       tag = "normal"

            diff_dead_str = f"+{diff_dead:.1f}" if diff_dead > 0 else f"{diff_dead:.1f}"
            diff_feed_str = f"+{diff_feed:.1f}" if diff_feed > 0 else f"{diff_feed:.1f}"

            self.tree.insert("", "end", iid=r["rec_date"], tags=(tag,),
                values=(r["rec_date"], day_num, f"أ{week}", fmt_num(active), actual_dead, f"{exp_dead:.1f}", diff_dead_str, fmt_num(actual_feed,1), f"{exp_feed:.1f}", diff_feed_str, cum_dead, fmt_num(cum_feed,1), r["notes"] or ""))

        mort_pct = cum_dead / chicks * 100 if chicks > 0 else 0
        feed_eff = f"  |  إجمالي فرق النافق: {cum_dead-total_exp_dead:+.0f}  |  إجمالي فرق العلف: {cum_feed-total_exp_feed:+.0f} كجم"
        self.lbl_summary.config(text=f"النافق الكلي: {cum_dead:,} ({mort_pct:.2f}%)  |  إجمالي العلف: {fmt_num(cum_feed,1)} كجم{feed_eff}")
        
        # التمرير التلقائي للسجل المختار دون اختياره (لتجنب تفعيل _on_select)
        if focus_date and self.tree.exists(focus_date):
            self.tree.see(focus_date)
        elif not focus_date and self.tree.get_children():
            last_id = self.tree.get_children()[-1]
            self.tree.see(last_id)

    def _on_select(self, _=None):
        sel = self.tree.selection()
        if not sel: return
        r = db.fetch_one("SELECT * FROM daily_records WHERE batch_id=? AND rec_date=?", (self.batch_id, sel[0]))
        if r:
            self.v_date.set(r["rec_date"])
            self.v_daynum.set(str(r["day_num"] or ""))
            self.v_dead.set(str(r["dead_count"]))
            self.v_feed.set(str(r["feed_kg"]))
            self.v_notes.set(r["notes"] or "")

    def _save_record(self):
        rec_date = self.v_date.get().strip()
        if not rec_date: return messagebox.showwarning("تنبيه", "يرجى إدخال التاريخ", parent=self)
        try:
            dead   = int(self.v_dead.get() or 0)
            feed   = float(self.v_feed.get() or 0)
            daynum = int(self.v_daynum.get() or 0)
        except ValueError: return messagebox.showerror("خطأ", "القيم يجب أن تكون أرقاماً", parent=self)
        
        notes_saved = self.v_notes.get()

        db.execute("""INSERT INTO daily_records (batch_id, rec_date, day_num, dead_count, feed_kg, notes) VALUES (?,?,?,?,?,?) ON CONFLICT(batch_id, rec_date) DO UPDATE SET day_num=excluded.day_num, dead_count=excluded.dead_count, feed_kg=excluded.feed_kg, notes=excluded.notes""", 
            (self.batch_id, rec_date, daynum, dead, feed, self.v_notes.get()))

        row = db.fetch_one("SELECT SUM(dead_count) AS s FROM daily_records WHERE batch_id=?", (self.batch_id,))
        total_dead = row["s"] if row and row["s"] else 0
        chicks = self.batch_info.get("chicks", 0) or 0
        mort_rate = round(total_dead / chicks * 100, 2) if chicks > 0 else 0
        db.execute("UPDATE batches SET total_dead=?, mort_rate=? WHERE id=?", (total_dead, mort_rate, self.batch_id))
        self._load(focus_date=rec_date)

        # ── تحديث الحقول فوراً لليوم التالي ────────────────────
        try:
            curr = datetime.strptime(rec_date, "%Y-%m-%d")
            self.v_date.set((curr + timedelta(days=1)).strftime("%Y-%m-%d"))
        except: pass
        try: self.v_daynum.set(str(daynum + 1))
        except: pass
        self.v_dead.set("0"); self.v_feed.set("0"); self.v_notes.set("")

        # ── تحذير ذكي عند تجاوز المعدل الطبيعي ──────────────────
        if daynum > 0:
            active  = self._calc_active_birds(rec_date)
            exp_dead = round(active * self.mort_std.get(daynum, 0.05) / 100, 1)
            exp_feed = round(active * self.feed_std.get(daynum, 100) / 1000, 1)
            warnings = []
            if dead > exp_dead * 1.5 and dead > 0:
                warnings.append(f"⚠️ النافق اليوم ({dead}) أعلى من الطبيعي ({exp_dead:.1f})")
            if feed > 0 and exp_feed > 0 and feed > exp_feed * 1.3:
                warnings.append(f"⚠️ العلف ({feed:.1f} كجم) أعلى من المطلوب ({exp_feed:.1f} كجم)")
            if chicks > 0 and mort_rate > 8:
                warnings.append(f"🚨 النافق الكلي ({mort_rate:.1f}%) تجاوز 8%")
            
            if warnings:
                messagebox.showwarning("تنبيه فني",
                    "\n".join(warnings) + f"\n\nالكتل النشطة: {active:,} طائر",
                    parent=self)
            
            # إرسال تنبيه تليجرام ذكي
            check_and_send_smart_alert(self.batch_id, rec_date, daynum, dead, feed, notes_saved)

    def _del_record(self):
        sel = self.tree.selection()
        if not sel: return
        if not messagebox.askyesno("تأكيد", f"حذف سجل يوم {sel[0]}؟", parent=self): return
        db.execute("DELETE FROM daily_records WHERE batch_id=? AND rec_date=?", (self.batch_id, sel[0]))
        self._load()

    def _open_standards(self):
        win = ToplevelBase(self)
        win.title("⚙️ معايير الأداء اليومي (Ross 308)")
        win.geometry("700x750")
        center_window(win)
        if not HAS_TTKB: win.configure(bg=CLR["bg"])
        win.grab_set()

        hdr = UIFrame(win, bg=CLR["header"], pady=8); hdr.pack(fill="x")
        UILabel(hdr, text="⚙️ معايير النافق اليومي والعلف المطلوب (1-65 يوم)", font=FT_HEADER, bg=CLR["header"], fg="white").pack(side="right", padx=14)
        
        main_frm = UIFrame(win, bg=CLR["bg"], padx=10, pady=5); main_frm.pack(fill="both", expand=True)

        # FCR المستهدف
        fcr_frm = UIFrame(main_frm, bg=CLR["bg"])
        fcr_frm.pack(fill="x", pady=5)
        UILabel(fcr_frm, text="FCR المستهدف للدفعة:", font=FT_BODY, bg=CLR["bg"]).pack(side="right", padx=5)
        v_fcr = tk.StringVar(value=str(self.target_fcr))
        UIEntry(fcr_frm, textvariable=v_fcr, width=8, font=FT_BODY, relief="solid").pack(side="right", padx=10)
        
        def _reset_ross():
            if messagebox.askyesno("تأكيد", "هل تريد حقاً استعادة كافة قيم معايير Ross 308 القياسية؟"):
                for d in range(1, 66):
                    mort_vars[d].set(str(ROSS_DAILY_M.get(d, 0.05)))
                    feed_vars[d].set(str(ROSS_DAILY_F.get(d, 100)))

        UIButton(fcr_frm, text="♻️ استعادة معايير Ross 308", font=FT_SMALL, command=_reset_ross, bootstyle="outline-primary").pack(side="left", padx=10)

        # حاوية التمرير للأيام
        canvas_frm = UIFrame(main_frm)
        canvas_frm.pack(fill="both", expand=True, pady=10)
        
        canvas = tk.Canvas(canvas_frm, bg=CLR["bg"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frm, orient="vertical", command=canvas.yview)
        scroll_content = UIFrame(canvas, bg=CLR["bg"])

        scroll_content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # العناوين
        th_frm = UIFrame(scroll_content, bg="#eef2f7")
        th_frm.pack(fill="x", pady=2)
        UILabel(th_frm, text="العمر (يوم)", font=FT_SMALL, width=12).pack(side="right")
        UILabel(th_frm, text="نافق طبيعي %", font=FT_SMALL, width=18, fg=CLR["loss"]).pack(side="right")
        UILabel(th_frm, text="علف مطلوب (غرام)", font=FT_SMALL, width=22, fg=CLR["accent"]).pack(side="right")

        mort_vars = {}; feed_vars = {}
        for d in range(1, 66):
            row_frm = UIFrame(scroll_content, bg=CLR["bg"])
            row_frm.pack(fill="x", pady=1)
            UILabel(row_frm, text=f"اليوم {d}", font=FT_SMALL, width=12, anchor="e").pack(side="right")
            
            mv = tk.StringVar(value=str(self.mort_std.get(d, 0.05)))
            fv = tk.StringVar(value=str(self.feed_std.get(d, 100)))
            
            UIEntry(row_frm, textvariable=mv, width=12, font=FT_SMALL, justify="center").pack(side="right", padx=15)
            UIEntry(row_frm, textvariable=fv, width=15, font=FT_SMALL, justify="center").pack(side="right", padx=10)
            
            mort_vars[d] = mv; feed_vars[d] = fv

        def _save_std():
            try:
                fcr = float(v_fcr.get())
                m_vals = {d: float(mort_vars[d].get()) for d in range(1, 66)}
                f_vals = {d: float(feed_vars[d].get()) for d in range(1, 66)}
            except ValueError:
                return messagebox.showerror("خطأ", "يجب إدخال قيم عددية صحيحة لكافة الحقول", parent=win)

            with db.get_conn() as conn:
                # حفظ FCR
                conn.execute("UPDATE batch_standards SET target_fcr=? WHERE batch_id=?", (fcr, self.batch_id)) or \
                conn.execute("INSERT OR IGNORE INTO batch_standards (batch_id, target_fcr) VALUES (?, ?)", (self.batch_id, fcr))
                
                # حفظ المعايير اليومية
                for d in range(1, 66):
                    conn.execute("""INSERT INTO batch_daily_standards (batch_id, day_num, mort_std, feed_std) 
                        VALUES (?, ?, ?, ?) ON CONFLICT(batch_id, day_num) 
                        DO UPDATE SET mort_std=excluded.mort_std, feed_std=excluded.feed_std""",
                        (self.batch_id, d, m_vals[d], f_vals[d]))
                conn.commit()

            self._load_standards()
            self._load()
            win.destroy()
            messagebox.showinfo("تم", "تم حفظ المعايير اليومية بنجاح!", parent=self)

        UIButton(win, text="💾 حفظ التعديلات", font=FT_HEADER, padding=10, command=_save_std).pack(fill="x", side="bottom")

    def _export_excel(self):
        if not HAS_OPENPYXL: return messagebox.showerror("خطأ", "مكتبة openpyxl غير مثبتة", parent=self)
        b_num = self.batch_info.get("batch_num") or self.batch_id
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile=f"السجل_الفني_دفعة_{str(b_num).replace("/","-")}.xlsx", parent=self)
        if not path: return

        rows = db.fetch_all("SELECT rec_date, day_num, dead_count, feed_kg, notes FROM daily_records WHERE batch_id=? ORDER BY rec_date", (self.batch_id,))
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "السجل الفني"; ws.sheet_view.rightToLeft = True

        hdrs = ["التاريخ","اليوم","الأسبوع","الكتل النشطة","نافق واقعي","نافق طبيعي","فرق النافق","علف واقعي كجم","علف مطلوب كجم","فرق العلف","تراكم النافق","تراكم العلف","ملاحظة"]
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        red_fill = PatternFill("solid", fgColor="FCE4D6")
        yel_fill = PatternFill("solid", fgColor="FFF2CC")
        grn_fill = PatternFill("solid", fgColor="E2EFDA")

        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(1, ci, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(ci)].width = 14

        chicks = self.batch_info.get("chicks", 0) or 0
        cum_dead = 0; cum_feed = 0.0

        for ri, r in enumerate(rows, 2):
            day_num = r["day_num"] or 1
            week    = self._get_week(day_num)
            active  = self._calc_active_birds(r["rec_date"])
            exp_dead = round(active * self.mort_std.get(day_num, 0.05) / 100, 1)
            exp_feed = round(active * self.feed_std.get(day_num, 100) / 1000, 1)
            cum_dead += r["dead_count"]; cum_feed += r["feed_kg"]
            diff_dead = round(r["dead_count"] - exp_dead, 1)
            diff_feed = round(r["feed_kg"] - exp_feed, 1)

            vals = [r["rec_date"], day_num, f"أسبوع {week}", active, r["dead_count"], exp_dead, diff_dead, round(r["feed_kg"],1), exp_feed, diff_feed, cum_dead, round(cum_feed,1), r["notes"] or ""]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(ri, ci, v)
                cell.alignment = Alignment(horizontal="center")

            dead_over = diff_dead > 0; feed_over = diff_feed > 0
            if dead_over and feed_over: row_fill = red_fill
            elif dead_over:             row_fill = PatternFill("solid", fgColor="FCE4D6")
            elif feed_over:             row_fill = yel_fill
            else:                       row_fill = grn_fill
            for ci in range(1, len(vals)+1):
                ws.cell(ri, ci).fill = row_fill

        try:
            wb.save(path)
            messagebox.showinfo("تم", "تم التصدير الفني بنجاح!", parent=self)
        except PermissionError:
            messagebox.showerror("خطأ", "الملف مفتوح في Excel! أغلقه أولاً.", parent=self)
        except Exception as _ex:
            messagebox.showerror("خطأ", str(_ex), parent=self)

    def _export_pdf(self):
        if not HAS_FPDF:
            return messagebox.showerror("خطأ", "مكتبة fpdf غير مثبتة.\nيرجى التثبيت عبر: pip install fpdf2", parent=self)
        
        b_num = self.batch_info.get("batch_num") or self.batch_id
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")], initialfile=f"السجل_الفني_دفعة_{str(b_num).replace('/','-')}.pdf", parent=self)
        if not path: return

        company  = db.get_setting("company_name", "شركة آفاق الريف للدواجن")
        font_path = AMIRI_FONT_PATH

        # إعداد ملف PDF (P: Portrait, unit: mm, format: A4)
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(left=10, top=10, right=10)
        pdf.add_page()
        
        try:
            pdf.add_font("Arabic", "", font_path, uni=True)
        except: pass
            
        # ── الرأس ──
        logo_path = LOGO_PATH
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=10, y=8, w=22)
        
        pdf.set_font("Arabic", "", 16)
        pdf.cell(0, 10, prepare_text(company), ln=True, align="C")
        pdf.set_font("Arabic", "", 14)
        pdf.cell(0, 8, prepare_text("السجل الفني اليومي للدفعة"), ln=True, align="C")
        
        pdf.set_font("Arabic", "", 11)
        wh_name = self.batch_info.get("warehouse_name", "")
        chicks = self.batch_info.get("chicks", 0) or 0
        info_str = f"العنبر: {wh_name}  |  الدفعة: {b_num}  |  العدد: {chicks:,} طائر"
        pdf.cell(0, 7, prepare_text(info_str), ln=True, align="C")
        pdf.ln(4)

        # ── إعدادات الأعمدة (RTL بترتيب من اليسار لليمين على الورقة ليظهر التاريخ يميناً) ──
        # الأعمدة المطلوبة: علف متراكم، نافق متراكم، فرق العلف، علف مطلوب، علف واقعي، فرق النافق، نافق طبيعي، نافق واقعي، كتل نشطة، أسبوع، يوم، التاريخ
        cols   = ["علف متراكم (كيس)", "نافق متراكم", "فرق العلف", "علف مطلوب", "علف واقعي", "فرق النافق", "نافق طبيعي", "نافق واقعي", "الكتل النشطة", "أسبوع", "يوم", "التاريخ"]
        widths = [18, 17, 13, 18, 18, 14, 14, 14, 18, 9, 9, 28] # الإجمالي 190mm
        
        # دالة لطباعة صف بعناوين ملتفة (Multi-line Headers)
        def draw_header_row():
            pdf.set_fill_color(31, 78, 121); pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arabic", "", 8)
            
            # حساب أقصى ارتفاع محتمل للعناوين
            max_h = 8
            x_start = pdf.get_x()
            y_start = pdf.get_y()
            
            # رسم العناوين وحساب الارتفاع
            for w, h in zip(widths, cols):
                # نستخدم روتين مؤقت لحساب عدد الأسطر
                lines = len(pdf.multi_cell(w, 4, prepare_text(h), split_only=True))
                if lines * 4 > max_h: max_h = lines * 4
            
            # الآن نقوم بالرسم الفعلي بارتفاع موحد
            pdf.set_xy(x_start, y_start)
            for w, h in zip(widths, cols):
                curr_x = pdf.get_x()
                curr_y = pdf.get_y()
                pdf.multi_cell(w, max_h/len(pdf.multi_cell(w, 4, prepare_text(h), split_only=True)) if len(pdf.multi_cell(w, 4, prepare_text(h), split_only=True))>0 else max_h, prepare_text(h), 1, "C", True)
                pdf.set_xy(curr_x + w, curr_y)
            pdf.ln(max_h)
            pdf.set_text_color(0, 0, 0)

        draw_header_row()
        
        # جلب البيانات
        rows = db.fetch_all("SELECT rec_date, day_num, dead_count, feed_kg, notes FROM daily_records WHERE batch_id=? ORDER BY rec_date", (self.batch_id,))
        
        cum_dead = 0; cum_feed = 0.0
        tot_active = 0
        tot_diff_dead = 0.0; tot_diff_feed = 0.0
        count_rows = 0

        for r in rows:
            day_num = r["day_num"] or 1
            week    = self._get_week(day_num)
            active  = self._calc_active_birds(r["rec_date"])
            exp_dead = round(active * self.mort_std.get(day_num, 0.05) / 100, 1)
            exp_feed = round(active * self.feed_std.get(day_num, 100) / 1000, 1)
            
            feed_val = r["feed_kg"] or 0
            dead_val = r["dead_count"] or 0
            
            cum_dead += dead_val
            cum_feed += feed_val
            diff_dead = round(dead_val - exp_dead, 1)
            diff_feed = round(feed_val - exp_feed, 1)
            tot_diff_dead += diff_dead
            tot_diff_feed += diff_feed
            
            # التلوين
            if diff_dead > 0 and diff_feed > 0: pdf.set_fill_color(244, 204, 204)
            elif diff_dead > 0: pdf.set_fill_color(252, 228, 214)
            elif diff_feed > 0: pdf.set_fill_color(255, 242, 204)
            else: pdf.set_fill_color(248, 252, 245)

            # القيم بترتيب العرض (يسار -> يمين على الورق)
            vals = [
                f"{cum_feed/50:,.1f}",
                f"{cum_dead:,}",
                f"{diff_feed:+.1f}",
                f"{exp_feed:.1f}",
                f"{feed_val:.1f}",
                f"{diff_dead:+.1f}",
                f"{exp_dead:.1f}",
                str(dead_val),
                f"{active:,}",
                str(week),
                str(day_num),
                str(r["rec_date"])
            ]
            
            pdf.set_font("Arabic", "", 8)
            for v, w in zip(vals, widths):
                pdf.cell(w, 7, prepare_text(v), 1, 0, "C", True)
            pdf.ln()
            
            tot_active += active
            count_rows += 1

        # ── الخلاصة النهائية ──
        pdf.ln(6)
        if count_rows > 0:
            pdf.set_fill_color(230, 240, 255)
            pdf.set_font("Arabic", "", 12)
            pdf.cell(0, 10, prepare_text("الخلاصة الفنية للدفعة"), 1, 1, "C", True)
            
            pdf.set_font("Arabic", "", 10)
            avg_active = tot_active / count_rows
            mortality_rate = (cum_dead / chicks * 100) if chicks > 0 else 0
            fcr_approx = (cum_feed / (chicks - cum_dead) / 1.8) if (chicks - cum_dead) > 0 else 0 # مجرد مثال
            
            summary_data = [
                ("إجمالي النافق", f"{cum_dead:,} طائر"),
                ("إجمالي فرق النافق", f"{tot_diff_dead:+.1f}"),
                ("إجمالي العلف المستهلك", f"{cum_feed/1000:,.3f} طن"),
                ("إجمالي فرق العلف", f"{tot_diff_feed:+.1f} كجم"),
                ("نسبة النافق الكلية", f"{mortality_rate:.2f}%"),
                ("متوسط استهلاك الطائر اليومي", f"{(cum_feed/tot_active*1000 if tot_active>0 else 0):,.1f} جرام")
            ]
            
            # رسم الخلاصة في سطرين أو عمودين
            col_w = 95
            for i in range(0, len(summary_data), 2):
                pdf.set_fill_color(250, 250, 250)
                # Left item
                lbl, val = summary_data[i+1]
                pdf.cell(35, 8, prepare_text(val), 1, 0, "C")
                pdf.cell(60, 8, prepare_text(lbl), 1, 0, "R", True)
                # Right item
                lbl, val = summary_data[i]
                pdf.cell(35, 8, prepare_text(val), 1, 0, "C")
                pdf.cell(60, 8, prepare_text(lbl), 1, 1, "R", True)

        # ── التذييل ──
        pdf.set_y(-15)
        pdf.set_font("Arabic", "", 7)
        pdf.set_text_color(120, 120, 120)
        footer_text = f"تاريخ الإصدار: {datetime.now().strftime('%Y-%m-%d %H:%M')} | تم الاستخراج بواسطة نظام إدارة الدواجن المتطور"
        pdf.cell(0, 10, prepare_text(footer_text), 0, 0, "C")

        try:
            pdf.output(path)
            messagebox.showinfo("تم", "تم تصدير السجل الفني بنجاح!", parent=self)
            try: os.startfile(path)
            except: pass
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل تصدير PDF:\n{str(e)}", parent=self)

# ════════════════════════════════════════════════════════════════
# نافذة إدخال / تعديل دفعة
# ════════════════════════════════════════════════════════════════
class BatchForm(ToplevelBase):
    def __init__(self, master, batch_id=None, on_save=None):
        super().__init__(master)
        self.batch_id = batch_id
        self.on_save  = on_save
        self.title("إدخال دفعة جديدة" if not batch_id else "تعديل دفعة")
        self.geometry("1250x820")
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.resizable(True, True)
        self.grab_set()

        self._vars = {}
        self._farm_sales = []
        self._market_sales = []
        self._build_ui()
        if batch_id: self._load_batch()
        center_window(self)

    def _build_ui(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=12)
        hdr.pack(fill="x")
        UILabel(hdr, text="الملف المالي للدفعة", font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)

        btn_frm = UIFrame(self, bg=CLR["bg"], pady=10)
        btn_frm.pack(side="bottom", fill="x")
        UIButton(btn_frm, text="💾 حفظ الدفعة", font=FT_HEADER, bg=CLR["nav"], fg="white", padx=30, pady=8, cursor="hand2", relief="flat", command=self._save).pack(side="right", padx=20)
        if self.batch_id:
            UIButton(btn_frm, text="📅 السجلات اليومية", font=FT_BODY, bg=CLR["daily_bg"], fg=CLR["accent"], padx=20, pady=8, cursor="hand2", relief="flat", bd=1, command=self._open_daily).pack(side="right", padx=4)
        UIButton(btn_frm, text="إلغاء وإغلاق", font=FT_BODY, bg="#e0e0e0", fg=CLR["text"], padx=20, pady=8, cursor="hand2", relief="solid", bd=1, command=self.destroy).pack(side="left", padx=20)

        style = ttk.Style()
        if not HAS_TTKB: style.theme_use('default')
        style.configure('TNotebook.Tab', font=FT_HEADER, padding=[15, 5])

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(side="top", fill="both", expand=True, padx=10, pady=10)

        self.tab_basic   = UIFrame(self.notebook, bg=CLR["bg"], padx=20, pady=20)
        self.tab_costs   = UIFrame(self.notebook, bg=CLR["bg"], padx=20, pady=10)
        self.tab_sales   = UIFrame(self.notebook, bg=CLR["bg"], padx=20, pady=10)
        self.tab_results = UIFrame(self.notebook, bg=CLR["bg"], padx=20, pady=20)

        self.notebook.add(self.tab_basic,   text="📋 البيانات الأساسية")
        self.notebook.add(self.tab_costs,   text="💰 سجل التكاليف")
        self.notebook.add(self.tab_sales,   text="📈 سجل المبيعات")
        self.notebook.add(self.tab_results, text="📊 الخلاصة والنتائج")

        self._build_basic_tab(self.tab_basic)
        self._build_costs_tab(self.tab_costs)
        self._build_sales_tab(self.tab_sales)
        self._build_results_tab(self.tab_results)

    def _suggest_batch_num(self):
        """اقتراح رقم دفعة تلقائي بصيغة: 2026-001 لكل عنبر على حده"""
        try:
            fy = int(self.v_fiscal_year.get())
        except:
            fy = date.today().year
        
        wh_name = self.wh_var.get().strip()
        if not wh_name:
            # إذا لم يتم اختيار عنبر، نترك الترقيم مفتوحاً أو نستخدم التنسيق الأساسي
            self._vars["batch_num"].set(f"{fy}-???")
            return

        # جلب ID العنبر المختار
        wh = db.fetch_one("SELECT id FROM warehouses WHERE name=?", (wh_name,))
        if not wh:
            # إذا كان عنبراً جديداً لم يُحفظ بعد في الدفعة
            self._vars["batch_num"].set(f"{fy}-001")
            return

        # عدد الدفعات لهذا العنبر في هذه السنة المالية (باستثناء الدفعة الحالية عند التعديل)
        row = db.fetch_one(
            "SELECT COUNT(*) AS c FROM batches WHERE fiscal_year=? AND warehouse_id=? AND id!=?",
            (fy, wh["id"], self.batch_id or 0))
        count = (row["c"] if row else 0) + 1
        suggested = f"{fy}-{count:03d}"
        self._vars["batch_num"].set(suggested)

    def _open_daily(self):
        if not self.batch_id: return
        batch = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (self.batch_id,))
        if batch: DailyRecordsWindow(self, self.batch_id, dict(batch))

    def _build_basic_tab(self, F):
        def _l_e(txt, key, r, c, w=16, ro=False):
            UILabel(F, text=txt, font=FT_SMALL, bg=CLR["bg"], fg=CLR["text2"]).grid(row=r, column=c, sticky="e", padx=(8,2), pady=8)
            v = tk.StringVar()
            self._vars[key] = v
            st = "readonly" if ro else "normal"
            bg_c = "#e9ecef" if ro else CLR["white"]
            e = UIEntry(F, textvariable=v, width=w, font=FT_BODY, state=st, bg=bg_c, relief="solid")
            e.grid(row=r, column=c+1, sticky="ew", padx=(2,12), pady=8)
            e.configure(justify="right")
            if not ro: v.trace_add("write", lambda *a: self._auto_calc())

        UILabel(F, text="اسم العنبر *", font=FT_SMALL, bg=CLR["bg"], fg=CLR["text2"]).grid(row=0, column=0, sticky="e", padx=(8,2), pady=10)
        
        wh_frm = UIFrame(F, bg=CLR["bg"])
        wh_frm.grid(row=0, column=1, sticky="ew", padx=(0,20), pady=10)
        self.wh_var = tk.StringVar()
        self.wh_combo = ttk.Combobox(wh_frm, textvariable=self.wh_var, width=18, font=FT_BODY)
        self.wh_combo["values"] = [r["name"] for r in db.fetch_all("SELECT name FROM warehouses ORDER BY name")]
        self.wh_combo.pack(side="right", fill="x", expand=True)
        self.wh_combo.bind("<<ComboboxSelected>>", lambda e: self._suggest_batch_num())
        UIButton(wh_frm, text="🗑", font=FT_SMALL, cursor="hand2", command=self._delete_selected_wh).pack(side="left", padx=5)

        _l_e("رقم الدفعة", "batch_num", 0, 2)

        # ── السنة المالية بجانب رقم الدفعة ──
        UILabel(F, text="السنة المالية:", font=FT_SMALL, bg=CLR["bg"], fg=CLR["text2"]).grid(
            row=0, column=4, sticky="e", padx=(8,2), pady=10)
        self.v_fiscal_year = tk.StringVar(value=str(date.today().year))
        fy_vals = [str(y) for y in range(date.today().year + 2, date.today().year - 6, -1)]
        self.fy_combo = ttk.Combobox(F, textvariable=self.v_fiscal_year,
                                      values=fy_vals, width=7, font=FT_BODY)
        self.fy_combo.grid(row=0, column=5, sticky="w", padx=(2,4), pady=10)
        self.fy_combo.bind("<<ComboboxSelected>>", lambda e: self._suggest_batch_num())
        UIButton(F, text="↺ ترقيم تلقائي", font=FT_TINY, padx=6, cursor="hand2",
                 command=self._suggest_batch_num).grid(row=0, column=6, padx=(0,8), pady=10)

        _l_e("تاريخ الدخول *", "date_in", 1, 0)
        _l_e("تاريخ الخروج *", "date_out", 1, 2)
        _l_e("عدد الأيام", "days", 1, 4, 10, ro=True)
        _l_e("عدد الكتاكيت المستلمة *", "chicks", 2, 0)
        _l_e("إجمالي قيمة الكتاكيت", "chick_val", 2, 2)
        _l_e("متوسط وزن الطائر (كجم)", "avg_weight", 3, 0)
        _l_e("معدل التحويل (FCR)", "fcr", 3, 2, 16, ro=True)

    def _delete_selected_wh(self):
        wh_name = self.wh_var.get().strip()
        if not wh_name: return messagebox.showwarning("تنبيه", "اختر عنبراً أولاً لحذفه", parent=self)
        wh = db.fetch_one("SELECT id FROM warehouses WHERE name=?", (wh_name,))
        if not wh: return
        cnt = db.fetch_one("SELECT COUNT(*) AS c FROM batches WHERE warehouse_id=?", (wh["id"],))["c"]
        if cnt > 0:
            return messagebox.showerror("خطأ", f"لا يمكن حذف العنبر لأنه مرتبط بـ {cnt} دفعة", parent=self)
        if messagebox.askyesno("تأكيد", f"حذف العنبر '{wh_name}' نهائياً؟", parent=self):
            db.execute("DELETE FROM warehouses WHERE id=?", (wh["id"],))
            self.wh_combo["values"] = [r["name"] for r in db.fetch_all("SELECT name FROM warehouses ORDER BY name")]
            self.wh_var.set("")
            if hasattr(self.master, '_load_batches'):
                self.master._load_batches()
            messagebox.showinfo("تم", "تم حذف العنبر بنجاح", parent=self)

    def _build_costs_tab(self, F):
        """
        يبني تبويب التكاليف ديناميكياً من جدول cost_types في قاعدة البيانات.
        أي تغيير في إدارة التكاليف ينعكس فوراً عند فتح نافذة دفعة جديدة.
        """
        v = self._vars
        # تنظيف محتوى التبويب (في حال إعادة البناء)
        for widget in F.winfo_children():
            widget.destroy()

        # جلب أنواع التكاليف النشطة من قاعدة البيانات
        active_costs = db.get_cost_types(active_only=True)

        # بناء قائمة الحقول: الثابتة (كمية+قيمة) + الديناميكية
        cost_fields = []
        STATIC_QTY = {
            "feed_val":    ("feed_qty",    "علف—كمية(طن)"),
            "sawdust_val": ("sawdust_qty", "نشارة—كمية"),
            "gas_val":     ("gas_qty",     "غاز—كمية"),
        }
        shown_qty_keys = set()

        for ct in active_costs:
            code = ct["code"]
            if code == "chick_val": continue  # موجود مسبقاً في البيانات الأساسية
            # إذا كان له كمية ثابتة مرتبطة أضفها أولاً
            if code in STATIC_QTY:
                qty_key, qty_lbl = STATIC_QTY[code]
                if qty_key not in shown_qty_keys:
                    cost_fields.append((qty_key, qty_lbl))
                    shown_qty_keys.add(qty_key)
            cost_fields.append((code, ct["name_ar"]))

        # إضافة أعمدة الكمية الباقية إذا لم تُعرض
        for qty_key, qty_lbl in [("feed_qty","علف—كمية(طن)"),("sawdust_qty","نشارة—كمية"),("gas_qty","غاز—كمية")]:
            if qty_key not in shown_qty_keys:
                cost_fields.insert(0, (qty_key, qty_lbl))

        row, col = 0, 0
        for i, (key, lbl) in enumerate(cost_fields):
            if i > 0 and i % 3 == 0: row += 1; col = 0
            UILabel(F, text=lbl, font=FT_SMALL, bg=CLR["bg"], fg=CLR["text2"]).grid(
                row=row, column=col, sticky="e", padx=(8,2), pady=6)
            if key not in v:
                v[key] = tk.StringVar()
            e = UIEntry(F, textvariable=v[key], width=16, font=FT_BODY,
                        relief="solid", highlightthickness=1, highlightbackground=CLR["border"])
            e.grid(row=row, column=col+1, sticky="ew", padx=(0,20), pady=6)
            e.configure(justify="right")
            v[key].trace_add("write", lambda *a: self._auto_calc())
            col += 2

        frm_tc = UIFrame(F, bg=CLR["loss_bg"], pady=8, padx=15, bd=1, relief="solid")
        frm_tc.grid(row=row+1, column=0, columnspan=6, sticky="ew", pady=(20,0))
        UILabel(frm_tc, text="إجمالي التكاليف والمصروفات:", font=FT_HEADER,
                bg=CLR["loss_bg"], fg=CLR["loss"]).pack(side="right")
        self.lbl_total_cost = UILabel(frm_tc, text="0", font=("Arial",16,"bold"),
                                       bg=CLR["loss_bg"], fg=CLR["loss"])
        self.lbl_total_cost.pack(side="right", padx=15)

    def _build_sales_tab(self, F):
        canvas = tk.Canvas(F, bg=CLR["bg"], highlightthickness=0)
        v_scroll = ttk.Scrollbar(F, orient="vertical", command=canvas.yview)
        scroll_frm = UIFrame(canvas, bg=CLR["bg"])
        canvas.configure(yscrollcommand=v_scroll.set)
        canvas.pack(side="right", fill="both", expand=True)
        v_scroll.pack(side="left", fill="y")
        
        # إنشاء نافذة الإطار داخل الكانفاس بدون عرض ثابت
        inner_id = canvas.create_window((0,0), window=scroll_frm, anchor="nw")

        def _on_cfg(e):
            # تحديث منطقة التمرير
            canvas.configure(scrollregion=canvas.bbox("all"))
            # تحديث عرض الإطار الداخلي ليطابق عرض الكانفاس (يساعد في التوسيط)
            canvas.itemconfig(inner_id, width=e.width)
            
        canvas.bind("<Configure>", _on_cfg)

        f_frm = UILabelFrame(scroll_frm, text="🐓 بيان مبيعات العنبر", font=FT_HEADER, bg=CLR["bg"], fg=CLR["nav"], padx=10, pady=10)
        f_frm.pack(fill="x", pady=(0,15), padx=20)

        inp_f = UIFrame(f_frm, bg=CLR["bg"])
        inp_f.pack(fill="x", pady=5)

        UILabel(inp_f, text="تاريخ البيع:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0,column=0, padx=2, pady=2)
        self.v_fs_date = tk.StringVar(value=date.today().isoformat())
        UIEntry(inp_f, textvariable=self.v_fs_date, width=12, font=FT_BODY, relief="solid").grid(row=0,column=1, padx=2, pady=2)

        UILabel(inp_f, text="النوع:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0,column=2, padx=2, pady=2)
        self.v_fs_type = tk.StringVar(value="آجل")
        cbo_type = ttk.Combobox(inp_f, textvariable=self.v_fs_type, values=["آجل", "نقداً"], width=8, font=FT_BODY, state="readonly")
        cbo_type.grid(row=0,column=3, padx=2, pady=2)

        UILabel(inp_f, text="اسم العميل:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0,column=4, padx=2, pady=2)
        self.v_fs_cust = tk.StringVar()
        UIEntry(inp_f, textvariable=self.v_fs_cust, width=15, font=FT_BODY, relief="solid").grid(row=0,column=5, padx=2, pady=2)

        UILabel(inp_f, text="الكمية:", font=FT_SMALL, bg=CLR["bg"]).grid(row=1,column=0, padx=2, pady=2)
        self.v_fs_qty = tk.StringVar()
        UIEntry(inp_f, textvariable=self.v_fs_qty, width=12, font=FT_BODY, justify="right", relief="solid").grid(row=1,column=1, padx=2, pady=2)

        UILabel(inp_f, text="السعر:", font=FT_SMALL, bg=CLR["bg"]).grid(row=1,column=2, padx=2, pady=2)
        self.v_fs_price = tk.StringVar()
        UIEntry(inp_f, textvariable=self.v_fs_price, width=15, font=FT_BODY, justify="right", relief="solid").grid(row=1,column=3, padx=2, pady=2)

        btn_f_frm = UIFrame(inp_f, bg=CLR["bg"])
        btn_f_frm.grid(row=2, column=0, columnspan=6, sticky="w", pady=(8,0))

        UIButton(btn_f_frm, text="➕ إضافة", font=FT_BODY, bg=CLR["nav"], fg="white", relief="flat", cursor="hand2", command=self._add_farm_sale).pack(side="right", padx=4)
        UIButton(btn_f_frm, text="✏️ تعديل", font=FT_BODY, relief="flat", cursor="hand2", command=self._edit_farm_sale).pack(side="right", padx=4)
        UIButton(btn_f_frm, text="🗑 حذف", font=FT_BODY, bg=CLR["loss_bg"], fg=CLR["loss"], relief="flat", cursor="hand2", command=self._del_farm_sale).pack(side="right", padx=4)

        f_cols = ("م", "تاريخ البيع", "نوع البيع", "اسم العميل", "الكمية", "السعر", "الإجمالي")
        self.tv_farm = ttk.Treeview(f_frm, columns=f_cols, show="headings", selectmode="browse", height=6)
        widths_f = [35, 100, 60, 180, 70, 70, 90]
        for c, w in zip(f_cols, widths_f):
            self.tv_farm.heading(c, text=c, anchor="center")
            self.tv_farm.column(c, width=w, anchor="center")
        self.tv_farm.pack(fill="both", expand=True, pady=5)

        sum_f = UIFrame(f_frm, bg=CLR["profit_bg"], pady=5, padx=15)
        sum_f.pack(fill="x")
        self.lbl_cust_tot = UILabel(sum_f, text="إجمالي مبيعات العنبر: 0 طائر | 0", font=FT_BODY, bg=CLR["profit_bg"], fg=CLR["profit"])
        self.lbl_cust_tot.pack(side="right")

        m_frm = UILabelFrame(scroll_frm, text="🏢 بيان مبيعات السوق (المكاتب)", font=FT_HEADER, bg=CLR["bg"], fg=CLR["accent"], padx=10, pady=10)
        m_frm.pack(fill="x", pady=15, padx=20)

        inp_m = UIFrame(m_frm, bg=CLR["bg"])
        inp_m.pack(fill="x", pady=5)
        
        UILabel(inp_m, text="تاريخ البيع:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0,column=0, padx=2, pady=2)
        self.v_ms_date = tk.StringVar(value=date.today().isoformat())
        UIEntry(inp_m, textvariable=self.v_ms_date, width=12, font=FT_BODY, relief="solid").grid(row=0,column=1, padx=2, pady=2)

        UILabel(inp_m, text="المكتب:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0,column=2, padx=2, pady=2)
        self.v_ms_office = tk.StringVar()
        UIEntry(inp_m, textvariable=self.v_ms_office, width=15, font=FT_BODY, relief="solid").grid(row=0,column=3, padx=2, pady=2)

        UILabel(inp_m, text="رقم الفاتورة:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0,column=4, padx=2, pady=2)
        self.v_ms_inv = tk.StringVar()
        UIEntry(inp_m, textvariable=self.v_ms_inv, width=12, font=FT_BODY, relief="solid").grid(row=0,column=5, padx=2, pady=2)

        UILabel(inp_m, text="الكمية:", font=FT_SMALL, bg=CLR["bg"]).grid(row=1,column=0, padx=2, pady=2)
        self.v_ms_qty = tk.StringVar()
        UIEntry(inp_m, textvariable=self.v_ms_qty, width=12, font=FT_BODY, justify="right", relief="solid").grid(row=1,column=1, padx=2, pady=2)

        UILabel(inp_m, text="الوفيات:", font=FT_SMALL, bg=CLR["bg"]).grid(row=1,column=2, padx=2, pady=2)
        self.v_ms_dead = tk.StringVar(value="0")
        UIEntry(inp_m, textvariable=self.v_ms_dead, width=15, font=FT_BODY, justify="right", relief="solid").grid(row=1,column=3, padx=2, pady=2)

        UILabel(inp_m, text="صافي الفاتورة:", font=FT_SMALL, bg=CLR["bg"]).grid(row=1,column=4, padx=2, pady=2)
        self.v_ms_net = tk.StringVar()
        UIEntry(inp_m, textvariable=self.v_ms_net, width=12, font=FT_BODY, justify="right", relief="solid").grid(row=1,column=5, padx=2, pady=2)

        btn_m_frm = UIFrame(inp_m, bg=CLR["bg"])
        btn_m_frm.grid(row=2, column=0, columnspan=6, sticky="w", pady=(8,0))
        UIButton(btn_m_frm, text="➕ إضافة", font=FT_BODY, bg=CLR["nav"], fg="white", relief="flat", cursor="hand2", command=self._add_market_sale).pack(side="right", padx=4)
        UIButton(btn_m_frm, text="✏️ تعديل", font=FT_BODY, relief="flat", cursor="hand2", command=self._edit_market_sale).pack(side="right", padx=4)
        UIButton(btn_m_frm, text="🗑 حذف", font=FT_BODY, bg=CLR["loss_bg"], fg=CLR["loss"], relief="flat", cursor="hand2", command=self._del_market_sale).pack(side="right", padx=4)

        m_cols = ("م", "تاريخ البيع", "مكتب التسويق", "الكمية", "الوفيات", "المباع", "صافي الفاتورة", "رقم الفاتورة")
        self.tv_mkt = ttk.Treeview(m_frm, columns=m_cols, show="headings", selectmode="browse", height=6)
        widths_m = [35, 100, 170, 70, 70, 70, 100, 100]
        for c, w in zip(m_cols, widths_m):
            self.tv_mkt.heading(c, text=c, anchor="center")
            self.tv_mkt.column(c, width=w, anchor="center")
        self.tv_mkt.pack(fill="both", expand=True, pady=5)

        sum_m = UIFrame(m_frm, bg=CLR["profit_bg"], pady=5, padx=15)
        sum_m.pack(fill="x")
        self.lbl_mkt_tot = UILabel(sum_m, text="إجمالي مبيعات السوق: 0 طائر | 0", font=FT_BODY, bg=CLR["profit_bg"], fg=CLR["profit"])
        self.lbl_mkt_tot.pack(side="right")

        o_frm = UILabelFrame(scroll_frm, text="💰 إيرادات وضبط مالي إضافي", font=FT_HEADER, bg=CLR["bg"], fg="#607d8b", padx=10, pady=10)
        o_frm.pack(fill="x", pady=(15,0))

        v = self._vars
        def _l_e_o(txt, key, r, c):
            UILabel(o_frm, text=txt, font=FT_SMALL, bg=CLR["bg"]).grid(row=r, column=c, sticky="e", padx=(8,2), pady=8)
            v[key] = tk.StringVar()
            e = UIEntry(o_frm, textvariable=v[key], width=14, font=FT_BODY, relief="solid")
            e.grid(row=r, column=c+1, sticky="w", padx=(0,15), pady=8); e.configure(justify="right")
            v[key].trace_add("write", lambda *a: self._auto_calc())

        _l_e_o("مبيعات ذبيل (قيمة):", "offal_val", 0, 0)
        _l_e_o("مرتجع علاجات (قيمة):", "drug_return", 0, 2)
        _l_e_o("نقل غاز/نشارة (قيمة):", "gas_return", 0, 4)
        _l_e_o("مبيعات علف (كمية/كيس):", "feed_sale_qty", 1, 0)
        _l_e_o("مبيعات علف (قيمة):", "feed_sale", 1, 2)
        _l_e_o("علف منقول لعنابر (كمية/كيس):", "feed_trans_r_qty", 2, 0)
        _l_e_o("علف منقول لعنابر (قيمة):", "feed_trans_r", 2, 2)
        _l_e_o("علف متبقي (كمية/كيس):", "feed_rem_qty", 3, 0)
        _l_e_o("علف متبقي (قيمة):", "feed_rem_val", 3, 2)

        sum_total = UIFrame(F, bg=CLR["profit_bg"], pady=10, padx=20, bd=1, relief="ridge")
        sum_total.pack(fill="x", side="bottom")
        UILabel(sum_total, text="إجمالي الإيرادات والمبيعات (بيان موحد):", font=FT_HEADER, bg=CLR["profit_bg"], fg=CLR["profit"]).pack(side="right")
        self.lbl_total_rev = UILabel(sum_total, text="0", font=("Arial",18,"bold"), bg=CLR["profit_bg"], fg=CLR["profit"])
        self.lbl_total_rev.pack(side="right", padx=20)

    def _add_farm_sale(self):
        t  = self.v_fs_type.get().strip()
        c  = self.v_fs_cust.get().strip()
        sd = self.v_fs_date.get().strip()
        try:
            q = int(self.v_fs_qty.get() or 0)
            p = float(self.v_fs_price.get() or 0)
        except ValueError:
            return messagebox.showerror("خطأ", "الكمية والسعر يجب أن تكون أرقاماً", parent=self)
        if not c or q <= 0:
            return messagebox.showwarning("تنبيه", "يرجى إدخال اسم العميل والكمية", parent=self)

        # ── تحقق من الكتل النشطة ──────────────────────────────────
        if self.batch_id:
            chicks = int(self._n("chicks"))
            already_sold_farm = sum(s.get("qty",0) for s in self._farm_sales)
            already_sold_mkt  = sum(s.get("qty_sold",0) for s in self._market_sales)
            total_dead = int(self._n("total_dead") or 0)
            consumed   = int(self._n("consumed_birds") or 0)
            active = max(0, chicks - total_dead - already_sold_farm - already_sold_mkt - consumed)
            if q > active:
                return messagebox.showerror(
                    "⚠️ تجاوز الكتل النشطة",
                    f"الكمية المدخلة ({q:,}) أكبر من الكتل النشطة المتاحة ({active:,})\n\n"
                    f"الكتاكيت: {chicks:,}\n"
                    f"النافق: {total_dead:,}\n"
                    f"مباع عنبر: {already_sold_farm:,}\n"
                    f"مباع سوق: {already_sold_mkt:,}\n"
                    f"مستهلك: {consumed:,}\n"
                    f"المتاح: {active:,}",
                    parent=self)

        self._farm_sales.append({"sale_date": sd, "sale_type": t, "customer": c, "qty": q, "price": p, "total_val": q * p})
        self.v_fs_cust.set(""); self.v_fs_qty.set(""); self.v_fs_price.set("")
        self._refresh_sales_views(); self._auto_calc()

    def _del_farm_sale(self):
        sel = self.tv_farm.selection()
        if not sel: return
        idx = self.tv_farm.index(sel[0])
        del self._farm_sales[idx]
        self._refresh_sales_views(); self._auto_calc()

    def _edit_farm_sale(self):
        sel = self.tv_farm.selection()
        if not sel: return
        idx = self.tv_farm.index(sel[0])
        item = self._farm_sales[idx]
        self.v_fs_date.set(item.get("sale_date", ""))
        self.v_fs_type.set(item.get("sale_type", "آجل"))
        self.v_fs_cust.set(item["customer"]); self.v_fs_qty.set(str(item["qty"])); self.v_fs_price.set(str(item["price"]))
        del self._farm_sales[idx]
        self._refresh_sales_views(); self._auto_calc()

    def _add_market_sale(self):
        off = self.v_ms_office.get().strip()
        sd  = self.v_ms_date.get().strip()
        try:
            q = int(self.v_ms_qty.get() or 0)
            d = int(self.v_ms_dead.get() or 0)
            n_val = float(self.v_ms_net.get() or 0)
        except ValueError:
            return messagebox.showerror("خطأ", "الكمية، الوفيات، وصافي الفاتورة يجب أن تكون أرقاماً", parent=self)
        if not off or q <= 0:
            return messagebox.showwarning("تنبيه", "يرجى إدخال مكتب التسويق والكمية", parent=self)
        if d > q:
            return messagebox.showerror("خطأ", "الوفيات لا يمكن أن تتجاوز الكمية المرسلة", parent=self)

        # ── تحقق من الكتل النشطة ──────────────────────────────────
        if self.batch_id:
            chicks = int(self._n("chicks"))
            already_sold_farm = sum(s.get("qty",0) for s in self._farm_sales)
            already_sold_mkt  = sum(s.get("qty_sold",0) for s in self._market_sales)
            total_dead = int(self._n("total_dead") or 0)
            consumed   = int(self._n("consumed_birds") or 0)
            active = max(0, chicks - total_dead - already_sold_farm - already_sold_mkt - consumed)
            sold_new = q - d
            if q > active:
                return messagebox.showerror(
                    "⚠️ تجاوز الكتل النشطة",
                    f"الكمية المرسلة ({q:,}) أكبر من الكتل النشطة المتاحة ({active:,})\n\n"
                    f"الكتاكيت: {chicks:,}  |  النافق: {total_dead:,}\n"
                    f"مباع عنبر: {already_sold_farm:,}  |  مباع سوق: {already_sold_mkt:,}\n"
                    f"المتاح الآن: {active:,}",
                    parent=self)

        sold = q - d
        self._market_sales.append({"sale_date": sd, "office": off, "qty_sent": q, "deaths": d, "qty_sold": sold, "net_val": n_val, "inv_num": self.v_ms_inv.get().strip()})
        self.v_ms_office.set(""); self.v_ms_qty.set(""); self.v_ms_dead.set("0"); self.v_ms_net.set(""); self.v_ms_inv.set("")
        self._refresh_sales_views(); self._auto_calc()

    def _del_market_sale(self):
        sel = self.tv_mkt.selection()
        if not sel: return
        idx = self.tv_mkt.index(sel[0])
        del self._market_sales[idx]
        self._refresh_sales_views(); self._auto_calc()

    def _edit_market_sale(self):
        sel = self.tv_mkt.selection()
        if not sel: return
        idx = self.tv_mkt.index(sel[0])
        item = self._market_sales[idx]
        self.v_ms_date.set(item.get("sale_date", ""))
        self.v_ms_office.set(item["office"]); self.v_ms_qty.set(str(item["qty_sent"]))
        self.v_ms_dead.set(str(item["deaths"])); self.v_ms_net.set(str(item["net_val"])); self.v_ms_inv.set(item.get("inv_num", ""))
        del self._market_sales[idx]
        self._refresh_sales_views(); self._auto_calc()

    def _refresh_sales_views(self):
        self.tv_farm.delete(*self.tv_farm.get_children())
        for i, s in enumerate(self._farm_sales, 1):
            self.tv_farm.insert("", "end", values=(i, s.get("sale_date",""), s.get("sale_type", "آجل"), s["customer"], fmt_num(s["qty"]), fmt_num(s["price"],2), fmt_num(s["total_val"])))
        self.tv_mkt.delete(*self.tv_mkt.get_children())
        for i, s in enumerate(self._market_sales, 1):
            self.tv_mkt.insert("", "end", values=(i, s.get("sale_date",""), s["office"], fmt_num(s["qty_sent"]), fmt_num(s["deaths"]), fmt_num(s["qty_sold"]), fmt_num(s["net_val"]), s.get("inv_num","")))

    def _build_results_tab(self, F):
        def _l_e(txt, key, r, c, w=16, ro=False, cs=1):
            UILabel(F, text=txt, font=FT_SMALL, bg=CLR["bg"]).grid(row=r, column=c, sticky="e", padx=(8,2), pady=8)
            v = tk.StringVar(); self._vars[key] = v
            e = UIEntry(F, textvariable=v, width=w, font=FT_BODY, state="readonly" if ro else "normal", relief="solid")
            e.grid(row=r, column=c+1, sticky="ew", padx=(2,15), pady=8, columnspan=cs); e.configure(justify="right")
            if not ro: v.trace_add("write", lambda *a: self._auto_calc())

        _l_e("إجمالي الطيور المباعة (حبة)", "total_sold", 0, 0, ro=True)
        _l_e("النافق الكلي (حبة)", "total_dead", 0, 2)
        _l_e("نسبة النافق الكلية %", "mort_rate", 0, 4, ro=True)
        _l_e("طيور ضيافة / مستهلكة (حبة)", "consumed_birds", 1, 0)
        _l_e("متوسط سعر البيع للطائر", "avg_price", 1, 2, ro=True)
        _l_e("نصيب الشركة من الأرباح %", "share_pct", 2, 0); self._vars["share_pct"].set("65")
        _l_e("نصيب الشركة (ريال)", "share_val", 2, 2, ro=True)
        _l_e("اسم الشريك", "partner_name", 3, 0)
        _l_e("ملاحظات إضافية على الدفعة", "notes", 4, 0, w=40, cs=5)

        frm_epef = UIFrame(F, pady=5, padx=20, bg=CLR["info_bg"])
        frm_epef.grid(row=5, column=0, columnspan=6, sticky="ew", pady=(10,0))
        UILabel(frm_epef, text="مؤشر الكفاءة الأوروبي (EPEF):", font=("Arial",12,"bold"), bg=CLR["info_bg"]).pack(side="right")
        self.lbl_epef = UILabel(frm_epef, text="0", font=("Arial",14,"bold"), bg=CLR["info_bg"], fg=CLR["accent"])
        self.lbl_epef.pack(side="right", padx=20)

        frm_net = UIFrame(F, pady=15, padx=20, bd=2, relief="groove")
        frm_net.grid(row=6, column=0, columnspan=6, sticky="ew", pady=(20,0))
        UILabel(frm_net, text="صافي النتيجة للدفعة:", font=("Arial",18,"bold")).pack(side="right")
        self.lbl_net = UILabel(frm_net, text="0", font=("Arial",24,"bold"))
        self.lbl_net.pack(side="right", padx=20)
        self._net_frame = frm_net

    def _n(self, key):
        try: return float(self._vars[key].get()) if self._vars.get(key) and self._vars[key].get() else 0.0
        except: return 0.0

    def _auto_calc(self):
        v = self._vars
        curr = db.get_setting("currency", "ريال")
        
        d_in, d_out = None, None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"):
            if not d_in:
                try: d_in = datetime.strptime(v.get("date_in", tk.StringVar()).get().strip(), fmt)
                except: pass
            if not d_out:
                try: d_out = datetime.strptime(v.get("date_out", tk.StringVar()).get().strip(), fmt)
                except: pass
        if d_in and d_out:
            days = (d_out - d_in).days
            v["days"].set(days if days > 0 else "")
        else: v["days"].set("")

        chick_val = self._n("chick_val")
        # قراءة التكاليف ديناميكياً من قاعدة البيانات (تعكس التغييرات فوراً)
        active_cost_codes = [ct["code"] for ct in db.get_cost_types(active_only=True) if ct["code"] != "chick_val"]
        static_extra = ["feed_trans"]  # أجور نقل علف — قد لا تكون في cost_types
        all_cost_keys = list(set(active_cost_codes + static_extra))
        total_cost = chick_val + sum(self._n(k) for k in all_cost_keys)
        if hasattr(self, 'lbl_total_cost'): self.lbl_total_cost.config(text=f"{fmt_num(total_cost)} {curr}")

        total_cust_qty = sum(float(x.get('qty', 0)) for x in self._farm_sales)
        total_cust_val = sum(float(x.get('total_val', 0)) for x in self._farm_sales)
        total_mkt_qty  = sum(float(x.get('qty_sold', 0)) for x in self._market_sales)
        total_mkt_val  = sum(float(x.get('net_val', 0)) for x in self._market_sales)

        if hasattr(self, 'lbl_cust_tot'): self.lbl_cust_tot.config(text=f"إجمالي مبيعات العنبر: {fmt_num(total_cust_qty)} طائر | {fmt_num(total_cust_val)} {curr}")
        if hasattr(self, 'lbl_mkt_tot'): self.lbl_mkt_tot.config(text=f"إجمالي مبيعات السوق: {fmt_num(total_mkt_qty)} طائر | {fmt_num(total_mkt_val)} {curr}")

        rev_keys = ["offal_val","feed_sale","feed_trans_r","feed_rem_val","drug_return","gas_return"]
        total_rev = total_cust_val + total_mkt_val + sum(self._n(k) for k in rev_keys)
        if hasattr(self, 'lbl_total_rev'): self.lbl_total_rev.config(text=f"{fmt_num(total_rev)} {curr}")

        chicks = self._n("chicks"); dead = self._n("total_dead")
        days = self._n("days"); avg_weight = self._n("avg_weight"); feed_tons = self._n("feed_qty")

        if "mort_rate" in v: v["mort_rate"].set(f"{(dead / chicks * 100):.2f}" if chicks > 0 else "")
        sold_qty = total_cust_qty + total_mkt_qty
        if "total_sold" in v: v["total_sold"].set(fmt_num(sold_qty))
        if sold_qty > 0 and "avg_price" in v: v["avg_price"].set(fmt_num((total_cust_val + total_mkt_val) / sold_qty, 2))
        if sold_qty > 0 and avg_weight > 0 and feed_tons > 0 and "fcr" in v: v["fcr"].set(f"{(feed_tons * 1000) / (sold_qty * avg_weight):.3f}")

        if hasattr(self, 'lbl_epef'):
            if days > 0 and chicks > 0 and avg_weight > 0:
                epef = ((100 - (dead/chicks*100)) * avg_weight / days) * 100
                self.lbl_epef.config(text=f"{epef:.0f}", foreground=CLR["profit"] if epef >= 300 else CLR["loss"])
            else: self.lbl_epef.config(text="0", foreground=CLR["text2"])

        net = total_rev - total_cost
        if hasattr(self, 'lbl_net'):
            self.lbl_net.config(text=f"{fmt_num(net)} {curr}", foreground=CLR["profit"] if net >= 0 else CLR["loss"])
        if not HAS_TTKB and hasattr(self, '_net_frame'):
            try: self._net_frame.config(bg=CLR["profit_bg"] if net >= 0 else CLR["loss_bg"])
            except: pass
        try: v["share_val"].set(fmt_num(net * float(v["share_pct"].get()) / 100))
        except: pass

    def _load_batch(self):
        row = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (self.batch_id,))
        if not row: return
        self.wh_var.set(row["warehouse_name"])
        # استرداد السنة المالية
        fy = row["fiscal_year"] if "fiscal_year" in row.keys() and row["fiscal_year"] else 0
        if not fy:
            try: fy = int(str(row["date_in"])[:4])
            except: fy = date.today().year
        if hasattr(self, 'v_fiscal_year'): self.v_fiscal_year.set(str(fy))
        for k in [x for x in row.keys() if x in self._vars]:
            val = row[k]
            if val is not None:
                # تحويل الطن إلى أكياس عند العرض في الواجهة (لإيرادات العلف فقط)
                if k in ["feed_sale_qty", "feed_trans_r_qty", "feed_rem_qty"]:
                    try: val = float(val) * 20
                    except: pass
                self._vars[k].set(str(val))
            else:
                self._vars[k].set("")

        f_sales = db.fetch_all("SELECT * FROM farm_sales WHERE batch_id=?", (self.batch_id,))
        if f_sales:
            self._farm_sales = [dict(r) for r in f_sales]
            for s in self._farm_sales:
                if "sale_date" not in s: s["sale_date"] = ""
        elif row["cust_qty"] or row["cust_val"]:
            old_qty = row["cust_qty"] or 0
            self._farm_sales.append({"sale_date":"", "sale_type":"آجل", "customer": "مبيعات سابقة (مرحلة)", "qty": old_qty, "price": (row["cust_val"]/old_qty) if old_qty else 0, "total_val": row["cust_val"] or 0})

        m_sales = db.fetch_all("SELECT * FROM market_sales WHERE batch_id=?", (self.batch_id,))
        if m_sales:
            self._market_sales = [dict(r) for r in m_sales]
            for s in self._market_sales:
                if "sale_date" not in s: s["sale_date"] = ""
        elif row["mkt_qty"] or row["mkt_val"]:
            old_m_qty = row["mkt_qty"] or 0
            self._market_sales.append({"sale_date":"", "office": "مبيعات سابقة (مرحلة)", "qty_sent": old_m_qty, "deaths": 0, "qty_sold": old_m_qty, "net_val": row["mkt_val"] or 0, "inv_num": ""})

        self._refresh_sales_views(); self._auto_calc()

    def _collect(self):
        v = self._vars
        n = lambda k: float(v[k].get()) if v.get(k) and v[k].get() else 0.0
        s = lambda k: v[k].get().strip() if v.get(k) else ""

        chicks_count = int(n("chicks")); chick_val_total = n("chick_val")
        active_cost_codes = [ct["code"] for ct in db.get_cost_types(active_only=True) if ct["code"] != "chick_val"]
        all_cost_keys = list(set(active_cost_codes + ["feed_trans"]))
        total_cost = chick_val_total + sum(n(k) for k in all_cost_keys)

        total_cust_qty = sum(float(x.get('qty', 0)) for x in self._farm_sales)
        total_cust_val = sum(float(x.get('total_val', 0)) for x in self._farm_sales)
        total_mkt_qty  = sum(float(x.get('qty_sold', 0)) for x in self._market_sales)
        total_mkt_val  = sum(float(x.get('net_val', 0)) for x in self._market_sales)
        total_rev  = total_cust_val + total_mkt_val + sum(n(k) for k in ["offal_val","feed_sale","feed_trans_r","feed_rem_val","drug_return","gas_return"])

        net = total_rev - total_cost; sold_qty = total_cust_qty + total_mkt_qty

        return {
            "batch_num":s("batch_num"), "date_in":s("date_in"), "date_out":s("date_out"), "days":int(n("days")),
            "chicks":chicks_count, "chick_price":chick_val_total/chicks_count if chicks_count>0 else 0, "chick_val":chick_val_total,
            "feed_qty":n("feed_qty"), "feed_val":n("feed_val"), "feed_trans":n("feed_trans"),
            "sawdust_qty":n("sawdust_qty"), "sawdust_val":n("sawdust_val"), "water_val":n("water_val"),
            "gas_qty":n("gas_qty"), "gas_val":n("gas_val"), "drugs_val":n("drugs_val"),
            "wh_expenses":n("wh_expenses"), "house_exp":n("house_exp"), "breeders_pay":n("breeders_pay"),
            "qat_pay":n("qat_pay"), "rent_val":n("rent_val"), "light_val":n("light_val"),
            "sup_wh_pay":n("sup_wh_pay"), "sup_co_pay":n("sup_co_pay"), "sup_sale_pay":n("sup_sale_pay"),
            "admin_val":n("admin_val"), "vaccine_pay":n("vaccine_pay"), "delivery_val":n("delivery_val"),
            "mixing_val":n("mixing_val"), "wash_val":n("wash_val"), "other_costs":n("other_costs"),
            "total_cost":total_cost, "cust_qty":int(total_cust_qty), "cust_val":total_cust_val,
            "mkt_qty":int(total_mkt_qty), "mkt_val":total_mkt_val,
            "offal_val":n("offal_val"), "feed_sale":n("feed_sale"), "feed_trans_r":n("feed_trans_r"), "drug_return":n("drug_return"), "gas_return":n("gas_return"),
            "total_rev":total_rev, "total_sold":int(sold_qty), "total_dead":int(n("total_dead")), "mort_rate": round(n("total_dead")/chicks_count*100, 2) if chicks_count>0 else 0,
            "avg_weight":n("avg_weight"), "fcr":n("fcr"), "avg_price": round((total_cust_val+total_mkt_val)/sold_qty,2) if sold_qty>0 else 0, "net_result":net,
            "share_pct": n("share_pct") or 65, "share_val": net * (n("share_pct") or 65) / 100, "notes":s("notes"),
            "consumed_birds": int(n("consumed_birds")), "partner_name": s("partner_name"),
            # تحويل الأكياس إلى أطنان قبل الحفظ في قاعدة البيانات
            "feed_sale_qty": n("feed_sale_qty") / 20, 
            "feed_trans_r_qty": n("feed_trans_r_qty") / 20, 
            "feed_rem_qty": n("feed_rem_qty") / 20, 
            "feed_rem_val": n("feed_rem_val"),
            "fiscal_year": int(self.v_fiscal_year.get()) if hasattr(self, 'v_fiscal_year') else date.today().year,
        }

    def _save(self):
        wh_name = self.wh_var.get().strip()
        if not wh_name: return messagebox.showwarning("تنبيه", "يرجى تحديد اسم العنبر", parent=self)
        d = self._collect()
        if not d["date_in"] or not d["date_out"] or not d["chicks"]: return messagebox.showwarning("تنبيه", "يرجى ملء: تاريخ الدخول والخروج وعدد الكتاكيت", parent=self)

        wh = db.fetch_one("SELECT id FROM warehouses WHERE name=?", (wh_name,))
        if not wh:
            db.execute("INSERT INTO warehouses (name) VALUES (?)", (wh_name,))
            wh = db.fetch_one("SELECT id FROM warehouses WHERE name=?", (wh_name,))

        vals = list(d.values())
        if self.batch_id:
            db.execute(f"UPDATE batches SET {','.join(f'{k}=?' for k in d)},fiscal_year=? WHERE id=?",
                       vals + [d.get("fiscal_year", date.today().year), self.batch_id])
            b_id = self.batch_id
        else:
            b_id = db.execute(
                f"INSERT INTO batches (warehouse_id,{','.join(k for k in d)},fiscal_year,created_at) "
                f"VALUES (?,{','.join('?' for _ in d)},?,datetime('now'))",
                [wh["id"]] + vals + [d.get("fiscal_year", date.today().year)])

        db.execute("DELETE FROM farm_sales WHERE batch_id=?", (b_id,))
        for fs in self._farm_sales:
            db.execute("INSERT INTO farm_sales (batch_id, sale_date, sale_type, customer, qty, price, total_val) VALUES (?,?,?,?,?,?,?)",
                       (b_id, fs.get("sale_date",""), fs.get("sale_type", "آجل"), fs["customer"], fs["qty"], fs["price"], fs["total_val"]))

        db.execute("DELETE FROM market_sales WHERE batch_id=?", (b_id,))
        for ms in self._market_sales:
            db.execute("INSERT INTO market_sales (batch_id, sale_date, office, qty_sent, deaths, qty_sold, net_val, inv_num) VALUES (?,?,?,?,?,?,?,?)",
                       (b_id, ms.get("sale_date",""), ms["office"], ms["qty_sent"], ms["deaths"], ms["qty_sold"], ms["net_val"], ms.get("inv_num","")))

        costs_dict = {
            "chick_val":    {"amount": d["chick_val"],    "qty": d["chicks"]},
            "feed_val":     {"amount": d["feed_val"],     "qty": d["feed_qty"]},
            "feed_trans":   {"amount": d["feed_trans"],   "qty": 0},
            "sawdust_val":  {"amount": d["sawdust_val"],  "qty": d["sawdust_qty"]},
            "water_val":    {"amount": d["water_val"],    "qty": 0},
            "gas_val":      {"amount": d["gas_val"],      "qty": d["gas_qty"]},
            "drugs_val":    {"amount": d["drugs_val"],    "qty": 0},
            "wh_expenses":  {"amount": d["wh_expenses"],  "qty": 0},
            "house_exp":    {"amount": d["house_exp"],    "qty": 0},
            "breeders_pay": {"amount": d["breeders_pay"], "qty": 0},
            "qat_pay":      {"amount": d["qat_pay"],      "qty": 0},
            "rent_val":     {"amount": d["rent_val"],     "qty": 0},
            "light_val":    {"amount": d["light_val"],    "qty": 0},
            "sup_wh_pay":   {"amount": d["sup_wh_pay"],   "qty": 0},
            "sup_co_pay":   {"amount": d["sup_co_pay"],   "qty": 0},
            "sup_sale_pay": {"amount": d["sup_sale_pay"], "qty": 0},
            "admin_val":    {"amount": d["admin_val"],    "qty": 0},
            "vaccine_pay":  {"amount": d["vaccine_pay"],  "qty": 0},
            "delivery_val": {"amount": d["delivery_val"], "qty": 0},
            "mixing_val":   {"amount": d["mixing_val"],   "qty": 0},
            "wash_val":     {"amount": d["wash_val"],     "qty": 0},
            "other_costs":  {"amount": d["other_costs"],  "qty": 0},
        }
        revenues_dict = {
            "offal_val":    {"amount": d["offal_val"],    "qty": 0},
            "feed_sale":    {"amount": d["feed_sale"],    "qty": d["feed_sale_qty"]},
            "feed_trans_r": {"amount": d["feed_trans_r"], "qty": d["feed_trans_r_qty"]},
            "feed_rem_val": {"amount": d["feed_rem_val"], "qty": d["feed_rem_qty"]},
            "drug_return":  {"amount": d["drug_return"],  "qty": 0},
            "gas_return":   {"amount": d["gas_return"],   "qty": 0},
        }
        db.save_batch_costs(b_id, costs_dict)
        db.save_batch_revenues(b_id, revenues_dict)

        messagebox.showinfo("تم", "تم حفظ الدفعة بنجاح", parent=self)
        if self.on_save: self.on_save()
        self.destroy()

class DashboardWindow(ToplevelBase):
    def __init__(self, master):
        super().__init__(master)
        self.title("لوحة القياس والرسوم البيانية (Dashboard)")
        self.geometry("1100x650")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.grab_set(); self._build()

    def _build(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=10); hdr.pack(fill="x")
        UILabel(hdr, text="📈 لوحة القياس التفاعلية (Dashboard)", font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)

        if not HAS_MATPLOTLIB: return UILabel(self, text="مكتبة الرسوم البيانية (matplotlib) غير مثبتة", font=FT_HEADER, fg="red").pack(pady=50)

        batches = db.fetch_all("SELECT * FROM v_batches ORDER BY date_in ASC")
        if not batches: return UILabel(self, text="لا توجد بيانات كافية لعرض الرسوم البيانية.", font=FT_HEADER).pack(pady=50)

        labels = [f"دفعة {b['batch_num'] or b['id']}" for b in batches]; nets = [b['net_result'] or 0 for b in batches]; morts = [b['mort_rate'] or 0 for b in batches]
        colors = [CLR["profit"] if n >= 0 else CLR["loss"] for n in nets]

        fig = Figure(figsize=(12, 5), dpi=100); fig.patch.set_facecolor(CLR["bg"])
        ax1 = fig.add_subplot(121)
        ax1.bar(labels, nets, color=colors); ax1.set_title(prepare_text("صافي الأرباح والخسائر لكل دفعة"), fontsize=14, pad=10); ax1.axhline(0, color='black', linewidth=1.2); ax1.tick_params(axis='x', rotation=45)

        ax2 = fig.add_subplot(122)
        ax2.plot(labels, morts, marker='o', color=CLR["nav"], linestyle='-', linewidth=2.5, markersize=8)
        ax2.set_title(prepare_text("معدل النافق الكلي (%)"), fontsize=14, pad=10); ax2.set_ylim(bottom=0); ax2.grid(True, linestyle='--', alpha=0.6); ax2.tick_params(axis='x', rotation=45)

        fig.tight_layout(pad=3.0)
        canvas = FigureCanvasTkAgg(fig, master=self); canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=15, pady=15)

class WarehousesReportWindow(ToplevelBase):
    def __init__(self, master):
        super().__init__(master)
        self.title("تقرير العنابر الشامل")
        self.geometry("1200x700")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self._build(); self._load()

    def _build(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=10); hdr.pack(fill="x")
        UILabel(hdr, text="📊 تقرير العنابر الشامل", font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)
        btn_frm = UIFrame(self, bg=CLR["nav"], pady=6); btn_frm.pack(fill="x")
        UIButton(btn_frm, text="📥 تصدير Excel (تحليلي شامل)", font=FT_BODY, bg=CLR["white"], fg=CLR["profit"], padx=10, pady=4, cursor="hand2", relief="flat", command=self._export_excel).pack(side="right", padx=6)

        nb = ttk.Notebook(self); nb.pack(fill="both", expand=True, padx=8, pady=8)
        self.tab_by_wh = UIFrame(nb, bg=CLR["bg"]); self.tab_overall = UIFrame(nb, bg=CLR["bg"])
        nb.add(self.tab_by_wh, text="📦 ملخص حسب العنبر"); nb.add(self.tab_overall, text="🏭 الملخص الإجمالي")

        cols_wh = ("العنبر","عدد الدفعات","إجمالي الكتاكيت","إجمالي التكاليف","إجمالي الإيرادات","صافي الربح/الخسارة","متوسط النافق%","متوسط سعر البيع")
        self.tree_wh = ttk.Treeview(self.tab_by_wh, columns=cols_wh, show="headings", selectmode="browse")
        for c, w in zip(cols_wh, [160, 100, 120, 140, 140, 150, 120, 130]): self.tree_wh.heading(c, text=c, anchor="center"); self.tree_wh.column(c, width=w, anchor="center")
        self.tree_wh.tag_configure("profit", background="#f0f9ea"); self.tree_wh.tag_configure("loss", background="#fff0f0"); self.tree_wh.pack(fill="both", expand=True)

        cols_all = ("رقم الدفعة","العنبر","تاريخ الدخول","تاريخ الخروج","الأيام","الكتاكيت","التكاليف","الإيرادات","الربح/الخسارة","النافق%","متوسط البيع","نصيب الشركة")
        self.tree_all = ttk.Treeview(self.tab_overall, columns=cols_all, show="headings", selectmode="browse")
        for c, w in zip(cols_all, [80, 130, 95, 95, 55, 85, 110, 110, 120, 70, 100, 110]): self.tree_all.heading(c, text=c, anchor="center"); self.tree_all.column(c, width=w, anchor="center")
        self.tree_all.tag_configure("profit", background="#f0f9ea"); self.tree_all.tag_configure("loss", background="#fff0f0"); self.tree_all.pack(fill="both", expand=True)

        self.sum_frame = UIFrame(self, bg=CLR["info_bg"], pady=8, padx=12); self.sum_frame.pack(fill="x")

    def _load(self):
        batches = db.fetch_all("SELECT * FROM v_batches ORDER BY warehouse_name, date_in")
        wh_data = {}
        for b in batches:
            wh = b["warehouse_name"]
            if wh not in wh_data: wh_data[wh] = {"count":0,"chicks":0,"cost":0,"rev":0,"net":0,"mort_sum":0,"sold_sum":0,"cust_mkt_val_sum":0}
            d = wh_data[wh]; d["count"] += 1; d["chicks"] += b["chicks"] or 0; d["cost"] += b["total_cost"] or 0; d["rev"] += b["total_rev"] or 0; d["net"] += b["net_result"] or 0; d["mort_sum"] += b["mort_rate"] or 0; d["sold_sum"] += (b["total_sold"] or 0); d["cust_mkt_val_sum"] += (b["total_rev"] or 0)

        self.tree_wh.delete(*self.tree_wh.get_children())
        for wh, d in wh_data.items():
            self.tree_wh.insert("", "end", tags=("profit" if d["net"] >= 0 else "loss",), values=(wh, d["count"], fmt_num(d["chicks"]), fmt_num(d["cost"]), fmt_num(d["rev"]), f"{'+'if d['net']>=0 else ''}{fmt_num(d['net'])}", f"{(d['mort_sum']/d['count'] if d['count']>0 else 0):.1f}%", fmt_num(d["cust_mkt_val_sum"]/d["sold_sum"] if d["sold_sum"]>0 else 0)))

        self.tree_all.delete(*self.tree_all.get_children())
        T = {"chicks":0,"cost":0,"rev":0,"net":0,"share":0}
        for b in batches:
            b_num = b["batch_num"] if b["batch_num"] else str(b["id"])
            self.tree_all.insert("", "end", iid=str(b["id"]), tags=("profit" if (b["net_result"] or 0) >= 0 else "loss",), values=(b_num, b["warehouse_name"], b["date_in"], b["date_out"], b["days"] or "", fmt_num(b["chicks"]), fmt_num(b["total_cost"]), fmt_num(b["total_rev"]), f"{'+'if (b['net_result'] or 0)>=0 else ''}{fmt_num(b['net_result'])}", f"{b['mort_rate'] or 0:.1f}%", fmt_num(b["avg_price"]), fmt_num(b["share_val"])))
            T["chicks"] += b["chicks"] or 0; T["cost"] += b["total_cost"] or 0; T["rev"] += b["total_rev"] or 0; T["net"] += b["net_result"] or 0; T["share"] += b["share_val"] or 0

        for w in self.sum_frame.winfo_children(): w.destroy()
        for lbl, val in [("الدفعات", str(len(batches))), ("إجمالي الكتاكيت", fmt_num(T["chicks"])), ("إجمالي التكاليف", fmt_num(T["cost"])), ("إجمالي الإيرادات", fmt_num(T["rev"])), ("صافي النتيجة", f"{'+'if T['net']>=0 else ''}{fmt_num(T['net'])}"), ("نصيب الشركة", fmt_num(T["share"]))]:
            f = UIFrame(self.sum_frame, bg=CLR["white"], padx=12, pady=4, relief="solid", bd=1); f.pack(side="right", padx=4)
            UILabel(f, text=lbl, font=FT_TINY, bg=CLR["white"], fg=CLR["text2"]).pack()
            UILabel(f, text=val, font=("Arial",11,"bold"), bg=CLR["white"], fg=CLR["profit"] if "الإيرادات" in lbl or ("صافي" in lbl and "+" in val) else CLR["loss"] if "التكاليف" in lbl else CLR["header"]).pack()

    def _export_excel(self):
        if not HAS_OPENPYXL:
            return messagebox.showerror("خطأ", "يرجى تثبيت مكتبة openpyxl:\npip install openpyxl", parent=self)
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"تقرير_الدفعات_الشامل_{datetime.now().strftime('%Y%m%d')}.xlsx", parent=self)
        if not path: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "التقرير الشامل"; ws.sheet_view.rightToLeft = True
            
            hdr_fill  = PatternFill("solid", fgColor="1F4E79")
            cost_fill = PatternFill("solid", fgColor="FCE4D6")
            rev_fill  = PatternFill("solid", fgColor="E2EFDA")
            p_fill    = PatternFill("solid", fgColor="E2EFDA")
            l_fill    = PatternFill("solid", fgColor="FCE4D6")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            brd = Border(left=Side(style="thin",color="BBBBBB"), right=Side(style="thin",color="BBBBBB"),
                         top=Side(style="thin",color="BBBBBB"), bottom=Side(style="thin",color="BBBBBB"))

            cost_types = db.get_cost_types(active_only=True)
            rev_types  = db.get_revenue_types(active_only=True)
            batches    = db.fetch_all("SELECT * FROM v_batches ORDER BY warehouse_name, date_in")

            # ── بناء الرؤوس ─────────────────────────────────────
            headers = ["اسم العنبر", "رقم الدفعة", "تاريخ الدخول", "تاريخ الخروج", "الأيام"]
            # تكاليف تحليلية
            for ct in cost_types:
                if ct.get("has_qty"):
                    headers.append(f"{ct['name_ar']} (كمية)")
                headers.append(f"{ct['name_ar']} (قيمة)")
            headers.append("إجمالي التكاليف")
            # إيرادات تحليلية
            for rt in rev_types:
                if rt.get("has_qty"):
                    headers.append(f"{rt['name_ar']} (كمية)")
                headers.append(f"{rt['name_ar']} (قيمة)")
            headers.append("إجمالي المبيعات")
            # النتائج والطيور
            headers += ["نتيجة الدفعة", "نصيب الشركة", "المباع في العنبر", "المباع في السوق", "الوفيات", "الاستهلاك (ضيافة)"]

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(1, ci, h)
                cell.font = header_font; cell.fill = hdr_fill; cell.alignment = center
                ws.column_dimensions[get_column_letter(ci)].width = 15

            # ── تعبئة البيانات ──────────────────────────────────
            for ri, b in enumerate(batches, 2):
                bid = b["id"]
                b_costs = db.get_batch_costs(bid)
                b_revs  = db.get_batch_revenues(bid)
                
                row = [b["warehouse_name"], b["batch_num"] or str(bid), b["date_in"], b["date_out"], b["days"]]
                
                # إضافة التكاليف
                for ct in cost_types:
                    c_data = b_costs.get(ct["code"], {})
                    amt = c_data.get("amount", 0)
                    qty = c_data.get("qty", 0)
                    if not amt and not qty:
                        amt = b[ct["code"]] if ct["code"] in b.keys() else 0
                        qty = b[ct["code"].replace("_val", "_qty")] if ct["code"].replace("_val", "_qty") in b.keys() else 0
                    
                    if ct.get("has_qty"):
                        row.append(qty or 0)
                    row.append(amt or 0)
                
                row.append(b["total_cost"] or 0)
                
                # إضافة الإيرادات
                for rt in rev_types:
                    r_data = b_revs.get(rt["code"], {})
                    amt = r_data.get("amount", 0)
                    qty = r_data.get("qty", 0)
                    if not amt and not qty:
                        amt = b[rt["code"]] if rt["code"] in b.keys() else 0
                        qty = b[rt["code"].replace("_val", "_qty")] if rt["code"].replace("_val", "_qty") in b.keys() else 0
                    
                    if rt.get("has_qty"):
                        row.append(qty or 0)
                    row.append(amt or 0)

                row.append(b["total_rev"] or 0)
                
                net = b["net_result"] or 0
                row += [net, b["share_val"] or 0, b["cust_qty"] or 0, b["mkt_qty"] or 0, b["total_dead"] or 0, b["consumed_birds"] or 0]
                
                for ci, v in enumerate(row, 1):
                    cell = ws.cell(ri, ci, v); cell.alignment = center; cell.border = brd
                    if ci > 5: # تنسيق أرقام (تخطي أول 5 أعمدة وصفية)
                        if isinstance(v, (int, float)):
                            cell.number_format = "#,##0.##" if ci > (len(headers)-4) else "#,##0"
                
                # تلوين الخلفية للأعمدة (ديناميكي)
                curr_col = 6
                # تلوين التكاليف
                for ct in cost_types:
                    if ct.get("has_qty"):
                        ws.cell(ri, curr_col).fill = cost_fill
                        curr_col += 1
                    ws.cell(ri, curr_col).fill = cost_fill
                    curr_col += 1
                ws.cell(ri, curr_col).fill = cost_fill # إجمالي التكاليف
                curr_col += 1
                
                # تلوين الإيرادات
                for rt in rev_types:
                    if rt.get("has_qty"):
                        ws.cell(ri, curr_col).fill = rev_fill
                        curr_col += 1
                    ws.cell(ri, curr_col).fill = rev_fill
                    curr_col += 1
                ws.cell(ri, curr_col).fill = rev_fill # إجمالي المبيعات
                curr_col += 1
                
                # النتيجة ونصيب الشركة
                color = p_fill if net >= 0 else l_fill
                ws.cell(ri, curr_col).fill = color
                ws.cell(ri, curr_col + 1).fill = color

            wb.save(path)
            messagebox.showinfo("تم ✅", f"تم تصدير التقرير الشامل بنجاح!\n\nكل دفعة في سطر واحد مع كامل التفاصيل.\n{path}", parent=self)
            try: os.startfile(path)
            except: pass

        except PermissionError:
            messagebox.showerror("خطأ ❌", f"الملف مفتوح حالياً في Excel!\nيرجى إغلاقه أولاً.\n{path}", parent=self)
        except Exception as _ex:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء التصدير:\n{_ex}", parent=self)
            messagebox.showerror("خطأ في التصدير", f"{type(_ex).__name__}: {_ex}", parent=self)

# ════════════════════════════════════════════════════════════════
class AdvancedAnalyticsWindow(ToplevelBase):
    """
    نافذة التحليلات المتقدمة v4.6
    تشمل:
      - تبويب 1: مقارنة العنابر (FCR / نافق% / ربحية / كفاءة)
      - تبويب 2: التحليل السنوي والموسمي (ربح سنوي + موسمية)
      - تبويب 3: رسوم بيانية (Matplotlib مدمج)
    """

    MONTHS_AR = {1:"يناير",2:"فبراير",3:"مارس",4:"أبريل",
                 5:"مايو",6:"يونيو",7:"يوليو",8:"أغسطس",
                 9:"سبتمبر",10:"أكتوبر",11:"نوفمبر",12:"ديسمبر"}
    SEASONS   = {(12,1,2):"شتاء ❄️",(3,4,5):"ربيع 🌸",
                 (6,7,8):"صيف ☀️",(9,10,11):"خريف 🍂"}

    def __init__(self, master):
        super().__init__(master)
        self.title("📊 التحليلات المتقدمة — مقارنة العنابر والتحليل السنوي")
        self.geometry("1300x750")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.resizable(True, True)
        self.grab_set()
        self._build()

    # ──────────────────────────────────────────────────────────────
    def _get_season(self, month):
        for months, name in self.SEASONS.items():
            if month in months:
                return name
        return "غير محدد"

    def _build(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=10)
        hdr.pack(fill="x")
        UILabel(hdr, text="📊 التحليلات المتقدمة — مقارنة العنابر والتحليل السنوي",
                font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)
        UIButton(hdr, text="📥 تصدير Excel", font=FT_SMALL,
                 bg="#2c6fad", fg="white", relief="flat", cursor="hand2",
                 command=self._export_excel).pack(side="left", padx=10)

        style = ttk.Style()
        if not HAS_TTKB: style.theme_use("default")
        style.configure("TNotebook.Tab", font=FT_HEADER, padding=[14, 6])

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=8, pady=8)

        self.tab_compare = UIFrame(nb, bg=CLR["bg"])
        self.tab_annual  = UIFrame(nb, bg=CLR["bg"])
        self.tab_charts  = UIFrame(nb, bg=CLR["bg"])

        nb.add(self.tab_compare, text="🏆 مقارنة العنابر")
        nb.add(self.tab_annual,  text="📅 التحليل السنوي والموسمي")
        nb.add(self.tab_charts,  text="📈 الرسوم البيانية")

        self._build_compare_tab()
        self._build_annual_tab()
        self._build_charts_tab()

    # ══════════════════════════════════════════════════════════════
    # تبويب 1: مقارنة العنابر
    # ══════════════════════════════════════════════════════════════
    def _build_compare_tab(self):
        F = self.tab_compare

        # ── شريط الفلترة ──
        fbar = UIFrame(F, bg=CLR["info_bg"], pady=6, padx=10)
        fbar.pack(fill="x")
        UILabel(fbar, text="عرض:", font=FT_SMALL, bg=CLR["info_bg"]).pack(side="right", padx=4)
        self.v_compare_mode = tk.StringVar(value="all")
        for val, lbl in [("all","كل الدفعات"),("by_wh","حسب العنبر"),("by_year","حسب السنة")]:
            tk.Radiobutton(fbar, text=lbl, variable=self.v_compare_mode,
                          value=val, bg=CLR["info_bg"], font=FT_SMALL,
                          command=self._load_compare).pack(side="right", padx=6)
        UILabel(fbar, text="💡 اضغط على رأس العمود للترتيب",
                font=FT_TINY, bg=CLR["info_bg"], fg=CLR["text2"]).pack(side="left", padx=8)

        # ── الجدول ──
        cols = (
            "العنبر","رقم الدفعة","تاريخ الدخول","الأيام","الكتاكيت",
            "FCR","نافق%","متوسط وزن(كجم)",
            "سعر بيع متوسط","تكلفة/طائر","إيراد/طائر","ربح/طائر",
            "إجمالي ربح","نصيب الشركة",
            "تصنيف الأداء"
        )
        frm = UIFrame(F, bg=CLR["bg"])
        frm.pack(fill="both", expand=True, padx=8, pady=6)

        self.tree_cmp = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")
        widths = [120,80,95,55,80, 65,65,90, 90,90,90,90, 100,100, 100]
        for c, w in zip(cols, widths):
            self.tree_cmp.heading(c, text=c, anchor="center",
                                  command=lambda _c=c: self._sort_compare(_c))
            self.tree_cmp.column(c, width=w, anchor="center")

        self.tree_cmp.tag_configure("excellent", background="#e2efda", foreground="#27680a")
        self.tree_cmp.tag_configure("good",      background="#f0f9ea", foreground="#3B6D11")
        self.tree_cmp.tag_configure("average",   background="#fff2cc", foreground="#bf9000")
        self.tree_cmp.tag_configure("poor",      background="#fce4d6", foreground="#c00000")

        sbx = ttk.Scrollbar(frm, orient="horizontal", command=self.tree_cmp.xview)
        sby = ttk.Scrollbar(frm, orient="vertical",   command=self.tree_cmp.yview)
        self.tree_cmp.configure(xscrollcommand=sbx.set, yscrollcommand=sby.set)
        sbx.pack(side="bottom", fill="x")
        sby.pack(side="left",   fill="y")
        self.tree_cmp.pack(fill="both", expand=True)

        # ── ملخص إحصائي ──
        self.sum_frm_cmp = UIFrame(F, bg=CLR["info_bg"], pady=6, padx=12)
        self.sum_frm_cmp.pack(fill="x")

        self._sort_col = None
        self._sort_rev = False
        self._load_compare()

    def _load_compare(self):
        self.tree_cmp.delete(*self.tree_cmp.get_children())
        for w in self.sum_frm_cmp.winfo_children():
            w.destroy()

        batches = db.fetch_all(
            "SELECT * FROM v_batches ORDER BY warehouse_name, date_in")
        if not batches:
            UILabel(self.sum_frm_cmp, text="لا توجد بيانات",
                    font=FT_BODY, bg=CLR["info_bg"]).pack()
            return

        rows = []
        for b in batches:
            bid     = b["id"]
            chicks  = b["chicks"] or 1
            days    = b["days"]   or 1
            sold    = b["total_sold"] or 0
            dead    = b["total_dead"] or 0
            cost    = b["total_cost"] or 0
            rev     = b["total_rev"]  or 0
            net     = b["net_result"] or 0
            share   = b["share_val"]  or 0
            fcr     = b["fcr"]        or 0
            mort    = b["mort_rate"]  or 0
            avg_w   = b["avg_weight"] or 0
            avg_p   = b["avg_price"]  or 0

            cost_per = round(cost / max(sold,1), 2)
            rev_per  = round(rev  / max(sold,1), 2)
            net_per  = round(net  / max(sold,1), 2)

            # تصنيف الأداء بناءً على 3 معايير
            score = 0
            if fcr   > 0 and fcr   <= 1.7: score += 2
            elif fcr > 0 and fcr   <= 2.0: score += 1
            if mort  <= 3:   score += 2
            elif mort <= 5:  score += 1
            if net_per >= 5: score += 2
            elif net_per >= 2: score += 1

            if score >= 5:   perf = "ممتاز ⭐⭐⭐"
            elif score >= 3: perf = "جيد ⭐⭐"
            elif score >= 1: perf = "متوسط ⭐"
            else:            perf = "ضعيف"

            tag = {"ممتاز ⭐⭐⭐":"excellent","جيد ⭐⭐":"good",
                   "متوسط ⭐":"average","ضعيف":"poor"}[perf]

            rows.append({
                "wh":       b["warehouse_name"],
                "bnum":     b["batch_num"] or str(bid),
                "date_in":  b["date_in"],
                "days":     b["days"] or 0,
                "chicks":   chicks,
                "fcr":      fcr,
                "mort":     mort,
                "avg_w":    avg_w,
                "avg_p":    avg_p,
                "cost_per": cost_per,
                "rev_per":  rev_per,
                "net_per":  net_per,
                "net":      net,
                "share":    share,
                "perf":     perf,
                "tag":      tag,
                "net_raw":  net,
            })

        # فلترة حسب الوضع
        mode = self.v_compare_mode.get()
        if mode == "by_wh":
            # تجميع حسب العنبر
            wh_data = {}
            for r in rows:
                wh = r["wh"]
                if wh not in wh_data:
                    wh_data[wh] = {"count":0,"fcr_sum":0,"mort_sum":0,
                                    "net":0,"share":0,"chicks":0,"sold":0}
                d = wh_data[wh]
                d["count"]    += 1
                d["fcr_sum"]  += r["fcr"]
                d["mort_sum"] += r["mort"]
                d["net"]      += r["net"]
                d["share"]    += r["share"]
                d["chicks"]   += r["chicks"]

            rows = []
            for wh, d in wh_data.items():
                cnt = d["count"] or 1
                avg_fcr  = round(d["fcr_sum"]  / cnt, 3)
                avg_mort = round(d["mort_sum"] / cnt, 2)
                avg_net_per = round(d["net"] / max(d["chicks"],1), 2)
                score = 0
                if avg_fcr  <= 1.7: score += 2
                elif avg_fcr <= 2.0: score += 1
                if avg_mort <= 3: score += 2
                elif avg_mort <= 5: score += 1
                if avg_net_per >= 5: score += 2
                elif avg_net_per >= 2: score += 1
                perf = ("ممتاز ⭐⭐⭐" if score>=5 else "جيد ⭐⭐" if score>=3
                        else "متوسط ⭐" if score>=1 else "ضعيف")
                tag  = {"ممتاز ⭐⭐⭐":"excellent","جيد ⭐⭐":"good",
                        "متوسط ⭐":"average","ضعيف":"poor"}[perf]
                rows.append({"wh":wh,"bnum":f"{cnt} دفعة","date_in":"—","days":"—",
                              "chicks":d["chicks"],"fcr":avg_fcr,"mort":avg_mort,
                              "avg_w":0,"avg_p":0,"cost_per":0,"rev_per":0,
                              "net_per":avg_net_per,"net":d["net"],"share":d["share"],
                              "perf":perf,"tag":tag,"net_raw":d["net"]})

        # عرض الصفوف
        for r in rows:
            self.tree_cmp.insert("", "end", tags=(r["tag"],), values=(
                r["wh"], r["bnum"], r["date_in"], r["days"],
                f"{r['chicks']:,}",
                f"{r['fcr']:.3f}" if r["fcr"] else "—",
                f"{r['mort']:.2f}%",
                f"{r['avg_w']:.2f}" if r["avg_w"] else "—",
                f"{r['avg_p']:,.2f}" if r["avg_p"] else "—",
                f"{r['cost_per']:,.2f}",
                f"{r['rev_per']:,.2f}",
                f"{r['net_per']:,.2f}",
                f"{r['net']:,.0f}",
                f"{r['share']:,.0f}",
                r["perf"],
            ))

        # إحصائيات ملخصة
        total_net  = sum(r["net_raw"] for r in rows)
        best  = max(rows, key=lambda x: x["net_raw"], default=None)
        worst = min(rows, key=lambda x: x["net_raw"], default=None)
        best_fcr  = min((r for r in rows if r["fcr"] > 0),
                        key=lambda x: x["fcr"], default=None)
        best_mort = min(rows, key=lambda x: x["mort"], default=None)

        kpis = [
            ("إجمالي الربح/خسارة", f"{total_net:+,.0f}", "#dce6f1"),
            ("أفضل دفعة ربحاً", f"{best['wh']} ({best['bnum']})" if best else "—", "#e2efda"),
            ("أعلى خسارة",       f"{worst['wh']} ({worst['bnum']})" if worst else "—", "#fce4d6"),
            ("أفضل FCR",         f"{best_fcr['fcr']:.3f} ({best_fcr['wh']})" if best_fcr else "—", "#fff2cc"),
            ("أقل نفوق%",        f"{best_mort['mort']:.2f}% ({best_mort['wh']})" if best_mort else "—", "#e2efda"),
        ]
        for lbl, val, bg in kpis:
            f = UIFrame(self.sum_frm_cmp, bg=bg, padx=10, pady=4, relief="solid", bd=1)
            f.pack(side="right", padx=4)
            UILabel(f, text=lbl, font=FT_TINY,  bg=bg, fg=CLR["text2"]).pack()
            UILabel(f, text=val, font=("Arial",10,"bold"), bg=bg, fg=CLR["text"]).pack()

    def _sort_compare(self, col):
        """ترتيب الجدول عند الضغط على رأس العمود"""
        items = [(self.tree_cmp.set(k, col), k)
                 for k in self.tree_cmp.get_children("")]
        rev = (self._sort_col == col and not self._sort_rev)
        try:
            items.sort(key=lambda x: float(x[0].replace(",","").replace("%","").replace("+","")) , reverse=rev)
        except ValueError:
            items.sort(reverse=rev)
        for i, (_, k) in enumerate(items):
            self.tree_cmp.move(k, "", i)
        self._sort_col = col; self._sort_rev = rev

    # ══════════════════════════════════════════════════════════════
    # تبويب 2: التحليل السنوي والموسمي
    # ══════════════════════════════════════════════════════════════
    def _build_annual_tab(self):
        F = self.tab_annual

        # ── فلتر السنة ──
        fbar = UIFrame(F, bg=CLR["info_bg"], pady=6, padx=10)
        fbar.pack(fill="x")
        UILabel(fbar, text="السنة:", font=FT_SMALL, bg=CLR["info_bg"]).pack(side="right", padx=4)
        years = [str(r["yr"]) for r in db.fetch_all(
            "SELECT DISTINCT strftime('%Y', date_in) AS yr FROM batches ORDER BY yr DESC") if r["yr"]]
        self.v_year = tk.StringVar(value=years[0] if years else "")
        self.cbo_year = ttk.Combobox(fbar, textvariable=self.v_year,
                                      values=["الكل"] + years, width=8,
                                      state="readonly", font=FT_BODY)
        self.cbo_year.pack(side="right", padx=4)
        if years: self.cbo_year.set(years[0])
        UIButton(fbar, text="🔄 تحديث", font=FT_SMALL, relief="flat",
                 cursor="hand2", command=self._load_annual).pack(side="right", padx=6)

        # ── تبويب فرعي: شهري / موسمي ──
        sub_nb = ttk.Notebook(F)
        sub_nb.pack(fill="both", expand=True, padx=6, pady=6)

        self.sub_monthly  = UIFrame(sub_nb, bg=CLR["bg"])
        self.sub_seasonal = UIFrame(sub_nb, bg=CLR["bg"])
        self.sub_wh_year  = UIFrame(sub_nb, bg=CLR["bg"])
        sub_nb.add(self.sub_monthly,  text="📅 شهري")
        sub_nb.add(self.sub_seasonal, text="🍂 موسمي")
        sub_nb.add(self.sub_wh_year,  text="🏭 حسب العنبر × السنة")

        # جدول شهري
        m_cols = ("الشهر","السنة","عدد الدفعات","إجمالي الكتاكيت",
                  "إجمالي التكاليف","إجمالي الإيرادات",
                  "صافي الربح/خسارة","نصيب الشركة",
                  "متوسط FCR","متوسط النافق%")
        self.tree_monthly = ttk.Treeview(self.sub_monthly,
                                          columns=m_cols, show="headings")
        mw = [90,60,90,110,120,120,120,110,90,100]
        for c, w in zip(m_cols, mw):
            self.tree_monthly.heading(c, text=c, anchor="center")
            self.tree_monthly.column(c, width=w, anchor="center")
        self.tree_monthly.tag_configure("profit", background="#f0f9ea")
        self.tree_monthly.tag_configure("loss",   background="#fff0f0")
        sb = ttk.Scrollbar(self.sub_monthly, command=self.tree_monthly.yview)
        self.tree_monthly.configure(yscrollcommand=sb.set)
        sb.pack(side="left", fill="y")
        self.tree_monthly.pack(fill="both", expand=True)

        # جدول موسمي
        s_cols = ("الموسم","السنة","عدد الدفعات",
                  "إجمالي التكاليف","إجمالي الإيرادات",
                  "صافي الربح/خسارة","نصيب الشركة",
                  "متوسط FCR","متوسط النافق%","أفضل عنبر")
        self.tree_seasonal = ttk.Treeview(self.sub_seasonal,
                                           columns=s_cols, show="headings")
        sw = [100,60,90,120,120,120,110,90,100,120]
        for c, w in zip(s_cols, sw):
            self.tree_seasonal.heading(c, text=c, anchor="center")
            self.tree_seasonal.column(c, width=w, anchor="center")
        self.tree_seasonal.tag_configure("profit", background="#f0f9ea")
        self.tree_seasonal.tag_configure("loss",   background="#fff0f0")
        sb2 = ttk.Scrollbar(self.sub_seasonal, command=self.tree_seasonal.yview)
        self.tree_seasonal.configure(yscrollcommand=sb2.set)
        sb2.pack(side="left", fill="y")
        self.tree_seasonal.pack(fill="both", expand=True)

        # جدول عنبر × سنة
        wy_cols = ("العنبر","السنة","عدد الدفعات","إجمالي الكتاكيت",
                   "إجمالي التكاليف","إجمالي الإيرادات",
                   "صافي الربح/خسارة","متوسط FCR","متوسط النافق%")
        self.tree_wh_year = ttk.Treeview(self.sub_wh_year,
                                          columns=wy_cols, show="headings")
        yw = [130,60,90,110,120,120,120,90,100]
        for c, w in zip(wy_cols, yw):
            self.tree_wh_year.heading(c, text=c, anchor="center")
            self.tree_wh_year.column(c, width=w, anchor="center")
        self.tree_wh_year.tag_configure("profit", background="#f0f9ea")
        self.tree_wh_year.tag_configure("loss",   background="#fff0f0")
        sb3 = ttk.Scrollbar(self.sub_wh_year, command=self.tree_wh_year.yview)
        self.tree_wh_year.configure(yscrollcommand=sb3.set)
        sb3.pack(side="left", fill="y")
        self.tree_wh_year.pack(fill="both", expand=True)

        # ── إجماليات في الأسفل ──
        self.sum_frm_ann = UIFrame(F, bg=CLR["info_bg"], pady=6, padx=12)
        self.sum_frm_ann.pack(fill="x")

        self._load_annual()

    def _load_annual(self):
        yr_filter = self.v_year.get()
        if yr_filter == "الكل" or not yr_filter:
            batches = db.fetch_all("SELECT * FROM v_batches ORDER BY date_in")
        else:
            batches = db.fetch_all(
                "SELECT * FROM v_batches WHERE strftime('%Y', date_in)=? ORDER BY date_in",
                (yr_filter,))

        # ── تجميع شهري ──
        monthly = {}
        for b in batches:
            try:
                dt  = datetime.strptime(b["date_in"], "%Y-%m-%d")
                key = (dt.year, dt.month)
            except: continue
            if key not in monthly:
                monthly[key] = {"count":0,"chicks":0,"cost":0,"rev":0,
                                "net":0,"share":0,"fcr_sum":0,"mort_sum":0,"fcr_cnt":0}
            d = monthly[key]
            d["count"]   += 1
            d["chicks"]  += b["chicks"] or 0
            d["cost"]    += b["total_cost"] or 0
            d["rev"]     += b["total_rev"]  or 0
            d["net"]     += b["net_result"] or 0
            d["share"]   += b["share_val"]  or 0
            if b["fcr"] and b["fcr"] > 0:
                d["fcr_sum"] += b["fcr"]; d["fcr_cnt"] += 1
            d["mort_sum"] += b["mort_rate"] or 0

        self.tree_monthly.delete(*self.tree_monthly.get_children())
        for (yr, mo) in sorted(monthly.keys()):
            d = monthly[(yr, mo)]
            cnt = d["count"] or 1
            avg_fcr  = round(d["fcr_sum"] / d["fcr_cnt"], 3) if d["fcr_cnt"] else 0
            avg_mort = round(d["mort_sum"] / cnt, 2)
            net = d["net"]
            tag = "profit" if net >= 0 else "loss"
            self.tree_monthly.insert("", "end", tags=(tag,), values=(
                self.MONTHS_AR.get(mo, str(mo)), yr, d["count"],
                f"{d['chicks']:,}", f"{d['cost']:,.0f}",
                f"{d['rev']:,.0f}",
                f"{'+' if net>=0 else ''}{net:,.0f}",
                f"{d['share']:,.0f}",
                f"{avg_fcr:.3f}" if avg_fcr else "—",
                f"{avg_mort:.2f}%",
            ))

        # ── تجميع موسمي ──
        seasonal = {}
        for b in batches:
            try:
                dt  = datetime.strptime(b["date_in"], "%Y-%m-%d")
                mo  = dt.month; yr = dt.year
                season = self._get_season(mo)
                key = (yr, season)
            except: continue
            if key not in seasonal:
                seasonal[key] = {"count":0,"cost":0,"rev":0,"net":0,"share":0,
                                  "fcr_sum":0,"mort_sum":0,"fcr_cnt":0,
                                  "wh_net":{}}
            d = seasonal[key]
            d["count"]   += 1
            d["cost"]    += b["total_cost"] or 0
            d["rev"]     += b["total_rev"]  or 0
            net_b         = b["net_result"] or 0
            d["net"]     += net_b
            d["share"]   += b["share_val"]  or 0
            if b["fcr"] and b["fcr"] > 0:
                d["fcr_sum"] += b["fcr"]; d["fcr_cnt"] += 1
            d["mort_sum"] += b["mort_rate"] or 0
            wh = b["warehouse_name"]
            d["wh_net"][wh] = d["wh_net"].get(wh, 0) + net_b

        self.tree_seasonal.delete(*self.tree_seasonal.get_children())
        for (yr, season) in sorted(seasonal.keys()):
            d = seasonal[(yr, season)]
            cnt = d["count"] or 1
            avg_fcr  = round(d["fcr_sum"]/d["fcr_cnt"],3) if d["fcr_cnt"] else 0
            avg_mort = round(d["mort_sum"]/cnt, 2)
            net = d["net"]
            best_wh = max(d["wh_net"], key=d["wh_net"].get) if d["wh_net"] else "—"
            tag = "profit" if net >= 0 else "loss"
            self.tree_seasonal.insert("", "end", tags=(tag,), values=(
                season, yr, d["count"],
                f"{d['cost']:,.0f}", f"{d['rev']:,.0f}",
                f"{'+' if net>=0 else ''}{net:,.0f}",
                f"{d['share']:,.0f}",
                f"{avg_fcr:.3f}" if avg_fcr else "—",
                f"{avg_mort:.2f}%", best_wh,
            ))

        # ── تجميع عنبر × سنة ──
        wh_year = {}
        for b in batches:
            try:
                yr  = datetime.strptime(b["date_in"], "%Y-%m-%d").year
                wh  = b["warehouse_name"]
                key = (wh, yr)
            except: continue
            if key not in wh_year:
                wh_year[key] = {"count":0,"chicks":0,"cost":0,"rev":0,"net":0,
                                 "fcr_sum":0,"mort_sum":0,"fcr_cnt":0}
            d = wh_year[key]
            d["count"]  += 1
            d["chicks"] += b["chicks"] or 0
            d["cost"]   += b["total_cost"] or 0
            d["rev"]    += b["total_rev"]  or 0
            d["net"]    += b["net_result"] or 0
            if b["fcr"] and b["fcr"] > 0:
                d["fcr_sum"] += b["fcr"]; d["fcr_cnt"] += 1
            d["mort_sum"] += b["mort_rate"] or 0

        self.tree_wh_year.delete(*self.tree_wh_year.get_children())
        for (wh, yr) in sorted(wh_year.keys()):
            d = wh_year[(wh, yr)]
            cnt = d["count"] or 1
            avg_fcr  = round(d["fcr_sum"]/d["fcr_cnt"],3) if d["fcr_cnt"] else 0
            avg_mort = round(d["mort_sum"]/cnt, 2)
            net = d["net"]
            tag = "profit" if net >= 0 else "loss"
            self.tree_wh_year.insert("", "end", tags=(tag,), values=(
                wh, yr, d["count"], f"{d['chicks']:,}",
                f"{d['cost']:,.0f}", f"{d['rev']:,.0f}",
                f"{'+' if net>=0 else ''}{net:,.0f}",
                f"{avg_fcr:.3f}" if avg_fcr else "—",
                f"{avg_mort:.2f}%",
            ))

        # ── ملخص الإجماليات ──
        for w in self.sum_frm_ann.winfo_children(): w.destroy()
        total_cost = sum(b["total_cost"] or 0 for b in batches)
        total_rev  = sum(b["total_rev"]  or 0 for b in batches)
        total_net  = sum(b["net_result"] or 0 for b in batches)
        total_share= sum(b["share_val"]  or 0 for b in batches)
        total_chk  = sum(b["chicks"]     or 0 for b in batches)
        yr_lbl = yr_filter if yr_filter and yr_filter != "الكل" else "كل السنوات"
        kpis = [
            (f"الفترة: {yr_lbl}",           f"{len(batches)} دفعة", "#dce6f1"),
            ("إجمالي الكتاكيت",              f"{total_chk:,}",        "#dce6f1"),
            ("إجمالي التكاليف",              f"{total_cost:,.0f}",     "#fce4d6"),
            ("إجمالي الإيرادات",             f"{total_rev:,.0f}",      "#e2efda"),
            ("صافي الربح/خسارة",             f"{'+' if total_net>=0 else ''}{total_net:,.0f}",
             "#e2efda" if total_net >= 0 else "#fce4d6"),
            ("نصيب الشركة",                  f"{total_share:,.0f}",    "#fff2cc"),
        ]
        for lbl, val, bg in kpis:
            f = UIFrame(self.sum_frm_ann, bg=bg, padx=10, pady=4, relief="solid", bd=1)
            f.pack(side="right", padx=3)
            UILabel(f, text=lbl, font=FT_TINY,  bg=bg, fg=CLR["text2"]).pack()
            UILabel(f, text=val, font=("Arial",11,"bold"), bg=bg, fg=CLR["text"]).pack()

    # ══════════════════════════════════════════════════════════════
    # تبويب 3: الرسوم البيانية
    # ══════════════════════════════════════════════════════════════
    def _build_charts_tab(self):
        F = self.tab_charts

        if not HAS_MATPLOTLIB:
            UILabel(F, text="مكتبة matplotlib غير مثبتة\npip install matplotlib",
                    font=FT_HEADER, bg=CLR["bg"], fg=CLR["loss"]).pack(pady=80)
            return

        fbar = UIFrame(F, bg=CLR["info_bg"], pady=6, padx=10)
        fbar.pack(fill="x")
        UILabel(fbar, text="نوع الرسم:", font=FT_SMALL, bg=CLR["info_bg"]).pack(side="right", padx=4)
        self.v_chart_type = tk.StringVar(value="net")
        charts = [
            ("net",   "صافي الربح لكل دفعة"),
            ("fcr",   "مقارنة FCR"),
            ("mort",  "معدل النافق%"),
            ("annual","الربح السنوي التراكمي"),
            ("season","الربحية الموسمية"),
        ]
        for val, lbl in charts:
            tk.Radiobutton(fbar, text=lbl, variable=self.v_chart_type,
                          value=val, bg=CLR["info_bg"], font=FT_TINY,
                          command=self._draw_chart).pack(side="right", padx=5)

        self.chart_frm = UIFrame(F, bg=CLR["bg"])
        self.chart_frm.pack(fill="both", expand=True)
        self._draw_chart()

    def _draw_chart(self):
        if not HAS_MATPLOTLIB: return
        for w in self.chart_frm.winfo_children(): w.destroy()

        chart_type = self.v_chart_type.get()
        batches = db.fetch_all("SELECT * FROM v_batches ORDER BY date_in")
        if not batches:
            UILabel(self.chart_frm, text="لا توجد بيانات",
                    font=FT_BODY, bg=CLR["bg"]).pack(pady=50)
            return

        fig = Figure(figsize=(13, 5.5), dpi=95)
        fig.patch.set_facecolor(CLR["bg"])
        ax = fig.add_subplot(111)
        ax.set_facecolor("#f8fafc")

        if chart_type == "net":
            labels = [f"{b['warehouse_name']}\n{b['batch_num'] or b['id']}" for b in batches]
            vals   = [b["net_result"] or 0 for b in batches]
            colors = [CLR["profit"] if v >= 0 else CLR["loss"] for v in vals]
            bars = ax.bar(range(len(labels)), vals, color=colors, width=0.6, edgecolor="white", linewidth=0.5)
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, fontsize=8, rotation=30, ha="right")
            ax.axhline(0, color="gray", linewidth=0.8, linestyle="--")
            ax.set_title(prepare_text("صافي الربح / الخسارة لكل دفعة"), fontsize=13)
            ax.set_ylabel(prepare_text("الربح (ريال)"))
            for bar, val in zip(bars, vals):
                ax.text(bar.get_x()+bar.get_width()/2, bar.get_height() + max(abs(v) for v in vals)*0.01,
                        f"{val:,.0f}", ha="center", va="bottom", fontsize=7)

        elif chart_type == "fcr":
            data = [(b["warehouse_name"], b["fcr"] or 0) for b in batches if b["fcr"] and b["fcr"] > 0]
            if not data:
                ax.text(0.5, 0.5, prepare_text("لا توجد بيانات FCR"), transform=ax.transAxes, ha="center")
            else:
                labels, vals = zip(*data)
                labels_short = [f"{l}\n{i+1}" for i,l in enumerate(labels)]
                colors = [CLR["profit"] if v <= 1.8 else CLR["warn"] if v <= 2.2 else CLR["loss"] for v in vals]
                ax.bar(range(len(labels_short)), vals, color=colors, width=0.6)
                ax.axhline(1.8, color=CLR["profit"], linewidth=1.2, linestyle="--", label="هدف 1.8")
                ax.axhline(2.2, color=CLR["loss"],   linewidth=1.2, linestyle="--", label="حد أعلى 2.2")
                ax.set_xticks(range(len(labels_short)))
                ax.set_xticklabels(labels_short, fontsize=8, rotation=30)
                ax.set_title(prepare_text("مقارنة معدل التحويل FCR"), fontsize=13)
                ax.legend(fontsize=9)

        elif chart_type == "mort":
            labels = [f"{b['warehouse_name']}\n{b['batch_num'] or b['id']}" for b in batches]
            vals   = [b["mort_rate"] or 0 for b in batches]
            colors = [CLR["profit"] if v <= 3 else CLR["warn"] if v <= 6 else CLR["loss"] for v in vals]
            ax.bar(range(len(labels)), vals, color=colors, width=0.6)
            ax.axhline(3, color=CLR["profit"], linewidth=1.2, linestyle="--", label="طبيعي ≤3%")
            ax.axhline(6, color=CLR["loss"],   linewidth=1.2, linestyle="--", label="تحذير >6%")
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, fontsize=8, rotation=30, ha="right")
            ax.set_title(prepare_text("معدل النافق% لكل دفعة"), fontsize=13)
            ax.set_ylabel(prepare_text("النافق%"))
            ax.legend(fontsize=9)

        elif chart_type == "annual":
            yr_net = {}
            for b in batches:
                try:
                    yr = datetime.strptime(b["date_in"], "%Y-%m-%d").year
                except: continue
                yr_net[yr] = yr_net.get(yr, 0) + (b["net_result"] or 0)
            if yr_net:
                years = sorted(yr_net.keys())
                vals  = [yr_net[y] for y in years]
                colors = [CLR["profit"] if v >= 0 else CLR["loss"] for v in vals]
                ax.bar(years, vals, color=colors, width=0.5)
                ax.plot(years, vals, "o-", color=CLR["nav"], linewidth=2, markersize=6)
                ax.axhline(0, color="gray", linewidth=0.8, linestyle="--")
                ax.set_xticks(years)
                ax.set_title(prepare_text("الربح السنوي الإجمالي"), fontsize=13)
                ax.set_ylabel(prepare_text("الربح (ريال)"))
                for yr, val in zip(years, vals):
                    ax.text(yr, val + max(abs(v) for v in vals)*0.02,
                            f"{val:,.0f}", ha="center", va="bottom", fontsize=9)

        elif chart_type == "season":
            season_net = {}
            for b in batches:
                try:
                    mo = datetime.strptime(b["date_in"], "%Y-%m-%d").month
                    s  = self._get_season(mo)
                except: continue
                season_net[s] = season_net.get(s, 0) + (b["net_result"] or 0)
            if season_net:
                labels = list(season_net.keys())
                vals   = [season_net[s] for s in labels]
                colors_s = ["#3B8BD4","#1D9E75","#EF9F27","#D85A30"][:len(labels)]
                wedges, texts, autotexts = ax.pie(
                    [abs(v) for v in vals], labels=labels,
                    colors=colors_s, autopct="%1.1f%%",
                    startangle=90, pctdistance=0.8)
                ax.set_title(prepare_text("توزيع الربحية حسب الموسم"), fontsize=13)

        fig.tight_layout(pad=2.0)
        canvas = FigureCanvasTkAgg(fig, master=self.chart_frm)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    # ══════════════════════════════════════════════════════════════
    # تصدير Excel شامل للتحليلات
    # ══════════════════════════════════════════════════════════════
    def _export_excel(self):
        if not HAS_OPENPYXL:
            return messagebox.showerror("خطأ", "يرجى تثبيت openpyxl", parent=self)
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"التحليلات_المتقدمة_{datetime.now().strftime('%Y%m%d')}.xlsx",
            parent=self)
        if not path: return

        wb  = openpyxl.Workbook()
        hdr_fill   = PatternFill("solid", fgColor="1F4E79")
        profit_fill= PatternFill("solid", fgColor="E2EFDA")
        loss_fill  = PatternFill("solid", fgColor="FCE4D6")
        warn_fill  = PatternFill("solid", fgColor="FFF2CC")
        center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
        brd = Border(
            left=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
            top=Side(style="thin", color="AAAAAA"),
            bottom=Side(style="thin", color="AAAAAA"))

        batches = db.fetch_all("SELECT * FROM v_batches ORDER BY warehouse_name, date_in")

        # ── ورقة 1: مقارنة الدفعات ──
        ws1 = wb.active; ws1.title = "مقارنة الدفعات"; ws1.sheet_view.rightToLeft = True
        cmp_hdrs = ["العنبر","رقم الدفعة","تاريخ الدخول","الأيام","الكتاكيت",
                    "FCR","نافق%","متوسط وزن",
                    "سعر/طائر","تكلفة/طائر","إيراد/طائر","ربح/طائر",
                    "إجمالي ربح","نصيب الشركة","تصنيف الأداء"]
        for ci, h in enumerate(cmp_hdrs, 1):
            cell = ws1.cell(1, ci, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill; cell.alignment = center
            ws1.column_dimensions[get_column_letter(ci)].width = 14

        for ri, b in enumerate(batches, 2):
            sold = max(b["total_sold"] or 1, 1)
            cost = b["total_cost"] or 0; rev = b["total_rev"] or 0; net = b["net_result"] or 0
            fcr = b["fcr"] or 0; mort = b["mort_rate"] or 0; avg_w = b["avg_weight"] or 0
            mort_pct = mort
            score = sum([
                2 if fcr <= 1.7 else (1 if fcr <= 2.0 else 0),
                2 if mort <= 3  else (1 if mort <= 5   else 0),
                2 if (net/sold) >= 5 else (1 if (net/sold) >= 2 else 0),
            ])
            perf = ("ممتاز" if score>=5 else "جيد" if score>=3
                    else "متوسط" if score>=1 else "ضعيف")
            row_vals = [
                b["warehouse_name"], b["batch_num"] or str(b["id"]),
                b["date_in"], b["days"] or 0, b["chicks"] or 0,
                round(fcr,3), round(mort_pct,2), round(avg_w,2),
                round(b["avg_price"] or 0, 2),
                round(cost/sold, 2), round(rev/sold, 2), round(net/sold, 2),
                round(net, 0), round(b["share_val"] or 0, 0), perf,
            ]
            for ci, v in enumerate(row_vals, 1):
                cell = ws1.cell(ri, ci, v); cell.alignment = center; cell.border = brd
            fill = profit_fill if net >= 0 else loss_fill
            for ci in range(1, len(row_vals)+1):
                ws1.cell(ri, ci).fill = fill

        # ── ورقة 2: التحليل الشهري ──
        ws2 = wb.create_sheet("التحليل الشهري"); ws2.sheet_view.rightToLeft = True
        m_hdrs = ["الشهر","السنة","عدد الدفعات","إجمالي الكتاكيت",
                  "إجمالي التكاليف","إجمالي الإيرادات",
                  "صافي الربح/خسارة","نصيب الشركة",
                  "متوسط FCR","متوسط النافق%"]
        for ci, h in enumerate(m_hdrs, 1):
            cell = ws2.cell(1, ci, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill; cell.alignment = center
            ws2.column_dimensions[get_column_letter(ci)].width = 15

        monthly = {}
        for b in batches:
            try:
                dt  = datetime.strptime(b["date_in"], "%Y-%m-%d")
                key = (dt.year, dt.month)
            except: continue
            if key not in monthly:
                monthly[key] = {"count":0,"chicks":0,"cost":0,"rev":0,
                                "net":0,"share":0,"fcr_sum":0,"mort_sum":0,"fcr_cnt":0}
            d = monthly[key]
            d["count"]   += 1; d["chicks"] += b["chicks"] or 0
            d["cost"]    += b["total_cost"] or 0; d["rev"] += b["total_rev"] or 0
            d["net"]     += b["net_result"] or 0; d["share"] += b["share_val"] or 0
            if b["fcr"] and b["fcr"] > 0:
                d["fcr_sum"] += b["fcr"]; d["fcr_cnt"] += 1
            d["mort_sum"] += b["mort_rate"] or 0

        for ri, (yr, mo) in enumerate(sorted(monthly.keys()), 2):
            d = monthly[(yr, mo)]
            cnt = d["count"] or 1
            avg_fcr  = round(d["fcr_sum"]/d["fcr_cnt"],3) if d["fcr_cnt"] else 0
            avg_mort = round(d["mort_sum"]/cnt, 2)
            net = d["net"]
            row = [self.MONTHS_AR.get(mo,str(mo)), yr, d["count"],
                   d["chicks"], round(d["cost"],0), round(d["rev"],0),
                   round(net,0), round(d["share"],0),
                   avg_fcr if avg_fcr else "—", avg_mort]
            for ci, v in enumerate(row, 1):
                cell = ws2.cell(ri, ci, v); cell.alignment = center; cell.border = brd
                if ci in [5,6,7,8]: cell.number_format = "#,##0"
            fill = profit_fill if net >= 0 else loss_fill
            for ci in range(1, len(row)+1):
                ws2.cell(ri, ci).fill = fill

        # ── ورقة 3: التحليل الموسمي ──
        ws3 = wb.create_sheet("التحليل الموسمي"); ws3.sheet_view.rightToLeft = True
        s_hdrs = ["الموسم","السنة","عدد الدفعات","إجمالي التكاليف",
                  "إجمالي الإيرادات","صافي الربح/خسارة",
                  "نصيب الشركة","متوسط FCR","متوسط النافق%"]
        for ci, h in enumerate(s_hdrs, 1):
            cell = ws3.cell(1, ci, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill; cell.alignment = center
            ws3.column_dimensions[get_column_letter(ci)].width = 15

        seasonal = {}
        for b in batches:
            try:
                dt = datetime.strptime(b["date_in"], "%Y-%m-%d")
                s  = self._get_season(dt.month)
                key = (dt.year, s)
            except: continue
            if key not in seasonal:
                seasonal[key] = {"count":0,"cost":0,"rev":0,"net":0,"share":0,
                                  "fcr_sum":0,"mort_sum":0,"fcr_cnt":0}
            d = seasonal[key]
            d["count"] += 1; d["cost"] += b["total_cost"] or 0
            d["rev"]   += b["total_rev"] or 0; d["net"] += b["net_result"] or 0
            d["share"] += b["share_val"] or 0
            if b["fcr"] and b["fcr"] > 0:
                d["fcr_sum"] += b["fcr"]; d["fcr_cnt"] += 1
            d["mort_sum"] += b["mort_rate"] or 0

        for ri, (yr, s) in enumerate(sorted(seasonal.keys()), 2):
            d = seasonal[(yr, s)]
            cnt = d["count"] or 1
            avg_fcr  = round(d["fcr_sum"]/d["fcr_cnt"],3) if d["fcr_cnt"] else 0
            avg_mort = round(d["mort_sum"]/cnt, 2)
            net = d["net"]
            row = [s, yr, d["count"], round(d["cost"],0), round(d["rev"],0),
                   round(net,0), round(d["share"],0),
                   avg_fcr if avg_fcr else "—", avg_mort]
            for ci, v in enumerate(row, 1):
                cell = ws3.cell(ri, ci, v); cell.alignment = center; cell.border = brd
                if ci in [4,5,6,7]: cell.number_format = "#,##0"
            fill = profit_fill if net >= 0 else loss_fill
            for ci in range(1, len(row)+1):
                ws3.cell(ri, ci).fill = fill

        try:
            wb.save(path)
            messagebox.showinfo("تم", f"تم تصدير التحليلات بنجاح!\n{path}", parent=self)
            try: os.startfile(path)
            except: pass
        except PermissionError:
            messagebox.showerror("خطأ", f"الملف مفتوح في Excel! أغلقه أولاً.\n{path}", parent=self)
        except Exception as _ex:
            messagebox.showerror("خطأ", f"فشل الحفظ:\n{_ex}", parent=self)



# ════════════════════════════════════════════════════════════════
# مركز الإدخالات — كل عمليات الإدخال في نافذة واحدة
# ════════════════════════════════════════════════════════════════
class BatchSalesReportWindow(ToplevelBase):
    """نافذة تقرير مبيعات دفعة — آجل + نقداً + سوق"""
    def __init__(self, master, batch_id=None):
        super().__init__(master)
        self.title("📋 تقرير مبيعات الدفعة")
        self.geometry("1050x650")
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.resizable(True, True)
        self.grab_set()
        self._bid = batch_id
        self._build()
        if batch_id: self._load(batch_id)

    def _build(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=10); hdr.pack(fill="x")
        UILabel(hdr, text="📋 تقرير مبيعات الدفعة — آجل + نقداً + سوق",
                font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)
        UIButton(hdr, text="🖨 PDF", font=FT_BODY, bg="#ffd966", fg="#333",
                 relief="flat", cursor="hand2", padx=12,
                 command=self._export_pdf).pack(side="left", padx=10)

        # فلتر الدفعة
        fbar = UIFrame(self, bg=CLR["info_bg"], pady=6, padx=12); fbar.pack(fill="x")
        UILabel(fbar, text="الدفعة:", font=FT_SMALL, bg=CLR["info_bg"]).pack(side="right", padx=4)
        batches = db.fetch_all("SELECT id, batch_num, warehouse_name, date_in FROM v_batches ORDER BY date_in DESC")
        items = [f"{b['warehouse_name']} — {b['batch_num'] or b['id']} ({b['date_in']})" for b in batches]
        self._bmap = {items[i]: batches[i]["id"] for i in range(len(items))}
        self.v_batch = tk.StringVar()
        cbo = ttk.Combobox(fbar, textvariable=self.v_batch, values=items, width=55, font=FT_BODY, state="readonly")
        cbo.pack(side="right", padx=4)
        if self._bid:
            for item, bid in self._bmap.items():
                if bid == self._bid: self.v_batch.set(item); break
        cbo.bind("<<ComboboxSelected>>", lambda e: self._load(self._bmap.get(self.v_batch.get())))
        UIButton(fbar, text="🔄", font=FT_SMALL, bg=CLR["info_bg"], relief="flat",
                 command=lambda: self._load(self._bmap.get(self.v_batch.get()))).pack(side="right", padx=2)

        nb = ttk.Notebook(self); nb.pack(fill="both", expand=True, padx=8, pady=6)
        self.tab_ajl  = UIFrame(nb, bg=CLR["bg"]); nb.add(self.tab_ajl,  text="📄 آجل")
        self.tab_cash = UIFrame(nb, bg=CLR["bg"]); nb.add(self.tab_cash, text="💵 نقداً")
        self.tab_mkt  = UIFrame(nb, bg=CLR["bg"]); nb.add(self.tab_mkt,  text="🏪 السوق")
        self.tab_sum  = UIFrame(nb, bg=CLR["bg"]); nb.add(self.tab_sum,  text="📊 الملخص")

        def make_tree(parent, cols, widths):
            frm = UIFrame(parent, bg=CLR["bg"]); frm.pack(fill="both", expand=True)
            t = ttk.Treeview(frm, columns=cols, show="headings")
            for c, w in zip(cols, widths):
                t.heading(c, text=c, anchor="center"); t.column(c, width=w, anchor="center")
            sb = ttk.Scrollbar(frm, command=t.yview); t.configure(yscrollcommand=sb.set)
            sb.pack(side="left", fill="y"); t.pack(fill="both", expand=True)
            return t

        cols_f = ("#","تاريخ","العميل","الكمية","السعر","الإجمالي")
        self.tree_ajl  = make_tree(self.tab_ajl,  cols_f, [35,100,220,80,90,110])
        self.lbl_ajl   = UILabel(self.tab_ajl,  text="", font=FT_BODY, bg=CLR["profit_bg"], fg=CLR["profit"])
        self.lbl_ajl.pack(fill="x", padx=8, pady=4)
        self.tree_cash = make_tree(self.tab_cash, cols_f, [35,100,220,80,90,110])
        self.lbl_cash  = UILabel(self.tab_cash, text="", font=FT_BODY, bg=CLR["profit_bg"], fg=CLR["profit"])
        self.lbl_cash.pack(fill="x", padx=8, pady=4)
        cols_m = ("#","تاريخ","المكتب","المرسل","الوفيات","المباع","صافي الفاتورة","الفاتورة")
        self.tree_mkt  = make_tree(self.tab_mkt,  cols_m, [35,100,160,90,80,80,110,100])
        self.lbl_mkt   = UILabel(self.tab_mkt,  text="", font=FT_BODY, bg=CLR["profit_bg"], fg=CLR["profit"])
        self.lbl_mkt.pack(fill="x", padx=8, pady=4)
        self.sum_frm   = UIFrame(self.tab_sum, bg=CLR["bg"]); self.sum_frm.pack(fill="both", expand=True, padx=20, pady=20)

    def _load(self, batch_id):
        if not batch_id: return
        self._bid = batch_id
        curr = db.get_setting("currency", "ريال")
        all_farm = db.fetch_all("SELECT * FROM farm_sales WHERE batch_id=? ORDER BY sale_date", (batch_id,))
        ajl=[]; cash=[]
        for r in all_farm:
            r=dict(r)
            stype=r.get("sale_type") or "آجل"
            if "(نقداً)" in str(r.get("customer","")) and stype=="آجل": stype="نقداً"
            (cash if stype=="نقداً" else ajl).append(r)

        def fill_farm(tree, lbl, rows):
            tree.delete(*tree.get_children())
            tq=tv=0
            for i,r in enumerate(rows,1):
                tree.insert("","end",values=(i,r.get("sale_date",""),r["customer"] or "",
                    f"{r['qty']:,}",f"{r['price']:,.2f}",f"{r['total_val']:,.0f}"))
                tq+=r["qty"]; tv+=r["total_val"]
            lbl.config(text=f"  الإجمالي: {tq:,} طائر | {tv:,.0f} {curr} | {len(rows)} فاتورة")
            return tq, tv

        aq,av=fill_farm(self.tree_ajl, self.lbl_ajl, ajl)
        cq,cv=fill_farm(self.tree_cash,self.lbl_cash,cash)

        self.tree_mkt.delete(*self.tree_mkt.get_children())
        mkt=db.fetch_all("SELECT * FROM market_sales WHERE batch_id=? ORDER BY sale_date",(batch_id,))
        ts=td=tq2=tv2=0
        for i,r in enumerate(mkt,1):
            r=dict(r)
            self.tree_mkt.insert("","end",values=(i,r.get("sale_date",""),r["office"] or "",
                f"{r['qty_sent']:,}",f"{r['deaths']:,}",f"{r['qty_sold']:,}",
                f"{r['net_val']:,.0f}",r.get("inv_num","") or ""))
            ts+=r["qty_sent"]; td+=r["deaths"]; tq2+=r["qty_sold"]; tv2+=r["net_val"]
        self.lbl_mkt.config(text=f"  مرسل: {ts:,} | وفيات: {td:,} | مباع: {tq2:,} | صافي: {tv2:,.0f} {curr}")

        for w in self.sum_frm.winfo_children(): w.destroy()
        rows_s=[
            ("مبيعات العنبر — آجل",  aq,    av,    "#e8f5e9"),
            ("مبيعات العنبر — نقداً", cq,    cv,    "#e3f2fd"),
            ("إجمالي مبيعات العنبر",  aq+cq, av+cv, "#c8e6c9"),
            ("مبيعات السوق",          tq2,   tv2,   "#fff9c4"),
            ("الإجمالي العام",         aq+cq+tq2, av+cv+tv2, "#1F4E79"),
        ]
        tk.Label(self.sum_frm,text="ملخص المبيعات",font=FT_TITLE).grid(row=0,column=0,columnspan=3,pady=(0,10))
        for ci,lbl in enumerate(["البند",f"الكمية",f"القيمة ({curr})"]):
            tk.Label(self.sum_frm,text=lbl,font=FT_HEADER,bg=CLR["header"],fg="white",
                     padx=10,pady=5).grid(row=1,column=ci,sticky="nsew",padx=1,pady=1)
        for ri,(lbl,qty,val,bg) in enumerate(rows_s,2):
            fg="white" if bg=="#1F4E79" else CLR["text"]
            tk.Label(self.sum_frm,text=lbl,font=FT_BODY,bg=bg,fg=fg,padx=10,pady=6).grid(row=ri,column=0,sticky="nsew",padx=1,pady=1)
            tk.Label(self.sum_frm,text=f"{qty:,}",font=("Arial",11,"bold"),bg=bg,fg=fg,padx=10).grid(row=ri,column=1,sticky="nsew",padx=1,pady=1)
            tk.Label(self.sum_frm,text=f"{val:,.0f}",font=("Arial",11,"bold"),bg=bg,fg=fg,padx=10).grid(row=ri,column=2,sticky="nsew",padx=1,pady=1)
        self.sum_frm.columnconfigure(0,weight=3); self.sum_frm.columnconfigure(1,weight=1); self.sum_frm.columnconfigure(2,weight=2)

    def _export_pdf(self):
        if not self._bid: return messagebox.showwarning("تنبيه","اختر دفعة أولاً",parent=self)
        root=self.master
        while not isinstance(root,MainWindow): root=root.master
        root._export_sales_pdf(_batch_id=self._bid)


class DataEntryHub(ToplevelBase):
    """
    مركز الإدخالات الموحد — يضم:
      - إدارة الدفعات (جديد / تعديل / حذف)
      - السجلات اليومية
      - إدارة العنابر
      - إدارة أنواع التكاليف والإيرادات
    """
    def __init__(self, master, on_refresh=None):
        super().__init__(master)
        self.on_refresh = on_refresh
        self.title("📝 مركز الإدخالات")
        self.geometry("1350x800")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.resizable(True, True)
        self._build()

    def _build(self):
        # ── رأس ──
        hdr = UIFrame(self, bg=CLR["header"], pady=10)
        hdr.pack(fill="x")
        UILabel(hdr, text="📝 مركز الإدخالات — الدفعات والسجلات اليومية وإدارة النظام",
                font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)

        style = ttk.Style()
        if not HAS_TTKB: style.theme_use("default")
        style.configure("TNotebook.Tab", font=FT_HEADER, padding=[16, 7])

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=8, pady=8)

        self.tab_batches  = UIFrame(nb, bg=CLR["bg"])
        self.tab_daily    = UIFrame(nb, bg=CLR["bg"])
        self.tab_wh       = UIFrame(nb, bg=CLR["bg"])
        self.tab_costs    = UIFrame(nb, bg=CLR["bg"])

        nb.add(self.tab_batches, text="🐔 الدفعات")
        nb.add(self.tab_daily,   text="📅 السجلات اليومية")
        nb.add(self.tab_wh,      text="🏚 إدارة العنابر")
        nb.add(self.tab_costs,   text="💰 أنواع التكاليف والإيرادات")

        self._build_batches_tab()
        self._build_daily_tab()
        self._build_wh_tab()
        self._build_costs_tab()

    # ══ تبويب الدفعات ══════════════════════════════════════════
    def _build_batches_tab(self):
        F = self.tab_batches

        # شريط الأدوات
        tb = UIFrame(F, bg=CLR["nav"], pady=6)
        tb.pack(fill="x")
        for txt, cmd, bg in [
            ("＋ دفعة جديدة", self._new_batch,  "#ffffff"),
            ("✏ تعديل",       self._edit_batch, "#d0e8ff"),
            ("🗑 حذف دفعة",  self._del_batch,   "#fce4d6"),
            ("📅 سجل يومي",  self._open_daily,  "#fff2cc"),
        ]:
            UIButton(tb, text=txt, command=cmd, font=FT_SMALL, bg=bg,
                     fg=CLR["text"], padx=10, pady=5, cursor="hand2",
                     relief="flat").pack(side="right", padx=3)

        # فلتر
        fbar = UIFrame(F, bg=CLR["bg"], pady=4)
        fbar.pack(fill="x", padx=8)
        UILabel(fbar, text="فلتر العنبر:", font=FT_SMALL, bg=CLR["bg"]).pack(side="right", padx=4)
        self.filter_wh_hub = ttk.Combobox(fbar, width=18, font=FT_BODY)
        self.filter_wh_hub.pack(side="right", padx=4)
        self.filter_wh_hub.bind("<<ComboboxSelected>>", lambda e: self._load_batches())
        UILabel(fbar, text="السنة:", font=FT_SMALL, bg=CLR["bg"]).pack(side="right", padx=(8,2))
        self.filter_fy_hub = ttk.Combobox(fbar, width=8, font=FT_BODY, state="readonly")
        self.filter_fy_hub.pack(side="right", padx=2)
        self.filter_fy_hub.bind("<<ComboboxSelected>>", lambda e: self._load_batches())
        UIButton(fbar, text="عرض الكل", font=FT_SMALL,
                 command=lambda: [self.filter_wh_hub.set(""), self.filter_fy_hub.set(""), self._load_batches()],
                 bg=CLR["bg"], relief="flat").pack(side="right", padx=4)

        # جدول الدفعات
        frm = UIFrame(F, bg=CLR["bg"])
        frm.pack(fill="both", expand=True, padx=8, pady=4)
        cols = ("رقم الدفعة","العنبر","السنة المالية","تاريخ الدخول","تاريخ الخروج",
                "الأيام","الكتاكيت","التكاليف","الإيرادات","صافي النتيجة","النافق%","FCR","نصيب الشركة")
        self.tree_b = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")
        widths = [90,140,90,100,100,50,85,110,110,120,70,75,110]
        for c, w in zip(cols, widths):
            self.tree_b.heading(c, text=c, anchor="center")
            self.tree_b.column(c, width=w, anchor="center")
        self.tree_b.tag_configure("profit", background="#f0f9ea")
        self.tree_b.tag_configure("loss",   background="#fff0f0")
        self.tree_b.bind("<Double-1>", lambda e: self._edit_batch())
        sb = ttk.Scrollbar(frm, command=self.tree_b.yview)
        self.tree_b.configure(yscrollcommand=sb.set)
        sb.pack(side="left", fill="y"); self.tree_b.pack(fill="both", expand=True)

        # شريط الإجماليات
        self.kpi_hub = UIFrame(F, bg=CLR["info_bg"], pady=6, padx=10)
        self.kpi_hub.pack(fill="x")

        self._load_batches()

    def _load_batches(self):
        wh_filter = self.filter_wh_hub.get().strip()
        fy_filter = self.filter_fy_hub.get().strip()

        self.filter_wh_hub["values"] = [""] + [r["name"] for r in db.fetch_all("SELECT name FROM warehouses ORDER BY name")]
        fy_rows = db.fetch_all("SELECT DISTINCT COALESCE(fiscal_year,CAST(strftime('%Y',date_in) AS INTEGER)) AS fy FROM batches ORDER BY fy DESC")
        self.filter_fy_hub["values"] = [""] + [str(r["fy"]) for r in fy_rows if r["fy"]]

        where, params = [], []
        if wh_filter: where.append("warehouse_name=?"); params.append(wh_filter)
        if fy_filter:
            where.append("COALESCE(fiscal_year,CAST(strftime('%Y',date_in) AS INTEGER))=?")
            params.append(int(fy_filter))
        q = "SELECT * FROM v_batches" + (" WHERE " + " AND ".join(where) if where else "") + " ORDER BY date_in DESC"
        rows = db.fetch_all(q, params)

        self.tree_b.delete(*self.tree_b.get_children())
        T = {"cost":0,"rev":0,"net":0,"chicks":0,"share":0}
        for b in rows:
            profit = (b["net_result"] or 0) >= 0
            fy = b["fiscal_year"] if "fiscal_year" in b.keys() and b["fiscal_year"] else ""
            self.tree_b.insert("", "end", iid=str(b["id"]),
                tags=("profit" if profit else "loss",),
                values=(b["batch_num"] or str(b["id"]), b["warehouse_name"], fy,
                        b["date_in"], b["date_out"], b["days"],
                        f"{b['chicks']:,}" if b["chicks"] else "",
                        f"{b['total_cost']:,.0f}" if b["total_cost"] else "0",
                        f"{b['total_rev']:,.0f}"  if b["total_rev"]  else "0",
                        f"{'+' if profit else ''}{b['net_result']:,.0f}" if b["net_result"] is not None else "0",
                        f"{b['mort_rate']:.1f}%" if b["mort_rate"] else "0%",
                        f"{b['fcr']:.3f}" if b["fcr"] else "—",
                        f"{b['share_val']:,.0f}" if b["share_val"] else "0"))
            T["cost"]   += b["total_cost"]  or 0
            T["rev"]    += b["total_rev"]   or 0
            T["net"]    += b["net_result"]  or 0
            T["chicks"] += b["chicks"]      or 0
            T["share"]  += b["share_val"]   or 0

        for w in self.kpi_hub.winfo_children(): w.destroy()
        sign = "+" if T["net"] >= 0 else ""
        for lbl, val, bg, fg in [
            ("الدفعات",          str(len(rows)),             "#dce6f1", CLR["header"]),
            ("إجمالي الكتاكيت", f"{T['chicks']:,}",         "#dce6f1", CLR["header"]),
            ("إجمالي التكاليف", f"{T['cost']:,.0f}",        CLR["loss_bg"],   CLR["loss"]),
            ("إجمالي الإيرادات",f"{T['rev']:,.0f}",         CLR["profit_bg"], CLR["profit"]),
            ("صافي النتيجة",    f"{sign}{T['net']:,.0f}",
             CLR["profit_bg"] if T["net"]>=0 else CLR["loss_bg"],
             CLR["profit"]    if T["net"]>=0 else CLR["loss"]),
            ("نصيب الشركة",    f"{T['share']:,.0f}",        "#fff2cc", CLR["warn"]),
        ]:
            f = UIFrame(self.kpi_hub, bg=bg, padx=12, pady=5, relief="solid", bd=1)
            f.pack(side="right", padx=3)
            UILabel(f, text=lbl, font=FT_TINY,  bg=bg, fg=CLR["text2"]).pack()
            UILabel(f, text=val, font=("Arial",11,"bold"), bg=bg, fg=fg).pack()

        if self.on_refresh: self.on_refresh()

    def _selected_batch_id(self):
        sel = self.tree_b.selection()
        if not sel: messagebox.showwarning("تنبيه","يرجى تحديد دفعة أولاً",parent=self); return None
        return int(sel[0])

    def _new_batch(self):  BatchForm(self, on_save=self._load_batches)
    def _edit_batch(self):
        bid = self._selected_batch_id()
        if bid: BatchForm(self, batch_id=bid, on_save=self._load_batches)
    def _del_batch(self, bid=None):
        if bid is None: bid = self._selected_batch_id()
        if not bid: return
        b = db.fetch_one("SELECT batch_num, warehouse_name FROM v_batches WHERE id=?", (bid,))
        if not b: return
        label = f"{b['warehouse_name']} — {b['batch_num'] or bid}"

        sales = db.fetch_one("SELECT COUNT(*) AS c FROM farm_sales   WHERE batch_id=?", (bid,))["c"]
        mkt   = db.fetch_one("SELECT COUNT(*) AS c FROM market_sales  WHERE batch_id=?", (bid,))["c"]
        daily = db.fetch_one("SELECT COUNT(*) AS c FROM daily_records WHERE batch_id=?", (bid,))["c"]

        details = ""
        if sales or mkt or daily:
            parts = []
            if sales: parts.append(f"{sales} فاتورة مبيعات")
            if mkt:   parts.append(f"{mkt} سجل سوق")
            if daily: parts.append(f"{daily} سجل يومي")
            details = "\nسيتم حذف أيضاً: " + " | ".join(parts)

        if messagebox.askyesno("تأكيد الحذف",
                f"حذف الدفعة «{label}» نهائياً؟{details}\n\nلا يمكن التراجع عن هذا الإجراء.",
                parent=self):
            db.execute("DELETE FROM batches WHERE id=?", (bid,))
            self._load_batches()

    def _open_daily(self):
        bid = self._selected_batch_id()
        if bid:
            b = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (bid,))
            if b: DailyRecordsWindow(self, bid, dict(b))

    # ══ تبويب السجلات اليومية — اختصار يفتح نافذة فرعية ══════
    def _build_daily_tab(self):
        F = self.tab_daily
        UILabel(F, text="📅 السجلات اليومية", font=FT_TITLE, bg=CLR["bg"]).pack(pady=20)
        UILabel(F, text="حدد دفعة من تبويب «الدفعات» ثم اضغط الزر أدناه", font=FT_BODY, bg=CLR["bg"], fg=CLR["text2"]).pack()
        UIButton(F, text="📅 فتح السجلات اليومية للدفعة المحددة",
                 font=FT_HEADER, cursor="hand2", relief="flat", pady=12,
                 command=self._open_daily_from_tab).pack(pady=20)
        # قائمة سريعة بآخر السجلات
        UILabel(F, text="آخر السجلات المدخلة:", font=FT_HEADER, bg=CLR["bg"]).pack(anchor="e", padx=20)
        frm = UIFrame(F, bg=CLR["bg"]); frm.pack(fill="both", expand=True, padx=16, pady=8)
        r_cols = ("العنبر","رقم الدفعة","التاريخ","اليوم رقم","النافق","العلف كجم","ملاحظة")
        self.tree_recent = ttk.Treeview(frm, columns=r_cols, show="headings", height=18)
        for c,w in zip(r_cols,[130,100,100,80,80,90,200]):
            self.tree_recent.heading(c,text=c,anchor="center")
            self.tree_recent.column(c,width=w,anchor="center")
        sb=ttk.Scrollbar(frm,command=self.tree_recent.yview); self.tree_recent.configure(yscrollcommand=sb.set)
        sb.pack(side="left",fill="y"); self.tree_recent.pack(fill="both",expand=True)
        self._load_recent_records()

    def _load_recent_records(self):
        self.tree_recent.delete(*self.tree_recent.get_children())
        rows = db.fetch_all(""" SELECT dr.*, b.batch_num, w.name AS wh_name
            FROM daily_records dr
            JOIN batches b ON dr.batch_id=b.id
            JOIN warehouses w ON b.warehouse_id=w.id
            ORDER BY dr.rec_date DESC, dr.id DESC LIMIT 80""")
        for r in rows:
            self.tree_recent.insert("","end",values=(
                r["wh_name"], r["batch_num"] or str(r["batch_id"]),
                r["rec_date"], r["day_num"] or "",
                r["dead_count"], f"{r['feed_kg']:.1f}", r["notes"] or ""))

    def _open_daily_from_tab(self):
        self._open_daily()
        self._load_recent_records()

    # ══ تبويب إدارة العنابر ════════════════════════════════════
    def _build_wh_tab(self):
        F = self.tab_wh
        UILabel(F, text="🏚 إدارة العنابر", font=FT_TITLE, bg=CLR["bg"]).pack(pady=10)

        inp = UILabelFrame(F, text="إضافة عنبر جديد", font=FT_HEADER, bg=CLR["bg"], padx=15, pady=10)
        inp.pack(fill="x", padx=20, pady=8)
        UILabel(inp, text="اسم العنبر:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0,column=0,sticky="e",padx=6)
        self.v_wh_name = tk.StringVar()
        UIEntry(inp, textvariable=self.v_wh_name, width=24, font=FT_BODY, relief="solid").grid(row=0,column=1,padx=6)
        UILabel(inp, text="ملاحظات:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0,column=2,sticky="e",padx=6)
        self.v_wh_notes = tk.StringVar()
        UIEntry(inp, textvariable=self.v_wh_notes, width=30, font=FT_BODY, relief="solid").grid(row=0,column=3,padx=6)
        UIButton(inp, text="➕ إضافة عنبر", font=FT_BODY, cursor="hand2", command=self._add_wh).grid(row=0,column=4,padx=10)

        frm = UIFrame(F, bg=CLR["bg"]); frm.pack(fill="both", expand=True, padx=20, pady=6)
        wh_cols = ("الاسم","عدد الدفعات","ملاحظات")
        self.tree_wh = ttk.Treeview(frm, columns=wh_cols, show="headings", selectmode="browse")
        for c,w in zip(wh_cols,[200,100,300]):
            self.tree_wh.heading(c,text=c,anchor="center")
            self.tree_wh.column(c,width=w,anchor="center")
        sb=ttk.Scrollbar(frm,command=self.tree_wh.yview); self.tree_wh.configure(yscrollcommand=sb.set)
        sb.pack(side="left",fill="y"); self.tree_wh.pack(fill="both",expand=True)

        btn_frm = UIFrame(F, bg=CLR["bg"], pady=6); btn_frm.pack(fill="x", padx=20)
        UIButton(btn_frm, text="🗑 حذف العنبر المحدد", font=FT_BODY,
                 cursor="hand2", command=self._del_wh).pack(side="right", padx=4)
        UIButton(btn_frm, text="🔄 تحديث", font=FT_SMALL,
                 cursor="hand2", command=self._load_wh).pack(side="right", padx=4)
        self._load_wh()

    def _load_wh(self):
        self.tree_wh.delete(*self.tree_wh.get_children())
        rows = db.fetch_all("""SELECT w.id, w.name, w.notes, COUNT(b.id) AS bc FROM warehouses w LEFT JOIN batches b ON b.warehouse_id=w.id
            GROUP BY w.id ORDER BY w.name""")
        for r in rows:
            self.tree_wh.insert("","end",iid=str(r["id"]),
                values=(r["name"], r["bc"], r["notes"] or ""))

    def _add_wh(self):
        name = self.v_wh_name.get().strip()
        if not name: return messagebox.showwarning("تنبيه","اسم العنبر مطلوب",parent=self)
        try:
            db.execute("INSERT INTO warehouses (name,notes) VALUES (?,?)",
                       (name, self.v_wh_notes.get().strip()))
            self.v_wh_name.set(""); self.v_wh_notes.set("")
            self._load_wh(); self._load_batches()
            messagebox.showinfo("تم",f"تمت إضافة العنبر «{name}» بنجاح",parent=self)
        except Exception as e:
            messagebox.showerror("خطأ", f"الاسم مستخدم مسبقاً: {e}", parent=self)

    def _del_wh(self):
        sel = self.tree_wh.selection()
        if not sel: return messagebox.showwarning("تنبيه","اختر عنبراً أولاً",parent=self)
        wid = int(sel[0])
        r = db.fetch_one("SELECT name FROM warehouses WHERE id=?",(wid,))
        cnt = db.fetch_one("SELECT COUNT(*) AS c FROM batches WHERE warehouse_id=?",(wid,))["c"]
        wn = r["name"] if r else "؟"
        if cnt > 0:
            return messagebox.showerror("لا يمكن الحذف",
                f"العنبر «{wn}» مرتبط بـ {cnt} دفعة\nاحذف الدفعات أولاً", parent=self)
        if messagebox.askyesno("تأكيد",f"حذف «{wn}» نهائياً؟",parent=self):
            db.execute("DELETE FROM warehouses WHERE id=?",(wid,))
            self._load_wh(); self._load_batches()

    # ══ تبويب التكاليف والإيرادات — يضمّن CostTypesManager ════
    def _build_costs_tab(self):
        F = self.tab_costs
        UILabel(F, text="💰 إدارة أنواع التكاليف والإيرادات",
                font=FT_TITLE, bg=CLR["bg"]).pack(pady=8)
        UILabel(F, text="أي تغيير هنا ينعكس فوراً على نموذج إدخال الدفعة",
                font=FT_SMALL, bg=CLR["bg"], fg=CLR["text2"]).pack()
        UIButton(F, text="⚙️ فتح مدير التكاليف والإيرادات",
                 font=FT_HEADER, cursor="hand2", relief="flat", pady=12,
                 command=lambda: CostTypesManager(self)).pack(pady=16)


# ════════════════════════════════════════════════════════════════
# مركز التقارير — كل التقارير والتصديرات في نافذة واحدة
# ════════════════════════════════════════════════════════════════
class ReportsHub(ToplevelBase):
    """
    مركز التقارير الموحد — يضم:
      - تصدير Excel الشامل
      - التحليلات المتقدمة (مقارنة / سنوي / موسمي / رسوم بيانية)
      - تقرير ملخص العنابر
      - PDF تصفية الدفعة
      - PDF مبيعات الدفعة
      - PDF التقرير اليومي
      - إرسال Telegram
    """
    def __init__(self, master):
        super().__init__(master)
        self.title("📊 مركز التقارير والتصديرات")
        self.geometry("1350x820")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.resizable(True, True)
        self._build()

    def _build(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=10)
        hdr.pack(fill="x")
        UILabel(hdr, text="📊 مركز التقارير والتحليلات — جميع التصديرات في مكان واحد",
                font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)

        style = ttk.Style()
        if not HAS_TTKB: style.theme_use("default")
        style.configure("TNotebook.Tab", font=FT_HEADER, padding=[14, 7])

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=8, pady=8)

        self.tab_overview   = UIFrame(nb, bg=CLR["bg"])
        self.tab_analytics  = UIFrame(nb, bg=CLR["bg"])
        self.tab_batch_rpt  = UIFrame(nb, bg=CLR["bg"])
        self.tab_daily_rpt  = UIFrame(nb, bg=CLR["bg"])
        self.tab_export     = UIFrame(nb, bg=CLR["bg"])
        self.tab_telegram   = UIFrame(nb, bg=CLR["bg"])

        nb.add(self.tab_overview,  text="📋 ملخص العنابر")
        nb.add(self.tab_analytics, text="🔬 التحليلات المتقدمة")
        nb.add(self.tab_batch_rpt, text="🖨 تقارير الدفعة PDF")
        nb.add(self.tab_daily_rpt, text="📄 التقرير اليومي PDF")
        nb.add(self.tab_export,    text="📥 تصدير Excel")
        nb.add(self.tab_telegram,  text="📱 Telegram")

        self._build_overview_tab()
        self._build_analytics_tab()
        self._build_batch_rpt_tab()
        self._build_daily_rpt_tab()
        self._build_export_tab()
        self._build_telegram_tab()

    # ══ تبويب ملخص العنابر ══════════════════════════════════════
    def _build_overview_tab(self):
        F = self.tab_overview
        # يُضمّن محتوى WarehousesReportWindow مباشرة
        UILabel(F, text="📋 ملخص العنابر الشامل", font=FT_TITLE, bg=CLR["bg"]).pack(pady=8)

        nb2 = ttk.Notebook(F); nb2.pack(fill="both", expand=True, padx=6, pady=4)
        tab_wh   = UIFrame(nb2, bg=CLR["bg"])
        tab_all  = UIFrame(nb2, bg=CLR["bg"])
        nb2.add(tab_wh,  text="📦 ملخص حسب العنبر")
        nb2.add(tab_all, text="🏭 كل الدفعات")

        cols_wh = ("العنبر","عدد الدفعات","إجمالي الكتاكيت","إجمالي التكاليف","إجمالي الإيرادات","صافي الربح/الخسارة","متوسط النافق%")
        self.tree_ov_wh = ttk.Treeview(tab_wh, columns=cols_wh, show="headings")
        for c,w in zip(cols_wh,[160,100,130,140,140,150,120]):
            self.tree_ov_wh.heading(c,text=c,anchor="center"); self.tree_ov_wh.column(c,width=w,anchor="center")
        self.tree_ov_wh.tag_configure("profit",background="#f0f9ea"); self.tree_ov_wh.tag_configure("loss",background="#fff0f0")
        sb=ttk.Scrollbar(tab_wh,command=self.tree_ov_wh.yview); self.tree_ov_wh.configure(yscrollcommand=sb.set)
        sb.pack(side="left",fill="y"); self.tree_ov_wh.pack(fill="both",expand=True)

        cols_all = ("رقم الدفعة","العنبر","السنة","تاريخ الدخول","تاريخ الخروج","الأيام","الكتاكيت","التكاليف","الإيرادات","الربح/الخسارة","النافق%","FCR","نصيب الشركة")
        self.tree_ov_all = ttk.Treeview(tab_all, columns=cols_all, show="headings")
        for c,w in zip(cols_all,[85,130,70,95,95,50,80,110,110,120,65,70,110]):
            self.tree_ov_all.heading(c,text=c,anchor="center"); self.tree_ov_all.column(c,width=w,anchor="center")
        self.tree_ov_all.tag_configure("profit",background="#f0f9ea"); self.tree_ov_all.tag_configure("loss",background="#fff0f0")
        sb2=ttk.Scrollbar(tab_all,command=self.tree_ov_all.yview); self.tree_ov_all.configure(yscrollcommand=sb2.set)
        sb2.pack(side="left",fill="y"); self.tree_ov_all.pack(fill="both",expand=True)

        self.sum_ov = UIFrame(F, bg=CLR["info_bg"], pady=6, padx=10); self.sum_ov.pack(fill="x")
        UIButton(F, text="🔄 تحديث البيانات", font=FT_SMALL, cursor="hand2",
                 command=self._load_overview, relief="flat").pack(pady=4)
        self._load_overview()

    def _load_overview(self):
        batches = db.fetch_all("SELECT * FROM v_batches ORDER BY warehouse_name, date_in")
        wh_data = {}
        for b in batches:
            wh = b["warehouse_name"]
            if wh not in wh_data:
                wh_data[wh]={"count":0,"chicks":0,"cost":0,"rev":0,"net":0,"mort_sum":0,"sold":0}
            d=wh_data[wh]; d["count"]+=1; d["chicks"]+=b["chicks"] or 0
            d["cost"]+=b["total_cost"] or 0; d["rev"]+=b["total_rev"] or 0
            d["net"]+=b["net_result"] or 0; d["mort_sum"]+=b["mort_rate"] or 0
            d["sold"]+=b["total_sold"] or 0

        self.tree_ov_wh.delete(*self.tree_ov_wh.get_children())
        for wh,d in wh_data.items():
            cnt=d["count"] or 1
            self.tree_ov_wh.insert("","end",tags=("profit" if d["net"]>=0 else "loss",),
                values=(wh,d["count"],f"{d['chicks']:,}",f"{d['cost']:,.0f}",f"{d['rev']:,.0f}",
                        f"{'+' if d['net']>=0 else ''}{d['net']:,.0f}",f"{d['mort_sum']/cnt:.1f}%"))

        self.tree_ov_all.delete(*self.tree_ov_all.get_children())
        T={"chicks":0,"cost":0,"rev":0,"net":0,"share":0}
        for b in batches:
            fy = b["fiscal_year"] if "fiscal_year" in b.keys() and b["fiscal_year"] else ""
            net=b["net_result"] or 0
            self.tree_ov_all.insert("","end",iid=str(b["id"]),
                tags=("profit" if net>=0 else "loss",),
                values=(b["batch_num"] or str(b["id"]),b["warehouse_name"],fy,
                        b["date_in"],b["date_out"],b["days"] or "",
                        f"{b['chicks']:,}" if b["chicks"] else "",
                        f"{b['total_cost']:,.0f}" if b["total_cost"] else "0",
                        f"{b['total_rev']:,.0f}"  if b["total_rev"]  else "0",
                        f"{'+' if net>=0 else ''}{net:,.0f}",
                        f"{b['mort_rate'] or 0:.1f}%",f"{b['fcr'] or 0:.3f}",
                        f"{b['share_val'] or 0:,.0f}"))
            T["chicks"]+=b["chicks"] or 0; T["cost"]+=b["total_cost"] or 0
            T["rev"]+=b["total_rev"] or 0; T["net"]+=b["net_result"] or 0; T["share"]+=b["share_val"] or 0

        for w in self.sum_ov.winfo_children(): w.destroy()
        sign="+" if T["net"]>=0 else ""
        for lbl,val,bg,fg in [
            ("الدفعات",str(len(batches)),"#dce6f1",CLR["header"]),
            ("إجمالي الكتاكيت",f"{T['chicks']:,}","#dce6f1",CLR["header"]),
            ("إجمالي التكاليف",f"{T['cost']:,.0f}",CLR["loss_bg"],CLR["loss"]),
            ("إجمالي الإيرادات",f"{T['rev']:,.0f}",CLR["profit_bg"],CLR["profit"]),
            ("صافي النتيجة",f"{sign}{T['net']:,.0f}",
             CLR["profit_bg"] if T["net"]>=0 else CLR["loss_bg"],
             CLR["profit"] if T["net"]>=0 else CLR["loss"]),
            ("نصيب الشركة",f"{T['share']:,.0f}","#fff2cc",CLR["warn"]),
        ]:
            fr=UIFrame(self.sum_ov,bg=bg,padx=12,pady=4,relief="solid",bd=1); fr.pack(side="right",padx=3)
            UILabel(fr,text=lbl,font=FT_TINY,bg=bg,fg=CLR["text2"]).pack()
            UILabel(fr,text=val,font=("Arial",11,"bold"),bg=bg,fg=fg).pack()

    # ══ تبويب التحليلات المتقدمة ════════════════════════════════
    def _build_analytics_tab(self):
        F = self.tab_analytics
        UILabel(F, text="🔬 التحليلات المتقدمة", font=FT_TITLE, bg=CLR["bg"]).pack(pady=8)
        UILabel(F, text="تحليلات شاملة: مقارنة العنابر — تحليل سنوي — رسوم بيانية",
                font=FT_BODY, bg=CLR["bg"], fg=CLR["text2"]).pack()
        UIButton(F, text="🔬 فتح نافذة التحليلات المتقدمة",
                 font=FT_HEADER, cursor="hand2", relief="flat", pady=12,
                 command=lambda: AdvancedAnalyticsWindow(self)).pack(pady=20)

    # ══ تبويب تقارير الدفعة PDF ═════════════════════════════════
    def _build_batch_rpt_tab(self):
        F = self.tab_batch_rpt

        hdr2 = UIFrame(F, bg=CLR["info_bg"], pady=8, padx=14)
        hdr2.pack(fill="x")
        UILabel(hdr2, text="اختر دفعة من القائمة أدناه ثم اضغط نوع التقرير",
                font=FT_BODY, bg=CLR["info_bg"], fg=CLR["text2"]).pack(side="right")

        # قائمة الدفعات
        frm = UIFrame(F, bg=CLR["bg"]); frm.pack(fill="both", expand=True, padx=8, pady=6)
        cols = ("رقم الدفعة","العنبر","السنة","تاريخ الدخول","تاريخ الخروج","صافي النتيجة","النافق%")
        self.tree_batch_rpt = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")
        for c,w in zip(cols,[100,150,70,110,110,120,80]):
            self.tree_batch_rpt.heading(c,text=c,anchor="center")
            self.tree_batch_rpt.column(c,width=w,anchor="center")
        self.tree_batch_rpt.tag_configure("profit",background="#f0f9ea")
        self.tree_batch_rpt.tag_configure("loss",  background="#fff0f0")
        sb=ttk.Scrollbar(frm,command=self.tree_batch_rpt.yview)
        self.tree_batch_rpt.configure(yscrollcommand=sb.set)
        sb.pack(side="left",fill="y"); self.tree_batch_rpt.pack(fill="both",expand=True)

        # أزرار التقارير
        btn_frm = UIFrame(F, bg=CLR["bg"], pady=10); btn_frm.pack(fill="x", padx=8)
        UIButton(btn_frm, text="🖨 PDF تصفية الدفعة",
                 font=FT_BODY, cursor="hand2", relief="flat", padx=16, pady=8,
                 command=self._rpt_pdf_tafyia).pack(side="right", padx=6)
        UIButton(btn_frm, text="📋 PDF مبيعات الدفعة",
                 font=FT_BODY, cursor="hand2", relief="flat", padx=16, pady=8,
                 command=self._rpt_pdf_sales).pack(side="right", padx=6)
        UIButton(btn_frm, text="📊 تقرير المبيعات التحليلي",
                 font=FT_BODY, cursor="hand2", relief="flat", padx=14, pady=8,
                 command=self._open_sales_report_win).pack(side="right", padx=6)
        UIButton(btn_frm, text="🔄 تحديث", font=FT_SMALL, cursor="hand2",
                 command=self._load_batch_rpt_list).pack(side="left", padx=6)

        self._load_batch_rpt_list()

    def _load_batch_rpt_list(self):
        self.tree_batch_rpt.delete(*self.tree_batch_rpt.get_children())
        for b in db.fetch_all("SELECT * FROM v_batches ORDER BY date_in DESC"):
            fy = b["fiscal_year"] if "fiscal_year" in b.keys() and b["fiscal_year"] else ""
            net = b["net_result"] or 0
            self.tree_batch_rpt.insert("","end",iid=str(b["id"]),
                tags=("profit" if net>=0 else "loss",),
                values=(b["batch_num"] or str(b["id"]),b["warehouse_name"],fy,
                        b["date_in"],b["date_out"],
                        f"{'+' if net>=0 else ''}{net:,.0f}",
                        f"{b['mort_rate'] or 0:.1f}%"))

    def _get_selected_batch_for_rpt(self):
        sel = self.tree_batch_rpt.selection()
        if not sel:
            messagebox.showwarning("تنبيه","اختر دفعة من القائمة أولاً",parent=self)
            return None, None
        bid = int(sel[0])
        b = db.fetch_one("SELECT * FROM v_batches WHERE id=?",(bid,))
        return bid, dict(b) if b else None

    def _rpt_pdf_tafyia(self):
        bid, b = self._get_selected_batch_for_rpt()
        if not bid: return
        # استدعاء مباشر للدالة في MainWindow
        self.master._export_pdf_for_batch(bid, b)

    def _rpt_pdf_sales(self):
        bid, b = self._get_selected_batch_for_rpt()
        if not bid: return
        self.master._export_sales_pdf_for_batch(bid, b)

    def _open_sales_report_win(self):
        sel = self.tree_batch_rpt.selection()
        bid = int(sel[0]) if sel else None
        BatchSalesReportWindow(self.master, batch_id=bid)

    # ══ تبويب التقرير اليومي PDF ════════════════════════════════
    def _build_daily_rpt_tab(self):
        F = self.tab_daily_rpt
        UILabel(F, text="📄 التقرير الفني اليومي", font=FT_TITLE, bg=CLR["bg"]).pack(pady=16)
        UILabel(F,
            text="تقرير PDF يعرض لكل عنبر نشط: الكتل النشطة — النافق (واقعي/طبيعي/فرق) — العلف (واقعي/مطلوب/فرق). مناسب للإدارة الفنية لاتخاذ القرار.",
            font=FT_BODY, bg=CLR["bg"], fg=CLR["text2"], justify="center").pack(padx=30)
        UIButton(F, text="📄 إنشاء وتصدير التقرير الفني اليومي",
                 font=FT_HEADER, cursor="hand2", relief="flat", pady=14,
                 command=self.master._export_daily_pdf).pack(pady=24)

    # ══ تبويب تصدير Excel ═══════════════════════════════════════
    def _build_export_tab(self):
        F = self.tab_export
        UILabel(F, text="📥 تصدير Excel الشامل", font=FT_TITLE, bg=CLR["bg"]).pack(pady=12)

        cards = [
            ("📊 التقرير الشامل للدفعات (4 أوراق: ملخص + تكاليف + مبيعات عنبر + مبيعات سوق)",
             self.master._export_wh_excel, CLR["info_bg"]),
            ("🔬 تصدير التحليلات المتقدمة (مقارنة دفعات + تحليل شهري + موسمي)",
             self._export_analytics, CLR["profit_bg"]),
            ("📅 تصدير السجلات اليومية لدفعة (اختر الدفعة من تبويب تقارير الدفعة)",
             self._export_daily_records, "#f0f7ff"),
            ("📦 تصدير الدفعة بالكامل للنقل (اختر الدفعة من تبويب تقارير الدفعة)",
             self._export_batch_portable, CLR["warn_bg"]),
            ("📥 استيراد دفعة كاملة من ملف Excel",
             self.master._import_batch_full_excel, CLR["profit_bg"]),
        ]
        for lbl, cmd, bg in cards:
            f = UIFrame(F, bg=bg, pady=16, padx=20, relief="solid", bd=1)
            f.pack(fill="x", padx=30, pady=8)
            UILabel(f, text=lbl, font=FT_BODY, bg=bg, justify="right").pack(side="right")
            UIButton(f, text="📥 تصدير", font=FT_BODY, cursor="hand2",
                     relief="flat", padx=16, command=cmd).pack(side="left")

    def _do_wh_excel(self):
        """يفتح نافذة تقرير العنابر مباشرة ثم يشغل التصدير"""
        win = WarehousesReportWindow(self.master)
        self.after(300, win._export_excel)

    def _export_analytics(self):
        AdvancedAnalyticsWindow(self)

    def _export_daily_records(self):
        bid, b = self._get_selected_batch_for_rpt()
        if not bid: return
        if not HAS_OPENPYXL:
            return messagebox.showerror("خطأ","يرجى تثبيت openpyxl",parent=self)
        b_num = b.get("batch_num") or str(bid)
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"سجل_يومي_دفعة_{str(b_num).replace("/","-")}.xlsx", parent=self)
        if not path: return
        rows = db.fetch_all(
            "SELECT rec_date,day_num,dead_count,feed_kg,notes FROM daily_records WHERE batch_id=? ORDER BY rec_date",
            (bid,))
        hdr_fill = PatternFill("solid",fgColor="1F4E79")
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="السجل اليومي"; ws.sheet_view.rightToLeft=True
        hdrs=["التاريخ","اليوم","النافق","تراكم النافق","العلف كجم","إجمالي العلف","ملاحظة"]
        for ci,h in enumerate(hdrs,1):
            cell=ws.cell(1,ci,h); cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=hdr_fill; cell.alignment=Alignment(horizontal="center")
        cum_dead=cum_feed=0
        for ri,r in enumerate(rows,2):
            cum_dead+=r["dead_count"]; cum_feed+=r["feed_kg"]
            for ci,v in enumerate([r["rec_date"],r["day_num"],r["dead_count"],
                cum_dead,round(r["feed_kg"],2),round(cum_feed,2),r["notes"] or ""],1):
                ws.cell(ri,ci,v).alignment=Alignment(horizontal="center")
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=14
        try:
            wb.save(path)
            messagebox.showinfo("تم","تم تصدير السجل اليومي بنجاح!",parent=self)
            try: os.startfile(path)
            except: pass
        except PermissionError:
            messagebox.showerror("خطأ","الملف مفتوح! أغلقه أولاً.",parent=self)

    def _export_batch_portable(self):
        bid, b = self._get_selected_batch_for_rpt()
        if not bid: return
        self.master._export_batch_full_excel(bid)

    # ══ تبويب Telegram ══════════════════════════════════════════
    def _build_telegram_tab(self):
        F = self.tab_telegram
        UILabel(F, text="📱 تنبيهات Telegram", font=FT_TITLE, bg=CLR["bg"]).pack(pady=12)

        cards2 = [
            ("📤 إرسال تقرير يومي الآن يرسل ملخص كل العنابر النشطة فوراً",
             self.master._send_telegram_now, CLR["info_bg"]),
            ("⚙ إعدادات Telegram (Bot Token, Chat ID, الإرسال التلقائي)",
             self.master._open_settings, "#ede7f6"),
        ]
        for lbl, cmd, bg in cards2:
            f = UIFrame(F, bg=bg, pady=14, padx=20, relief="solid", bd=1)
            f.pack(fill="x", padx=30, pady=8)
            UILabel(f, text=lbl, font=FT_BODY, bg=bg, justify="right").pack(side="right")
            UIButton(f, text="تنفيذ", font=FT_BODY, cursor="hand2",
                     relief="flat", padx=16, command=cmd).pack(side="left")

        UILabel(F,
            text="💡 التقرير التلقائي يُرسل مرة واحدة يومياً عند فتح البرنامج إذا كان مفعلاً في الإعدادات",
            font=FT_TINY, bg=CLR["bg"], fg=CLR["text2"]).pack(pady=10)

        # ══ قسم الاستيراد من OnyxPro ══
        sep = ttk.Separator(F, orient="horizontal")
        sep.pack(fill="x", padx=40, pady=20)
        
        UILabel(F, text="🔗 الربط مع OnyxPro ERP", font=FT_HEADER, bg=CLR["bg"]).pack()
        f_onyx = UIFrame(F, bg="#fffde7", pady=16, padx=20, relief="solid", bd=1)
        f_onyx.pack(fill="x", padx=30, pady=10)
        UILabel(f_onyx, text="📥 استيراد كشف حساب أونكس برو (Excel)\nلجلب المصاريف آلياً وتوزيعها على الدفعة",
                font=FT_BODY, bg="#fffde7", justify="right").pack(side="right")
        UIButton(f_onyx, text="استيراد الآن", font=FT_BODY, cursor="hand2",
                 relief="flat", padx=16, command=self.master._open_onyx_importer).pack(side="left")

        f_v4 = UIFrame(F, bg="#eaf5ff", pady=16, padx=20, relief="solid", bd=1)
        f_v4.pack(fill="x", padx=30, pady=10)
        UILabel(
            f_v4,
            text="📥 استيراد poultry_v4 الذكي (ملف أو مجلد)\nمع شاشة مراجعة وتصنيف قبل الحفظ النهائي",
            font=FT_BODY,
            bg="#eaf5ff",
            justify="right",
        ).pack(side="right")
        UIButton(
            f_v4,
            text="فتح معالج الاستيراد",
            font=FT_BODY,
            cursor="hand2",
            relief="flat",
            padx=16,
            command=self.master._open_poultry_v4_importer,
        ).pack(side="left")


class CostTypesManager(ToplevelBase):
    CATEGORIES_COST = ["مواد", "نقل", "مرافق", "صحة", "تشغيل", "رواتب", "إشراف", "إدارة", "عقارات", "أخرى"]
    CATEGORIES_REV  = ["مبيعات", "تحويل", "مخزون", "مرتجعات", "أخرى"]
    UNITS           = ["", "طن", "كجم", "حبة", "كيس", "أسطوانة", "م³", "لتر", "رحلة", "يوم"]

    def __init__(self, master):
        super().__init__(master)
        self.title("⚙️ إدارة أنواع التكاليف والإيرادات")
        self.geometry("900x600")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.resizable(True, True)
        self.grab_set()
        self._mode = "cost"
        self._build()
        self._load()

    def _build(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=10)
        hdr.pack(fill="x")
        UILabel(hdr, text="⚙️ إدارة أنواع التكاليف والإيرادات", font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)
        UILabel(hdr, text="أضف أو عدّل أنواع التكاليف التي تظهر في نموذج الدفعة", font=FT_SMALL, bg=CLR["header"], fg="#aad4f5").pack(side="right", padx=4)

        tab_bar = UIFrame(self, bg=CLR["nav"], pady=4)
        tab_bar.pack(fill="x")
        self.btn_cost = UIButton(tab_bar, text="💰 أنواع التكاليف", font=FT_BODY, bg=CLR["white"], fg=CLR["header"], relief="flat", padx=14, pady=4, cursor="hand2", command=lambda: self._switch("cost"))
        self.btn_cost.pack(side="right", padx=4)
        self.btn_rev = UIButton(tab_bar, text="📈 أنواع الإيرادات الإضافية", font=FT_BODY, bg=CLR["nav"], fg=CLR["white"], relief="flat", padx=14, pady=4, cursor="hand2", command=lambda: self._switch("revenue"))
        self.btn_rev.pack(side="right", padx=4)

        inp = UILabelFrame(self, text="إضافة / تعديل نوع", font=FT_HEADER, bg=CLR["bg"], padx=12, pady=8)
        inp.pack(fill="x", padx=10, pady=8)

        UILabel(inp, text="الاسم العربي *:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0, column=0, sticky="e", padx=4, pady=6)
        self.v_name = tk.StringVar()
        UIEntry(inp, textvariable=self.v_name, width=20, font=FT_BODY, relief="solid").grid(row=0, column=1, padx=4, sticky="ew")

        UILabel(inp, text="الرمز البرمجي *:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0, column=2, sticky="e", padx=4)
        self.v_code = tk.StringVar()
        UIEntry(inp, textvariable=self.v_code, width=16, font=FT_BODY, relief="solid").grid(row=0, column=3, padx=4, sticky="ew")

        UILabel(inp, text="التصنيف:", font=FT_SMALL, bg=CLR["bg"]).grid(row=0, column=4, sticky="e", padx=4)
        self.v_cat = tk.StringVar()
        self.cbo_cat = ttk.Combobox(inp, textvariable=self.v_cat, width=12, font=FT_BODY, state="readonly", values=self.CATEGORIES_COST)
        self.cbo_cat.grid(row=0, column=5, padx=4)
        self.cbo_cat.set(self.CATEGORIES_COST[0])

        UILabel(inp, text="له كمية؟:", font=FT_SMALL, bg=CLR["bg"]).grid(row=1, column=0, sticky="e", padx=4, pady=6)
        self.v_has_qty = tk.BooleanVar(value=False)
        chk = tk.Checkbutton(inp, variable=self.v_has_qty, bg=CLR["bg"], activebackground=CLR["bg"], command=self._toggle_unit)
        chk.grid(row=1, column=1, sticky="w", padx=4)

        UILabel(inp, text="الوحدة:", font=FT_SMALL, bg=CLR["bg"]).grid(row=1, column=2, sticky="e", padx=4)
        self.v_unit = tk.StringVar()
        self.cbo_unit = ttk.Combobox(inp, textvariable=self.v_unit, width=10, font=FT_BODY, values=self.UNITS, state="disabled")
        self.cbo_unit.grid(row=1, column=3, padx=4)

        btn_frm = UIFrame(inp, bg=CLR["bg"])
        btn_frm.grid(row=1, column=4, columnspan=2, padx=4)
        UIButton(btn_frm, text="➕ إضافة جديد", font=FT_BODY, padx=8, cursor="hand2", command=self._add).pack(side="right", padx=3)
        UIButton(btn_frm, text="💾 حفظ تعديل", font=FT_BODY, padx=8, cursor="hand2", command=self._update).pack(side="right", padx=3)
        UIButton(btn_frm, text="🗑 حذف", font=FT_BODY, padx=8, cursor="hand2", command=self._delete).pack(side="right", padx=3)
        UIButton(btn_frm, text="🔄 تفريغ", font=FT_BODY, padx=8, cursor="hand2", command=self._clear).pack(side="right", padx=3)

        tbl_frm = UIFrame(self, bg=CLR["bg"])
        tbl_frm.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        cols = ("الترتيب", "الرمز", "الاسم العربي", "التصنيف", "له كمية", "الوحدة", "الحالة")
        self.tree = ttk.Treeview(tbl_frm, columns=cols, show="headings", selectmode="browse")
        widths = [60, 130, 180, 100, 70, 80, 70]
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c, anchor="center")
            self.tree.column(c, width=w, anchor="center")
        self.tree.tag_configure("inactive", foreground="#aaaaaa")

        sb = ttk.Scrollbar(tbl_frm, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="left", fill="y")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        bot = UIFrame(self, bg=CLR["bg"], pady=6)
        bot.pack(fill="x", padx=10)
        UIButton(bot, text="⬆ أعلى", font=FT_SMALL, padx=8, cursor="hand2", command=self._move_up).pack(side="right", padx=3)
        UIButton(bot, text="⬇ أسفل", font=FT_SMALL, padx=8, cursor="hand2", command=self._move_down).pack(side="right", padx=3)
        UIButton(bot, text="✅ تفعيل / إيقاف", font=FT_SMALL, padx=8, cursor="hand2", command=self._toggle_active).pack(side="right", padx=3)
        UILabel(bot, text="💡 الأنواع المُوقفة لا تظهر في نموذج الدفعة", font=FT_TINY, bg=CLR["bg"], fg=CLR["text2"]).pack(side="left", padx=6)

    def _switch(self, mode):
        self._mode = mode
        is_cost = (mode == "cost")
        if HAS_TTKB:
            self.btn_cost.configure(bootstyle="primary" if is_cost else "secondary-outline")
            self.btn_rev.configure(bootstyle="secondary-outline" if is_cost else "primary")
        else:
            self.btn_cost.configure(bg=CLR["white"] if is_cost else CLR["nav"], fg=CLR["header"] if is_cost else CLR["white"])
            self.btn_rev.configure(bg=CLR["nav"] if is_cost else CLR["white"], fg=CLR["white"] if is_cost else CLR["header"])
        cats = self.CATEGORIES_COST if is_cost else self.CATEGORIES_REV
        self.cbo_cat["values"] = cats
        self.cbo_cat.set(cats[0])
        self._clear()
        self._load()

    def _load(self):
        self.tree.delete(*self.tree.get_children())
        tbl = "cost_types" if self._mode == "cost" else "revenue_types"
        rows = db.fetch_all(f"SELECT * FROM {tbl} ORDER BY sort_order")
        for r in rows:
            active_lbl = "✅ نشط" if r["is_active"] else "⛔ موقف"
            has_qty_lbl = "نعم" if r["has_qty"] else "—"
            tag = "" if r["is_active"] else "inactive"
            self.tree.insert("", "end", iid=str(r["id"]), tags=(tag,), values=(r["sort_order"], r["code"], r["name_ar"], r["category"], has_qty_lbl, r["unit"] or "—", active_lbl))

    def _on_select(self, _=None):
        sel = self.tree.selection()
        if not sel: return
        rid = int(sel[0])
        tbl = "cost_types" if self._mode == "cost" else "revenue_types"
        r = db.fetch_one(f"SELECT * FROM {tbl} WHERE id=?", (rid,))
        if not r: return
        self.v_name.set(r["name_ar"])
        self.v_code.set(r["code"])
        self.v_cat.set(r["category"])
        self.v_has_qty.set(bool(r["has_qty"]))
        self.v_unit.set(r["unit"] or "")
        self._toggle_unit()

    def _toggle_unit(self):
        state = "readonly" if self.v_has_qty.get() else "disabled"
        self.cbo_unit.configure(state=state)
        if not self.v_has_qty.get(): self.v_unit.set("")

    def _validate(self):
        name = self.v_name.get().strip()
        code = self.v_code.get().strip().replace(" ", "_")
        if not name:
            messagebox.showwarning("تنبيه", "الاسم العربي مطلوب", parent=self)
            return None, None
        if not code:
            messagebox.showwarning("تنبيه", "الرمز البرمجي مطلوب\nمثال: solar_val", parent=self)
            return None, None
        import re
        if not re.match(r'^[a-zA-Z][a-zA-Z0-9_]*$', code):
            messagebox.showwarning("تنبيه", "الرمز البرمجي يجب أن يبدأ بحرف إنجليزي\nويحتوي أحرف وأرقام وشرطة سفلية فقط", parent=self)
            return None, None
        return name, code

    def _add(self):
        name, code = self._validate()
        if not name: return
        tbl = "cost_types" if self._mode == "cost" else "revenue_types"
        exists = db.fetch_one(f"SELECT id FROM {tbl} WHERE code=?", (code,))
        if exists:
            messagebox.showwarning("تنبيه", f"الرمز «{code}» موجود مسبقاً", parent=self)
            return
        max_sort = db.fetch_one(f"SELECT COALESCE(MAX(sort_order),0)+1 AS s FROM {tbl}")["s"]
        db.execute(f"INSERT INTO {tbl} (code, name_ar, category, has_qty, unit, sort_order, is_active) VALUES (?,?,?,?,?,?,1)", (code, name, self.v_cat.get(), int(self.v_has_qty.get()), self.v_unit.get() or None, max_sort))
        self._load()
        self._clear()
        messagebox.showinfo("تم", f"تمت إضافة «{name}» بنجاح", parent=self)

    def _update(self):
        sel = self.tree.selection()
        if not sel: return messagebox.showwarning("تنبيه", "حدد نوعاً من القائمة أولاً", parent=self)
        name, code = self._validate()
        if not name: return
        rid = int(sel[0])
        tbl = "cost_types" if self._mode == "cost" else "revenue_types"
        exists = db.fetch_one(f"SELECT id FROM {tbl} WHERE code=? AND id!=?", (code, rid))
        if exists: return messagebox.showwarning("تنبيه", f"الرمز «{code}» مستخدم في سجل آخر", parent=self)
        db.execute(f"UPDATE {tbl} SET name_ar=?, code=?, category=?, has_qty=?, unit=? WHERE id=?", (name, code, self.v_cat.get(), int(self.v_has_qty.get()), self.v_unit.get() or None, rid))
        self._load()
        self._refresh_open_batch_forms()
        messagebox.showinfo("تم", "تم حفظ التعديل بنجاح", parent=self)

    def _delete(self):
        sel = self.tree.selection()
        if not sel: return messagebox.showwarning("تنبيه", "حدد نوعاً من القائمة أولاً", parent=self)
        rid = int(sel[0])
        tbl     = "cost_types"    if self._mode == "cost" else "revenue_types"
        link_tbl = "batch_costs"  if self._mode == "cost" else "batch_revenues"
        fk_col  = "cost_type_id"  if self._mode == "cost" else "revenue_type_id"
        used = db.fetch_one(f"SELECT COUNT(*) AS c FROM {link_tbl} WHERE {fk_col}=?", (rid,))
        if used and used["c"] > 0:
            if messagebox.askyesno("تحذير", f"هذا النوع مرتبط بـ {used['c']} دفعة.\nهل تريد إيقاف تشغيله بدلاً من الحذف؟", parent=self):
                db.execute(f"UPDATE {tbl} SET is_active=0 WHERE id=?", (rid,))
                self._load()
            return
        if not messagebox.askyesno("تأكيد", "حذف هذا النوع نهائياً؟", parent=self): return
        db.execute(f"DELETE FROM {tbl} WHERE id=?", (rid,))
        self._load()
        self._clear()

    def _toggle_active(self):
        sel = self.tree.selection()
        if not sel: return messagebox.showwarning("تنبيه", "حدد نوعاً أولاً", parent=self)
        rid = int(sel[0])
        tbl = "cost_types" if self._mode == "cost" else "revenue_types"
        r = db.fetch_one(f"SELECT is_active FROM {tbl} WHERE id=?", (rid,))
        db.execute(f"UPDATE {tbl} SET is_active=? WHERE id=?", (0 if r["is_active"] else 1, rid))
        self._load()
        self._refresh_open_batch_forms()

    def _move_up(self):
        sel = self.tree.selection()
        if not sel: return
        rid = int(sel[0])
        tbl = "cost_types" if self._mode == "cost" else "revenue_types"
        r = db.fetch_one(f"SELECT sort_order FROM {tbl} WHERE id=?", (rid,))
        if not r or r["sort_order"] <= 1: return
        prev = db.fetch_one(f"SELECT id, sort_order FROM {tbl} WHERE sort_order < ? ORDER BY sort_order DESC LIMIT 1", (r["sort_order"],))
        if prev:
            db.execute(f"UPDATE {tbl} SET sort_order=? WHERE id=?", (r["sort_order"], prev["id"]))
            db.execute(f"UPDATE {tbl} SET sort_order=? WHERE id=?", (prev["sort_order"], rid))
        self._load()
        self.tree.selection_set(str(rid))

    def _move_down(self):
        sel = self.tree.selection()
        if not sel: return
        rid = int(sel[0])
        tbl = "cost_types" if self._mode == "cost" else "revenue_types"
        r = db.fetch_one(f"SELECT sort_order FROM {tbl} WHERE id=?", (rid,))
        if not r: return
        nxt = db.fetch_one(f"SELECT id, sort_order FROM {tbl} WHERE sort_order > ? ORDER BY sort_order ASC LIMIT 1", (r["sort_order"],))
        if nxt:
            db.execute(f"UPDATE {tbl} SET sort_order=? WHERE id=?", (r["sort_order"], nxt["id"]))
            db.execute(f"UPDATE {tbl} SET sort_order=? WHERE id=?", (nxt["sort_order"], rid))
        self._load()
        self.tree.selection_set(str(rid))

    def _refresh_open_batch_forms(self):
        """تحديث تبويب التكاليف في كل نوافذ BatchForm المفتوحة حالياً"""
        try:
            for widget in self.master.winfo_children():
                if isinstance(widget, BatchForm) and hasattr(widget, "tab_costs"):
                    widget._build_costs_tab(widget.tab_costs)
                    # تحديث القيم المحفوظة
                    if widget.batch_id:
                        widget._load_batch()
        except Exception:
            pass

    def _clear(self):
        self.v_name.set(""); self.v_code.set(""); self.v_has_qty.set(False); self.v_unit.set("")
        self._toggle_unit()
        self.tree.selection_remove(*self.tree.selection())

class OnyxImporterWindow(ToplevelBase):
    """نافذة استيراد كشوف الحساب من نظام OnyxPro"""
    AUTO_MAP = {
        'feed_val': ['علف', 'بادي', 'نامي', 'ناهي', 'مركز', 'صومعة', 'كيس'],
        'chick_val': ['صوص', 'كتكوت', 'كتاكيت', 'فقاسة'],
        'drugs_val': ['أدوية', 'علاج', 'علاجات', 'لقاح', 'تحصين', 'فيتامين', 'مضاد', 'جاليميسين', 'أوكسي'],
        'gas_val': ['غاز', 'ديزل', 'وقود', 'بترول', 'أسطوانة'],
        'sawdust_val': ['نشارة', 'خشب'],
        'water_val': ['ماء', 'وايت', 'تحلية'],
        'breeders_pay': ['أجور', 'راتب', 'رواتب', 'حساب المربي', 'عامل'],
        'rent_val': ['إيجار', 'ايجار'],
    }

    def __init__(self, master):
        super().__init__(master)
        self.title("📥 مستورد أونكس برو (OnyxPro Importer)")
        self.geometry("1100x700")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.grab_set()
        
        # بيانات العمل الحالية
        self.cost_types = {ct['code']: ct['name_ar'] for ct in db.get_cost_types()}
        self.cost_codes = list(self.cost_types.keys())
        self.extracted_data = [] # [(date, desc, amount, suggested_code)]
        
        self._build()

    def _build(self):
        # Header
        hdr = UIFrame(self, bg=CLR["header"], pady=12)
        hdr.pack(fill="x")
        UILabel(hdr, text="📥 استيراد كشف حساب OnyxPro", font=FT_TITLE, bg=CLR["header"], fg="white").pack()

        # Step 1: File & Batch Selection
        top = UIFrame(self, padding=20, relief="solid", bd=1)
        top.pack(fill="x", padx=15, pady=10)
        
        UILabel(top, text="1️⃣ إعدادات الاستيراد:", font=FT_HEADER).pack(pady=(0,10))
        
        row1 = UIFrame(top)
        row1.pack(fill="x", pady=5)
        self.v_file = tk.StringVar(value="لم يتم اختيار ملف")
        UILabel(row1, text="الملف المختارة:", font=FT_SMALL).pack(side="right", padx=5)
        UILabel(row1, textvariable=self.v_file, font=FT_TINY, fg=CLR["info"]).pack(side="right", padx=10)
        tk.Button(row1, text="📁 اختيار ملف الإكسل...", command=self._pick_file,
                  font=FT_SMALL, cursor="hand2", relief="flat",
                  bg="#17a2b8", fg="white", padx=10, pady=4).pack(side="left", padx=5)

        row2 = UIFrame(top)
        row2.pack(fill="x", pady=5)
        UILabel(row2, text="الدفعة المستهدفة:", font=FT_SMALL).pack(side="right", padx=5)
        self.batches = db.fetch_all("SELECT id, batch_num, warehouse_name FROM v_batches ORDER BY date_in DESC LIMIT 50")
        self.batch_options = [f"{b['batch_num']} - {b['warehouse_name']} (ID:{b['id']})" for b in self.batches]
        self.v_batch_sel = ttk.Combobox(row2, values=self.batch_options, width=50, font=FT_SMALL, state="readonly")
        self.v_batch_sel.pack(side="right", padx=10)
        if self.batch_options: self.v_batch_sel.current(0)

        # Step 2: Review Table
        mid = UIFrame(self, padding=15)
        mid.pack(fill="both", expand=True)
        UILabel(mid, text="2️⃣ مراجعة البيانات المكتشفة وتصنيفها:", font=FT_HEADER).pack(anchor="w", pady=(0,5))
        
        cols = ("التاريخ", "البيان", "المبلغ", "التصنيف المقترح")
        self.tree = ttk.Treeview(mid, columns=cols, show="headings", selectmode="browse")
        for c, w in zip(cols, [100, 500, 100, 150]):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="center" if c != "البيان" else "e")
        
        sb = ttk.Scrollbar(mid, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        # Tool Section (Category Assignment)
        tool = UIFrame(self, padding=10, relief="solid", bd=1)
        tool.pack(fill="x", padx=15, pady=10)
        
        UILabel(tool, text="تعديل التصنيف للسطر المختار:", font=FT_SMALL).pack(side="right", padx=10)
        self.v_cat = ttk.Combobox(tool, values=["-- غير مصنف --"] + list(self.cost_types.values()), state="readonly", width=30)
        self.v_cat.pack(side="right", padx=5)
        UIButton(tool, text="✅ اعتماد التصنيف", command=self._apply_cat).pack(side="right", padx=5)
        
        # Step 3: Save
        bot = UIFrame(self, padding=15)
        bot.pack(fill="x")
        tk.Button(bot, text="💾 حفظ التكاليف المختارة في الدفعة", command=self._save_to_db,
                  font=FT_BODY, cursor="hand2", relief="flat",
                  bg="#28a745", fg="white", padx=14, pady=6).pack(side="left", padx=10)
        UIButton(bot, text="❌ إلغاء", command=self.destroy).pack(side="left")

    def _pick_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path: return
        self.v_file.set(path)
        self._process_file(path)

    def _process_file(self, path):
        if not HAS_OPENPYXL: return messagebox.showerror("خطأ", "يرجى تثبيت openpyxl")
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            self.extracted_data = []
            
            # محاولة العثور على بداية البيانات (البحث عن تاريخ في أول عمود)
            for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
                if not row or len(row) < 7: continue
                # Onyx Excel عادة يبدأ بوصف التاريخ في العمود الأول ككائن datetime أو نص تاريخ
                dt = row[0]
                if isinstance(dt, (datetime, date)) or (isinstance(dt, str) and "/" in dt):
                    try:
                        desc = str(row[3] or "")
                        debit = float(row[5] or 0)
                        if debit > 0:
                            # اقتراح تصنيف
                            suggested = "-- غير مصنف --"
                            for code, keys in self.AUTO_MAP.items():
                                if any(k in desc for k in keys):
                                    suggested = self.cost_types.get(code, suggested)
                                    break
                            
                            self.extracted_data.append({
                                'date': dt if isinstance(dt, str) else dt.strftime("%Y-%m-%d"),
                                'desc': desc,
                                'amount': debit,
                                'cat': suggested
                            })
                    except: continue
            
            self._refresh_tree()
            if not self.extracted_data:
                messagebox.showwarning("تنبيه", "لم يتم العثور على أي حركات مالية (مدين) في الملف المختار.")

        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء معالجة الملف:\n{e}")

    def _refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        for i, d in enumerate(self.extracted_data):
            self.tree.insert("", "end", iid=str(i), values=(d['date'], d['desc'], fmt_num(d['amount']), d['cat']))

    def _apply_cat(self):
        sel = self.tree.selection()
        if not sel: return messagebox.showwarning("تنبيه", "اختر سطراً من الجدول أولاً")
        new_val = self.v_cat.get()
        if not new_val: return
        
        idx = int(sel[0])
        self.extracted_data[idx]['cat'] = new_val
        self._refresh_tree()
        self.tree.selection_set(str(idx))

    def _save_to_db(self):
        if not self.extracted_data: return
        sel_idx = self.v_batch_sel.current()
        if sel_idx < 0: return messagebox.showwarning("خطأ", "يرجى اختيار الدفعة المستهدفة")
        batch_id = self.batches[sel_idx]['id']
        
        # تجميع المبالغ حسب الفئة
        consolidated = {} # {code: sum}
        rev_types = {v: k for k, v in self.cost_types.items()}
        
        count = 0
        for d in self.extracted_data:
            cat_name = d['cat']
            if cat_name == "-- غير مصنف --": continue
            
            code = rev_types.get(cat_name)
            if code:
                consolidated[code] = consolidated.get(code, 0) + d['amount']
                count += 1
        
        if not consolidated:
            return messagebox.showwarning("تنبيه", "لم يتم اختيار أي تصنيفات صالحة للحفظ.")
            
        if not messagebox.askyesno("تأكيد", f"سيتم إضافة {count} عملية مجمعة إلى {len(consolidated)} بنود تكاليف في الدفعة.\nهل تريد المتابعة؟"):
            return
            
        # تحديث قاعدة البيانات
        current_costs = db.get_batch_costs(batch_id)
        for code, amount in consolidated.items():
            old_data = current_costs.get(code, {})
            new_amount = float(old_data.get('amount', 0)) + amount
            db.save_batch_costs(batch_id, {code: {'amount': new_amount, 'qty': old_data.get('qty', 0)}})
            
        messagebox.showinfo("نجاح", "تم استيراد وحفظ التكاليف بنجاح في الدفعة المختارة.")
        self.destroy()



class SystemSettingsWindow(ToplevelBase):
    def __init__(self, master):
        super().__init__(master)
        self.title("⚙️ إعدادات النظام")
        self.geometry("700x750")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.grab_set()
        self._build()

    def _build(self):
        hdr = UIFrame(self, bg=CLR["header"], pady=12)
        hdr.pack(fill="x")
        UILabel(hdr, text="⚙️ الإعدادات العامة للنظام", font=FT_TITLE, bg=CLR["header"], fg="white").pack(side="right", padx=16)

        # ── Canvas + Scrollbar لدعم التمرير ──
        # ── إطار قابل للتمرير يعمل مع ttkbootstrap والوضع الكلاسيكي ──
        outer = tk.Frame(self)
        outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(outer, highlightthickness=0)
        vbar   = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vbar.set)
        vbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas)
        canvas_window = canvas.create_window((0,0), window=inner, anchor="nw")
        def _on_cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _resize_inner(e):
            canvas.itemconfig(canvas_window, width=e.width)
        inner.bind("<Configure>", _on_cfg)
        canvas.bind("<Configure>", _resize_inner)
        def _on_scroll(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas.bind("<MouseWheel>", _on_scroll)
        inner.bind("<MouseWheel>", _on_scroll)

        # ─── نموذج الإعدادات ─────────────────────────────────────
        frm = UILabelFrame(inner, text="بيانات التقارير والعملة", font=FT_HEADER, bg=CLR["bg"], padx=20, pady=10)
        frm.pack(fill="x", padx=15, pady=10)

        def _add_setting(row, label_text, default_val, key):
            UILabel(frm, text=label_text, font=FT_SMALL, bg=CLR["bg"]).grid(row=row, column=0, sticky="e", pady=5, padx=5)
            var = tk.StringVar(value=db.get_setting(key, default_val))
            UIEntry(frm, textvariable=var, width=38, font=FT_BODY).grid(row=row, column=1, pady=5, sticky="w")
            return var

        self.v_company  = _add_setting(0, "اسم الشركة:", "شركة جيداء الوطنية للدواجن", "company_name")
        self.v_address  = _add_setting(1, "العنوان:", "المملكة العربية السعودية - جدة", "farm_address")
        self.v_contact  = _add_setting(2, "رقم التواصل:", "", "contact_number")
        self.v_currency = _add_setting(3, "العملة:", "ريال", "currency")
        self.v_manager  = _add_setting(4, "المدير العام:", "", "manager_name")
        self.v_finance  = _add_setting(5, "المدير المالي:", "", "finance_name")
        self.v_auditor  = _add_setting(6, "المراجع:", "", "auditor_name")

        UIButton(frm, text="💾 حفظ البيانات أعلاه", font=FT_BODY, cursor="hand2", command=self._save, padx=20).grid(row=7, column=1, sticky="e", pady=10)

        # ─── إدارة قاعدة البيانات ─────────────────────────────────
        db_frm = UILabelFrame(inner, text="💾 إدارة قاعدة البيانات (نسخ واسترجاع)", font=FT_HEADER, bg=CLR["bg"], padx=20, pady=15)
        db_frm.pack(fill="x", padx=15, pady=10)

        UILabel(db_frm, text="تحذير: استرجاع نسخة قديمة سيؤدي لاستبدال كافة البيانات الحالية.", font=FT_TINY, bg=CLR["bg"], fg=CLR["loss"]).pack(pady=(0,10))

        btn_grid = UIFrame(db_frm, bg=CLR["bg"])
        btn_grid.pack(fill="x")

        UIButton(btn_grid, text="📁 نسخ احتياطي سريع", font=FT_SMALL, cursor="hand2", command=self._backup_fast).pack(side="right", padx=5)
        UIButton(btn_grid, text="📤 تصدير نسخة للخارج", font=FT_SMALL, cursor="hand2", command=self._backup_export).pack(side="right", padx=5)
        UIButton(btn_grid, text="📥 استرجاع نسخة", font=FT_SMALL, cursor="hand2", command=self._restore).pack(side="right", padx=5)

        # ─── إعدادات Telegram ──────────────────────────────────────
        tg_frm = UILabelFrame(inner, text="📱 تنبيهات Telegram اليومية", font=FT_HEADER, bg=CLR["bg"], padx=20, pady=12)
        tg_frm.pack(fill="x", padx=15, pady=10)

        UILabel(tg_frm, text="أرسل تقريراً يومياً تلقائياً لصاحب المزرعة عبر Telegram.",
                font=FT_TINY, bg=CLR["bg"], fg=CLR["text2"]).grid(row=0, column=0, columnspan=4, sticky="e", pady=(0,8))

        UILabel(tg_frm, text="Bot Token:", font=FT_SMALL, bg=CLR["bg"]).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.v_tg_token = tk.StringVar(value=db.get_setting("tg_token",""))
        UIEntry(tg_frm, textvariable=self.v_tg_token, width=46, font=FT_BODY, relief="solid").grid(row=1, column=1, columnspan=3, padx=4, sticky="ew")

        UILabel(tg_frm, text="Chat ID:", font=FT_SMALL, bg=CLR["bg"]).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.v_tg_chat = tk.StringVar(value=db.get_setting("tg_chat_id",""))
        UIEntry(tg_frm, textvariable=self.v_tg_chat, width=20, font=FT_BODY, relief="solid").grid(row=2, column=1, padx=4, sticky="w")

        UILabel(tg_frm, text="إرسال تلقائي:", font=FT_SMALL, bg=CLR["bg"]).grid(row=2, column=2, sticky="e", padx=5)
        self.v_tg_auto = tk.BooleanVar(value=db.get_setting("tg_auto","0")=="1")
        tk.Checkbutton(tg_frm, variable=self.v_tg_auto, text="تفعيل (عند فتح البرنامج يومياً)",
                       bg=CLR["bg"], font=FT_SMALL).grid(row=2, column=3, sticky="w", padx=4)

        btn_tg = UIFrame(tg_frm, bg=CLR["bg"])
        btn_tg.grid(row=3, column=0, columnspan=4, sticky="e", pady=8)
        UIButton(btn_tg, text="💾 حفظ إعدادات Telegram", font=FT_SMALL, cursor="hand2",
                 command=self._save_telegram).pack(side="right", padx=4)
        UIButton(btn_tg, text="📤 إرسال تقرير تجريبي الآن", font=FT_SMALL, cursor="hand2",
                 command=self._send_test_telegram).pack(side="right", padx=4)

        UILabel(tg_frm,
                text="💡 للحصول على Bot Token: افتح @BotFather في Telegram ← /newbot ← انسخ الـ Token\n"
                     "   للحصول على Chat ID: أرسل رسالة للـ bot ثم افتح: api.telegram.org/bot<TOKEN>/getUpdates",
                font=FT_TINY, bg=CLR["bg"], fg=CLR["text2"]).grid(row=4, column=0, columnspan=4, sticky="e", pady=(4,0))

    def _save(self):
        db.set_setting("company_name", self.v_company.get().strip())
        db.set_setting("farm_address", self.v_address.get().strip())
        db.set_setting("contact_number", self.v_contact.get().strip())
        db.set_setting("currency", self.v_currency.get().strip())
        db.set_setting("manager_name", self.v_manager.get().strip())
        db.set_setting("finance_name", self.v_finance.get().strip())
        db.set_setting("auditor_name", self.v_auditor.get().strip())
        messagebox.showinfo("تم", "تم حفظ الإعدادات بنجاح.", parent=self)

    def _save_telegram(self):
        db.set_setting("tg_token",   self.v_tg_token.get().strip())
        db.set_setting("tg_chat_id", self.v_tg_chat.get().strip())
        db.set_setting("tg_auto",    "1" if self.v_tg_auto.get() else "0")
        messagebox.showinfo("تم", "تم حفظ إعدادات Telegram بنجاح.", parent=self)

    def _send_test_telegram(self):
        token   = self.v_tg_token.get().strip()
        chat_id = self.v_tg_chat.get().strip()
        if not token or not chat_id:
            return messagebox.showwarning("تنبيه", "يرجى إدخال Bot Token و Chat ID أولاً", parent=self)
        msg = build_daily_telegram_report()
        ok, err = send_telegram(token, chat_id, msg)
        if ok:
            messagebox.showinfo("تم الإرسال ✅",
                "تم إرسال التقرير التجريبي بنجاح!\nتحقق من محادثة Telegram.", parent=self)
        else:
            messagebox.showerror("فشل الإرسال ❌",
                f"تعذّر الإرسال:\n{err}\n\nتأكد من صحة Token و Chat ID والاتصال بالإنترنت.",
                parent=self)

    def _backup_fast(self):
        path = make_backup()
        if path:
            messagebox.showinfo("نجاح", f"تم إنشاء نسخة احتياطية سريعة في:\n{path}", parent=self)
        else:
            messagebox.showerror("خطأ", "فشل إنشاء النسخة الاحتياطية", parent=self)

    def _backup_export(self):
        dest = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("SQLite Database","*.db")], 
                                          initialfile=f"poultry_backup_{datetime.now().strftime('%Y%m%d')}.db", parent=self)
        if not dest: return
        try:
            shutil.copy2(DB_PATH, dest)
            messagebox.showinfo("نجاح", "تم تصدير نسخة من قاعدة البيانات بنجاح", parent=self)
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل التصدير: {e}", parent=self)

    def _restore(self):
        if not messagebox.askyesno("تأكيد خطير", "هل أنت متأكد من استرجاع نسخة احتياطية؟\nسيتم حذف كافة البيانات الحالية واستبدالها بالنسخة المختارة.", parent=self):
            return
        
        path = filedialog.askopenfilename(filetypes=[("SQLite Database","*.db")], parent=self)
        if not path: return
        
        success, msg = restore_db(path)
        if success:
            messagebox.showinfo("تم الاسترجاع", "تم استرجاع البيانات بنجاح.\n\n⚠️ يرجى إغلاق البرنامج الآن وإعادة فتحه لضمان تحديث كافة البيانات.", parent=self)
        else:
            messagebox.showerror("خطأ", f"فشل الاسترجاع: {msg}", parent=self)


# ════════════════════════════════════════════════════════════════
# شاشة البداية — Splash Screen
# ════════════════════════════════════════════════════════════════
class SplashScreen(tk.Toplevel):
    """شاشة بداية تظهر 3 ثواني عند فتح البرنامج"""
    DEV_NAME = "Eng. Saleem Homi"
    DEV_PHONE = "770199865"

    def __init__(self, master):
        super().__init__(master)
        self.overrideredirect(True)          # بدون شريط عنوان
        self.resizable(False, False)
        self.configure(bg="#1F4E79")

        # ── تمركز في منتصف الشاشة ──
        w, h = 480, 300
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.lift(); self.attributes("-topmost", True)

        # ── شعار / أيقونة ──
        tk.Label(self, text="🐔", font=("Arial",52),
                 bg="#1F4E79", fg="white").pack(pady=(30,5))

        # ── اسم النظام ──
        tk.Label(self,
                 text="نظام إدارة عنابر الدجاج اللاحم",
                 font=("Arial",16,"bold"),
                 bg="#1F4E79", fg="white").pack()
        tk.Label(self,
                 text="Poultry Farm Management System",
                 font=("Arial",11),
                 bg="#1F4E79", fg="#aad4f5").pack(pady=(2,16))

        # ── خط فاصل ──
        tk.Frame(self, bg="#2e6da4", height=1).pack(fill="x", padx=40)

        # ── بيانات المطوّر ──
        tk.Label(self,
                 text=f"Developed & Engineered by:  {self.DEV_NAME}",
                 font=("Arial",10),
                 bg="#1F4E79", fg="#ffd966").pack(pady=(14,2))
        tk.Label(self,
                 text=f"📞  {self.DEV_PHONE}",
                 font=("Arial",10),
                 bg="#1F4E79", fg="#aad4f5").pack()

        # ── شريط تحميل بصري ──
        self._bar_frm = tk.Frame(self, bg="#1F4E79")
        self._bar_frm.pack(pady=(18,0))
        tk.Label(self._bar_frm, text="Loading...",
                 font=("Arial",8), bg="#1F4E79", fg="#6699bb").pack(side="right", padx=6)
        self._bar_bg = tk.Frame(self._bar_frm, bg="#2e6da4",
                                width=260, height=5)
        self._bar_bg.pack(side="right")
        self._bar_fg = tk.Frame(self._bar_bg, bg="#ffd966",
                                width=0, height=5)
        self._bar_fg.place(x=0, y=0, height=5)
        self._step = 0
        self._animate()

        # ── إغلاق تلقائي بعد 3 ثواني ──
        self.after(3000, self.destroy)

    def _animate(self):
        self._step += 1
        w = min(260, self._step * 9)
        self._bar_fg.place(x=0, y=0, width=w, height=5)
        if w < 260:
            self.after(100, self._animate)


# ════════════════════════════════════════════════════════════════
# نافذة "حول البرنامج" — About
# ════════════════════════════════════════════════════════════════
class AboutWindow(ToplevelBase):
    DEV_NAME  = "المهندس سليم حومي  /  Eng. Saleem Homi"
    DEV_PHONE = "770199865"

    def __init__(self, master):
        super().__init__(master)
        self.title("حول البرنامج — About")
        self.geometry("440x360")
        self.resizable(False, False)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.grab_set()
        self._build()

    def _build(self):
        # ── رأس ملوّن ──
        hdr = tk.Frame(self, bg="#1F4E79", pady=18)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🐔", font=("Arial",36),
                 bg="#1F4E79", fg="white").pack()
        tk.Label(hdr,
                 text="نظام إدارة عنابر الدجاج اللاحم",
                 font=("Arial",13,"bold"),
                 bg="#1F4E79", fg="white").pack(pady=(4,2))
        tk.Label(hdr,
                 text="Poultry Farm Management System  v4.6",
                 font=("Arial",9),
                 bg="#1F4E79", fg="#aad4f5").pack()

        # ── تفاصيل ──
        body = UIFrame(self, bg=CLR["bg"], pady=16)
        body.pack(fill="both", expand=True)

        rows = [
            ("التصميم والتطوير:", "المهندس سليم حومي"),
            ("Developed by:",      "Eng. Saleem Homi"),
            ("رقم التواصل:",       self.DEV_PHONE),
            ("Contact:",           self.DEV_PHONE),
            ("الإصدار:",           "4.6  —  2025"),
            ("قاعدة البيانات:",    "SQLite 3  (WAL Mode)"),
            ("واجهة المستخدم:",   "Python / Tkinter / ttkbootstrap"),
        ]
        for lbl, val in rows:
            row_frm = UIFrame(body, bg=CLR["bg"])
            row_frm.pack(fill="x", padx=30, pady=3)
            UILabel(row_frm, text=val, font=("Arial",10,"bold"),
                    bg=CLR["bg"], fg=CLR["header"]).pack(side="right")
            UILabel(row_frm, text=lbl, font=("Arial",9),
                    bg=CLR["bg"], fg=CLR["text2"]).pack(side="right", padx=(0,8))

        tk.Frame(self, bg="#dddddd", height=1).pack(fill="x", padx=20, pady=(0,8))
        UIButton(self, text="إغلاق", font=FT_BODY,
                 cursor="hand2", command=self.destroy).pack(pady=8)

class MainWindow(WindowBase):
    def __init__(self):
        saved_theme = db.get_setting("theme", "lumen")
        if HAS_TTKB:
            try: super().__init__(themename=saved_theme)
            except: super().__init__()
        else: super().__init__()
        self.title("نظام إدارة عنابر الدجاج اللاحم — النسخة المطورة 4.6")
        self.geometry("1200x700")
        center_window(self)
        if not HAS_TTKB: self.configure(bg=CLR["bg"])
        self.resizable(True, True)
        self._build(); self._load_batches()
        # إرسال تقرير Telegram تلقائي (إذا كان مفعلاً)
        self.after(3000, self._auto_telegram)

    def _change_theme(self, event):
        if HAS_TTKB:
            try: 
                new_theme = self.theme_cbo.get()
                self.style.theme_use(new_theme)
                db.set_setting("theme", new_theme)
            except: pass

    def _build(self):
        if HAS_TTKB:
            hdr = ttkb.Frame(self, padding=10, bootstyle="primary")
            hdr.pack(fill="x")
            ttkb.Label(hdr, text="🐔 نظام إدارة عنابر الدجاج اللاحم", font=("Arial",16,"bold"), bootstyle="inverse-primary").pack(side="right", padx=20)
            self.lbl_count = ttkb.Label(hdr, text="", font=FT_BODY, bootstyle="inverse-primary"); self.lbl_count.pack(side="right", padx=20)
            self.theme_cbo = ttkb.Combobox(hdr, values=self.style.theme_names() if hasattr(self, 'style') else ttkb.Style().theme_names(), width=15, state="readonly", bootstyle="primary")
            self.theme_cbo.pack(side="left", padx=10)
            
            saved_theme = db.get_setting("theme", "lumen")
            self.theme_cbo.set(saved_theme)
            
            self.theme_cbo.bind("<<ComboboxSelected>>", self._change_theme)
            ttkb.Label(hdr, text="المظهر (Theme):", font=FT_SMALL, bootstyle="inverse-primary").pack(side="left", padx=5)

            # ── شريط الأدوات المبسط — نافذتان رئيسيتان ──
            tb1 = ttkb.Frame(self, padding=(8,6), bootstyle="secondary"); tb1.pack(fill="x")
            ttkb.Button(tb1, text="📝  مركز الإدخالات  — الدفعات والسجلات اليومية والعنابر",
                command=self._open_entry_hub, bootstyle="success",
                cursor="hand2", padding=(20,8)).pack(side="right", padx=6)
            ttkb.Button(tb1, text="📊  مركز التقارير  — التصدير والتحليلات والـ PDF",
                command=self._open_reports_hub, bootstyle="primary",
                cursor="hand2", padding=(20,8)).pack(side="right", padx=6)
            ttkb.Separator(tb1, orient="vertical").pack(side="right", fill="y", padx=10, pady=2)
            for txt, cmd, bstyle in [
                ("📱 Telegram",   self._send_telegram_now, "info-outline"),
                ("⚙ الإعدادات",  self._open_settings,     "dark"),
                ("ℹ حول البرنامج", self._open_about,       "secondary"),
            ]:
                ttkb.Button(tb1, text=txt, command=cmd, bootstyle=bstyle,
                            cursor="hand2", padding=(10,6)).pack(side="right", padx=3)

            fbar = ttkb.Frame(self, padding=4); fbar.pack(fill="x", padx=8)
            self.filter_wh = ttkb.Combobox(fbar, width=20, font=FT_BODY); self.filter_wh.pack(side="right", padx=4); self.filter_wh.bind("<<ComboboxSelected>>", lambda e: self._load_batches())
            ttkb.Label(fbar, text="السنة المالية:", font=FT_SMALL, bootstyle="secondary").pack(side="right", padx=(8,2))
            self.filter_fy = ttkb.Combobox(fbar, width=8, font=FT_BODY, state="readonly"); self.filter_fy.pack(side="right", padx=2); self.filter_fy.bind("<<ComboboxSelected>>", lambda e: self._load_batches())
            ttkb.Button(fbar, text="عرض الكل", command=lambda: [self.filter_wh.set(""), self.filter_fy.set(""), self._load_batches()], bootstyle="link").pack(side="right", padx=4)

            self.kpi_frame = ttkb.Frame(self); self.kpi_frame.pack(fill="x", padx=8, pady=4)
            frm = ttkb.Frame(self); frm.pack(fill="both", expand=True, padx=8, pady=(0,8))
            cols = ("رقم الدفعة","العنبر","تاريخ الدخول","تاريخ الخروج","الأيام","الكتاكيت","التكاليف","الإيرادات","صافي النتيجة","النافق%","معدل FCR","نصيب الشركة")
            self.tree = ttkb.Treeview(frm, columns=cols, show="headings", selectmode="browse", bootstyle="primary")
        else:
            hdr = UIFrame(self, bg=CLR["header"], pady=10); hdr.pack(fill="x")
            UILabel(hdr, text="🐔 نظام إدارة عنابر الدجاج اللاحم", font=("Arial",16,"bold"), bg=CLR["header"], fg="white").pack(side="right", padx=20)
            self.lbl_count = UILabel(hdr, text="", font=FT_BODY, bg=CLR["header"], fg="#aad4f5"); self.lbl_count.pack(side="left", padx=20)

            tb = UIFrame(self, bg=CLR["nav"], pady=8); tb.pack(fill="x")
            UIButton(tb, text="📝  مركز الإدخالات  — الدفعات والسجلات اليومية والعنابر",
                command=self._open_entry_hub, font=FT_BODY,
                bg="#e8f5e9", fg="#27680a", padx=20, pady=8,
                cursor="hand2", relief="flat").pack(side="right", padx=8)
            UIButton(tb, text="📊  مركز التقارير  — التصدير والتحليلات والـ PDF",
                command=self._open_reports_hub, font=FT_BODY,
                bg="#e3f2fd", fg="#1F4E79", padx=20, pady=8,
                cursor="hand2", relief="flat").pack(side="right", padx=8)
            UIFrame(tb, bg="#1a5296", width=2).pack(side="right", fill="y", padx=8, pady=4)
            for txt, cmd, bg in [("📱 Telegram",self._send_telegram_now,"#e1f5fe"),
                                  ("⚙ الإعدادات",self._open_settings,"#cfd8dc"),
                                  ("ℹ حول",self._open_about,"#e8eaf6")]:
                UIButton(tb, text=txt, command=cmd, font=FT_SMALL, bg=bg,
                         fg=CLR["text"], padx=10, pady=6,
                         cursor="hand2", relief="flat").pack(side="right", padx=4)

            fbar = UIFrame(self, bg=CLR["bg"], pady=4); fbar.pack(fill="x", padx=8)
            self.filter_wh = ttk.Combobox(fbar, width=20, font=FT_BODY); self.filter_wh.pack(side="right", padx=4); self.filter_wh.bind("<<ComboboxSelected>>", lambda e: self._load_batches())
            UILabel(fbar, text="السنة:", font=FT_SMALL, bg=CLR["bg"]).pack(side="right", padx=(8,2))
            self.filter_fy = ttk.Combobox(fbar, width=8, font=FT_BODY, state="readonly"); self.filter_fy.pack(side="right", padx=2); self.filter_fy.bind("<<ComboboxSelected>>", lambda e: self._load_batches())
            UIButton(fbar, text="عرض الكل", font=FT_SMALL, command=lambda: [self.filter_wh.set(""), self.filter_fy.set(""), self._load_batches()], bg=CLR["bg"], relief="flat").pack(side="right", padx=4)

            self.kpi_frame = UIFrame(self, bg=CLR["bg"]); self.kpi_frame.pack(fill="x", padx=8, pady=4)
            frm = UIFrame(self, bg=CLR["bg"]); frm.pack(fill="both", expand=True, padx=8, pady=(0,8))
            cols = ("رقم الدفعة","العنبر","تاريخ الدخول","تاريخ الخروج","الأيام","الكتاكيت","التكاليف","الإيرادات","صافي النتيجة","النافق%","معدل FCR","نصيب الشركة")
            self.tree = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")

        for c, w in zip(cols, [80, 150, 100, 100, 50, 90, 110, 110, 120, 70, 80, 110]):
            self.tree.heading(c, text=c, anchor="center"); self.tree.column(c, width=w, anchor="e" if c=="العنبر" else "center")

        if not HAS_TTKB:
            self.tree.tag_configure("profit", background="#f0f9ea"); self.tree.tag_configure("loss", background="#fff0f0")

        self.tree.bind("<Double-1>", lambda e: self._edit_batch())
        sb_y = ttk.Scrollbar(frm, command=self.tree.yview); self.tree.configure(yscrollcommand=sb_y.set); sb_y.pack(side="left", fill="y"); self.tree.pack(fill="both", expand=True)

    def _load_batches(self):
        wh_filter = self.filter_wh.get().strip()
        fy_filter = self.filter_fy.get().strip() if hasattr(self, "filter_fy") else ""

        # تحديث قوائم الفلاتر
        self.filter_wh["values"] = [""] + [r["name"] for r in db.fetch_all("SELECT name FROM warehouses ORDER BY name")]
        # جلب السنوات المالية المتاحة
        fy_rows = db.fetch_all(
            "SELECT DISTINCT COALESCE(fiscal_year, CAST(strftime('%Y',date_in) AS INTEGER)) AS fy "
            "FROM batches ORDER BY fy DESC")
        fy_vals = [""] + [str(r["fy"]) for r in fy_rows if r["fy"]]
        if hasattr(self, "filter_fy"): self.filter_fy["values"] = fy_vals

        # بناء الاستعلام حسب الفلاتر
        where, params = [], []
        if wh_filter:
            where.append("warehouse_name=?"); params.append(wh_filter)
        if fy_filter:
            where.append("COALESCE(fiscal_year, CAST(strftime('%Y',date_in) AS INTEGER))=?")
            params.append(int(fy_filter))
        q = "SELECT * FROM v_batches"
        if where: q += " WHERE " + " AND ".join(where)
        q += " ORDER BY date_in DESC"
        rows = db.fetch_all(q, params)

        self.tree.delete(*self.tree.get_children())

        T = {"cost":0,"rev":0,"net":0,"chicks":0,"share":0}
        for b in rows:
            profit = (b["net_result"] or 0) >= 0; mort = b["mort_rate"] or 0; b_num = b["batch_num"] if b["batch_num"] else str(b["id"])
            self.tree.insert("", "end", iid=str(b["id"]), tags=("profit" if profit else "loss",), values=(b_num, b["warehouse_name"], b["date_in"], b["date_out"], b["days"], fmt_num(b["chicks"]), fmt_num(b["total_cost"]), fmt_num(b["total_rev"]), f"{'+' if profit else ''}{fmt_num(b['net_result'])}", f"{mort:.1f}%", b["fcr"] or "0", fmt_num(b["share_val"])))
            T["cost"] += b["total_cost"] or 0; T["rev"] += b["total_rev"] or 0; T["net"] += b["net_result"] or 0; T["chicks"] += b["chicks"] or 0; T["share"] += b["share_val"] or 0

        for w in self.kpi_frame.winfo_children(): w.destroy()
        sign_net = "+" if T['net'] >= 0 else ""

        if HAS_TTKB:
            for lbl, val, val_bstyle, frm_bstyle in [("الدفعات", str(len(rows)), "primary", "secondary"), ("إجمالي الكتاكيت", fmt_num(T["chicks"]), "info", "secondary"), ("إجمالي التكاليف", fmt_num(T["cost"]), "danger", "secondary"), ("إجمالي الإيرادات",fmt_num(T["rev"]), "success", "secondary"), ("صافي النتيجة", f"{sign_net}{fmt_num(T['net'])}", "success" if T["net"]>=0 else "danger", "success" if T["net"]>=0 else "danger"), ("نصيب الشركة", fmt_num(T["share"]), "warning", "secondary")]:
                lfrm = ttkb.Frame(self.kpi_frame, padding=8); lfrm.pack(side="right", padx=5)
                ttkb.Label(lfrm, text=lbl, font=FT_SMALL, bootstyle=frm_bstyle).pack()
                ttkb.Label(lfrm, text=val, font=("Arial",14,"bold"), bootstyle=val_bstyle).pack()
        else:
            for lbl, val, bg, fg in [("الدفعات", str(len(rows)), "#dce6f1", CLR["header"]), ("إجمالي الكتاكيت", fmt_num(T["chicks"]), "#dce6f1", CLR["header"]), ("إجمالي التكاليف", fmt_num(T["cost"]), CLR["loss_bg"], CLR["loss"]), ("إجمالي الإيرادات",fmt_num(T["rev"]), CLR["profit_bg"], CLR["profit"]), ("صافي النتيجة", f"{sign_net}{fmt_num(T['net'])}", CLR["profit_bg"] if T["net"]>=0 else CLR["loss_bg"], CLR["profit"] if T["net"]>=0 else CLR["loss"]), ("نصيب الشركة", fmt_num(T["share"]), "#fff2cc", CLR["warn"])]:
                frm = UIFrame(self.kpi_frame, bg=bg, padx=12, pady=6, relief="solid", bd=1); frm.pack(side="right", padx=3)
                UILabel(frm, text=lbl, font=FT_SMALL, bg=bg, fg=CLR["text2"]).pack()
                UILabel(frm, text=val, font=("Arial",12,"bold"), bg=bg, fg=fg).pack()
        self.lbl_count.config(text=f"{len(rows)} دفعة مسجلة")

    def _selected_id(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("تنبيه", "يرجى تحديد دفعة من الجدول أولاً")
            return None
        return int(sel[0])

    def _new_batch(self): BatchForm(self, on_save=self._load_batches)
    def _edit_batch(self):
        bid = self._selected_id()
        if bid: BatchForm(self, batch_id=bid, on_save=self._load_batches)
    def _del_batch(self, bid=None):
        if bid is None: bid = self._selected_id()
        if not bid: return
        b = db.fetch_one("SELECT batch_num, warehouse_name FROM v_batches WHERE id=?", (bid,))
        if not b: return
        label = f"{b['warehouse_name']} — {b['batch_num'] or bid}"

        sales = db.fetch_one("SELECT COUNT(*) AS c FROM farm_sales   WHERE batch_id=?", (bid,))["c"]
        mkt   = db.fetch_one("SELECT COUNT(*) AS c FROM market_sales  WHERE batch_id=?", (bid,))["c"]
        daily = db.fetch_one("SELECT COUNT(*) AS c FROM daily_records WHERE batch_id=?", (bid,))["c"]

        details = ""
        if sales or mkt or daily:
            parts = []
            if sales: parts.append(f"{sales} فاتورة مبيعات")
            if mkt:   parts.append(f"{mkt} سجل سوق")
            if daily: parts.append(f"{daily} سجل يومي")
            details = "\nسيتم حذف أيضاً: " + " | ".join(parts)

        if messagebox.askyesno("تأكيد الحذف",
                f"حذف الدفعة «{label}» نهائياً؟{details}\n\nلا يمكن التراجع عن هذا الإجراء.",
                parent=self):
            db.execute("DELETE FROM batches WHERE id=?", (bid,))
            self._load_batches()

    def _del_warehouse(self):
        win = ToplevelBase(self)
        win.title("🗑 حذف عنبر")
        win.geometry("500x400")
        if not HAS_TTKB: win.configure(bg=CLR["bg"])
        win.grab_set()
        hdr = UIFrame(win, bg=CLR["header"], pady=8); hdr.pack(fill="x")
        UILabel(hdr, text="🗑 حذف / إدارة العنابر", font=FT_HEADER, bg=CLR["header"], fg="white").pack(side="right", padx=14)
        frm = UIFrame(win, bg=CLR["bg"], padx=15, pady=10); frm.pack(fill="both", expand=True)
        UILabel(frm, text="اختر العنبر:", font=FT_BODY, bg=CLR["bg"]).pack(anchor="e", pady=(0,4))
        whs = db.fetch_all("SELECT w.id, w.name, COUNT(b.id) AS batch_count FROM warehouses w LEFT JOIN batches b ON b.warehouse_id=w.id GROUP BY w.id ORDER BY w.name")
        cols = ("الاسم", "عدد الدفعات")
        tree = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse", height=10)
        tree.heading("الاسم", text="الاسم", anchor="center")
        tree.heading("عدد الدفعات", text="عدد الدفعات", anchor="center")
        tree.column("الاسم", width=250, anchor="center")
        tree.column("عدد الدفعات", width=120, anchor="center")
        for w in whs:
            tree.insert("", "end", iid=str(w["id"]), values=(w["name"], w["batch_count"]))
        tree.pack(fill="both", expand=True)
        def _do_del():
            sel = tree.selection()
            if not sel: return messagebox.showwarning("تنبيه", "اختر عنبراً أولاً", parent=win)
            wid = int(sel[0])
            wrow = db.fetch_one("SELECT name FROM warehouses WHERE id=?", (wid,))
            cnt  = db.fetch_one("SELECT COUNT(*) AS c FROM batches WHERE warehouse_id=?", (wid,))["c"]
            wh_name_str = wrow['name'] if wrow else "غير معروف"
            if cnt > 0:
                messagebox.showerror("لا يمكن الحذف", f"العنبر «{wh_name_str}» مرتبط بـ {cnt} دفعة.\nاحذف الدفعات أولاً ثم احذف العنبر.", parent=win)
                return
            if messagebox.askyesno("تأكيد", f"حذف العنبر «{wh_name_str}» نهائياً؟", parent=win):
                db.execute("DELETE FROM warehouses WHERE id=?", (wid,))
                self._load_batches()
                win.destroy()
                messagebox.showinfo("تم", "تم حذف العنبر بنجاح")
        UIButton(frm, text="🗑 حذف العنبر المحدد", font=FT_BODY, cursor="hand2", command=_do_del).pack(pady=10)
        
    def _open_daily(self):
        bid = self._selected_id()
        if bid:
            batch = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (bid,))
            if batch: DailyRecordsWindow(self, bid, dict(batch))
            
    def _send_telegram_now(self):
        """إرسال تقرير Telegram فوري بضغطة زر"""
        token   = db.get_setting("tg_token",   "")
        chat_id = db.get_setting("tg_chat_id", "")
        if not token or not chat_id:
            return messagebox.showwarning("تنبيه",
                "يرجى إعداد Bot Token و Chat ID في\n⚙️ الإعدادات → Telegram أولاً")
        msg = build_daily_telegram_report()
        ok, err = send_telegram(token, chat_id, msg)
        if ok:
            messagebox.showinfo("تم ✅", "تم إرسال التقرير إلى Telegram بنجاح!")
        else:
            messagebox.showerror("فشل ❌",
                f"تعذّر الإرسال:\n{err}\n\nتحقق من الإنترنت والإعدادات.")

    def _auto_telegram(self):
        """إرسال تقرير Telegram تلقائي عند بدء البرنامج إذا كان الإرسال مفعلاً"""
        if db.get_setting("tg_auto", "0") != "1": return
        token   = db.get_setting("tg_token",   "")
        chat_id = db.get_setting("tg_chat_id", "")
        if not token or not chat_id: return
        # تحقق: هل أُرسل اليوم مسبقاً؟
        today = date.today().isoformat()
        if db.get_setting("tg_last_sent", "") == today: return
        msg = build_daily_telegram_report()
        ok, _ = send_telegram(token, chat_id, msg)
        if ok:
            db.set_setting("tg_last_sent", today)


    def _export_daily_pdf(self):
        """تقرير PDF يومي لحركة كل العنابر النشطة — للإدارة الفنية"""
        if not HAS_FPDF:
            messagebox.showerror("خطأ", "مكتبة fpdf غير مثبتة.\nنفّذ: pip install fpdf2")
            return
        font_path = AMIRI_FONT_PATH
        if not os.path.exists(font_path):
            messagebox.showerror("خطأ", "ملف الخط Amiri-Regular.ttf غير موجود!")
            return

        # اختيار تاريخ التقرير
        from tkinter.simpledialog import askstring
        report_date = askstring("تاريخ التقرير",
            "أدخل التاريخ (YYYY-MM-DD)\nاتركه فارغاً لليوم الحالي:",
            initialvalue=date.today().isoformat())
        if report_date is None: return
        report_date = report_date.strip() or date.today().isoformat()
        try:
            datetime.strptime(report_date, "%Y-%m-%d")
        except:
            return messagebox.showerror("خطأ", "صيغة التاريخ غير صحيحة. استخدم: YYYY-MM-DD")

        # جلب الدفعات النشطة في ذلك التاريخ
        active_batches = db.fetch_all(
            "SELECT * FROM v_batches WHERE date_in<=? AND date_out>=? ORDER BY warehouse_name",
            (report_date, report_date))

        company  = db.get_setting("company_name", "شركة آفاق الريف للدواجن")
        currency = db.get_setting("currency", "ريال")

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.add_page()
        pdf.add_font("Arabic", "", font_path, uni=True)

        # ── رأس الصفحة ──
        logo_path = LOGO_PATH
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=10, y=5, w=25)

        pdf.set_font("Arabic","",16)
        pdf.cell(0,10,prepare_text(company),ln=True,align="C")
        pdf.set_font("Arabic","",13)
        pdf.cell(0,8,prepare_text("التقرير الفني اليومي للعنابر"),ln=True,align="C")
        pdf.set_font("Arabic","",11)
        pdf.cell(0,7,prepare_text(f"تاريخ التقرير: {report_date}    |    عدد العنابر النشطة: {len(active_batches)}"),ln=True,align="C")
        pdf.ln(2)

        if not active_batches:
            pdf.set_font("Arabic","",12)
            pdf.cell(0,10,prepare_text("لا توجد عنابر نشطة في هذا التاريخ"),ln=True,align="C")
        else:
            # ── رؤوس الجدول ──
            # عرض الأعمدة: عنبر(30) دفعة(20) يوم(15) كتل نشطة(20) نافق واقعي(22) نافق طبيعي(22) فرق نافق(20) علف واقعي(22) علف مطلوب(22) فرق علف(20)
            cols   = ["العنبر","دفعة","يوم","كتل نشطة","نافق واقعي","نافق طبيعي","فرق نافق","علف واقعي كجم","علف مطلوب كجم","فرق علف","ملاحظات"]
            widths = [30,18,12,20,18,18,16,20,20,16,22]

            pdf.set_fill_color(31,78,121); pdf.set_text_color(255,255,255)
            pdf.set_font("Arabic","",8)
            for w,h in zip(widths,cols):
                pdf.cell(w,10,prepare_text(h),1,0,"C",True)
            pdf.ln(); pdf.set_text_color(0,0,0)

            # معدلات النافق والعلف الافتراضية (أسبوع 1-8)
            MORT_STD = {1:0.15,2:0.08,3:0.06,4:0.05,5:0.05,6:0.05,7:0.05,8:0.05}
            FEED_STD = {1:20,2:45,3:80,4:115,5:145,6:165,7:175,8:180}

            total_dead_all=0; total_exp_dead_all=0
            total_feed_all=0; total_exp_feed_all=0
            total_active_all=0

            for i, b in enumerate(active_batches):
                bid    = b["id"]
                chicks = b["chicks"] or 0
                # حساب رقم اليوم
                try:
                    d_in = datetime.strptime(b["date_in"],"%Y-%m-%d")
                    d_rep= datetime.strptime(report_date,"%Y-%m-%d")
                    day_num = (d_rep - d_in).days + 1
                except: day_num = 0

                week = min(max(1,(day_num-1)//7+1),8)

                # الكتل النشطة حتى تاريخ التقرير
                active = calc_active_birds_for_batch(bid, chicks, report_date)

                # جلب سجل ذلك اليوم
                rec = db.fetch_one(
                    "SELECT dead_count, feed_kg, notes FROM daily_records WHERE batch_id=? AND rec_date=?",
                    (bid, report_date))
                dead_actual = rec["dead_count"] if rec else None
                feed_actual = rec["feed_kg"]    if rec else None
                notes_day   = rec["notes"]      if rec else ""

                # المعدل الطبيعي
                exp_dead = round(active * MORT_STD[week] / 100, 1)
                exp_feed = round(active * FEED_STD[week] / 1000, 1)

                # الفرق
                if dead_actual is not None:
                    diff_dead = round(dead_actual - exp_dead, 1)
                    diff_feed = round((feed_actual or 0) - exp_feed, 1)
                    dead_str = str(dead_actual)
                    feed_str = f"{feed_actual:.1f}" if feed_actual else "—"
                    diff_d_str = f"{'+' if diff_dead>0 else ''}{diff_dead:.1f}"
                    diff_f_str = f"{'+' if diff_feed>0 else ''}{diff_feed:.1f}"
                    # تلوين حسب الحالة
                    dead_alert = diff_dead > exp_dead * 0.5 and diff_dead > 0
                    feed_alert = diff_feed > exp_feed * 0.3 and diff_feed > 0
                    if dead_alert and feed_alert: pdf.set_fill_color(252,228,214)
                    elif dead_alert:              pdf.set_fill_color(252,228,214)
                    elif feed_alert:              pdf.set_fill_color(255,242,204)
                    else:                         pdf.set_fill_color(226,239,218)
                    fill=True
                    total_dead_all  += dead_actual
                    total_feed_all  += feed_actual or 0
                else:
                    diff_d_str = "—"; diff_f_str = "—"
                    dead_str = "لم يُسجل"; feed_str = "لم يُسجل"
                    pdf.set_fill_color(245,245,245); fill=True

                total_exp_dead_all += exp_dead
                total_exp_feed_all += exp_feed
                total_active_all   += active

                pdf.set_font("Arabic","",8)
                row_vals = [
                    b["warehouse_name"],
                    b["batch_num"] or str(bid),
                    str(day_num),
                    f"{active:,}",
                    dead_str,
                    f"{exp_dead:.1f}",
                    diff_d_str,
                    feed_str,
                    f"{exp_feed:.1f}",
                    diff_f_str,
                    notes_day or ""
                ]
                for w, v in zip(widths, row_vals):
                    pdf.cell(w,8,prepare_text(str(v)),1,0,"C",fill)
                pdf.ln()

            # ── سطر المجاميع ──
            pdf.set_fill_color(31,78,121); pdf.set_text_color(255,255,255)
            pdf.set_font("Arabic","",9)
            diff_dead_total = round(total_dead_all - total_exp_dead_all, 1)
            diff_feed_total = round(total_feed_all - total_exp_feed_all, 1)
            totals = [
                "الإجمالي","","",f"{total_active_all:,}",
                str(total_dead_all),f"{total_exp_dead_all:.1f}",
                f"{'+' if diff_dead_total>0 else ''}{diff_dead_total:.1f}",
                f"{total_feed_all:.1f}",f"{total_exp_feed_all:.1f}",
                f"{'+' if diff_feed_total>0 else ''}{diff_feed_total:.1f}",""]
            for w,v in zip(widths,totals):
                pdf.cell(w,9,prepare_text(v),1,0,"C",True)
            pdf.ln()
            pdf.set_text_color(0,0,0)

            # ── مفتاح الألوان ──
            pdf.ln(4)
            pdf.set_font("Arabic","",9)
            pdf.set_fill_color(226,239,218); pdf.cell(6,5,"",1,0,"C",True)
            pdf.cell(40,5,prepare_text("  أداء طبيعي"),0,0,"R")
            pdf.set_fill_color(255,242,204); pdf.cell(6,5,"",1,0,"C",True)
            pdf.cell(40,5,prepare_text("  علف مرتفع"),0,0,"R")
            pdf.set_fill_color(252,228,214); pdf.cell(6,5,"",1,0,"C",True)
            pdf.cell(40,5,prepare_text("  نافق مرتفع / كلاهما"),0,0,"R")
            pdf.set_fill_color(245,245,245); pdf.cell(6,5,"",1,0,"C",True)
            pdf.cell(50,5,prepare_text("  لم يُسجل بعد"),0,0,"R")
            pdf.ln(8)

        # ── التوقيعات ──
        pdf.set_font("Arabic","",10)
        pdf.set_fill_color(240,240,240)
        pdf.cell(0,7,prepare_text("التوقيعات"),1,1,"C",True)
        pdf.cell(63,12,prepare_text("المشرف الفني"),1,0,"C")
        pdf.cell(63,12,prepare_text("مدير الإنتاج"),1,0,"C")
        pdf.cell(64,12,prepare_text("المدير التنفيذي"),1,1,"C")

        # ── تذييل المطوّر ──
        pdf.set_y(-10)
        pdf.set_font("Arabic", "", 7)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 5, "Developed & Engineered by: Eng. Saleem Homi  |  Tel: 770199865", 0, 0, 'C')
        pdf.set_text_color(0, 0, 0)
        save_path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF","*.pdf")],
            initialfile=f"التقرير_اليومي_{report_date}.pdf",
            title="حفظ التقرير الفني اليومي")
        if save_path:
            pdf.output(save_path)
            messagebox.showinfo("تم ✅", f"تم تصدير التقرير الفني اليومي بنجاح!\n{save_path}")
            try: os.startfile(save_path)
            except: pass

    def _open_entry_hub(self):
        DataEntryHub(self, on_refresh=self._load_batches)

    def _open_reports_hub(self):
        ReportsHub(self)

    def _export_wh_excel(self):
        """تصدير شامل لكل الدفعات — سطر واحد لكل دفعة"""
        if not HAS_OPENPYXL:
            return messagebox.showerror("خطأ", "يرجى تثبيت openpyxl:\npip install openpyxl")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"تقرير_الدفعات_{datetime.now().strftime('%Y%m%d')}.xlsx")
        if not path: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "الدفعات الشاملة"
            ws.sheet_view.rightToLeft = True
            ws.freeze_panes = "F3"

            H  = PatternFill("solid", fgColor="1F4E79")
            CF = PatternFill("solid", fgColor="FCE4D6")
            RF = PatternFill("solid", fgColor="E2EFDA")
            PF = PatternFill("solid", fgColor="E2EFDA")
            LF = PatternFill("solid", fgColor="FCE4D6")
            BF = PatternFill("solid", fgColor="FFF2CC")
            GF = {
                "info": PatternFill("solid", fgColor="2E75B6"),
                "cost": PatternFill("solid", fgColor="C00000"),
                "rev":  PatternFill("solid", fgColor="375623"),
                "net":  PatternFill("solid", fgColor="7030A0"),
                "bird": PatternFill("solid", fgColor="833C00"),
            }
            ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
            brd = Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                         top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
            hf  = Font(bold=True, color="FFFFFF", size=9)
            gf  = Font(bold=True, color="FFFFFF", size=10)
            pf  = Font(bold=True, color="375623")
            lf  = Font(bold=True, color="C00000")

            cost_types = db.get_cost_types(active_only=True)
            rev_types  = db.get_revenue_types(active_only=True)
            def col_count(types): return sum(2 if t["has_qty"] else 1 for t in types)
            cc = col_count(cost_types); rc = col_count(rev_types)

            # صف 1 — مجموعات
            for grp, span, key in [
                ("معلومات الدفعة",5,"info"),("التكاليف التحليلية",cc,"cost"),
                ("إجمالي التكاليف",1,"cost"),("الإيرادات التحليلية",rc,"rev"),
                ("إجمالي الإيرادات",1,"rev"),("النتائج المالية",2,"net"),
                ("حركة الطيور",4,"bird")]:
                if span == 0: continue
                col = 1 + sum(s for _,s,_ in [
                    ("معلومات الدفعة",5,"info"),("التكاليف التحليلية",cc,"cost"),
                    ("إجمالي التكاليف",1,"cost"),("الإيرادات التحليلية",rc,"rev"),
                    ("إجمالي الإيرادات",1,"rev"),("النتائج المالية",2,"net"),
                    ("حركة الطيور",4,"bird")][:["معلومات الدفعة","التكاليف التحليلية",
                    "إجمالي التكاليف","الإيرادات التحليلية","إجمالي الإيرادات",
                    "النتائج المالية","حركة الطيور"].index(grp)])
                ws.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+span-1)
                c = ws.cell(1,col,grp); c.font=gf; c.fill=GF[key]; c.alignment=ctr

            # صف 2 — رؤوس
            hdrs=["العنبر","رقم الدفعة","تاريخ الدخول","تاريخ الخروج","الأيام"]
            wdths=[18,14,13,13,7]
            for ct in cost_types:
                if ct["has_qty"]: hdrs.append(f"{ct['name_ar']}\n({ct['unit'] or 'كمية'})"); wdths.append(11)
                hdrs.append(ct["name_ar"]); wdths.append(max(len(ct["name_ar"])+2,12))
            hdrs.append("إجمالي التكاليف"); wdths.append(16)
            for rt in rev_types:
                if rt["has_qty"]: hdrs.append(f"{rt['name_ar']}\n({rt['unit'] or 'كمية'})"); wdths.append(11)
                hdrs.append(rt["name_ar"]); wdths.append(max(len(rt["name_ar"])+2,12))
            hdrs.append("إجمالي الإيرادات"); wdths.append(17)
            hdrs+=["نتيجة الدفعة","نصيب الشركة","مباع العنبر","مباع السوق","الوفيات","الاستهلاك"]
            wdths+=[16,14,13,13,12,12]
            for ci,(h,w) in enumerate(zip(hdrs,wdths),1):
                c=ws.cell(2,ci,h); c.font=hf; c.fill=H; c.alignment=ctr
                ws.column_dimensions[get_column_letter(ci)].width=w
            ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=36

            # حدود الأعمدة
            cs=6; ce=cs+cc; rs=ce+1; re=rs+rc; nc=re+1

            batches=db.fetch_all("SELECT * FROM v_batches ORDER BY warehouse_name, date_in")
            for ri,b in enumerate(batches,3):
                bid=b["id"]; net=b["net_result"] or 0; share=b["share_val"] or 0
                fs=db.fetch_one("SELECT COALESCE(SUM(qty),0) AS q FROM farm_sales WHERE batch_id=?",(bid,))
                ms=db.fetch_one("SELECT COALESCE(SUM(qty_sold),0) AS q,COALESCE(SUM(deaths),0) AS d FROM market_sales WHERE batch_id=?",(bid,))
                fq=fs["q"] if fs else 0; mq=ms["q"] if ms else 0; md=ms["d"] if ms else 0
                td=(b["total_dead"] or 0)+md; co=b["consumed_birds"] or 0
                bc=db.get_batch_costs(bid); br=db.get_batch_revenues(bid)
                row=[b["warehouse_name"],b["batch_num"] or str(bid),b["date_in"],b["date_out"],b["days"] or 0]
                for ct in cost_types:
                    rec=bc.get(ct["code"],{}); amt=rec.get("amount",0) or (b[ct["code"]] if ct["code"] in b.keys() else 0) or 0
                    if ct["has_qty"]:
                        qty=rec.get("qty",0) or 0
                        if not qty:
                            qk=ct["code"].replace("_val","_qty").replace("_pay","_qty")
                            qty=b[qk] if qk in b.keys() else 0
                        row.append(qty or 0)
                    row.append(amt or 0)
                row.append(b["total_cost"] or 0)
                for rt in rev_types:
                    rec=br.get(rt["code"],{}); amt=rec.get("amount",0) or 0
                    if rt["has_qty"]: row.append(rec.get("qty",0) or 0)
                    row.append(amt or 0)
                row.append(b["total_rev"] or 0)
                row+=[net,share,fq,mq,td,co]
                for ci,v in enumerate(row,1):
                    c=ws.cell(ri,ci,v); c.alignment=ctr; c.border=brd
                    if ci>5: c.number_format="#,##0"
                for ci in range(cs,ce): ws.cell(ri,ci).fill=CF
                ws.cell(ri,ce).fill=PatternFill("solid",fgColor="F4CCCC")
                for ci in range(rs,re): ws.cell(ri,ci).fill=RF
                ws.cell(ri,re).fill=PatternFill("solid",fgColor="C6EFCE")
                ws.cell(ri,nc).fill=PF if net>=0 else LF
                ws.cell(ri,nc).font=pf if net>=0 else lf
                for ci in range(nc+2,len(row)+1): ws.cell(ri,ci).fill=BF

            # صف الإجماليات
            lr=len(batches)+3
            ws.cell(lr,1,"الإجمالي").font=Font(bold=True); ws.cell(lr,1).alignment=ctr
            for ci in [ce,re,nc,nc+1,nc+2,nc+3,nc+4,nc+5]:
                if ci<=len(hdrs):
                    cl=get_column_letter(ci)
                    ws.cell(lr,ci,f"=SUM({cl}3:{cl}{lr-1})")
                    ws.cell(lr,ci).number_format="#,##0"
                    ws.cell(lr,ci).font=Font(bold=True)
                    ws.cell(lr,ci).fill=PatternFill("solid",fgColor="FFFACD")
                    ws.cell(lr,ci).alignment=ctr

            wb.save(path)
            messagebox.showinfo("تم ✅", f"تم التصدير!\n{len(batches)} دفعة | {len(hdrs)} عمود\n{path}")
            try: os.startfile(path)
            except: pass
        except PermissionError:
            messagebox.showerror("خطأ", f"الملف مفتوح في Excel! أغلقه أولاً.")
        except Exception as ex:
            messagebox.showerror("خطأ", f"{type(ex).__name__}: {ex}")

    def _export_pdf_for_batch(self, batch_id, b):
        """تصدير PDF تصفية — يعمل مباشرة بدون الاعتماد على tree"""
        # نُضيف الدفعة للجدول مؤقتاً إن لم تكن موجودة ثم نستدعي الدالة
        existing = self.tree.exists(str(batch_id))
        if not existing:
            # أضف سطراً مؤقتاً حتى يعمل _selected_id
            b_row = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (batch_id,))
            if b_row:
                net = b_row["net_result"] or 0
                self.tree.insert("", "end", iid=str(batch_id),
                    tags=("profit" if net >= 0 else "loss",),
                    values=(b_row["batch_num"] or str(batch_id), b_row["warehouse_name"],
                            "", b_row["date_in"], b_row["date_out"], "", "", "", "", "", "", ""))
        self.tree.selection_set(str(batch_id))
        self._export_pdf()
        if not existing:
            try: self.tree.delete(str(batch_id))
            except: pass

    def _export_sales_pdf_for_batch(self, batch_id, b):
        self._export_sales_pdf(_batch_id=batch_id)

    def _open_wh_report(self): WarehousesReportWindow(self)
    def _open_analytics(self): AdvancedAnalyticsWindow(self)
    def _open_cost_manager(self): CostTypesManager(self)
    def _open_about(self): AboutWindow(self)
    def _open_settings(self): SystemSettingsWindow(self)
    def _open_onyx_importer(self): OnyxImporterWindow(self)
    def _open_poultry_v4_importer(self): open_import_wizard(self)
    def _backup(self): make_backup(); messagebox.showinfo("نسخ احتياطي", "تم حفظ النسخة الاحتياطية بنجاح")

    # ══ ميزة نقل الدفعة (Batch Portability) ══════════════════════
    def _export_batch_full_excel(self, batch_id):
        """تصدير كافة بيانات الدفعة إلى ملف إكسل واحد متعدد الصفحات"""
        if not HAS_OPENPYXL:
            return messagebox.showerror("خطأ", "يرجى تثبيت openpyxl", parent=self)
            
        b = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (batch_id,))
        if not b: return
        
        b_num = str(b['batch_num'] or b['id']).replace('/', '-')
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel Portability File", "*.xlsx")],
            initialfile=f"نقل_دفعة_{b_num}.xlsx", parent=self)
        if not path: return
        
        wb = openpyxl.Workbook()
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        font_white = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center")
        
        # 1. الورقة العامة (البيانات الأساسية)
        ws_gen = wb.active; ws_gen.title = "BatchInfo"; ws_gen.sheet_view.rightToLeft = True
        cols_b = [k for k in b.keys() if k not in ['id', 'warehouse_id', 'fy']]
        for ci, h in enumerate(cols_b, 1):
            cell = ws_gen.cell(1, ci, h); cell.fill = hdr_fill; cell.font = font_white
            ws_gen.cell(2, ci, b[h])
            
        # 2. السجلات اليومية
        ws_daily = wb.create_sheet("DailyRecords"); ws_daily.sheet_view.rightToLeft = True
        daily = db.fetch_all("SELECT rec_date, day_num, dead_count, feed_kg, water_ltr, notes FROM daily_records WHERE batch_id=? ORDER BY day_num", (batch_id,))
        if daily:
            hdrs = list(daily[0].keys())
            for ci, h in enumerate(hdrs, 1):
                cell = ws_daily.cell(1, ci, h); cell.fill = hdr_fill; cell.font = font_white
            for ri, row in enumerate(daily, 2):
                for ci, h in enumerate(hdrs, 1): ws_daily.cell(ri, ci, row[h])
                
        # 3. مبيعات العنبر
        ws_fs = wb.create_sheet("FarmSales"); ws_fs.sheet_view.rightToLeft = True
        fs = db.fetch_all("SELECT sale_date, sale_type, customer, qty, price, total_val FROM farm_sales WHERE batch_id=?", (batch_id,))
        if fs:
            hdrs = list(fs[0].keys())
            for ci, h in enumerate(hdrs, 1):
                cell = ws_fs.cell(1, ci, h); cell.fill = hdr_fill; cell.font = font_white
            for ri, row in enumerate(fs, 2):
                for ci, h in enumerate(hdrs, 1): ws_fs.cell(ri, ci, row[h])
                
        # 4. مبيعات السوق
        ws_ms = wb.create_sheet("MarketSales"); ws_ms.sheet_view.rightToLeft = True
        ms = db.fetch_all("SELECT sale_date, office, qty_sent, deaths, qty_sold, net_val, inv_num FROM market_sales WHERE batch_id=?", (batch_id,))
        if ms:
            hdrs = list(ms[0].keys())
            for ci, h in enumerate(hdrs, 1):
                cell = ws_ms.cell(1, ci, h); cell.fill = hdr_fill; cell.font = font_white
            for ri, row in enumerate(ms, 2):
                for ci, h in enumerate(hdrs, 1): ws_ms.cell(ri, ci, row[h])
                
        # 5. التكاليف الإضافية
        ws_c = wb.create_sheet("Costs"); ws_c.sheet_view.rightToLeft = True
        costs = db.fetch_all("SELECT ct.code, bc.qty, bc.amount, bc.notes FROM batch_costs bc JOIN cost_types ct ON bc.cost_type_id=ct.id WHERE bc.batch_id=?", (batch_id,))
        if costs:
            hdrs = ["code", "qty", "amount", "notes"]
            for ci, h in enumerate(hdrs, 1):
                cell = ws_c.cell(1, ci, h); cell.fill = hdr_fill; cell.font = font_white
            for ri, row in enumerate(costs, 2):
                for ci, h in enumerate(hdrs, 1): ws_c.cell(ri, ci, row[h])
                
        # 6. الإيرادات الإضافية
        ws_r = wb.create_sheet("Revenues"); ws_r.sheet_view.rightToLeft = True
        revs = db.fetch_all("SELECT rt.code, br.qty, br.amount, br.notes FROM batch_revenues br JOIN revenue_types rt ON br.revenue_type_id=rt.id WHERE br.batch_id=?", (batch_id,))
        if revs:
            hdrs = ["code", "qty", "amount", "notes"]
            for ci, h in enumerate(hdrs, 1):
                cell = ws_r.cell(1, ci, h); cell.fill = hdr_fill; cell.font = font_white
            for ri, row in enumerate(revs, 2):
                for ci, h in enumerate(hdrs, 1): ws_r.cell(ri, ci, row[h])
                
        for s in wb.worksheets:
            for col in s.columns: s.column_dimensions[col[0].column_letter].width = 15

        try:
            wb.save(path)
            messagebox.showinfo("تم التصدير", f"تم تصدير الدفعة بنجاح إلى:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل الحفظ: {e}", parent=self)

    def _import_batch_full_excel(self):
        """استيراد دفعة كاملة من ملف إكسل"""
        if not HAS_OPENPYXL:
            return messagebox.showerror("خطأ", "يرجى تثبيت openpyxl", parent=self)
            
        path = filedialog.askopenfilename(filetypes=[("Excel Portability File", "*.xlsx")], parent=self)
        if not path: return
        
        try:
            wb = openpyxl.load_workbook(path)
            if "BatchInfo" not in wb.sheetnames:
                return messagebox.showerror("خطأ", "الملف غير متوافق (صفحة BatchInfo مفقودة)", parent=self)
                
            # 1. قراءة البيانات الأساسية
            ws_gen = wb["BatchInfo"]; hdrs_gen = [c.value for c in ws_gen[1]]
            vals_gen = [c.value for c in ws_gen[2]]
            b_data = dict(zip(hdrs_gen, vals_gen))
            
            wh_name = b_data.pop('warehouse_name', 'مستودع مستورد')
            # البحث عن المستودع أو إنشاؤه
            wh = db.fetch_one("SELECT id FROM warehouses WHERE name=?", (wh_name,))
            if wh: wh_id = wh['id']
            else: wh_id = db.execute("INSERT INTO warehouses (name) VALUES (?)", (wh_name,))
            
            b_data['warehouse_id'] = wh_id
            b_data.pop('id', None); b_data.pop('fy', None)
            
            # معالجة تكرار رقم الدفعة
            orig_num = b_data.get('batch_num', '')
            if db.fetch_one("SELECT id FROM batches WHERE batch_num=? AND warehouse_id=?", (orig_num, wh_id)):
                b_data['batch_num'] = f"{orig_num}_مستورد_{datetime.now().strftime('%H%M')}"
            
            # إدراج الدفعة
            keys = list(b_data.keys()); placeholders = ",".join(["?"]*len(keys))
            new_bid = db.execute(f"INSERT INTO batches ({','.join(keys)}) VALUES ({placeholders})", tuple(b_data.values()))
            
            # 2. السجلات اليومية
            if "DailyRecords" in wb.sheetnames:
                ws = wb["DailyRecords"]; hdrs = [c.value for c in ws[1]]
                for ri in range(2, ws.max_row + 1):
                    row_vals = [c.value for c in ws[ri]]
                    if not any(row_vals): continue
                    d = dict(zip(hdrs, row_vals)); d['batch_id'] = new_bid
                    ks = list(d.keys()); ps = ",".join(["?"]*len(ks))
                    db.execute(f"INSERT OR IGNORE INTO daily_records ({','.join(ks)}) VALUES ({ps})", tuple(d.values()))
                    
            # 3. مبيعات العنبر
            if "FarmSales" in wb.sheetnames:
                ws = wb["FarmSales"]; hdrs = [c.value for c in ws[1]]
                for ri in range(2, ws.max_row + 1):
                    row_vals = [c.value for c in ws[ri]]
                    if not any(row_vals): continue
                    d = dict(zip(hdrs, row_vals)); d['batch_id'] = new_bid
                    ks = list(d.keys()); ps = ",".join(["?"]*len(ks))
                    db.execute(f"INSERT INTO farm_sales ({','.join(ks)}) VALUES ({ps})", tuple(d.values()))

            # 4. مبيعات السوق
            if "MarketSales" in wb.sheetnames:
                ws = wb["MarketSales"]; hdrs = [c.value for c in ws[1]]
                for ri in range(2, ws.max_row + 1):
                    row_vals = [c.value for c in ws[ri]]
                    if not any(row_vals): continue
                    d = dict(zip(hdrs, row_vals)); d['batch_id'] = new_bid
                    ks = list(d.keys()); ps = ",".join(["?"]*len(ks))
                    db.execute(f"INSERT INTO market_sales ({','.join(ks)}) VALUES ({ps})", tuple(d.values()))

            # 5. التكاليف الإضافية
            if "Costs" in wb.sheetnames:
                ws = wb["Costs"]; hdrs = [c.value for c in ws[1]]
                for ri in range(2, ws.max_row + 1):
                    row_vals = [c.value for c in ws[ri]]
                    if not any(row_vals): continue
                    d = dict(zip(hdrs, row_vals))
                    ct = db.fetch_one("SELECT id FROM cost_types WHERE code=?", (d['code'],))
                    if ct:
                        db.execute("INSERT OR REPLACE INTO batch_costs (batch_id, cost_type_id, qty, amount, notes) VALUES (?,?,?,?,?)",
                                   (new_bid, ct['id'], d.get('qty', 0), d.get('amount', 0), d.get('notes', '')))

            # 6. الإيرادات الإضافية
            if "Revenues" in wb.sheetnames:
                ws = wb["Revenues"]; hdrs = [c.value for c in ws[1]]
                for ri in range(2, ws.max_row + 1):
                    row_vals = [c.value for c in ws[ri]]
                    if not any(row_vals): continue
                    d = dict(zip(hdrs, row_vals))
                    rt = db.fetch_one("SELECT id FROM revenue_types WHERE code=?", (d['code'],))
                    if rt:
                        db.execute("INSERT OR REPLACE INTO batch_revenues (batch_id, revenue_type_id, qty, amount, notes) VALUES (?,?,?,?,?)",
                                   (new_bid, rt['id'], d.get('qty', 0), d.get('amount', 0), d.get('notes', '')))
            
            messagebox.showinfo("تم الاستيراد", "تم استيراد الدفعة بنجاح.\nيرجى تحديث القائمة لرؤيتها.", parent=self)
            if hasattr(self, '_load_batches'): self._load_batches()
            
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل الاستيراد: {e}", parent=self)

    def _export_pdf(self):
        if not HAS_FPDF:
            messagebox.showerror("خطأ", "مكتبة fpdf غير مثبتة.\nنفّذ: pip install fpdf2", parent=self)
            return
        batch_id = self._selected_id()
        if not batch_id: return
        batch_row = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (batch_id,))
        if not batch_row: return
        b = dict(batch_row)
        pdf = FPDF()
        pdf.add_page()
        font_path = AMIRI_FONT_PATH
        if not os.path.exists(font_path):
            messagebox.showerror("خطأ", "ملف الخط Amiri-Regular.ttf غير موجود!", parent=self)
            return
        pdf.add_font('Arabic', '', font_path, uni=True)
        logo_path = LOGO_PATH
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=10, y=5, w=30)
            
        curr = db.get_setting("currency", "ريال")
        company_name = db.get_setting("company_name", "شركة جيداء الوطنية للدواجن")
        farm_address = db.get_setting("farm_address", "المملكة العربية السعودية - جدة")
        contact_number = db.get_setting("contact_number", "")
        
        pdf.set_font('Arabic', '', 18)
        pdf.cell(0, 10, prepare_text(company_name), ln=True, align='C')
        pdf.set_font('Arabic', '', 11)
        if farm_address or contact_number:
            addr_str = f"{farm_address}" + (f" | هاتف: {contact_number}" if contact_number else "")
            pdf.cell(0, 6, prepare_text(addr_str), ln=True, align='C')
            
        pdf.ln(1)
        pdf.set_font('Arabic', '', 14)
        b_num = b.get('batch_num') or str(b.get('id', ''))
        title_text = f"تقرير التصفية المالية — دفعة رقم ({b_num}) — {b.get('warehouse_name', '')}"
        pdf.cell(0, 8, prepare_text(title_text), ln=True, align='C')
        pdf.set_font('Arabic', '', 11)
        info1 = f"تاريخ الدخول: {b.get('date_in', '')}  |  تاريخ الخروج: {b.get('date_out', '')}  |  عدد الأيام: {b.get('days', 0)}  |  الكتاكيت: {int(b.get('chicks') or 0):,}"
        pdf.cell(0, 7, prepare_text(info1), ln=True, align='C')
        info2 = f"النافق: {int(b.get('total_dead') or 0):,} ({(b.get('mort_rate') or 0):.2f}%)  |  المباع: {int(b.get('total_sold') or 0):,}  |  متوسط السعر: {fmt_num(b.get('avg_price'))} {curr}"
        pdf.cell(0, 7, prepare_text(info2), ln=True, align='C')
        pdf.ln(2)
        
        pdf.set_font('Arabic', '', 12)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(198, 8, prepare_text("مطابقة حركة الطيور (بالعدد)"), 1, 1, 'C', True)
        pdf.set_font('Arabic', '', 10)
        pdf.set_text_color(0, 0, 0)
        headers_birds = ["الفارق (نقص/زيادة)", "إجمالي المنصرف", "ضيافة/استهلاك", "وفيات السوق", "وفيات العنبر", "مباع السوق", "مباع العنبر", "الكتاكيت"]
        w_b = [24.75, 24.75, 24.75, 24.75, 24.75, 24.75, 24.75, 24.75]
        pdf.set_fill_color(240, 246, 255)
        for w, h in zip(w_b, headers_birds):
            pdf.cell(w, 8, prepare_text(h), 1, 0, 'C', True)
        pdf.ln()
        f_sales_records = db.fetch_all("SELECT * FROM farm_sales WHERE batch_id=?", (batch_id,))
        f_sold_ajl = 0; f_val_ajl = 0; f_sold_cash = 0; f_val_cash = 0
        for r_obj in f_sales_records:
            r = dict(r_obj)
            stype = r.get('sale_type') or 'آجل'
            if '(نقداً)' in str(r.get('customer', '')) and stype == 'آجل': stype = 'نقداً'
            if stype == 'نقداً':
                f_sold_cash += r.get('qty') or 0; f_val_cash += r.get('total_val') or 0
            else:
                f_sold_ajl += r.get('qty') or 0; f_val_ajl += r.get('total_val') or 0
        f_sold_total = f_sold_ajl + f_sold_cash
        m_sales = db.fetch_one("SELECT SUM(qty_sold) as qs, SUM(deaths) as md FROM market_sales WHERE batch_id=?", (batch_id,))
        m_sold = m_sales['qs'] if m_sales and m_sales['qs'] else 0
        m_dead = m_sales['md'] if m_sales and m_sales['md'] else 0
        chicks = b.get('chicks') or 0; f_dead = b.get('total_dead') or 0; consumed_birds = b.get('consumed_birds') or 0
        total_out = f_sold_total + m_sold + f_dead + m_dead + consumed_birds
        variance = chicks - total_out
        vals_birds = [fmt_num(variance), fmt_num(total_out), fmt_num(consumed_birds), fmt_num(m_dead), fmt_num(f_dead), fmt_num(m_sold), fmt_num(f_sold_total), fmt_num(chicks)]
        for w, v in zip(w_b, vals_birds):
            pdf.cell(w, 8, prepare_text(v), 1, 0, 'C')
        pdf.ln(5)
        pdf.set_font('Arabic', '', 12)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(198, 8, prepare_text("حركة العلف (بالكيس) - الطن = 20 كيس / الكيس = 50 كجم"), 1, 1, 'C', True)
        pdf.set_font('Arabic', '', 10)
        pdf.set_text_color(0, 0, 0)
        headers_feed = ["الفارق", "الرصيد المتبقي", "المباع / المنقول", "المستهلك (يومي)", "الوارد للعنبر"]
        w_f = [39.6, 39.6, 39.6, 39.6, 39.6]
        pdf.set_fill_color(240, 246, 255)
        for w, h in zip(w_f, headers_feed):
            pdf.cell(w, 8, prepare_text(h), 1, 0, 'C', True)
        pdf.ln()
        feed_in_tons = b.get('feed_qty') or 0
        feed_in_bags = feed_in_tons * 20
        daily = db.fetch_one("SELECT SUM(feed_kg) as fk FROM daily_records WHERE batch_id=?", (batch_id,))
        consumed_kg = daily['fk'] if daily and daily['fk'] else 0
        consumed_bags = consumed_kg / 50
        sold_trans_tons = (b.get('feed_sale_qty') or 0) + (b.get('feed_trans_r_qty') or 0)
        sold_bags = sold_trans_tons * 20
        remaining_bags = feed_in_bags - consumed_bags - sold_bags
        vals_feed = ["0.0", fmt_num(remaining_bags, 1), fmt_num(sold_bags, 1), fmt_num(consumed_bags, 1), fmt_num(feed_in_bags, 1)]
        for w, v in zip(w_f, vals_feed):
            pdf.cell(w, 8, prepare_text(v), 1, 0, 'C')
        pdf.ln(5)
        pdf.set_font('Arabic', '', 12)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255,255,255)
        pdf.cell(198, 8, prepare_text(f"الملخص المالي التفصيلي (التكاليف والإيرادات بـ {curr})"), 1, 1, 'C', True)
        pdf.set_text_color(0,0,0)
        costs = []
        if b.get('chicks'):    costs.append(("الكتاكيت",    b.get('chicks'),      b.get('chick_val')))
        if b.get('feed_val'):  costs.append(("العلف",        b.get('feed_qty'),    b.get('feed_val')))
        if b.get('sawdust_val'):costs.append(("النشارة",     b.get('sawdust_qty'), b.get('sawdust_val')))
        if b.get('gas_val'):   costs.append(("الغاز",        b.get('gas_qty'),     b.get('gas_val')))
        if b.get('water_val'): costs.append(("الماء",        "",                   b.get('water_val')))
        if b.get('drugs_val'): costs.append(("العلاجات",     "",                   b.get('drugs_val')))
        if b.get('wh_expenses'):costs.append(("مصاريف عنبر","",                   b.get('wh_expenses')))
        if b.get('house_exp'): costs.append(("مصاريف بيت",  "",                   b.get('house_exp')))
        if b.get('breeders_pay'):costs.append(("أجور مربيين","",                  b.get('breeders_pay')))
        if b.get('qat_pay'):   costs.append(("قات مربيين",   "",                   b.get('qat_pay')))
        if b.get('rent_val'):  costs.append(("إيجار عنبر",   "",                   b.get('rent_val')))
        if b.get('light_val'): costs.append(("إضاءة",        "",                   b.get('light_val')))
        sup_tot = (b.get('sup_wh_pay') or 0) + (b.get('sup_co_pay') or 0) + (b.get('sup_sale_pay') or 0)
        if sup_tot: costs.append(("إشراف وتسويق", "", sup_tot))
        if b.get('admin_val'): costs.append(("إدارة وحسابات","",                  b.get('admin_val')))
        if b.get('vaccine_pay'):costs.append(("لقاحات",       "",                  b.get('vaccine_pay')))
        oth_tot = (b.get('delivery_val') or 0) + (b.get('mixing_val') or 0) + (b.get('wash_val') or 0) + (b.get('other_costs') or 0)
        if oth_tot: costs.append(("مصاريف أخرى متنوعة", "", oth_tot))
        revs = []
        if f_sold_ajl > 0: revs.append(("مبيعات العنبر (آجل)", f_sold_ajl, f_val_ajl))
        if f_sold_cash > 0: revs.append(("مبيعات العنبر (نقدي)", f_sold_cash, f_val_cash))
        if m_sold > 0: revs.append(("مبيعات السوق",  m_sold, (b.get('mkt_val') or 0)))
        if b.get('offal_val'): revs.append(("مبيعات ذبيل", "", b.get('offal_val')))
        fs_qty = b.get('feed_sale_qty') or 0; fs_val = b.get('feed_sale') or 0
        if fs_qty or fs_val: revs.append(("مبيعات علف", fs_qty if fs_qty else "", fs_val))
        ft_qty = b.get('feed_trans_r_qty') or 0; ft_val = b.get('feed_trans_r') or 0
        if ft_qty or ft_val: revs.append(("علف منقول / متبقي", ft_qty if ft_qty else "", ft_val))
        if b.get('drug_return'): revs.append(("مرتجع علاجات", "", b.get('drug_return')))
        if b.get('gas_return'): revs.append(("نقل غاز/نشارة", "", b.get('gas_return')))
        max_len = max(len(costs), len(revs), 1)
        costs += [("", "", "")] * (max_len - len(costs))
        revs  += [("", "", "")] * (max_len - len(revs))
        pdf.set_font('Arabic', '', 10)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255,255,255)
        pdf.cell(27, 8, prepare_text("القيمة"), 1, 0, 'C', True)
        pdf.cell(22, 8, prepare_text("الكمية"), 1, 0, 'C', True)
        pdf.cell(50, 8, prepare_text("الإيرادات"), 1, 0, 'C', True)
        pdf.cell(27, 8, prepare_text("القيمة"), 1, 0, 'C', True)
        pdf.cell(22, 8, prepare_text("الكمية"), 1, 0, 'C', True)
        pdf.cell(50, 8, prepare_text("التكاليف"), 1, 1, 'C', True)
        pdf.set_text_color(0,0,0)
        for i in range(max_len):
            r_name, r_qty, r_val = revs[i]; c_name, c_qty, c_val = costs[i]
            fill = (i % 2 == 0)
            if fill: pdf.set_fill_color(240,246,255)
            else: pdf.set_fill_color(255,255,255)
            pdf.cell(27, 7, prepare_text(fmt_num(r_val) if r_val else ""), 1, 0, 'C', fill)
            pdf.cell(22, 7, prepare_text(fmt_num(r_qty) if r_qty else ""), 1, 0, 'C', fill)
            pdf.cell(50, 7, prepare_text(r_name), 1, 0, 'R', fill)
            pdf.cell(27, 7, prepare_text(fmt_num(c_val) if c_val else ""), 1, 0, 'C', fill)
            pdf.cell(22, 7, prepare_text(fmt_num(c_qty) if c_qty else ""), 1, 0, 'C', fill)
            pdf.cell(50, 7, prepare_text(c_name), 1, 1, 'R', fill)
        pdf.ln(2)
        def draw_row(label, value, bg=(240,240,240), fg=(0,0,0)):
            pdf.set_fill_color(*bg); pdf.set_text_color(*fg)
            pdf.cell(27, 8, prepare_text(fmt_num(value)), 1, 0, 'C', True)
            pdf.cell(171, 8, prepare_text(label), 1, 1, 'R', True)
            pdf.set_text_color(0,0,0)
        # إجمالي التكاليف تحت عمود التكاليف وإجمالي الإيرادات تحت عمود الإيرادات — سطر واحد
        # هيكل الجدول: [27قيمة_إير][22كمية_إير][50اسم_إير][27قيمة_تكل][22كمية_تكل][50اسم_تكل] = 198
        pdf.set_font('Arabic', '', 10)
        pdf.set_fill_color(226,239,218); pdf.set_text_color(39,104,10)
        pdf.cell(27, 8, prepare_text(fmt_num(b.get("total_rev") or 0)), 1, 0, 'C', True)
        pdf.cell(72, 8, prepare_text("إجمالي الإيرادات"), 1, 0, 'R', True)
        pdf.set_fill_color(252,228,214); pdf.set_text_color(192,0,0)
        pdf.cell(27, 8, prepare_text(fmt_num(b.get("total_cost") or 0)), 1, 0, 'C', True)
        pdf.cell(72, 8, prepare_text("إجمالي التكاليف"), 1, 1, 'R', True)
        pdf.set_text_color(0,0,0)
        net = b.get("net_result") or 0
        net_bg = (226,239,218) if net >= 0 else (252,228,214)
        net_fg = (39,104,10) if net >= 0 else (192,0,0)
        w_status = "ربح" if net >= 0 else "خسارة"
        # التفقيط (التحويل لمبلغ نصي)
        words = num_to_words_ar(net)
        label = f"نتيجة الدفعة ({w_status}) — فقط: {words} {curr}"
        draw_row(label, abs(net), net_bg, net_fg)
        # نصيب الشركة ونصيب الشريك في سطر واحد مقسّم بالتساوي
        pct = b.get('share_pct') or 65
        share_v = b.get("share_val") or 0
        partner_pct = 100 - int(pct)
        partner_v = net - share_v
        p_name = b.get("partner_name", "")
        p_str = f" ({p_name})" if p_name else ""
        pdf.set_font('Arabic', '', 10)
        pdf.set_fill_color(255,242,204); pdf.set_text_color(191,144,0)
        pdf.cell(27, 8, prepare_text(fmt_num(share_v)),   1, 0, 'C', True)
        pdf.cell(72, 8, prepare_text(f"نصيب الشركة ({int(pct)}%)"), 1, 0, 'R', True)
        pdf.cell(27, 8, prepare_text(fmt_num(partner_v)), 1, 0, 'C', True)
        pdf.cell(72, 8, prepare_text(f"نصيب الشريك{p_str} ({partner_pct}%)"), 1, 1, 'R', True)
        pdf.set_text_color(0,0,0)
        raw_notes = b.get("notes", "")
        if raw_notes:
            pdf.ln(2)
            pdf.set_font('Arabic', '', 10)
            pdf.set_text_color(50, 50, 50)
            pdf.cell(0, 6, prepare_text("ملاحظات الدفعة الإضافية:"), ln=True, align='R')
            pdf.set_font('Arabic', '', 9)
            pdf.multi_cell(0, 5, prepare_text(raw_notes), align='R')
        # ── التواقيع: بدون جدول، خط صغير، مباشرة في نفس الصفحة ──
        pdf.ln(3)
        mgr_name = db.get_setting("manager_name", "")
        fin_name = db.get_setting("finance_name", "")
        aud_name = db.get_setting("auditor_name", "")
        pdf.set_font('Arabic', '', 8)
        pdf.set_text_color(60, 60, 60)
        pdf.cell(47.5, 5, prepare_text("المدير العام"), 0, 0, 'C')
        pdf.cell(47.5, 5, prepare_text("المدير المالي"), 0, 0, 'C')
        pdf.cell(47.5, 5, prepare_text("المراجع"), 0, 0, 'C')
        pdf.cell(47.5, 5, prepare_text("المحاسب"), 0, 1, 'C')
        pdf.set_font('Arabic', '', 7)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(47.5, 4, prepare_text(mgr_name), 0, 0, 'C')
        pdf.cell(47.5, 4, prepare_text(fin_name), 0, 0, 'C')
        pdf.cell(47.5, 4, prepare_text(aud_name), 0, 0, 'C')
        pdf.cell(47.5, 4, prepare_text(""), 0, 1, 'C')
        pdf.set_text_color(0, 0, 0)
        pdf.ln(1)
        # خط فاصل رفيع تحت الأسماء
        pdf.set_draw_color(180, 180, 180)
        x = pdf.get_x(); y = pdf.get_y()
        for i in range(4):
            pdf.line(10 + i*47.5 + 4, y, 10 + i*47.5 + 43.5, y)
        pdf.set_draw_color(0, 0, 0)
        # ── تذييل المطوّر ──
        pdf.set_y(-10)
        pdf.set_font("Arabic", "", 7)
        pdf.set_text_color(160, 160, 160)
        pdf.cell(0, 5, "Developed & Engineered by: Eng. Saleem Homi  |  Tel: 770199865", 0, 0, 'C')
        pdf.set_text_color(0, 0, 0)
        save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")], initialfile=f"تصفية_{b.get('warehouse_name', '')}_دفعة_{b_num}.pdf", title="حفظ تقرير PDF", parent=self)
        if save_path:
            pdf.output(save_path)
            messagebox.showinfo("تم", "تم تصدير تقرير التصفية بنجاح!", parent=self)
            try: os.startfile(save_path)
            except: pass


    def _export_sales_pdf(self, _batch_id=None):
        if not HAS_FPDF:
            messagebox.showerror("خطأ", "مكتبة fpdf غير مثبتة.\nنفّذ: pip install fpdf2", parent=self)
            return
        if _batch_id:
            batch_id = _batch_id
        else:
            batch_id = self._selected_id()
            if not batch_id: return
        batch_row = db.fetch_one("SELECT * FROM v_batches WHERE id=?", (batch_id,))
        if not batch_row: return
        b = dict(batch_row)
        font_path = AMIRI_FONT_PATH
        if not os.path.exists(font_path):
            messagebox.showerror("خطأ", "ملف الخط Amiri-Regular.ttf غير موجود!", parent=self)
            return
        
        # إعدادات الترويسة والعملة
        curr = db.get_setting("currency", "ريال")
        company_name = db.get_setting("company_name", "شركة آفاق الريف للدواجن")
        farm_address = db.get_setting("farm_address", "")
        contact_number = db.get_setting("contact_number", "")
        
        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_margins(left=10, top=5, right=10)
        pdf.add_page()
        pdf.add_font("Arabic", "", font_path, uni=True)
        
        b_num = b.get("batch_num") or str(b.get("id",""))
        wh    = b.get("warehouse_name","")
        
        # ── الترويسة ──
        pdf.set_font("Arabic","",16)
        pdf.cell(0,10,prepare_text(f"{company_name} - تقرير المبيعات التحليلي الشامل"),ln=True,align="C")
        pdf.set_font("Arabic","",11)
        if farm_address or contact_number:
            addr_str = f"{farm_address}" + (f" | هاتف: {contact_number}" if contact_number else "")
            pdf.cell(0, 6, prepare_text(addr_str), ln=True, align='C')
        pdf.ln(1)
        pdf.set_font("Arabic","",12)
        pdf.cell(0,7,prepare_text(f"العنبر: {wh}  |  الدفعة: {b_num}  |  من: {b.get('date_in','')} إلى: {b.get('date_out','')}"),ln=True,align="C")
        pdf.ln(2)

        # دوال مساعدة للرسم
        def draw_hdr(txt, r, g, bl):
            pdf.set_fill_color(r, g, bl); pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arabic","",12)
            pdf.cell(0, 8, prepare_text(txt), 1, 1, "C", True)
            pdf.set_text_color(0, 0, 0)

        def th(pdf, cols, widths, fill=(220, 230, 245)):
            pdf.set_fill_color(*fill); pdf.set_font("Arabic","",9)
            for w, h in zip(widths, cols):
                pdf.cell(w, 7, prepare_text(h), 1, 0, "C", True)
            pdf.ln()

        def td(pdf, vals, widths, fill=False, fc=(245, 248, 255)):
            pdf.set_font("Arabic","",9)
            if fill: pdf.set_fill_color(*fc)
            for w, v in zip(widths, vals):
                pdf.cell(w, 6, prepare_text(str(v)), 1, 0, "C", fill)
            pdf.ln()

        PW = 277  # عرض الصفحة الصافي

        # ── جلب وتصنيف مبيعات العنبر ──
        all_farm = db.fetch_all("SELECT * FROM farm_sales WHERE batch_id=? ORDER BY sale_date", (batch_id,))
        farm_cash = []; farm_ajl = []
        for s_row in all_farm:
            s_dict = dict(s_row)
            stype = s_dict.get('sale_type') or 'آجل'
            # توحيد المنطق: (نقداً) في اسم العميل يحولها لكاش
            if '(نقداً)' in str(s_dict.get('customer', '')) and stype == 'آجل':
                stype = 'نقداً'
            
            if stype == 'نقداً': farm_cash.append(s_dict)
            else: farm_ajl.append(s_dict)

        # ── أولاً: مبيعات العنبر (آجل) ──
        draw_hdr("أولاً: مبيعات العنبر — آجل", 39, 104, 10)
        ww = [35, 28, 84, 35, 40, 55]
        if farm_ajl:
            th(pdf, ["تاريخ البيع", "النوع", "اسم العميل", "الكمية", "السعر", "الإجمالي"], ww)
            tot_q = 0; tot_v = 0
            for i, s in enumerate(farm_ajl):
                td(pdf, [s.get("sale_date", ""), "آجل", s["customer"] or "",
                         f"{s['qty']:,}", f"{s['price']:,.2f}", f"{s['total_val']:,.0f}"], ww, (i % 2 == 0))
                tot_q += s["qty"]; tot_v += s["total_val"]
            pdf.set_fill_color(226, 239, 218); pdf.set_font("Arabic","",10)
            pdf.cell(PW - 55, 7, prepare_text(f"مجموع الآجل: {tot_q:,} طائر"), 1, 0, "R", True)
            pdf.cell(55, 7, prepare_text(f"{tot_v:,.0f} {curr}"), 1, 1, "C", True)
        else:
            pdf.set_font("Arabic","", 9); pdf.cell(PW, 7, prepare_text("لا توجد مبيعات آجلة"), 1, 1, "C")
        pdf.ln(3)

        # ── ثانياً: مبيعات العنبر (نقداً) ──
        draw_hdr("ثانياً: مبيعات العنبر — نقداً", 31, 78, 121)
        if farm_cash:
            th(pdf, ["تاريخ البيع", "النوع", "اسم العميل", "الكمية", "السعر", "الإجمالي"], ww)
            tot_q = 0; tot_v = 0
            for i, s in enumerate(farm_cash):
                td(pdf, [s.get("sale_date", ""), "نقداً", s["customer"] or "",
                         f"{s['qty']:,}", f"{s['price']:,.2f}", f"{s['total_val']:,.0f}"], ww, (i % 2 == 0))
                tot_q += s["qty"]; tot_v += s["total_val"]
            pdf.set_fill_color(226, 239, 218); pdf.set_font("Arabic","",10)
            pdf.cell(PW - 55, 7, prepare_text(f"مجموع النقدي: {tot_q:,} طائر"), 1, 0, "R", True)
            pdf.cell(55, 7, prepare_text(f"{tot_v:,.0f} {curr}"), 1, 1, "C", True)
        else:
            pdf.set_font("Arabic","", 9); pdf.cell(PW, 7, prepare_text("لا توجد مبيعات نقدية"), 1, 1, "C")
        pdf.ln(3)

        # ── ثالثاً: مبيعات السوق (المكاتب) ──
        draw_hdr("ثالثاً: مبيعات السوق (المكاتب) — تحليلي", 191, 144, 0)
        mkt = db.fetch_all("SELECT * FROM market_sales WHERE batch_id=? ORDER BY sale_date", (batch_id,))
        if mkt:
            ww_mkt = [35, 82, 30, 28, 30, 42, 30]
            th(pdf, ["تاريخ البيع", "مكتب التسويق", "المرسل", "الوفيات", "المباع", "صافي الفاتورة", "رقم الفاتورة"], ww_mkt)
            tot_sent = 0; tot_dead = 0; tot_sold = 0; tot_val = 0
            for i, s in enumerate(mkt):
                td(pdf, [s.get("sale_date", ""), s["office"] or "",
                         f"{s['qty_sent']:,}", f"{s['deaths']:,}", f"{s['qty_sold']:,}",
                         f"{s['net_val']:,.0f}", s.get("inv_num", "")], ww_mkt, (i % 2 == 0))
                tot_sent += s["qty_sent"]; tot_dead += s["deaths"]; tot_sold += s["qty_sold"]; tot_val += s["net_val"]
            pdf.set_fill_color(255, 242, 204); pdf.set_font("Arabic","",10)
            pdf.cell(35 + 82, 7, prepare_text("الإجمالي"), 1, 0, "R", True)
            pdf.cell(30, 7, prepare_text(f"{tot_sent:,}"), 1, 0, "C", True)
            pdf.cell(28, 7, prepare_text(f"{tot_dead:,}"), 1, 0, "C", True)
            pdf.cell(30, 7, prepare_text(f"{tot_sold:,}"), 1, 0, "C", True)
            pdf.cell(42, 7, prepare_text(f"{tot_val:,.0f}"), 1, 0, "C", True)
            pdf.cell(30, 7, prepare_text(""), 1, 1, "C", True)
        else:
            pdf.set_font("Arabic","", 9); pdf.cell(PW, 7, prepare_text("لا توجد مبيعات سوق"), 1, 1, "C")
        pdf.ln(3)

        # ── رابعاً: إيرادات أخرى (علف، ذبيل، مرتجعات) ──
        draw_hdr("رابعاً: إيرادات متنوعة (علف، ذبيل، ومرتجعات)", 83, 141, 213)
        other_revs = db.get_batch_revenues(batch_id)
        other_rev_list = []
        for code, r in other_revs.items():
            if r["amount"] > 0 or r["qty"] > 0:
                other_rev_list.append(r)
        
        if other_rev_list:
            ww_oth = [130, 47, 40, 60]
            th(pdf, ["بند الإيراد", "الكمية", "الوحدة", "القيمة الإجمالية"], ww_oth, (230, 240, 255))
            tot_oth_v = 0
            for i, r in enumerate(other_rev_list):
                td(pdf, [r["name_ar"], fmt_num(r["qty"], 1), r["unit"] or "—", f"{r['amount']:,.0f} {curr}"], ww_oth, (i % 2 == 0))
                tot_oth_v += r["amount"]
            pdf.set_fill_color(226, 239, 218); pdf.set_font("Arabic","",10)
            pdf.cell(PW - 60, 7, prepare_text(f"مجموع الإيرادات الأخرى:"), 1, 0, "R", True)
            pdf.cell(60, 7, prepare_text(f"{tot_oth_v:,.0f} {curr}"), 1, 1, "C", True)
        else:
            pdf.set_font("Arabic","", 9); pdf.cell(PW, 7, prepare_text("لا توجد إيرادات أخرى مسجلة"), 1, 1, "C")
        pdf.ln(4)

        # ── الملخص النهائي للمبيعات ──
        all_farm_q = sum(s["qty"] for s in farm_ajl + farm_cash)
        all_farm_v = sum(s["total_val"] for s in farm_ajl + farm_cash)
        mkt_sold_q = sum(s["qty_sold"] for s in (mkt or []))
        mkt_val    = sum(s["net_val"] for s in (mkt or []))
        total_birds = all_farm_q + mkt_sold_q
        grand_total_rev = all_farm_v + mkt_val + sum(r["amount"] for r in other_rev_list)

        pdf.set_fill_color(31, 78, 121); pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arabic","", 11)
        pdf.cell(0, 8, prepare_text("الملخص الإجمالي لكافة إيرادات الدفعة"), 1, 1, "C", True)
        pdf.set_text_color(0, 0, 0); pdf.set_font("Arabic","", 10)
        
        summary_rows = [
            ("إجمالي مبيعات العنبر (آجل + نقد)", f"{all_farm_q:,} طائر", f"{all_farm_v:,.0f} {curr}"),
            ("إجمالي مبيعات السوق (مكاتب)", f"{mkt_sold_q:,} طائر", f"{mkt_val:,.0f} {curr}"),
            ("إجمالي الإيرادات الأخرى (علف/ذبيل/مرتجعات)", "—", f"{sum(r['amount'] for r in other_rev_list):,.0f} {curr}"),
            ("الإجمالي العام للإيرادات", f"{total_birds:,} طائر", f"{grand_total_rev:,.0f} {curr}")
        ]
        
        for i, (lbl, qty, val) in enumerate(summary_rows):
            is_last = (i == len(summary_rows)-1)
            pdf.set_font("Arabic", "", 11 if is_last else 10)
            if is_last:
                pdf.set_fill_color(31, 78, 121); pdf.set_text_color(255, 255, 255)
            elif i % 2 == 0:
                pdf.set_fill_color(240, 246, 255); pdf.set_text_color(0, 0, 0)
            else:
                pdf.set_fill_color(226, 239, 218); pdf.set_text_color(0, 0, 0)
                
            # PW=277: label(130) qty(77) val(70)
            pdf.cell(130, 8, prepare_text(lbl), 1, 0, "R", True)
            pdf.cell(77, 8, prepare_text(qty), 1, 0, "C", True)
            pdf.cell(70, 8, prepare_text(val), 1, 1, "C", True)
        
        pdf.set_text_color(0, 0, 0)
        # ── تذييل المطوّر ──
        pdf.set_y(-10)
        pdf.set_font("Arabic", "", 7)
        pdf.set_text_color(160, 160, 160)
        pdf.cell(0, 5, "Developed & Engineered by: Eng. Saleem Homi  |  Tel: 770199865", 0, 0, 'C')
        pdf.set_text_color(0, 0, 0)

        save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")],
            initialfile=f"تقرير_مبيعات_{wh}_دفعة_{b_num.replace('/', '-') if b_num else 'unknown'}.pdf", title="حفظ تقرير المبيعات", parent=self)
        if save_path:
            pdf.output(save_path)
            messagebox.showinfo("تم", "تم تصدير تقرير المبيعات الشامل بنجاح!", parent=self)
            try: os.startfile(save_path)
            except: pass


if __name__ == "__main__":
    check_and_run()
