import openpyxl
import os
import sqlite3
import unicodedata
from datetime import datetime, date

class BatchImporter:
    """
    يستورد ملفات Excel الخاصة بدفعة عنبر دواجن مع دعم الربط الذكي والعمليات الحسابية التلقائية.
    """
    def __init__(self, db_manager):
        self.db = db_manager
        self.errors = []
        self.daily_rows = []
        self.farm_sales = []
        self.market_sales = []
        self.result = {}
        self.wh_name = ""
        self.filename = ""

    def _extract_wh_name(self, filename):
        """يستخرج اسم العنبر من اسم الملف بشكل ذكي"""
        clean = ''.join(c for c in filename if unicodedata.category(c) not in ('Cf',))
        clean = clean.strip()
        for sep in ['دفعة', 'دورة', 'batch', 'Batch']:
            if sep in clean:
                part = clean.split(sep)[0].strip()
                if part: return part
        return clean[:40]

    def _find_sheet_by_keywords(self, wb, keywords):
        best, best_score = None, 0
        for name in wb.sheetnames:
            score = sum(1 for kw in keywords if kw in name)
            if score > best_score:
                best, best_score = name, score
        return wb[best] if best else None

    def _sf(self, v):
        try:
            return float(str(v).replace(',', '').replace(' ', '')) if v not in (None, '', '#DIV/0!') else 0.0
        except: return 0.0

    def _si(self, v):
        try:
            return int(float(str(v).replace(',', '').replace(' ', ''))) if v not in (None, '', '#DIV/0!') else 0
        except: return 0

    def _is_date_row(self, row):
        for cell in row[:3]:
            if cell is None: continue
            if hasattr(cell, "year") and hasattr(cell, "month"): return True
            s = str(cell).strip()[:10]
            try:
                datetime.strptime(s, "%Y-%m-%d")
                return True
            except: pass
        return False

    def _parse_daily(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        hdr = -1
        for i, row in enumerate(rows):
            flat = " ".join(str(c) for c in row if c is not None)
            if "التاريخ" in flat and ("الوفيات" in flat or "النافق" in flat):
                for j in range(i+1, min(i+4, len(rows))):
                    if self._is_date_row(rows[j]):
                        hdr = i; break
                if hdr >= 0: break
        
        if hdr < 0: return

        col_date, col_age, col_dead, col_feed_daily = 0, 1, 3, 5
        hdr_row = rows[hdr]
        for ci, cell in enumerate(hdr_row):
            cv = str(cell or "").strip()
            if cv == "التاريخ" or (cv.startswith("تاريخ") and "عمر" not in cv): col_date = ci
            elif "العمر" in cv: col_age = ci
            elif "الوفيات" in cv or "النافق" in cv: col_dead = ci
            elif "مستهلك" in cv: col_feed_daily = ci

        daily = []
        for row in rows[hdr+1:]:
            if not row or len(row) <= col_date or not self._is_date_row(row): continue
            date_val = row[col_date]
            if date_val is None: continue
            
            if hasattr(date_val, "strftime"):
                rec_date = date_val.strftime("%Y-%m-%d")
            else:
                parsed = None
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        parsed = datetime.strptime(str(date_val).strip()[:10], fmt)
                        break
                    except: pass
                if not parsed: continue
                rec_date = parsed.strftime("%Y-%m-%d")

            day_num = self._si(row[col_age]) if len(row) > col_age else 0
            dead    = self._si(row[col_dead]) if len(row) > col_dead else 0
            feed_kg = self._sf(row[col_feed_daily]) if len(row) > col_feed_daily else 0
            daily.append({"rec_date": rec_date, "day_num": day_num, "dead_count": dead, "feed_kg": feed_kg})
        
        self.daily_rows = daily

    def _parse_sales(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        data_start = 0
        for i, row in enumerate(rows[:5]):
            flat = " ".join(str(c) for c in row if c is not None)
            if "اسم العميل" in flat or ("العدد" in flat and "السعر" in flat):
                data_start = i + 1; break
        
        SKIP_WORDS = {"الاجمالي", "إجمالي", "اجمالي", "المجموع", "البيان", "بيان", "None", "", "0", "اسم العميل"}
        for row in rows[data_start:]:
            if not row or all(c is None for c in row): continue
            ncols = len(row)
            cust = str(row[0] or "").strip()
            if cust and cust not in SKIP_WORDS and "#" not in cust:
                qty_ajl, price_ajl, total_ajl = self._si(row[1]), self._sf(row[2]), self._sf(row[3])
                if qty_ajl > 0:
                    self.farm_sales.append({"customer": cust, "qty": qty_ajl, "price": price_ajl or (total_ajl/qty_ajl if total_ajl else 0), "total_val": total_ajl or qty_ajl*price_ajl})
                
                if ncols > 4:
                    qty_nqd, price_nqd, total_nqd = self._si(row[4]), self._sf(row[5]), self._sf(row[6])
                    if qty_nqd > 0:
                        self.farm_sales.append({"customer": cust + " (نقداً)", "qty": qty_nqd, "price": price_nqd or (total_nqd/qty_nqd if total_nqd else 0), "total_val": total_nqd or qty_nqd*price_nqd})
            
            if ncols > 7:
                office = str(row[7] or "").strip()
                if office and office not in SKIP_WORDS:
                    ms_qty, ms_dead, ms_sold, ms_net = self._si(row[8]), self._si(row[9]), self._si(row[10]), self._sf(row[11])
                    if ms_qty > 0 or ms_net > 0:
                        self.market_sales.append({"office": office, "qty_sent": ms_qty, "deaths": ms_dead, "qty_sold": ms_sold or max(0, ms_qty-ms_dead), "net_val": ms_net, "inv_num": str(row[12] or "").strip()})

    def _parse_summary(self, ws):
        mapping = [
            (['الكتاكيت', 'عدد الكتاكيت'], 'chicks'),
            (['قيمة الكتاكيت', 'تكلفة الكتاكيت'], 'chick_val'),
            (['قيمة العلف', 'تكلفة العلف'], 'feed_val'),
            (['نقل علف', 'أجور نقل'], 'feed_trans'),
            (['علف—كمية', 'علف كمية'], 'feed_qty'),
            (['نشارة—قيمة', 'قيمة النشارة'], 'sawdust_val'),
            (['نشارة—كمية', 'كمية نشارة'], 'sawdust_qty'),
            (['غاز—قيمة', 'قيمة الغاز'], 'gas_val'),
            (['غاز—كمية', 'كمية غاز'], 'gas_qty'),
            (['مياه', 'مياة'], 'water_val'),
            (['كهرباء', 'إضاءة'], 'light_val'),
            (['العلاجات', 'ادوية', 'علاج'], 'drugs_val'),
            (['رواتب', 'أجور مربيين'], 'breeders_pay'),
            (['صيانة مباني', 'مصاريف العنبر'], 'wh_expenses'),
            (['مصاريف بيت', 'مصاريف المنزل'], 'house_exp'),
            (['قات مربيين'], 'qat_pay'),
            (['إيجار', 'ايجار'], 'rent_val'),
            (['مشرف عنبر'], 'sup_wh_pay'),
            (['مشرف شركة'], 'sup_co_pay'),
            (['مشرف بيع'], 'sup_sale_pay'),
            (['إدارة وحسابات', 'قرطاسية'], 'admin_val'),
            (['لقاحات', 'تطعيم'], 'vaccine_pay'),
            (['توصيل خدمات'], 'delivery_val'),
            (['حمالة وخلط'], 'mixing_val'),
            (['تغسيل عنبر', 'نظافة'], 'wash_val'),
            (['محروقات', 'مصاريف أخرى'], 'other_costs'),
            (['جمالي المصاريف', 'إجمالي المصاريف'], 'total_cost'),
            (['جمالي الايردات', 'إجمالي الإيرادات'], 'total_rev'),
            (['نتيجة الدفعة', 'صافي الربح'], 'net_result'),
            (['ذبيل', 'إيرادات ذبيل'], 'offal_val'),
        ]
        d = {}
        for row in ws.iter_rows(values_only=True):
            for ci, cell_label in enumerate(row):
                if not cell_label: continue
                label_str = str(cell_label).strip()
                val_candidate = None
                for cell_val in row[ci+1:]:
                    if cell_val is not None and str(cell_val) not in ('', '#DIV/0!'):
                        try:
                            val_candidate = float(str(cell_val).replace(',', '').replace(' ', ''))
                            break
                        except: pass
                for keywords, db_col in mapping:
                    if any(kw in label_str for kw in keywords):
                        if val_candidate is not None and db_col not in d: d[db_col] = val_candidate
        self.result = d

    def import_file(self, file_path):
        try:
            self.filename = os.path.splitext(os.path.basename(file_path))[0].strip()
            self.wh_name = self._extract_wh_name(self.filename)
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            ws_daily = self._find_sheet_by_keywords(wb, ["كرت", "يومي", "نافق"])
            if ws_daily: self._parse_daily(ws_daily)
            
            ws_sales = self._find_sheet_by_keywords(wb, ["مبيعات", "بيع", "بيان"])
            if ws_sales: self._parse_sales(ws_sales)
            
            ws_summ = self._find_sheet_by_keywords(wb, ["تصفية", "اجمالي", "ملخص"])
            if ws_summ: self._parse_summary(ws_summ)

            # Calculations
            chicks = self._si(self.result.get('chicks', 0))
            if chicks == 0 and self.daily_rows:
                # fallback: check first row for chicks
                pass 
            
            total_dead = sum(r['dead_count'] for r in self.daily_rows)
            mort_rate = round(total_dead / chicks * 100, 2) if chicks > 0 else 0
            
            total_cost = self.result.get('total_cost', 0)
            total_rev = self.result.get('total_rev', 0)
            if total_rev == 0:
                total_rev = sum(s['total_val'] for s in self.farm_sales) + sum(s['net_val'] for s in self.market_sales)
            
            net_result = self.result.get('net_result', total_rev - total_cost)
            total_sold = sum(s['qty'] for s in self.farm_sales) + sum(s['qty_sold'] for s in self.market_sales)
            
            # Save to DB
            wh = self.db.fetch_one("SELECT id FROM warehouses WHERE name=?", (self.wh_name,))
            if not wh:
                self.db.execute("INSERT INTO warehouses (name) VALUES (?)", (self.wh_name,))
                wh = self.db.fetch_one("SELECT id FROM warehouses WHERE name=?", (self.wh_name,))
            wh_id = wh['id']

            # Clean & Sync
            date_in = self.daily_rows[0]['rec_date'] if self.daily_rows else date.today().isoformat()
            existing = self.db.fetch_one("SELECT id FROM batches WHERE warehouse_id=? AND date_in=?", (wh_id, date_in))
            if existing:
                self.db.execute("DELETE FROM batches WHERE id=?", (existing['id'],))
                self.db.execute("DELETE FROM daily_records WHERE batch_id=?", (existing['id'],))
                self.db.execute("DELETE FROM farm_sales WHERE batch_id=?", (existing['id'],))
                self.db.execute("DELETE FROM market_sales WHERE batch_id=?", (existing['id'],))

            # Insert main record
            cols = ["warehouse_id", "batch_num", "date_in", "date_out", "days", "chicks", "total_dead", "mort_rate", "total_cost", "total_rev", "net_result", "total_sold", "created_at"]
            vals = [wh_id, self.filename, date_in, self.daily_rows[-1]['rec_date'] if self.daily_rows else date_in, 
                    len(self.daily_rows), chicks, total_dead, mort_rate, total_cost, total_rev, net_result, total_sold, datetime.now().isoformat()]
            
            # Map other summary fields
            for k, v in self.result.items():
                if k not in cols and k not in ["chicks", "total_cost", "total_rev", "net_result"]:
                    cols.append(k); vals.append(v)
            
            q_marks = ",".join(["?"] * len(vals))
            cols_str = ",".join(cols)
            res = self.db.execute(f"INSERT INTO batches ({cols_str}) VALUES ({q_marks})", vals)
            batch_id = res.lastrowid

            # Insert sub-records
            for r in self.daily_rows:
                self.db.execute("INSERT INTO daily_records (batch_id, rec_date, day_num, dead_count, feed_kg) VALUES (?,?,?,?,?)", (batch_id, r['rec_date'], r['day_num'], r['dead_count'], r['feed_kg']))
            for s in self.farm_sales:
                self.db.execute("INSERT INTO farm_sales (batch_id, customer, qty, price, total_val) VALUES (?,?,?,?,?)", (batch_id, s['customer'], s['qty'], s['price'], s['total_val']))
            for s in self.market_sales:
                self.db.execute("INSERT INTO market_sales (batch_id, office, qty_sent, deaths, qty_sold, net_val, inv_num) VALUES (?,?,?,?,?,?,?)", (batch_id, s['office'], s['qty_sent'], s['deaths'], s['qty_sold'], s['net_val'], s['inv_num']))

            return True, f"تم استيراد {self.wh_name} بنجاح"
        except Exception as e:
            return False, str(e)

    def import_folder(self, folder_path):
        results = []
        for f in os.listdir(folder_path):
            if f.lower().endswith(('.xlsx', '.xlsm')) and not f.startswith('~$'):
                success, msg = self.import_file(os.path.join(folder_path, f))
                results.append({'file': f, 'success': success, 'message': msg})
        return results
