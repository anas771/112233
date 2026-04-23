from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from core.database import DB_PATH, get_conn

REQUIRED_SHEETS = ["اجمالي التكاليف", "تكاليف العلف", "بيان المبيعات", "ورقة1"]
PROFILE_SOURCE_KEY = "poultry_v4"
GENERIC_SOURCE_KEY = "generic_excel"

COST_LEGACY_COLUMNS = {
    "chick_val",
    "feed_val",
    "feed_trans",
    "sawdust_val",
    "water_val",
    "gas_val",
    "drugs_val",
    "wh_expenses",
    "house_exp",
    "breeders_pay",
    "qat_pay",
    "rent_val",
    "light_val",
    "sup_wh_pay",
    "sup_co_pay",
    "sup_sale_pay",
    "admin_val",
    "vaccine_pay",
    "delivery_val",
    "mixing_val",
    "wash_val",
    "other_costs",
}

REVENUE_LEGACY_COLUMNS = {
    "offal_val",
    "feed_sale",
    "feed_trans_r",
    "feed_rem_val",
    "drug_return",
    "gas_return",
}

SUMMARY_SKIP_LABELS = {
    "الاجمالي",
    "اجمالي",
    "الإجمالي",
    "المجموع",
    "اجماليالمبيعات",
    "اجماليالتكاليف",
    "نتيجةالدفعة",
}


def normalize_source_label(value: str) -> str:
    raw = (value or "").strip().lower()
    compact = re.sub(r"\s+", "", raw)
    return "".join(ch for ch in compact if ch.isalnum())


def _safe_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "--"}:
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def _safe_int(value: Any) -> int:
    return int(round(_safe_float(value)))


def _to_iso_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return ""
    text = text.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except Exception:
            continue
    return ""


def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(8192)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _ensure_file_list(files: Any) -> list[Path]:
    if files is None:
        return []
    if isinstance(files, (str, Path)):
        candidates = [files]
    else:
        candidates = list(files)
    out: list[Path] = []
    for candidate in candidates:
        p = Path(candidate)
        if not p.exists() or not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in {".xlsm", ".xlsx", ".xls"}:
            continue
        out.append(p)
    return out


def _default_profile_id(conn, source_key: str = PROFILE_SOURCE_KEY) -> int:
    row = conn.execute(
        """
        SELECT id
        FROM import_profiles
        WHERE source_key=? AND is_active=1
        ORDER BY is_default DESC, id ASC
        LIMIT 1
        """,
        (source_key,),
    ).fetchone()
    if row:
        return int(row["id"])
    profile_name = "poultry_v4_default" if source_key == PROFILE_SOURCE_KEY else "generic_excel_default"
    row = conn.execute(
        """
        INSERT INTO import_profiles(name, source_key, is_default, is_active)
        VALUES (?, ?, 1, 1)
        RETURNING id
        """,
        (profile_name, source_key),
    ).fetchone()
    return int(row["id"])


def _profile_source_key(conn, profile_id: int | None) -> str:
    if profile_id:
        row = conn.execute("SELECT source_key FROM import_profiles WHERE id=?", (int(profile_id),)).fetchone()
        if row and row["source_key"]:
            return str(row["source_key"])
    return PROFILE_SOURCE_KEY


def detect_profile(files: Any, source_key: str | None = None) -> dict[str, Any]:
    file_list = _ensure_file_list(files)
    checks: list[dict[str, Any]] = []
    poultry_matched = bool(file_list)
    generic_matched = bool(file_list)
    for path in file_list:
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            try:
                sheetnames = list(wb.sheetnames)
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
            missing = [name for name in REQUIRED_SHEETS if name not in sheetnames]
            poultry_ok = len(missing) == 0
            checks.append(
                {
                    "file_name": path.name,
                    "file_path": str(path),
                    "ok": poultry_ok,
                    "poultry_ok": poultry_ok,
                    "generic_ok": True,
                    "missing_sheets": missing,
                    "sheetnames": sheetnames,
                }
            )
            if not poultry_ok:
                poultry_matched = False
        except Exception as exc:
            checks.append(
                {
                    "file_name": path.name,
                    "file_path": str(path),
                    "ok": False,
                    "poultry_ok": False,
                    "generic_ok": False,
                    "missing_sheets": REQUIRED_SHEETS[:],
                    "sheetnames": [],
                    "error": str(exc),
                }
            )
            poultry_matched = False
            generic_matched = False
    resolved_source = source_key or (PROFILE_SOURCE_KEY if poultry_matched else GENERIC_SOURCE_KEY)
    required_sheets = REQUIRED_SHEETS[:] if resolved_source == PROFILE_SOURCE_KEY else []
    if resolved_source == PROFILE_SOURCE_KEY:
        matched = poultry_matched and bool(file_list)
    else:
        matched = generic_matched and bool(file_list)
        for item in checks:
            item["ok"] = bool(item.get("generic_ok", False))
    profile_id = None
    with get_conn(DB_PATH) as conn:
        profile_id = _default_profile_id(conn, resolved_source)
    return {
        "matched": matched,
        "source_key": resolved_source,
        "profile_id": profile_id,
        "required_sheets": required_sheets,
        "files": checks,
    }


def _extract_daily_meta(ws) -> dict[str, Any]:
    meta: dict[str, Any] = {
        "warehouse_name": "",
        "batch_num": "",
        "date_in": "",
        "date_out": "",
        "chicks": 0,
    }
    for r in range(1, 7):
        row_values = [ws.cell(r, c).value for c in range(1, 15)]
        for idx, val in enumerate(row_values):
            text = str(val or "").strip()
            normalized = normalize_source_label(text)
            if normalized in {"اسمالمزرعه", "اسمالمزرعة"}:
                for n in row_values[idx + 1 :]:
                    name = str(n or "").strip()
                    if name:
                        meta["warehouse_name"] = name
                        break
            elif normalized in {"رقمالدفعة"}:
                for n in row_values[idx + 1 :]:
                    if n not in (None, ""):
                        meta["batch_num"] = str(n).strip()
                        break
            elif normalized in {"تاريخالتنزيل", "تاريحتنزيل", "تاريحالتنزيل"}:
                for n in row_values[idx + 1 :]:
                    iso = _to_iso_date(n)
                    if iso:
                        meta["date_in"] = iso
                        break
            elif normalized in {"تاريخاليوم"}:
                for n in row_values[idx + 1 :]:
                    iso = _to_iso_date(n)
                    if iso:
                        meta["date_out"] = iso
                        break
            elif normalized in {"عددالدجاج", "عددالكتاكيت"}:
                for n in row_values[idx + 1 :]:
                    v = _safe_int(n)
                    if v > 0:
                        meta["chicks"] = v
                        break
    if not meta["warehouse_name"]:
        fallback = str(ws.cell(1, 11).value or "").strip()
        if fallback:
            meta["warehouse_name"] = fallback
    if not meta["batch_num"]:
        fallback = str(ws.cell(2, 11).value or "").strip()
        if fallback:
            meta["batch_num"] = fallback
    if not meta["date_in"]:
        meta["date_in"] = _to_iso_date(ws.cell(1, 5).value)
    if not meta["date_out"]:
        meta["date_out"] = _to_iso_date(ws.cell(2, 5).value)
    if meta["chicks"] <= 0:
        meta["chicks"] = _safe_int(ws.cell(1, 2).value)
    return meta


def _parse_daily_sheet(ws) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    meta = _extract_daily_meta(ws)
    header_row = 0
    for r in range(1, min(ws.max_row, 40) + 1):
        values = [str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column, 16) + 1)]
        joined = " ".join(values)
        if "التاريخ" in joined and ("العمر" in joined or "الوفيات" in joined):
            header_row = r
            break
    if header_row == 0:
        return [], meta

    col_map = {
        "date": 1,
        "day": 2,
        "birds": 3,
        "dead": 4,
        "feed": 6,
        "water": 0,
    }
    for c in range(1, min(ws.max_column, 24) + 1):
        name = str(ws.cell(header_row, c).value or "").strip()
        normalized = normalize_source_label(name)
        if "تاريخ" in name:
            col_map["date"] = c
        elif "العمر" in name:
            col_map["day"] = c
        elif normalized in {"العدد"}:
            col_map["birds"] = c
        elif "الوفيات" in name or "النافق" in name:
            col_map["dead"] = c
        elif "العلف" in name:
            col_map["feed"] = c
        elif "الماء" in name:
            col_map["water"] = c

    rows: list[dict[str, Any]] = []
    trailing_blanks = 0
    for r in range(header_row + 1, ws.max_row + 1):
        rec_date = _to_iso_date(ws.cell(r, col_map["date"]).value)
        if not rec_date:
            if rows:
                trailing_blanks += 1
                if trailing_blanks >= 3:
                    break
            continue
        trailing_blanks = 0
        day_num = _safe_int(ws.cell(r, col_map["day"]).value)
        birds = _safe_int(ws.cell(r, col_map["birds"]).value)
        dead_raw = ws.cell(r, col_map["dead"]).value
        feed_raw = ws.cell(r, col_map["feed"]).value
        water_raw = ws.cell(r, col_map["water"]).value if col_map["water"] else None
        if dead_raw in (None, "") and feed_raw in (None, "") and birds <= 0 and day_num <= 0:
            continue
        dead_count = _safe_int(dead_raw)
        feed_bags = _safe_float(feed_raw)
        water_ltr = _safe_float(water_raw)
        rows.append(
            {
                "source_row": r,
                "rec_date": rec_date,
                "day_num": day_num,
                "birds": birds,
                "dead_count": dead_count,
                "feed_kg": round(feed_bags * 50.0, 4),
                "feed_bags": feed_bags,
                "water_ltr": water_ltr,
                "notes": "",
            }
        )
    if rows:
        if not meta["date_in"]:
            meta["date_in"] = rows[0]["rec_date"]
        if not meta["date_out"]:
            meta["date_out"] = rows[-1]["rec_date"]
        if meta["chicks"] <= 0:
            meta["chicks"] = rows[0]["birds"] if rows[0]["birds"] > 0 else 0
    return rows, meta


def _is_sales_summary_text(text: str) -> bool:
    normalized = normalize_source_label(text)
    if not normalized:
        return True
    return any(k in normalized for k in {"الاجمالي", "اجمالي", "المجموع", "اسمالعميل", "اسم", "بيان"})


def _parse_sales_sheet(ws) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    header_row = 0
    for r in range(1, min(ws.max_row, 20) + 1):
        values = [str(ws.cell(r, c).value or "") for c in range(1, 16)]
        joined = " ".join(values)
        if "اسم العميل" in joined and "العدد" in joined:
            header_row = r
            break
    if header_row == 0:
        return [], []

    farm_sales: list[dict[str, Any]] = []
    market_sales: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        customer = str(ws.cell(r, 1).value or "").strip()
        qty_ajl = _safe_int(ws.cell(r, 2).value)
        price_ajl = _safe_float(ws.cell(r, 3).value)
        total_ajl = _safe_float(ws.cell(r, 4).value)
        qty_cash = _safe_int(ws.cell(r, 5).value)
        price_cash = _safe_float(ws.cell(r, 6).value)
        total_cash = _safe_float(ws.cell(r, 7).value)
        office = str(ws.cell(r, 8).value or "").strip()
        qty_sent = _safe_int(ws.cell(r, 9).value)
        deaths = _safe_int(ws.cell(r, 10).value)
        qty_sold = _safe_int(ws.cell(r, 11).value)
        net_val = _safe_float(ws.cell(r, 12).value)
        inv_num = str(ws.cell(r, 13).value or "").strip()
        sale_date = _to_iso_date(ws.cell(r, 14).value)

        if customer and not _is_sales_summary_text(customer):
            if qty_ajl > 0 or total_ajl > 0:
                unit_price = price_ajl if price_ajl > 0 else (total_ajl / qty_ajl if qty_ajl > 0 else 0.0)
                total_value = total_ajl if total_ajl > 0 else unit_price * qty_ajl
                farm_sales.append(
                    {
                        "source_row": r,
                        "sale_date": sale_date,
                        "sale_type": "آجل",
                        "customer": customer,
                        "qty": qty_ajl,
                        "price": round(unit_price, 4),
                        "total_val": round(total_value, 4),
                    }
                )
            if qty_cash > 0 or total_cash > 0:
                unit_price = price_cash if price_cash > 0 else (total_cash / qty_cash if qty_cash > 0 else 0.0)
                total_value = total_cash if total_cash > 0 else unit_price * qty_cash
                farm_sales.append(
                    {
                        "source_row": r,
                        "sale_date": sale_date,
                        "sale_type": "نقداً",
                        "customer": customer,
                        "qty": qty_cash,
                        "price": round(unit_price, 4),
                        "total_val": round(total_value, 4),
                    }
                )

        if office and not _is_sales_summary_text(office):
            if qty_sent > 0 or net_val > 0:
                market_sales.append(
                    {
                        "source_row": r,
                        "sale_date": sale_date,
                        "office": office,
                        "qty_sent": qty_sent,
                        "deaths": deaths,
                        "qty_sold": qty_sold if qty_sold > 0 else max(0, qty_sent - deaths),
                        "net_val": round(net_val, 4),
                        "inv_num": inv_num,
                    }
                )
    return farm_sales, market_sales


def _build_candidate(source_sheet: str, source_row: int, source_label: str, qty: float, amount: float, payload: dict[str, Any]) -> dict[str, Any] | None:
    label_raw = str(source_label or "").strip()
    if not label_raw:
        return None
    normalized = normalize_source_label(label_raw)
    if not normalized:
        return None
    return {
        "source_sheet": source_sheet,
        "source_row": source_row,
        "source_label_raw": label_raw,
        "source_label": normalized,
        "qty": round(_safe_float(qty), 4),
        "amount": round(_safe_float(amount), 4),
        "payload": payload,
    }


def _parse_summary_sheet(ws) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for r in range(3, min(ws.max_row, 200) + 1):
        cost_label = ws.cell(r, 1).value
        cost_qty = _safe_float(ws.cell(r, 2).value)
        cost_a = _safe_float(ws.cell(r, 3).value)
        cost_b = _safe_float(ws.cell(r, 4).value)
        cost_amount = cost_a + cost_b
        if cost_label not in (None, "") and (cost_qty != 0 or cost_amount != 0):
            candidate = _build_candidate(
                "اجمالي التكاليف",
                r,
                str(cost_label),
                cost_qty,
                cost_amount,
                {
                    "side": "cost",
                    "company_amount": cost_a,
                    "partner_amount": cost_b,
                    "statement": str(ws.cell(r, 5).value or "").strip(),
                },
            )
            if candidate:
                candidates.append(candidate)

        rev_label = ws.cell(r, 9).value
        rev_amount = _safe_float(ws.cell(r, 10).value)
        if rev_label not in (None, "") and rev_amount != 0:
            candidate = _build_candidate(
                "اجمالي التكاليف",
                r,
                str(rev_label),
                0.0,
                rev_amount,
                {
                    "side": "revenue",
                    "statement": str(ws.cell(r, 11).value or "").strip(),
                },
            )
            if candidate:
                candidates.append(candidate)
    return candidates


def _parse_feed_sheet(ws) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for r in range(8, min(ws.max_row, 120) + 1):
        description = str(ws.cell(r, 3).value or "").strip()
        if not description:
            continue
        amount = _safe_float(ws.cell(r, 1).value)
        qty = _safe_float(ws.cell(r, 2).value)
        if amount == 0 and qty == 0:
            continue
        candidate = _build_candidate(
            "تكاليف العلف",
            r,
            description,
            qty,
            amount,
            {"source": "feed_sheet"},
        )
        if candidate:
            candidates.append(candidate)
    return candidates


GENERIC_LABEL_HINTS = {
    "label",
    "item",
    "name",
    "description",
    "desc",
    "account",
    "category",
    "type",
    "bayan",
    "band",
    "costitem",
    "revenueitem",
    "مصروف",
    "ايراد",
    "بيان",
    "بند",
    "نوع",
    "وصف",
    "اسم",
    "صنف",
}
GENERIC_AMOUNT_HINTS = {
    "amount",
    "value",
    "total",
    "price",
    "sum",
    "net",
    "balance",
    "cost",
    "expense",
    "revenue",
    "income",
    "مبلغ",
    "قيمة",
    "اجمالي",
    "الإجمالي",
    "السعر",
    "التكلفة",
    "تكلفة",
    "مصروف",
    "ايراد",
    "دخل",
    "مدين",
    "دائن",
}
GENERIC_QTY_HINTS = {
    "qty",
    "quantity",
    "count",
    "units",
    "pieces",
    "weight",
    "kg",
    "ton",
    "عدد",
    "كمية",
    "وزن",
    "كجم",
    "طن",
}
GENERIC_DATE_HINTS = {
    "date",
    "day",
    "datetime",
    "entrydate",
    "postdate",
    "تاريخ",
    "اليوم",
}
GENERIC_KIND_HINTS = {
    "kind",
    "type",
    "side",
    "direction",
    "category",
    "nature",
    "نوع",
    "تصنيف",
    "فئة",
    "جهة",
}
GENERIC_COST_HINTS = {
    "cost",
    "expense",
    "purchase",
    "debit",
    "مصروف",
    "تكلفة",
    "مدين",
    "مشتريات",
    "شراء",
}
GENERIC_REVENUE_HINTS = {
    "revenue",
    "income",
    "sale",
    "credit",
    "ايراد",
    "إيراد",
    "دخل",
    "مبيعات",
    "بيع",
    "دائن",
}


def _contains_any_hint(normalized_text: str, hints: set[str]) -> bool:
    if not normalized_text:
        return False
    for hint in hints:
        if normalize_source_label(hint) in normalized_text:
            return True
    return False


def _guess_generic_header_row(ws) -> int:
    max_probe_rows = min(max(ws.max_row, 1), 30)
    max_probe_cols = min(max(ws.max_column, 1), 40)
    best_row = 1
    best_score = -1
    for r in range(1, max_probe_rows + 1):
        row_score = 0
        text_cells = 0
        for c in range(1, max_probe_cols + 1):
            value = ws.cell(r, c).value
            text = str(value or "").strip()
            if not text:
                continue
            normalized = normalize_source_label(text)
            if not normalized:
                continue
            text_cells += 1
            if _contains_any_hint(normalized, GENERIC_LABEL_HINTS):
                row_score += 3
            if _contains_any_hint(normalized, GENERIC_AMOUNT_HINTS):
                row_score += 2
            if _contains_any_hint(normalized, GENERIC_QTY_HINTS):
                row_score += 1
            if _contains_any_hint(normalized, GENERIC_DATE_HINTS):
                row_score += 1
        if text_cells >= 2 and row_score > best_score:
            best_score = row_score
            best_row = r
    return best_row


def _detect_generic_columns(ws, header_row: int) -> dict[str, Any]:
    max_probe_cols = min(max(ws.max_column, 1), 80)
    columns: dict[str, Any] = {
        "label": 0,
        "qty": 0,
        "date": 0,
        "kind": 0,
        "amount": 0,
        "cost_cols": [],
        "revenue_cols": [],
    }
    for c in range(1, max_probe_cols + 1):
        raw = ws.cell(header_row, c).value
        text = str(raw or "").strip()
        normalized = normalize_source_label(text)
        if not normalized:
            continue
        if not columns["label"] and _contains_any_hint(normalized, GENERIC_LABEL_HINTS):
            columns["label"] = c
            continue
        if not columns["qty"] and _contains_any_hint(normalized, GENERIC_QTY_HINTS):
            columns["qty"] = c
            continue
        if not columns["date"] and _contains_any_hint(normalized, GENERIC_DATE_HINTS):
            columns["date"] = c
            continue
        if not columns["kind"] and _contains_any_hint(normalized, GENERIC_KIND_HINTS):
            columns["kind"] = c
            continue
        if _contains_any_hint(normalized, GENERIC_COST_HINTS):
            columns["cost_cols"].append(c)
            continue
        if _contains_any_hint(normalized, GENERIC_REVENUE_HINTS):
            columns["revenue_cols"].append(c)
            continue
        if not columns["amount"] and _contains_any_hint(normalized, GENERIC_AMOUNT_HINTS):
            columns["amount"] = c

    if not columns["label"]:
        columns["label"] = 1

    if not columns["amount"] and not columns["cost_cols"] and not columns["revenue_cols"]:
        best_col = 0
        best_score = 0
        max_probe_rows = min(ws.max_row, header_row + 50)
        for c in range(1, max_probe_cols + 1):
            numeric_count = 0
            for r in range(header_row + 1, max_probe_rows + 1):
                val = ws.cell(r, c).value
                if _safe_float(val) != 0:
                    numeric_count += 1
            if numeric_count > best_score:
                best_score = numeric_count
                best_col = c
        columns["amount"] = best_col
    return columns


def _guess_generic_side(label: str, kind_text: str, amount: float) -> str:
    kind_norm = normalize_source_label(kind_text)
    label_norm = normalize_source_label(label)
    if _contains_any_hint(kind_norm, GENERIC_REVENUE_HINTS) or _contains_any_hint(label_norm, GENERIC_REVENUE_HINTS):
        return "revenue"
    if _contains_any_hint(kind_norm, GENERIC_COST_HINTS) or _contains_any_hint(label_norm, GENERIC_COST_HINTS):
        return "cost"
    if amount < 0:
        return "cost"
    return "cost"


def _build_generic_meta(path: Path, candidates: list[dict[str, Any]], sheetnames: list[str]) -> dict[str, Any]:
    rec_dates: list[str] = []
    for item in candidates:
        payload = item.get("payload") or {}
        rec_date = _to_iso_date(payload.get("rec_date"))
        if rec_date:
            rec_dates.append(rec_date)

    warehouse_guess = path.stem.strip()
    for sep in ("دفعة", "دورة", "batch", "Batch"):
        if sep in warehouse_guess:
            warehouse_guess = warehouse_guess.split(sep)[0].strip()
            break
    warehouse_guess = warehouse_guess[:80] or "عنبر غير محدد"

    date_in = min(rec_dates) if rec_dates else ""
    date_out = max(rec_dates) if rec_dates else date_in
    if not date_in:
        date_in = date.today().isoformat()
    if not date_out:
        date_out = date_in

    return {
        "warehouse_name": warehouse_guess,
        "batch_num": path.stem[-30:],
        "date_in": date_in,
        "date_out": date_out,
        "chicks": 0,
        "sheetnames": sheetnames,
    }


def _parse_generic_sheet(ws, sheet_name: str) -> list[dict[str, Any]]:
    header_row = _guess_generic_header_row(ws)
    columns = _detect_generic_columns(ws, header_row)
    if not columns["amount"] and not columns["cost_cols"] and not columns["revenue_cols"]:
        return []

    max_probe_cols = min(max(ws.max_column, 1), 80)
    candidates: list[dict[str, Any]] = []
    trailing_blanks = 0

    for r in range(header_row + 1, ws.max_row + 1):
        label = ""
        if int(columns["label"] or 0) > 0:
            label = str(ws.cell(r, int(columns["label"])).value or "").strip()
        if not label:
            for c in range(1, max_probe_cols + 1):
                value = ws.cell(r, c).value
                text = str(value or "").strip()
                if not text:
                    continue
                if _safe_float(value) != 0:
                    continue
                if len(normalize_source_label(text)) < 2:
                    continue
                label = text
                break

        qty = _safe_float(ws.cell(r, int(columns["qty"])).value) if int(columns["qty"] or 0) > 0 else 0.0
        rec_date = _to_iso_date(ws.cell(r, int(columns["date"])).value) if int(columns["date"] or 0) > 0 else ""
        kind_text = str(ws.cell(r, int(columns["kind"])).value or "").strip() if int(columns["kind"] or 0) > 0 else ""

        added_any = False

        def append_candidate(amount_value: Any, side_hint: str, col_name: str = "") -> None:
            nonlocal added_any
            amount = _safe_float(amount_value)
            if amount == 0:
                return
            source_label = label or f"row_{r}"
            if col_name and col_name != source_label:
                source_label = f"{source_label} - {col_name}"
            guessed_side = _guess_generic_side(source_label, kind_text or side_hint, amount)
            normalized_label = normalize_source_label(source_label)
            if normalized_label in SUMMARY_SKIP_LABELS:
                return
            payload = {
                "side": guessed_side,
                "kind_hint": kind_text,
                "source_label_raw": source_label,
                "source_mode": "generic",
            }
            if rec_date:
                payload["rec_date"] = rec_date
            candidate = _build_candidate(sheet_name, r, source_label, qty, abs(amount), payload)
            if candidate:
                candidates.append(candidate)
                added_any = True

        for col in columns["cost_cols"]:
            col_name = str(ws.cell(header_row, int(col)).value or "").strip()
            append_candidate(ws.cell(r, int(col)).value, "cost", col_name)

        for col in columns["revenue_cols"]:
            col_name = str(ws.cell(header_row, int(col)).value or "").strip()
            append_candidate(ws.cell(r, int(col)).value, "revenue", col_name)

        if not columns["cost_cols"] and not columns["revenue_cols"] and int(columns["amount"] or 0) > 0:
            append_candidate(ws.cell(r, int(columns["amount"])).value, "", "")

        if added_any:
            trailing_blanks = 0
        else:
            if not label and qty == 0 and not rec_date:
                trailing_blanks += 1
                if trailing_blanks >= 20 and candidates:
                    break
            else:
                trailing_blanks = 0
    return candidates


def _parse_file_generic(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        sheetnames = list(wb.sheetnames)
        candidates: list[dict[str, Any]] = []
        seen: set[tuple[str, int, str, float, float]] = set()
        for sheet_name in sheetnames:
            ws = wb[sheet_name]
            for item in _parse_generic_sheet(ws, sheet_name):
                key = (
                    str(item.get("source_sheet") or ""),
                    int(item.get("source_row") or 0),
                    str(item.get("source_label") or ""),
                    round(_safe_float(item.get("qty")), 4),
                    round(_safe_float(item.get("amount")), 4),
                )
                if key in seen:
                    continue
                seen.add(key)
                candidates.append(item)
        if not candidates:
            raise ValueError("تعذر استخراج بنود قابلة للاستيراد من الملف. اختر بروفايل poultry_v4 أو عدل الملف.")
        meta = _build_generic_meta(path, candidates, sheetnames)
        return {
            "file_name": path.name,
            "file_path": str(path),
            "fingerprint_sha256": _file_sha256(path),
            "detected": meta,
            "daily_records": [],
            "farm_sales": [],
            "market_sales": [],
            "candidates": candidates,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _parse_file(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        sheetnames = list(wb.sheetnames)
        missing = [name for name in REQUIRED_SHEETS if name not in sheetnames]
        if missing:
            raise ValueError(f"ملف غير مطابق لنموذج poultry_v4، الشيتات المفقودة: {', '.join(missing)}")

        daily_rows, meta = _parse_daily_sheet(wb["ورقة1"])
        farm_sales, market_sales = _parse_sales_sheet(wb["بيان المبيعات"])
        candidates = _parse_summary_sheet(wb["اجمالي التكاليف"])
        candidates.extend(_parse_feed_sheet(wb["تكاليف العلف"]))

        if not meta.get("warehouse_name"):
            cleaned = path.stem.strip()
            for sep in ("دفعة", "دورة", "batch", "Batch"):
                if sep in cleaned:
                    cleaned = cleaned.split(sep)[0].strip()
                    break
            meta["warehouse_name"] = cleaned[:80]
        if not meta.get("batch_num"):
            meta["batch_num"] = path.stem[-30:]
        if not meta.get("date_in") and daily_rows:
            meta["date_in"] = daily_rows[0]["rec_date"]
        if not meta.get("date_out") and daily_rows:
            meta["date_out"] = daily_rows[-1]["rec_date"]
        if int(meta.get("chicks") or 0) <= 0 and daily_rows:
            meta["chicks"] = max(0, _safe_int(daily_rows[0].get("birds")))

        return {
            "file_name": path.name,
            "file_path": str(path),
            "fingerprint_sha256": _file_sha256(path),
            "detected": {
                "warehouse_name": str(meta.get("warehouse_name") or "").strip(),
                "batch_num": str(meta.get("batch_num") or "").strip(),
                "date_in": str(meta.get("date_in") or "").strip(),
                "date_out": str(meta.get("date_out") or "").strip(),
                "chicks": max(0, int(meta.get("chicks") or 0)),
                "sheetnames": sheetnames,
            },
            "daily_records": daily_rows,
            "farm_sales": farm_sales,
            "market_sales": market_sales,
            "candidates": candidates,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _parse_files_legacy_template_only(files: Any, profile_id: int | None) -> dict[str, Any]:
    file_list = _ensure_file_list(files)
    if not file_list:
        raise ValueError("لم يتم العثور على ملفات Excel صالحة للاستيراد.")
    parsed_files: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for path in file_list:
        try:
            parsed_files.append(_parse_file(path))
        except Exception as exc:
            errors.append({"file_name": path.name, "file_path": str(path), "error": str(exc)})
    if not parsed_files:
        raise ValueError("فشل تحليل كل الملفات. تأكد أن القالب هو poultry_v4.")
    return {
        "profile_id": profile_id,
        "source_key": PROFILE_SOURCE_KEY,
        "files": parsed_files,
        "errors": errors,
    }


def parse_files(files: Any, profile_id: int | None) -> dict[str, Any]:
    file_list = _ensure_file_list(files)
    if not file_list:
        raise ValueError("No valid Excel files were found for import.")

    with get_conn(DB_PATH) as conn:
        if profile_id:
            resolved_profile_id = int(profile_id)
            source_key = _profile_source_key(conn, resolved_profile_id)
        else:
            auto = detect_profile(file_list)
            resolved_profile_id = int(auto.get("profile_id") or _default_profile_id(conn))
            source_key = str(auto.get("source_key") or PROFILE_SOURCE_KEY)

    parsed_files: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for path in file_list:
        try:
            if source_key == GENERIC_SOURCE_KEY:
                parsed_files.append(_parse_file_generic(path))
            else:
                parsed_files.append(_parse_file(path))
        except Exception as exc:
            errors.append({"file_name": path.name, "file_path": str(path), "error": str(exc)})

    if not parsed_files:
        if source_key == GENERIC_SOURCE_KEY:
            raise ValueError("Failed to parse files in Generic Excel mode.")
        raise ValueError("Failed to parse files for poultry_v4 template.")

    return {
        "profile_id": resolved_profile_id,
        "source_key": source_key,
        "files": parsed_files,
        "errors": errors,
    }


def _load_profile_mappings(conn, profile_id: int) -> dict[tuple[str, str], dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT *
        FROM import_mappings
        WHERE profile_id=? AND is_active=1
        ORDER BY id DESC
        """,
        (profile_id,),
    ).fetchall()
    result: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (str(row["source_sheet"] or ""), str(row["source_label"] or ""))
        if key not in result:
            result[key] = dict(row)
    return result


def _run_status_from_lines(conn, run_id: int) -> str:
    unresolved = conn.execute(
        """
        SELECT COUNT(*) AS c
        FROM import_run_lines
        WHERE run_id=? AND line_kind='candidate' AND mapping_status='unmapped'
        """,
        (run_id,),
    ).fetchone()
    if unresolved and int(unresolved["c"] or 0) > 0:
        return "review"
    return "ready"


def build_staging(payload: dict[str, Any], profile_id: int | None, source_ui: str = "web", created_by: str = "") -> int:
    with get_conn(DB_PATH) as conn:
        source_key = str(payload.get("source_key") or PROFILE_SOURCE_KEY)
        resolved_profile_id = profile_id or _default_profile_id(conn, source_key)
        run_row = conn.execute(
            """
            INSERT INTO import_runs(source_ui, status, profile_id, batch_mode, merge_mode, created_by)
            VALUES (?, 'draft', ?, 'create', 'replace', ?)
            RETURNING id
            """,
            (source_ui, resolved_profile_id, created_by or ""),
        ).fetchone()
        run_id = int(run_row["id"])
        mappings = _load_profile_mappings(conn, resolved_profile_id)

        for parsed_file in payload.get("files", []):
            detected = parsed_file.get("detected", {})
            file_name = str(parsed_file.get("file_name") or "")
            file_path = str(parsed_file.get("file_path") or "")
            fingerprint = str(parsed_file.get("fingerprint_sha256") or "")
            detected_batch = str(detected.get("batch_num") or "")
            detected_date_in = str(detected.get("date_in") or "")
            duplicate = conn.execute(
                """
                SELECT id, import_count
                FROM import_file_fingerprints
                WHERE fingerprint_sha256=? AND batch_num=? AND date_in=?
                """,
                (fingerprint, detected_batch, detected_date_in),
            ).fetchone()
            file_status = "duplicate" if duplicate else "staged"
            file_reason = ""
            if duplicate:
                file_reason = f"تم استيراد الملف سابقًا ({int(duplicate['import_count'] or 0)} مرة)"

            run_file_row = conn.execute(
                """
                INSERT INTO import_run_files(
                    run_id, file_name, file_path, fingerprint_sha256,
                    detected_warehouse, detected_batch_num, detected_date_in,
                    status, reason
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                RETURNING id
                """,
                (
                    run_id,
                    file_name,
                    file_path,
                    fingerprint,
                    str(detected.get("warehouse_name") or ""),
                    detected_batch,
                    detected_date_in,
                    file_status,
                    file_reason,
                ),
            ).fetchone()
            run_file_id = int(run_file_row["id"])

            for item in parsed_file.get("daily_records", []):
                conn.execute(
                    """
                    INSERT INTO import_run_lines(
                        run_id, run_file_id, line_kind, source_sheet, source_row, source_label,
                        qty, amount, rec_date, payload_json, mapping_status, target_kind, apply_flag
                    )
                    VALUES (?, ?, 'daily', 'ورقة1', ?, '', ?, ?, ?, ?, 'fixed', 'daily', 1)
                    """,
                    (
                        run_id,
                        run_file_id,
                        int(item.get("source_row") or 0),
                        _safe_float(item.get("dead_count")),
                        _safe_float(item.get("feed_kg")),
                        str(item.get("rec_date") or ""),
                        json.dumps(item, ensure_ascii=False),
                    ),
                )

            for item in parsed_file.get("farm_sales", []):
                conn.execute(
                    """
                    INSERT INTO import_run_lines(
                        run_id, run_file_id, line_kind, source_sheet, source_row, source_label,
                        qty, amount, rec_date, payload_json, mapping_status, target_kind, apply_flag
                    )
                    VALUES (?, ?, 'sales_farm', 'بيان المبيعات', ?, ?, ?, ?, ?, ?, 'fixed', 'sales', 1)
                    """,
                    (
                        run_id,
                        run_file_id,
                        int(item.get("source_row") or 0),
                        normalize_source_label(item.get("customer", "")),
                        _safe_float(item.get("qty")),
                        _safe_float(item.get("total_val")),
                        str(item.get("sale_date") or ""),
                        json.dumps(item, ensure_ascii=False),
                    ),
                )

            for item in parsed_file.get("market_sales", []):
                conn.execute(
                    """
                    INSERT INTO import_run_lines(
                        run_id, run_file_id, line_kind, source_sheet, source_row, source_label,
                        qty, amount, rec_date, payload_json, mapping_status, target_kind, apply_flag
                    )
                    VALUES (?, ?, 'sales_market', 'بيان المبيعات', ?, ?, ?, ?, ?, ?, 'fixed', 'sales', 1)
                    """,
                    (
                        run_id,
                        run_file_id,
                        int(item.get("source_row") or 0),
                        normalize_source_label(item.get("office", "")),
                        _safe_float(item.get("qty_sold")),
                        _safe_float(item.get("net_val")),
                        str(item.get("sale_date") or ""),
                        json.dumps(item, ensure_ascii=False),
                    ),
                )

            for item in parsed_file.get("candidates", []):
                source_sheet = str(item.get("source_sheet") or "")
                source_label = str(item.get("source_label") or "")
                source_raw = str(item.get("source_label_raw") or "")
                mapping = mappings.get((source_sheet, source_label))
                mapping_status = "unmapped"
                target_kind = ""
                target_code = ""
                target_name = ""
                category = ""
                unit = ""
                has_qty = 0
                is_auto_created = 0
                apply_flag = 1
                if mapping:
                    target_kind = str(mapping.get("target_kind") or "")
                    target_code = str(mapping.get("target_code") or "")
                    target_name = str(mapping.get("target_name") or "")
                    category = str(mapping.get("category") or "")
                    unit = str(mapping.get("unit") or "")
                    has_qty = int(mapping.get("has_qty") or 0)
                    is_auto_created = int(mapping.get("is_auto_created") or 0)
                    if target_kind == "ignore":
                        mapping_status = "ignored"
                        apply_flag = 0
                    elif target_code:
                        mapping_status = "mapped_existing"
                    elif target_kind in {"cost", "revenue"} and target_name:
                        mapping_status = "mapped_new"
                    else:
                        mapping_status = "unmapped"

                payload_json = dict(item.get("payload") or {})
                payload_json["source_label_raw"] = source_raw
                conn.execute(
                    """
                    INSERT INTO import_run_lines(
                        run_id, run_file_id, line_kind, source_sheet, source_row, source_label,
                        qty, amount, payload_json, mapping_status,
                        target_kind, target_code, target_name, category, unit, has_qty, is_auto_created, apply_flag
                    )
                    VALUES (?, ?, 'candidate', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        run_id,
                        run_file_id,
                        source_sheet,
                        int(item.get("source_row") or 0),
                        source_label,
                        _safe_float(item.get("qty")),
                        _safe_float(item.get("amount")),
                        json.dumps(payload_json, ensure_ascii=False),
                        mapping_status,
                        target_kind,
                        target_code,
                        target_name,
                        category,
                        unit,
                        int(has_qty),
                        int(is_auto_created),
                        int(apply_flag),
                    ),
                )

        run_status = _run_status_from_lines(conn, run_id)
        conn.execute(
            "UPDATE import_runs SET status=?, profile_id=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (run_status, resolved_profile_id, run_id),
        )
        conn.commit()
        return run_id


def _upsert_profile_mapping(
    conn,
    profile_id: int,
    source_sheet: str,
    source_label: str,
    target_kind: str,
    target_code: str,
    target_name: str,
    category: str,
    unit: str,
    has_qty: int,
    is_auto_created: int,
) -> None:
    conn.execute(
        """
        INSERT INTO import_mappings(
            profile_id, source_sheet, source_label, target_kind, target_code, target_name,
            category, unit, has_qty, is_auto_created, is_active
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        ON CONFLICT(profile_id, source_sheet, source_label, target_kind)
        DO UPDATE SET
            target_code=excluded.target_code,
            target_name=excluded.target_name,
            category=excluded.category,
            unit=excluded.unit,
            has_qty=excluded.has_qty,
            is_auto_created=excluded.is_auto_created,
            is_active=1,
            updated_at=CURRENT_TIMESTAMP
        """,
        (
            profile_id,
            source_sheet,
            source_label,
            target_kind,
            target_code,
            target_name,
            category,
            unit,
            int(has_qty),
            int(is_auto_created),
        ),
    )


def apply_mapping_edits(run_id: int, edits: list[dict[str, Any]]) -> None:
    if not edits:
        return
    with get_conn(DB_PATH) as conn:
        run = conn.execute("SELECT id, profile_id FROM import_runs WHERE id=?", (run_id,)).fetchone()
        if not run:
            raise ValueError("run_id غير موجود.")
        profile_id = int(run["profile_id"] or _default_profile_id(conn))
        for edit in edits:
            line_id = int(edit.get("line_id") or 0)
            if line_id <= 0:
                continue
            line = conn.execute(
                """
                SELECT id, source_sheet, source_label
                FROM import_run_lines
                WHERE id=? AND run_id=? AND line_kind='candidate'
                """,
                (line_id, run_id),
            ).fetchone()
            if not line:
                continue

            target_kind = str(edit.get("target_kind") or "").strip()
            target_code = str(edit.get("target_code") or "").strip()
            target_name = str(edit.get("target_name") or "").strip()
            category = str(edit.get("category") or "").strip()
            unit = str(edit.get("unit") or "").strip()
            has_qty = int(edit.get("has_qty") or 0)
            action = str(edit.get("action") or "").strip().lower()

            if action == "ignore" or target_kind == "ignore":
                conn.execute(
                    """
                    UPDATE import_run_lines
                    SET mapping_status='ignored', target_kind='ignore', target_code='',
                        target_name='', category='', unit='', has_qty=0, is_auto_created=0, apply_flag=0
                    WHERE id=?
                    """,
                    (line_id,),
                )
                _upsert_profile_mapping(
                    conn,
                    profile_id,
                    str(line["source_sheet"] or ""),
                    str(line["source_label"] or ""),
                    "ignore",
                    "",
                    "",
                    "",
                    "",
                    0,
                    0,
                )
                continue

            if target_kind not in {"cost", "revenue"}:
                continue

            if action == "existing" or target_code:
                mapping_status = "mapped_existing"
                is_auto = 0
                if not target_name and target_code:
                    table = "cost_types" if target_kind == "cost" else "revenue_types"
                    row = conn.execute(
                        f"SELECT name_ar, category, unit, has_qty FROM {table} WHERE code=?",
                        (target_code,),
                    ).fetchone()
                    if row:
                        target_name = str(row["name_ar"] or "")
                        if not category:
                            category = str(row["category"] or "")
                        if not unit:
                            unit = str(row["unit"] or "")
                        has_qty = int(row["has_qty"] or 0)
            else:
                if not target_name:
                    continue
                mapping_status = "mapped_new"
                is_auto = 1
                target_code = ""

            conn.execute(
                """
                UPDATE import_run_lines
                SET mapping_status=?, target_kind=?, target_code=?, target_name=?,
                    category=?, unit=?, has_qty=?, is_auto_created=?, apply_flag=1
                WHERE id=?
                """,
                (
                    mapping_status,
                    target_kind,
                    target_code,
                    target_name,
                    category,
                    unit,
                    int(has_qty),
                    int(is_auto),
                    line_id,
                ),
            )
            _upsert_profile_mapping(
                conn,
                profile_id,
                str(line["source_sheet"] or ""),
                str(line["source_label"] or ""),
                target_kind,
                target_code,
                target_name,
                category,
                unit,
                int(has_qty),
                int(is_auto),
            )

        run_status = _run_status_from_lines(conn, run_id)
        conn.execute(
            "UPDATE import_runs SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (run_status, run_id),
        )
        conn.commit()


def _ensure_warehouse(conn, name: str) -> int:
    normalized = (name or "").strip() or "عنبر غير محدد"
    row = conn.execute("SELECT id FROM warehouses WHERE name=?", (normalized,)).fetchone()
    if row:
        return int(row["id"])
    row = conn.execute(
        "INSERT INTO warehouses(name, notes) VALUES (?, '') RETURNING id",
        (normalized,),
    ).fetchone()
    return int(row["id"])


def _create_batch(conn, run_file: dict[str, Any], file_meta: dict[str, Any]) -> int:
    warehouse_id = _ensure_warehouse(conn, str(run_file.get("detected_warehouse") or file_meta.get("warehouse_name") or ""))
    batch_num = str(run_file.get("detected_batch_num") or file_meta.get("batch_num") or "").strip()
    date_in = str(run_file.get("detected_date_in") or file_meta.get("date_in") or "")
    date_out = str(file_meta.get("date_out") or date_in)
    chicks = max(0, int(file_meta.get("chicks") or 0))
    if not date_in:
        date_in = date.today().isoformat()
    if not date_out:
        date_out = date_in
    fiscal_year = 0
    try:
        fiscal_year = int(date_in[:4])
    except Exception:
        fiscal_year = date.today().year
    row = conn.execute(
        """
        INSERT INTO batches(
            warehouse_id, batch_num, date_in, date_out, days, chicks,
            created_at, fiscal_year
        )
        VALUES (?, ?, ?, ?, 0, ?, ?, ?)
        RETURNING id
        """,
        (
            warehouse_id,
            batch_num,
            date_in,
            date_out,
            chicks,
            datetime.now().isoformat(timespec="seconds"),
            fiscal_year,
        ),
    ).fetchone()
    return int(row["id"])


def _find_or_assign_update_batch(conn, run_file: dict[str, Any], requested_batch_id: int | None) -> int:
    if requested_batch_id and requested_batch_id > 0:
        return requested_batch_id
    if run_file.get("target_batch_id"):
        return int(run_file["target_batch_id"])
    warehouse_name = str(run_file.get("detected_warehouse") or "").strip()
    batch_num = str(run_file.get("detected_batch_num") or "").strip()
    date_in = str(run_file.get("detected_date_in") or "").strip()
    if warehouse_name and date_in:
        row = conn.execute(
            """
            SELECT b.id
            FROM batches b
            JOIN warehouses w ON w.id=b.warehouse_id
            WHERE w.name=? AND b.date_in=?
            ORDER BY b.id DESC
            LIMIT 1
            """,
            (warehouse_name, date_in),
        ).fetchone()
        if row:
            return int(row["id"])
    if batch_num:
        row = conn.execute(
            "SELECT id FROM batches WHERE batch_num=? ORDER BY id DESC LIMIT 1",
            (batch_num,),
        ).fetchone()
        if row:
            return int(row["id"])
    raise ValueError("تعذر تحديد الدفعة المستهدفة لوضع Update.")


def _load_run_file_lines(conn, run_file_id: int) -> dict[str, list[dict[str, Any]]]:
    rows = conn.execute(
        """
        SELECT *
        FROM import_run_lines
        WHERE run_file_id=?
        ORDER BY id
        """,
        (run_file_id,),
    ).fetchall()
    out: dict[str, list[dict[str, Any]]] = {
        "daily": [],
        "sales_farm": [],
        "sales_market": [],
        "candidate": [],
    }
    for row in rows:
        row_dict = dict(row)
        payload = {}
        try:
            payload = json.loads(str(row_dict.get("payload_json") or "{}"))
        except Exception:
            payload = {}
        row_dict["payload"] = payload
        kind = str(row_dict.get("line_kind") or "")
        if kind in out:
            out[kind].append(row_dict)
    return out


def _farm_sale_exists(conn, batch_id: int, sale: dict[str, Any]) -> bool:
    row = conn.execute(
        """
        SELECT id
        FROM farm_sales
        WHERE batch_id=?
          AND COALESCE(sale_date,'')=?
          AND COALESCE(sale_type,'')=?
          AND COALESCE(customer,'')=?
          AND COALESCE(qty,0)=?
          AND ABS(COALESCE(price,0)-?) < 0.0001
          AND ABS(COALESCE(total_val,0)-?) < 0.0001
        LIMIT 1
        """,
        (
            batch_id,
            str(sale.get("sale_date") or ""),
            str(sale.get("sale_type") or ""),
            str(sale.get("customer") or ""),
            _safe_int(sale.get("qty")),
            _safe_float(sale.get("price")),
            _safe_float(sale.get("total_val")),
        ),
    ).fetchone()
    return bool(row)


def _market_sale_exists(conn, batch_id: int, sale: dict[str, Any]) -> bool:
    row = conn.execute(
        """
        SELECT id
        FROM market_sales
        WHERE batch_id=?
          AND COALESCE(sale_date,'')=?
          AND COALESCE(office,'')=?
          AND COALESCE(qty_sent,0)=?
          AND COALESCE(deaths,0)=?
          AND COALESCE(qty_sold,0)=?
          AND ABS(COALESCE(net_val,0)-?) < 0.0001
          AND COALESCE(inv_num,'')=?
        LIMIT 1
        """,
        (
            batch_id,
            str(sale.get("sale_date") or ""),
            str(sale.get("office") or ""),
            _safe_int(sale.get("qty_sent")),
            _safe_int(sale.get("deaths")),
            _safe_int(sale.get("qty_sold")),
            _safe_float(sale.get("net_val")),
            str(sale.get("inv_num") or ""),
        ),
    ).fetchone()
    return bool(row)


def _resolve_type(conn, line: dict[str, Any], created_types: list[dict[str, Any]]) -> tuple[int, str]:
    target_kind = str(line.get("target_kind") or "")
    if target_kind not in {"cost", "revenue"}:
        raise ValueError("target_kind must be cost or revenue.")
    table = "cost_types" if target_kind == "cost" else "revenue_types"
    code = str(line.get("target_code") or "").strip()
    target_name = str(line.get("target_name") or "").strip()
    category = str(line.get("category") or "").strip() or ("أخرى" if target_kind == "cost" else "مبيعات")
    unit = str(line.get("unit") or "").strip() or None
    has_qty = int(line.get("has_qty") or 0)

    if code:
        row = conn.execute(f"SELECT id, code FROM {table} WHERE code=?", (code,)).fetchone()
        if row:
            return int(row["id"]), str(row["code"])

    if target_name:
        row = conn.execute(
            f"SELECT id, code FROM {table} WHERE lower(name_ar)=lower(?) ORDER BY id LIMIT 1",
            (target_name,),
        ).fetchone()
        if row:
            return int(row["id"]), str(row["code"])

    if not target_name:
        raise ValueError("البند الجديد يحتاج target_name.")

    prefix = "cost" if target_kind == "cost" else "rev"
    base = hashlib.sha1(target_name.encode("utf-8")).hexdigest()[:10]
    generated = f"auto_{prefix}_{base}"
    probe = generated
    suffix = 1
    while conn.execute(f"SELECT 1 FROM {table} WHERE code=?", (probe,)).fetchone():
        suffix += 1
        probe = f"{generated}_{suffix}"
    max_sort = conn.execute(f"SELECT COALESCE(MAX(sort_order), 0) AS m FROM {table}").fetchone()
    sort_order = int(max_sort["m"] or 0) + 1
    row = conn.execute(
        f"""
        INSERT INTO {table}(code, name_ar, category, has_qty, unit, sort_order, is_active)
        VALUES (?, ?, ?, ?, ?, ?, 1)
        RETURNING id, code
        """,
        (probe, target_name, category, has_qty, unit, sort_order),
    ).fetchone()
    created_types.append(
        {
            "kind": target_kind,
            "code": str(row["code"]),
            "name_ar": target_name,
            "category": category,
            "unit": unit or "",
            "has_qty": has_qty,
        }
    )
    return int(row["id"]), str(row["code"])


def _apply_candidate_line(conn, batch_id: int, line: dict[str, Any], merge_mode: str, created_types: list[dict[str, Any]]) -> tuple[str, int]:
    type_id, resolved_code = _resolve_type(conn, line, created_types)
    qty = _safe_float(line.get("qty"))
    amount = _safe_float(line.get("amount"))
    target_kind = str(line.get("target_kind") or "")
    if target_kind == "cost":
        if merge_mode == "merge":
            conn.execute(
                """
                INSERT INTO batch_costs(batch_id, cost_type_id, qty, amount, notes)
                VALUES (?, ?, ?, ?, '')
                ON CONFLICT(batch_id, cost_type_id)
                DO UPDATE SET
                    qty=COALESCE(batch_costs.qty,0)+excluded.qty,
                    amount=COALESCE(batch_costs.amount,0)+excluded.amount
                """,
                (batch_id, type_id, qty, amount),
            )
        else:
            conn.execute(
                """
                INSERT INTO batch_costs(batch_id, cost_type_id, qty, amount, notes)
                VALUES (?, ?, ?, ?, '')
                ON CONFLICT(batch_id, cost_type_id)
                DO UPDATE SET qty=excluded.qty, amount=excluded.amount
                """,
                (batch_id, type_id, qty, amount),
            )
    else:
        if merge_mode == "merge":
            conn.execute(
                """
                INSERT INTO batch_revenues(batch_id, revenue_type_id, qty, amount, notes)
                VALUES (?, ?, ?, ?, '')
                ON CONFLICT(batch_id, revenue_type_id)
                DO UPDATE SET
                    qty=COALESCE(batch_revenues.qty,0)+excluded.qty,
                    amount=COALESCE(batch_revenues.amount,0)+excluded.amount
                """,
                (batch_id, type_id, qty, amount),
            )
        else:
            conn.execute(
                """
                INSERT INTO batch_revenues(batch_id, revenue_type_id, qty, amount, notes)
                VALUES (?, ?, ?, ?, '')
                ON CONFLICT(batch_id, revenue_type_id)
                DO UPDATE SET qty=excluded.qty, amount=excluded.amount
                """,
                (batch_id, type_id, qty, amount),
            )
    return resolved_code, type_id


def _recalculate_batch(conn, batch_id: int) -> None:
    row = conn.execute("SELECT * FROM batches WHERE id=?", (batch_id,)).fetchone()
    if not row:
        return
    b = dict(row)
    fixed_cost = sum(float(b.get(col) or 0) for col in COST_LEGACY_COLUMNS if col in b)
    fixed_rev = sum(float(b.get(col) or 0) for col in REVENUE_LEGACY_COLUMNS if col in b)

    extra_cost_row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS v FROM batch_costs WHERE batch_id=?",
        (batch_id,),
    ).fetchone()
    extra_rev_row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS v FROM batch_revenues WHERE batch_id=?",
        (batch_id,),
    ).fetchone()
    farm_row = conn.execute(
        "SELECT COALESCE(SUM(total_val), 0) AS val, COALESCE(SUM(qty), 0) AS qty FROM farm_sales WHERE batch_id=?",
        (batch_id,),
    ).fetchone()
    market_row = conn.execute(
        "SELECT COALESCE(SUM(net_val), 0) AS val, COALESCE(SUM(qty_sold), 0) AS sold, COALESCE(SUM(deaths), 0) AS deaths FROM market_sales WHERE batch_id=?",
        (batch_id,),
    ).fetchone()
    daily_row = conn.execute(
        "SELECT COALESCE(SUM(dead_count), 0) AS dead FROM daily_records WHERE batch_id=?",
        (batch_id,),
    ).fetchone()
    dates_row = conn.execute(
        "SELECT MIN(rec_date) AS dmin, MAX(rec_date) AS dmax, MAX(day_num) AS max_day FROM daily_records WHERE batch_id=?",
        (batch_id,),
    ).fetchone()

    extra_cost = float(extra_cost_row["v"] or 0)
    extra_rev = float(extra_rev_row["v"] or 0)
    farm_val = float(farm_row["val"] or 0)
    farm_qty = int(farm_row["qty"] or 0)
    market_val = float(market_row["val"] or 0)
    market_qty = int(market_row["sold"] or 0)
    market_deaths = int(market_row["deaths"] or 0)
    total_dead = int(daily_row["dead"] or 0) + market_deaths
    total_sold = farm_qty + market_qty

    total_cost = fixed_cost + extra_cost
    total_rev = fixed_rev + extra_rev + farm_val + market_val
    net_result = total_rev - total_cost
    chicks = int(b.get("chicks") or 0)
    mort_rate = (total_dead / chicks * 100) if chicks > 0 else 0.0
    avg_price = (farm_val + market_val) / total_sold if total_sold > 0 else 0.0
    share_pct = float(b.get("share_pct") or 65)
    share_val = net_result * share_pct / 100.0

    date_in = str(b.get("date_in") or "")
    date_out = str(b.get("date_out") or "")
    if dates_row and dates_row["dmin"]:
        date_in = str(dates_row["dmin"])
    if dates_row and dates_row["dmax"]:
        date_out = str(dates_row["dmax"])
    days = int(b.get("days") or 0)
    if date_in and date_out:
        try:
            d1 = datetime.strptime(date_in, "%Y-%m-%d").date()
            d2 = datetime.strptime(date_out, "%Y-%m-%d").date()
            days = max(days, (d2 - d1).days + 1)
        except Exception:
            pass
    if dates_row and dates_row["max_day"]:
        days = max(days, int(dates_row["max_day"] or 0))

    conn.execute(
        """
        UPDATE batches
        SET total_cost=?, total_rev=?, total_sold=?, total_dead=?, mort_rate=?,
            net_result=?, avg_price=?, share_val=?, date_in=?, date_out=?, days=?
        WHERE id=?
        """,
        (
            round(total_cost, 4),
            round(total_rev, 4),
            total_sold,
            total_dead,
            round(mort_rate, 4),
            round(net_result, 4),
            round(avg_price, 4),
            round(share_val, 4),
            date_in,
            date_out,
            days,
            batch_id,
        ),
    )


def _update_fingerprint(conn, run_id: int, profile_id: int, run_file: dict[str, Any]) -> None:
    fingerprint = str(run_file.get("fingerprint_sha256") or "")
    if not fingerprint:
        return
    batch_num = str(run_file.get("detected_batch_num") or "")
    date_in = str(run_file.get("detected_date_in") or "")
    row = conn.execute(
        """
        SELECT id, import_count
        FROM import_file_fingerprints
        WHERE fingerprint_sha256=? AND batch_num=? AND date_in=?
        """,
        (fingerprint, batch_num, date_in),
    ).fetchone()
    if row:
        conn.execute(
            """
            UPDATE import_file_fingerprints
            SET last_run_id=?, import_count=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (run_id, int(row["import_count"] or 0) + 1, int(row["id"])),
        )
        return
    conn.execute(
        """
        INSERT INTO import_file_fingerprints(
            fingerprint_sha256, profile_id, batch_num, date_in,
            first_run_id, last_run_id, import_count
        )
        VALUES (?, ?, ?, ?, ?, ?, 1)
        """,
        (fingerprint, profile_id, batch_num, date_in, run_id, run_id),
    )


def commit_run(
    run_id: int,
    batch_mode: str,
    merge_mode: str,
    target_batch_id: int | None = None,
) -> dict[str, Any]:
    if batch_mode not in {"create", "update"}:
        raise ValueError("batch_mode يجب أن يكون create أو update.")
    if merge_mode not in {"replace", "merge"}:
        raise ValueError("merge_mode يجب أن يكون replace أو merge.")

    with get_conn(DB_PATH) as conn:
        run_row = conn.execute("SELECT * FROM import_runs WHERE id=?", (run_id,)).fetchone()
        if not run_row:
            raise ValueError("run_id غير موجود.")
        profile_id = int(run_row["profile_id"] or _default_profile_id(conn))
        run_files_rows = conn.execute("SELECT * FROM import_run_files WHERE run_id=? ORDER BY id", (run_id,)).fetchall()
        run_files = [dict(r) for r in run_files_rows]
        conn.execute(
            """
            UPDATE import_runs
            SET status='committing', batch_mode=?, merge_mode=?, target_batch_id=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (batch_mode, merge_mode, target_batch_id, run_id),
        )
        conn.commit()

    created_types: list[dict[str, Any]] = []
    touched_batches: list[int] = []
    warnings: list[str] = []
    committed_files = 0
    failed_files = 0
    skipped_duplicates = 0

    for run_file in run_files:
        run_file_id = int(run_file["id"])
        try:
            with get_conn(DB_PATH) as conn:
                line_groups = _load_run_file_lines(conn, run_file_id)
                file_meta = {
                    "warehouse_name": str(run_file.get("detected_warehouse") or ""),
                    "batch_num": str(run_file.get("detected_batch_num") or ""),
                    "date_in": str(run_file.get("detected_date_in") or ""),
                    "date_out": "",
                    "chicks": 0,
                }
                daily = line_groups["daily"]
                if daily:
                    rec_dates = [str(x.get("rec_date") or "") for x in daily if x.get("rec_date")]
                    if rec_dates:
                        file_meta["date_in"] = min(rec_dates)
                        file_meta["date_out"] = max(rec_dates)
                    payload0 = daily[0].get("payload") or {}
                    file_meta["chicks"] = _safe_int(payload0.get("birds"))

                if batch_mode == "create":
                    batch_id = _create_batch(conn, run_file, file_meta)
                else:
                    batch_id = _find_or_assign_update_batch(conn, run_file, target_batch_id)

                if merge_mode == "replace" and batch_mode == "update":
                    conn.execute("DELETE FROM daily_records WHERE batch_id=?", (batch_id,))
                    conn.execute("DELETE FROM farm_sales WHERE batch_id=?", (batch_id,))
                    conn.execute("DELETE FROM market_sales WHERE batch_id=?", (batch_id,))
                    conn.execute("DELETE FROM batch_costs WHERE batch_id=?", (batch_id,))
                    conn.execute("DELETE FROM batch_revenues WHERE batch_id=?", (batch_id,))

                for line in daily:
                    payload = line.get("payload") or {}
                    rec_date = str(line.get("rec_date") or payload.get("rec_date") or "")
                    if not rec_date:
                        continue
                    conn.execute(
                        """
                        INSERT INTO daily_records(batch_id, rec_date, day_num, dead_count, feed_kg, water_ltr, notes)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(batch_id, rec_date)
                        DO UPDATE SET
                            day_num=excluded.day_num,
                            dead_count=excluded.dead_count,
                            feed_kg=excluded.feed_kg,
                            water_ltr=excluded.water_ltr,
                            notes=excluded.notes
                        """,
                        (
                            batch_id,
                            rec_date,
                            _safe_int(payload.get("day_num")),
                            _safe_int(payload.get("dead_count")),
                            _safe_float(payload.get("feed_kg")),
                            _safe_float(payload.get("water_ltr")),
                            str(payload.get("notes") or ""),
                        ),
                    )

                for line in line_groups["sales_farm"]:
                    payload = line.get("payload") or {}
                    sale = {
                        "sale_date": str(line.get("rec_date") or payload.get("sale_date") or ""),
                        "sale_type": str(payload.get("sale_type") or "آجل"),
                        "customer": str(payload.get("customer") or ""),
                        "qty": _safe_int(payload.get("qty")),
                        "price": _safe_float(payload.get("price")),
                        "total_val": _safe_float(payload.get("total_val")),
                    }
                    if merge_mode == "merge" and _farm_sale_exists(conn, batch_id, sale):
                        skipped_duplicates += 1
                        continue
                    if sale["total_val"] <= 0 and sale["qty"] > 0:
                        sale["total_val"] = round(sale["qty"] * sale["price"], 4)
                    conn.execute(
                        """
                        INSERT INTO farm_sales(batch_id, sale_date, sale_type, customer, qty, price, total_val)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            batch_id,
                            sale["sale_date"],
                            sale["sale_type"],
                            sale["customer"],
                            sale["qty"],
                            sale["price"],
                            sale["total_val"],
                        ),
                    )

                for line in line_groups["sales_market"]:
                    payload = line.get("payload") or {}
                    sale = {
                        "sale_date": str(line.get("rec_date") or payload.get("sale_date") or ""),
                        "office": str(payload.get("office") or ""),
                        "qty_sent": _safe_int(payload.get("qty_sent")),
                        "deaths": _safe_int(payload.get("deaths")),
                        "qty_sold": _safe_int(payload.get("qty_sold")),
                        "net_val": _safe_float(payload.get("net_val")),
                        "inv_num": str(payload.get("inv_num") or ""),
                    }
                    if sale["qty_sold"] <= 0 and sale["qty_sent"] > 0:
                        sale["qty_sold"] = max(0, sale["qty_sent"] - sale["deaths"])
                    if merge_mode == "merge" and _market_sale_exists(conn, batch_id, sale):
                        skipped_duplicates += 1
                        continue
                    conn.execute(
                        """
                        INSERT INTO market_sales(batch_id, sale_date, office, qty_sent, deaths, qty_sold, net_val, inv_num)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            batch_id,
                            sale["sale_date"],
                            sale["office"],
                            sale["qty_sent"],
                            sale["deaths"],
                            sale["qty_sold"],
                            sale["net_val"],
                            sale["inv_num"],
                        ),
                    )

                candidate_lines = [
                    line
                    for line in line_groups["candidate"]
                    if int(line.get("apply_flag") or 0) == 1
                    and str(line.get("mapping_status") or "") in {"mapped_existing", "mapped_new"}
                    and str(line.get("target_kind") or "") in {"cost", "revenue"}
                ]
                for line in candidate_lines:
                    resolved_code, _ = _apply_candidate_line(conn, batch_id, line, merge_mode, created_types)
                    if str(line.get("mapping_status") or "") == "mapped_new":
                        conn.execute(
                            """
                            UPDATE import_run_lines
                            SET target_code=?, mapping_status='mapped_existing', is_auto_created=0
                            WHERE id=?
                            """,
                            (resolved_code, int(line["id"])),
                        )
                        _upsert_profile_mapping(
                            conn,
                            profile_id,
                            str(line.get("source_sheet") or ""),
                            str(line.get("source_label") or ""),
                            str(line.get("target_kind") or ""),
                            resolved_code,
                            str(line.get("target_name") or ""),
                            str(line.get("category") or ""),
                            str(line.get("unit") or ""),
                            int(line.get("has_qty") or 0),
                            0,
                        )

                _recalculate_batch(conn, batch_id)
                _update_fingerprint(conn, run_id, profile_id, run_file)
                conn.execute(
                    """
                    UPDATE import_run_files
                    SET target_batch_id=?, status='committed', reason=''
                    WHERE id=?
                    """,
                    (batch_id, run_file_id),
                )
                conn.commit()
                touched_batches.append(batch_id)
                committed_files += 1
        except Exception as exc:
            failed_files += 1
            warnings.append(f"{run_file.get('file_name')}: {exc}")
            with get_conn(DB_PATH) as conn:
                conn.execute(
                    "UPDATE import_run_files SET status='failed', reason=? WHERE id=?",
                    (str(exc), run_file_id),
                )
                conn.commit()

    final_status = "committed"
    if failed_files > 0 and committed_files > 0:
        final_status = "partial_failed"
    elif failed_files > 0 and committed_files == 0:
        final_status = "failed"

    with get_conn(DB_PATH) as conn:
        conn.execute(
            """
            UPDATE import_runs
            SET status=?, committed_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP,
                batch_mode=?, merge_mode=?, target_batch_id=?
            WHERE id=?
            """,
            (final_status, batch_mode, merge_mode, target_batch_id, run_id),
        )
        conn.commit()

    return {
        "run_id": run_id,
        "status": final_status,
        "total_files": len(run_files),
        "committed_files": committed_files,
        "failed_files": failed_files,
        "touched_batches": sorted(set(touched_batches)),
        "created_types": created_types,
        "skipped_duplicates": skipped_duplicates,
        "warnings": warnings,
    }
