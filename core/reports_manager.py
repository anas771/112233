import sqlite3
import os
from datetime import datetime
from fpdf import FPDF
import arabic_reshaper
from bidi.algorithm import get_display
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ReportsManager:
    def __init__(self, db_manager, font_path=None, logo_path=None):
        self.db = db_manager
        self.font_path = font_path # Path to Amiri-Regular.ttf
        self.logo_path = logo_path

    def _prepare_arabic(self, text):
        if not text: return ""
        reshaped_text = arabic_reshaper.reshape(str(text))
        return get_display(reshaped_text)

    def get_customer_statement(self, customer_name, start_date=None, end_date=None):
        query = """
            SELECT b.date_out, w.name as wh_name, b.id as batch_num, 
                   fs.qty, fs.price, fs.total_val, fs.sale_date
            FROM farm_sales fs
            JOIN batches b ON fs.batch_id = b.id
            JOIN warehouses w ON b.warehouse_id = w.id
            WHERE fs.customer = ?
        """
        params = [customer_name]
        if start_date:
            query += " AND (fs.sale_date >= ? OR (fs.sale_date = '' AND b.date_out >= ?))"
            params.extend([start_date, start_date])
        if end_date:
            query += " AND (fs.sale_date <= ? OR (fs.sale_date = '' AND b.date_out <= ?))"
            params.extend([end_date, end_date])
        
        query += " ORDER BY b.date_out DESC"
        return self.db.fetch_all(query, params)

    def get_market_statement(self, office_name, start_date=None, end_date=None):
        query = """
            SELECT b.date_out, w.name as wh_name, b.id as batch_num, 
                   ms.qty_sent, ms.deaths, ms.qty_sold, ms.net_val, ms.inv_num
            FROM market_sales ms
            JOIN batches b ON ms.batch_id = b.id
            JOIN warehouses w ON b.warehouse_id = w.id
            WHERE ms.office = ?
        """
        params = [office_name]
        if start_date:
            query += " AND b.date_out >= ?"
            params.append(start_date)
        if end_date:
            query += " AND b.date_out <= ?"
            params.append(end_date)
            
        query += " ORDER BY b.date_out DESC"
        return self.db.fetch_all(query, params)

    def export_customer_pdf(self, data, customer_name, output_path):
        pdf = FPDF()
        pdf.add_page()
        
        if self.font_path and os.path.exists(self.font_path):
            pdf.add_font('Arabic', '', self.font_path, uni=True)
            pdf.set_font('Arabic', '', 16)
        else:
            pdf.set_font('Arial', 'B', 16)

        # Header
        pdf.cell(0, 10, self._prepare_arabic(f"كشف حساب عميل: {customer_name}"), ln=True, align='C')
        pdf.set_font('Arabic', '', 10) if self.font_path else pdf.set_font('Arial', '', 10)
        pdf.cell(0, 10, self._prepare_arabic(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), ln=True, align='L')
        pdf.ln(5)

        # Table Header
        pdf.set_fill_color(200, 220, 255)
        cols = [("الإجمالي", 30), ("السعر", 25), ("الكمية", 25), ("رقم الدفعة", 25), ("العنبر", 40), ("التاريخ", 40)]
        for title, width in cols:
            pdf.cell(width, 10, self._prepare_arabic(title), 1, 0, 'C', True)
        pdf.ln()

        # Data
        pdf.set_fill_color(245, 245, 245)
        total_q, total_v = 0, 0
        fill = False
        for row in data:
            # row: date_out, wh_name, batch_num, qty, price, total_val, sale_date
            display_date = row['sale_date'] if row['sale_date'] else row['date_out']
            vals = [
                f"{row['total_val']:,.0f}",
                f"{row['price']:,.2f}",
                f"{row['qty']:,}",
                str(row['batch_num']),
                row['wh_name'],
                display_date
            ]
            total_q += (row['qty'] or 0)
            total_v += (row['total_val'] or 0)
            
            for val, (_, width) in zip(vals, cols):
                pdf.cell(width, 8, self._prepare_arabic(val), 1, 0, 'C', fill)
            pdf.ln()
            fill = not fill

        # Footer
        pdf.set_font('Arabic', '', 12) if self.font_path else pdf.set_font('Arial', 'B', 12)
        pdf.set_fill_color(200, 255, 200)
        pdf.cell(30, 10, self._prepare_arabic(f"{total_v:,.0f}"), 1, 0, 'C', True)
        pdf.cell(25, 10, "", 1, 0, 'C', True)
        pdf.cell(25, 10, self._prepare_arabic(f"{total_q:,}"), 1, 0, 'C', True)
        pdf.cell(105, 10, self._prepare_arabic("الإجمالي العام"), 1, 1, 'C', True)

        pdf.output(output_path)

    def export_full_batch_pdf(self, batch_id, output_path):
        """تصدير تقرير تصفية الدفعة الشامل A4"""
        b_data = self.db.fetch_one("SELECT * FROM v_batches WHERE id=?", (batch_id,))
        if not b_data: return False
        b = dict(b_data)

        pdf = FPDF()
        pdf.add_page()
        
        if self.font_path and os.path.exists(self.font_path):
            pdf.add_font('Arabic', '', self.font_path, uni=True)
            pdf.set_font('Arabic', '', 20)
        else:
            pdf.set_font('Arial', 'B', 20)

        # Header with Logo
        if self.logo_path and os.path.exists(self.logo_path):
            pdf.image(self.logo_path, x=10, y=8, w=35)
        
        pdf.cell(0, 12, self._prepare_arabic("تقرير تصفية الدفعة النهائي"), ln=True, align='C')
        pdf.set_font('Arabic', '', 14) if self.font_path else pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, self._prepare_arabic(f"عنبر: {b['warehouse_name']} — رقم الدفعة: {b['batch_num'] or b['id']}"), ln=True, align='C')
        pdf.ln(5)

        # Basic Info Section
        pdf.set_font('Arabic', '', 11) if self.font_path else pdf.set_font('Arial', '', 11)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(0, 8, self._prepare_arabic("البيانات الأساسية"), 1, 1, 'R', True)
        
        info_rows = [
            (f"تاريخ الدخول: {b['date_in']}", f"تاريخ الخروج: {b['date_out']}"),
            (f"عدد الأيام: {b['days'] or 0}", f"عدد الكتاكيت: {b['chicks']:,}"),
            (f"إجمالي النافق: {b['total_dead']:,} ({b['mort_rate'] or 0:.2f}%)", f"إجمالي المباع: {b['total_sold']:,}"),
            (f"متوسط الوزن: {b['avg_weight'] or 0:.3f} كجم", f"معدل التحويل FCR: {b['fcr'] or 0}"),
            (f"مؤشر الكفاءة EPEF: {((100-(b['mort_rate'] or 0))*(b['avg_weight'] or 0)*10)/((b['days'] or 1)*(float(b['fcr']) if b['fcr'] else 1.5)):.0f}" if b['days'] and b['fcr'] else "EPEF: 0")
        ]
        
        for row in info_rows:
            if len(row) == 2:
                pdf.cell(95, 8, self._prepare_arabic(row[1]), 1, 0, 'R')
                pdf.cell(95, 8, self._prepare_arabic(row[0]), 1, 1, 'R')
            else:
                pdf.cell(190, 8, self._prepare_arabic(row[0]), 1, 1, 'R')
        
        pdf.ln(5)

        # Financial Summary Section
        pdf.set_fill_color(220, 230, 241)
        pdf.cell(0, 8, self._prepare_arabic("الملخص المالي بالتفصيل"), 1, 1, 'R', True)
        
        pdf.cell(63, 8, self._prepare_arabic("القيمة (ريال)"), 1, 0, 'C', True)
        pdf.cell(63, 8, self._prepare_arabic("البيان"), 1, 0, 'C', True)
        pdf.cell(64, 8, self._prepare_arabic("النوع"), 1, 1, 'C', True)

        m_data = [
            ("إجمالي الإيرادات", b['total_rev'], (226, 239, 218)),
            ("إجمالي التكاليف", b['total_cost'], (252, 228, 214)),
            ("صافي النتيجة", b['net_result'], (255, 242, 204))
        ]
        
        for label, val, color in m_data:
            pdf.set_fill_color(*color)
            pdf.cell(63, 10, f"{val:,.2f}", 1, 0, 'C', True)
            pdf.cell(127, 10, self._prepare_arabic(label), 1, 1, 'R', True)
        
        # Partner Shares
        pdf.ln(5)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(0, 8, self._prepare_arabic("توزيع الأرباح"), 1, 1, 'R', True)
        
        pct = b['share_pct'] or 65
        share_v = b['share_val'] or 0
        partner_v = (b['net_result'] or 0) - share_v
        p_name = b['partner_name'] or "الشريك"
        
        pdf.cell(95, 10, f"{share_v:,.2f}", 1, 0, 'C')
        pdf.cell(95, 10, self._prepare_arabic(f"نصيب الشركة ({int(pct)}%)"), 1, 1, 'R')
        pdf.cell(95, 10, f"{partner_v:,.2f}", 1, 0, 'C')
        pdf.cell(95, 10, self._prepare_arabic(f"نصيب {p_name} ({int(100-pct)}%)"), 1, 1, 'R')

        # Notes
        if b['notes']:
            pdf.ln(5)
            pdf.set_fill_color(240, 240, 240)
            pdf.cell(0, 8, self._prepare_arabic("ملاحظات إضافية"), 1, 1, 'R', True)
            pdf.set_font('Arabic', '', 10)
            pdf.multi_cell(0, 8, self._prepare_arabic(b['notes']), 1, 'R')

        # Signatures
        pdf.set_y(-40)
        pdf.set_font('Arabic', '', 11)
        pdf.cell(63, 8, self._prepare_arabic("المحاسب"), 0, 0, 'C')
        pdf.cell(63, 8, self._prepare_arabic("المراجع"), 0, 0, 'C')
        pdf.cell(64, 8, self._prepare_arabic("المدير العام"), 0, 1, 'C')
        pdf.ln(10)
        pdf.cell(63, 8, "...................", 0, 0, 'C')
        pdf.cell(63, 8, "...................", 0, 0, 'C')
        pdf.cell(64, 8, "...................", 0, 1, 'C')

        pdf.output(output_path)
        return True

    def export_daily_records_pdf(self, batch_id, output_path):
        """تصدير سجلات العنبر اليومية"""
        rows = self.db.fetch_all("SELECT * FROM daily_records WHERE batch_id=? ORDER BY rec_date", (batch_id,))
        b_data = self.db.fetch_one("SELECT warehouse_name, batch_num FROM v_batches WHERE id=?", (batch_id,))
        
        pdf = FPDF()
        pdf.add_page()
        if self.font_path:
            pdf.add_font('Arabic', '', self.font_path, uni=True)
            pdf.set_font('Arabic', '', 16)
        
        wh_name = b_data['warehouse_name'] if b_data else "غير معروف"
        pdf.cell(0, 10, self._prepare_arabic(f"السجل اليومي للفترة — {wh_name}"), ln=True, align='C')
        pdf.ln(5)
        
        pdf.set_font('Arabic', '', 10)
        # Table Header
        pdf.set_fill_color(200, 220, 255)
        headers = [("ملاحظات", 70), ("علف (كجم)", 30), ("النافق", 20), ("يوم", 20), ("التاريخ", 50)]
        for h, w in headers:
            pdf.cell(w, 10, self._prepare_arabic(h), 1, 0, 'C', True)
        pdf.ln()

        fill = False
        for r in rows:
            pdf.cell(70, 8, self._prepare_arabic(r['notes'] or ""), 1, 0, 'R', fill)
            pdf.cell(30, 8, f"{r['feed_kg'] or 0:,.1f}", 1, 0, 'C', fill)
            pdf.cell(20, 8, f"{r['dead_count'] or 0}", 1, 0, 'C', fill)
            pdf.cell(20, 8, f"{r['day_num'] or ''}", 1, 0, 'C', fill)
            pdf.cell(50, 8, r['rec_date'], 1, 1, 'C', fill)
            fill = not fill
        
        pdf.output(output_path)
        return True

    def get_customer_balances(self):
        """جلب مديونيات كافة العملاء"""
        query = """
            SELECT customer, SUM(total_val) as total_bought
            FROM farm_sales
            GROUP BY customer
            ORDER BY total_bought DESC
        """
        return self.db.fetch_all(query)

    def export_customer_excel(self, data, customer_name, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "كشف حساب"
        ws.sheet_view.rightToLeft = True

        headers = ["التاريخ", "العنبر", "رقم الدفعة", "الكمية", "السعر", "الإجمالي"]
        ws.append(headers)
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        total_q, total_v = 0, 0
        for row in data:
            display_date = row['sale_date'] if row['sale_date'] else row['date_out']
            ws.append([
                display_date,
                row['wh_name'],
                row['batch_num'],
                row['qty'],
                row['price'],
                row['total_val']
            ])
            total_q += (row['qty'] or 0)
            total_v += (row['total_val'] or 0)

        # Footer
        last_row = len(data) + 2
        ws.cell(row=last_row, column=1, value="الإجمالي")
        ws.cell(row=last_row, column=4, value=total_q)
        ws.cell(row=last_row, column=6, value=total_v)
        
        # Styling totals
        for col in range(1, 7):
            cell = ws.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        wb.save(output_path)

    def export_market_pdf(self, data, office_name, output_path):
        pdf = FPDF()
        pdf.add_page()
        
        if self.font_path and os.path.exists(self.font_path):
            pdf.add_font('Arabic', '', self.font_path, uni=True)
            pdf.set_font('Arabic', '', 16)
        else:
            pdf.set_font('Arial', 'B', 16)

        pdf.cell(0, 10, self._prepare_arabic(f"كشف حساب مكتب: {office_name}"), ln=True, align='C')
        pdf.ln(5)

        pdf.set_font('Arabic', '', 9) if self.font_path else pdf.set_font('Arial', '', 9)
        pdf.set_fill_color(200, 220, 255)
        # cols: التاريخ، العنبر، الدفعة، المرسل، وفيات، المباع، الصافي، الفاتورة
        cols = [("الصافي", 25), ("المباع", 20), ("وفيات", 15), ("المرسل", 20), ("الدفعة", 15), ("العنبر", 35), ("التاريخ", 30), ("الفاتورة", 30)]
        for title, width in cols:
            pdf.cell(width, 10, self._prepare_arabic(title), 1, 0, 'C', True)
        pdf.ln()

        total_s, total_v = 0, 0
        fill = False
        for row in data:
            vals = [
                f"{row['net_val']:,.0f}",
                str(row['qty_sold']),
                str(row['deaths']),
                str(row['qty_sent']),
                str(row['batch_num']),
                row['wh_name'],
                row['date_out'],
                str(row['inv_num'] or '')
            ]
            total_s += (row['qty_sold'] or 0)
            total_v += (row['net_val'] or 0)
            
            for val, (_, width) in zip(vals, cols):
                pdf.cell(width, 8, self._prepare_arabic(val), 1, 0, 'C', fill)
            pdf.ln()
            fill = not fill

        pdf.set_fill_color(200, 255, 200)
        pdf.cell(25, 10, self._prepare_arabic(f"{total_v:,.0f}"), 1, 0, 'C', True)
        pdf.cell(20, 10, self._prepare_arabic(f"{total_s:,}"), 1, 0, 'C', True)
        pdf.cell(145, 10, self._prepare_arabic("الإجمالي العام"), 1, 1, 'C', True)
        pdf.output(output_path)

    def get_batch_comparison(self, batch_ids):
        if not batch_ids: return []
        placeholders = ','.join(['?'] * len(batch_ids))
        query = f"""
            SELECT b.id, w.name as wh_name, b.date_out, b.chicks, b.total_sold,
                   b.mort_rate, b.fcr, b.avg_weight, b.total_cost, b.total_rev, b.net_result
            FROM batches b
            JOIN warehouses w ON b.warehouse_id = w.id
            WHERE b.id IN ({placeholders})
            ORDER BY b.date_out DESC
        """
        return self.db.fetch_all(query, batch_ids)

    def get_summary_stats(self):
        # إحصائيات عامة للوحة القيادة
        stats = {}
        res = self.db.fetch_one("SELECT COUNT(*) as cnt FROM batches")
        stats['total_batches'] = res['cnt'] if res else 0
        
        res = self.db.fetch_one("SELECT SUM(total_rev) as rev, SUM(total_cost) as cost, SUM(net_result) as net FROM batches")
        if res:
            stats['total_rev'] = res['rev'] or 0
            stats['total_cost'] = res['cost'] or 0
            stats['total_net'] = res['net'] or 0
        else:
            stats['total_rev'] = stats['total_cost'] = stats['total_net'] = 0
            
        return stats
    def export_nano_batch_pdf(self, batch_id, output_path):
        """تصدير تقرير نانو الاحترافي (Nanobanana Style)"""
        b_data = self.db.fetch_one("SELECT * FROM v_batches WHERE id=?", (batch_id,))
        if not b_data: return False
        b = dict(b_data)

        # 1. إعداد الرسم البياني للنافق (أو أي مؤشر آخر)
        import matplotlib.pyplot as plt
        import io
        
        # جلب بيانات السجل اليومي للرسم البياني
        daily_rows = self.db.fetch_all("SELECT day_num, dead_count, feed_kg FROM daily_records WHERE batch_id=? ORDER BY day_num", (batch_id,))
        
        plt.figure(figsize=(10, 4))
        if daily_rows:
            days = [r['day_num'] for r in daily_rows]
            dead = [r['dead_count'] for r in daily_rows]
            plt.plot(days, dead, color='#A80000', linewidth=2, marker='o', markersize=4, label='النافق اليومي')
            plt.fill_between(days, dead, color='#FDE7E9', alpha=0.5)
        
        plt.title(self._prepare_arabic("منحنى النافق اليومي"), fontsize=12)
        plt.xlabel(self._prepare_arabic("اليوم"), fontsize=10)
        plt.ylabel(self._prepare_arabic("العدد"), fontsize=10)
        plt.grid(True, linestyle='--', alpha=0.7)
        
        chart_buf = io.BytesIO()
        plt.savefig(chart_buf, format='png', dpi=100, bbox_inches='tight')
        chart_buf.seek(0)
        plt.close()

        # 2. إنشاء ملف PDF
        pdf = FPDF()
        pdf.add_page()
        
        if self.font_path and os.path.exists(self.font_path):
            pdf.add_font('Arabic', '', self.font_path, uni=True)
            pdf.set_font('Arabic', '', 14)
        
        # الهيدر الاحترافي (Nanobanana)
        nano_header = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets", "nano_header.png")
        if os.path.exists(nano_header):
            pdf.image(nano_header, x=0, y=0, w=210)
            pdf.ln(50) # مساحة تحت الهيدر
        else:
            pdf.ln(10)
            
        # العنوان
        pdf.set_font('Arabic', '', 24)
        pdf.set_text_color(0, 90, 158) # أزرق نانو
        pdf.cell(0, 15, self._prepare_arabic("تقرير الكفاءة الإنتاجية"), ln=True, align='C')
        pdf.ln(5)

        # بطاقات المعلومات الأساسية (Cards)
        pdf.set_font('Arabic', '', 11)
        pdf.set_text_color(50, 49, 48)
        
        # خلفية الكروت
        def draw_card(x, y, w, h, title, value, color_hex):
            pdf.set_fill_color(int(color_hex[1:3], 16), int(color_hex[3:5], 16), int(color_hex[5:7], 16))
            pdf.rect(x, y, w, h, 'F')
            pdf.set_xy(x, y + 2)
            pdf.set_font('Arabic', '', 10)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(w, 8, self._prepare_arabic(title), 0, 1, 'C')
            pdf.set_font('Arabic', '', 16)
            pdf.set_xy(x, y + 10)
            pdf.cell(w, 12, value, 0, 1, 'C')

        # رسم 4 كروت رئيسية
        draw_card(10, 85, 45, 25, "صافي الربح", f"{b['net_result_dynamic']:,.0f}", "#107C10")
        draw_card(60, 85, 45, 25, "معدل التحويل FCR", f"{b['fcr'] or 0:.2f}", "#005A9E")
        draw_card(110, 85, 45, 25, "نسبة النافق", f"{b['mort_rate'] or 0:.1f}%", "#A80000")
        draw_card(160, 85, 40, 25, "عدد الكتاكيت", f"{b['chicks']:,}", "#847545")

        pdf.ln(35)
        
        # جدول التفاصيل المالية بتصميم أنيق
        pdf.set_font('Arabic', '', 12)
        pdf.set_text_color(0, 0, 0)
        pdf.set_fill_color(243, 242, 241)
        pdf.cell(0, 10, self._prepare_arabic("ملخص البيانات التشغيلية"), 0, 1, 'R')
        
        pdf.set_font('Arabic', '', 10)
        details = [
            ("اسم العنبر", b['warehouse_name']),
            ("رقم الدفعة", b['batch_num'] or b['id']),
            ("تاريخ الدخول", b['date_in']),
            ("تاريخ الخروج", b['date_out']),
            ("مدة الدورة", f"{b['days'] or 0} يوم"),
            ("إجمالي التكاليف", f"{b['total_cost']:,.0f} ريال"),
            ("إجمالي الإيرادات", f"{b['total_rev']:,.0f} ريال"),
        ]
        
        for label, val in details:
            pdf.set_fill_color(249, 249, 249)
            pdf.cell(95, 8, str(val), 1, 0, 'C', True)
            pdf.cell(95, 8, self._prepare_arabic(label), 1, 1, 'R', True)

        pdf.ln(5)
        
        # إضافة الرسم البياني
        pdf.image(chart_buf, x=10, y=180, w=190)
        
        # تذييل الصفحة
        pdf.set_y(-30)
        pdf.set_font('Arabic', '', 8)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(0, 10, self._prepare_arabic("تم إنشاء هذا التقرير آلياً بواسطة منظومة دوجي برو - أسلوب نانو"), 0, 0, 'C')

        pdf.output(output_path)
        return True
