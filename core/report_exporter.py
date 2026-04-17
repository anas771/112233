import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class ReportExporter:
    def __init__(self, db_manager):
        self.db = db_manager

    def create_styled_header(self, ws):
        # ألوان Classic Plus
        header_fill = PatternFill(start_color="005A9E", end_color="005A9E", fill_type="solid")
        sub_fill = PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, name="Segoe UI")
        black_font = Font(color="323130", bold=True, name="Segoe UI")
        
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # الصف الأول: العناوين الرئيسية المدمجة
        ws.merge_cells('A1:D1')
        ws['A1'] = "بيانات الدورة"
        ws['A1'].fill = header_fill
        ws['A1'].font = white_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('E1:M1')
        ws['E1'] = "التكاليف المباشرة (كتاكيت + علف + ماء + نشارة)"
        ws['E1'].fill = header_fill
        ws['E1'].font = white_font
        ws['E1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('O1:AD1')
        ws['O1'] = "التكاليف التشغيلية (رواتب + كهرباء + غاز + أدوية ...)"
        ws['O1'].fill = header_fill
        ws['O1'].font = white_font
        ws['O1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('AF1:AO1')
        ws['AF1'] = "الإيرادات والمبيعات"
        ws['AF1'].fill = header_fill
        ws['AF1'].font = white_font
        ws['AF1'].alignment = Alignment(horizontal='center', vertical='center')

        # الصف الثاني والثالث: تفاصيل الأعمدة
        cols = [
            ("الاسم", "المزرعة"), ("الدخول", "التاريخ"), ("الخروج", "التاريخ"), ("الأيام", "العدد"),
            ("قيمة الكتاكيت", "إجمالي"), ("عدد الكتاكيت", "الرأس"), ("علف", "ريال"), ("علف", "طن"), ("نقل علف", "ريال"),
            ("", ""), ("نشارة", "قيمة"), ("نشارة", "كمية"), ("ماء", "قيمة"), ("إجمالي", "مباشرة"),
            ("مربيين", "رواتب"), ("مربيين", "قات"), ("مشرف", "إدارة"), ("إيجار", "عنبر"), ("إضاءة", "قيمة"),
            ("غاز", "قيمة"), ("ثابت", "-"), ("علاجات", "قيمة"), ("لقاحات", "قيمة"), ("مصاريف", "أخرى"),
            ("نقل", "أجور"), ("-", "-"), ("إدارة", "مصاريف"), ("أخرى 1", ""), ("أخرى 2", ""), ("أخرى 3", ""),
            ("إجمالي", "التكاليف"), ("المباع", "رأس"), ("المباع", "قيمة"), ("إيراد", "آخر"), ("إيراد", "آخر"),
            ("سوق 1", ""), ("سوق 2", ""), ("سوق 3", ""), ("سوق 4", ""), ("السوق", "إجمالي"),
            ("إجمالي", "الإيرادات"), ("صافي", "الربح"), ("نصيب", "الشريك"), ("نسبة", "الشريك"),
            ("إجمالي", "الكتاكيت"), ("إجمالي", "المباع"), ("متوسط", "الوزن"), ("معدل", "النفوق"),
            ("تحويل", "FCR"), ("كفاءة", "EPEF"), ("سعر", "التعادل"), ("تكلفة", "الطائر"),
            ("تاريخ", "الختام"), ("عمر", "الدورة")
        ]

        for i, (top, bottom) in enumerate(cols, 1):
            ws.cell(2, i, top).alignment = Alignment(horizontal='center')
            ws.cell(3, i, bottom).alignment = Alignment(horizontal='center')
            ws.cell(2, i).fill = sub_fill
            ws.cell(3, i).fill = sub_fill
            ws.cell(1, i).border = border
            ws.cell(2, i).border = border
            ws.cell(3, i).border = border

    def export_all(self, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "التقرير الشامل"
        ws.sheet_view.rightToLeft = True

        self.create_styled_header(ws)

        # جلب البيانات
        query = """
        SELECT b.*, w.name as warehouse_name 
        FROM batches b 
        JOIN warehouses w ON b.warehouse_id = w.id
        ORDER BY b.date_in DESC
        """
        batches = self.db.fetch_all(query)

        row_num = 4
        for b in batches:
            # تعبئة الأعمدة الـ 54
            data = [
                b['warehouse_name'], b['date_in'], b['date_out'], b['days'],
                b['chick_val'], b['chicks'], b['feed_val'], b['feed_qty'], b['feed_trans'],
                0, b['sawdust_val'], b['sawdust_qty'], b['water_val'], 
                (b['chick_val'] + b['feed_val'] + b['water_val'] + b['sawdust_val']), # إجمالي مباشرة
                b['breeders_pay'], b['qat_pay'], b['sup_co_pay'], b['rent_val'], b['light_val'],
                b['gas_val'], 0, b['drugs_val'], b['vaccine_pay'], b['wh_expenses'],
                b['delivery_val'], 0, b['admin_val'], b['mixing_val'], b['wash_val'], b['other_costs'],
                b['total_cost'], b['total_sold'], b['cust_val'], 0, 0,
                0, 0, 0, 0, b['mkt_val'],
                b['total_rev'], b['net_result'], b['share_val'], b['share_pct'],
                b['chicks'], b['total_sold'], b['avg_weight'], b['mort_rate'],
                b['fcr'], 350, b['avg_price'], (b['total_cost']/b['chicks'] if b['chicks'] else 0),
                b['date_out'], b['days']
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row_num, col_idx, value)
                cell.alignment = Alignment(horizontal='center')
                if col_idx == 42: # صافي الربح
                    if value and value > 0:
                        cell.font = Font(color="107C10", bold=True)
                    elif value and value < 0:
                        cell.font = Font(color="A80000", bold=True)
            
            row_num += 1

        # صف الإجماليات
        last_row = row_num
        ws.cell(last_row, 1, "الإجمالي العام").font = Font(bold=True)
        for col_idx in range(4, 55):
            if col_idx in [2,3, 53]: continue # تخطي التواريخ
            col_letter = get_column_letter(col_idx)
            ws.cell(last_row, col_idx, f"=SUM({col_letter}4:{col_letter}{last_row-1})").font = Font(bold=True)

        wb.save(output_path)
        return True, f"تم تصدير التقرير بنجاح إلى: {output_path}"
